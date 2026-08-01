# ====================================================
# Complete ITSM – All Modules + Dark Mode (theme column)
# ====================================================

from contextlib import asynccontextmanager
from datetime import datetime, timedelta, date
import enum
import os
import re
import uuid
import smtplib

# Sentry error monitoring — initialise before anything else
try:
    import sentry_sdk
    from sentry_sdk.integrations.fastapi import FastApiIntegration
    from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration
    SENTRY_DSN = os.getenv("SENTRY_DSN", "")
    if SENTRY_DSN:
        sentry_sdk.init(
            dsn=SENTRY_DSN,
            integrations=[FastApiIntegration(), SqlalchemyIntegration()],
            traces_sample_rate=0.1,   # 10% of requests traced
            profiles_sample_rate=0.1,
            environment=os.getenv("SENTRY_ENV", "production"),
            send_default_pii=False,   # don't send user PII to Sentry
        )
        print("✅ Sentry initialised")
    else:
        print("ℹ️ SENTRY_DSN not set — error monitoring disabled")
except ImportError:
    print("ℹ️ sentry-sdk not installed — skipping Sentry")
import json
import urllib.request
import urllib.error
import urllib.parse
import base64
import hashlib
import cloudinary
import cloudinary.uploader
import hmac as hmac_lib

# Cloudinary configuration
CLOUDINARY_CLOUD_NAME = os.getenv("CLOUDINARY_CLOUD_NAME", "")
CLOUDINARY_API_KEY    = os.getenv("CLOUDINARY_API_KEY", "")
CLOUDINARY_API_SECRET = os.getenv("CLOUDINARY_API_SECRET", "")
# Set CLOUDINARY_FOLDER_MODE=fixed if your account uses fixed folder mode
# Set CLOUDINARY_FOLDER_MODE=dynamic (default) for newer accounts with dynamic folders
CLOUDINARY_FOLDER_MODE = os.getenv("CLOUDINARY_FOLDER_MODE", "dynamic")

# Product prefix — root folder in Cloudinary for this product.
# ─────────────────────────────────────────────────────────────────────────────
# All files are stored under: {product_prefix}/tenants/{tenant_id}/...
#
# When deploying a second product (e.g. DodoHR) to the same Cloudinary account,
# set CLOUDINARY_PRODUCT_PREFIX=dodohr on that server so files stay separated:
#
#   dodesk/tenants/1/tickets/42/   ← DodoDesk files
#   dodohr/tenants/1/payslips/     ← DodoHR files (future)
#
# To export one tenant's files: search {product_prefix}/tenants/{tenant_id}/*
# in Cloudinary Media Library.
# ─────────────────────────────────────────────────────────────────────────────
CLOUDINARY_PRODUCT_PREFIX = os.getenv("CLOUDINARY_PRODUCT_PREFIX", "dodesk")

def _cloudinary_folder(tenant_id: int, entity_type: str, entity_id: int | str | None = None) -> str:
    """Return a fully-scoped Cloudinary folder path.

    Format: {product_prefix}/tenants/{tenant_id}/{entity_type}[/{entity_id}]

    Examples:
      dodesk/tenants/1/logos/
      dodesk/tenants/1/avatars/
      dodesk/tenants/1/tickets/42/
    """
    base = f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{tenant_id}/{entity_type}"
    if entity_id is not None:
        base = f"{base}/{entity_id}"
    return base

def _detect_resource_type(filename: str) -> str:
    """Return the Cloudinary resource_type for a given filename."""
    image_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp", ".ico"}
    ext = os.path.splitext(filename)[1].lower()
    return "image" if ext in image_exts else "raw"

def _configure_cloudinary():
    """Configure the Cloudinary SDK from env vars."""
    cloudinary.config(
        cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME", ""),
        api_key=os.getenv("CLOUDINARY_API_KEY", ""),
        api_secret=os.getenv("CLOUDINARY_API_SECRET", ""),
        secure=True,
    )

# Cache of folders already created this server lifetime — avoids redundant API calls
_cloudinary_folders_created: set = set()

def _ensure_cloudinary_folder(folder_path: str) -> None:
    """Create a Cloudinary folder via the Admin API if it doesn't exist yet.
    POST /folders/:folder with Basic auth (api_key:api_secret).
    Cloudinary requires folders to be explicitly created on newer account types
    before files can be placed in them — uploading with public_id alone isn't enough.
    Folders are created recursively: creating 'a/b/c' also creates 'a' and 'a/b'.
    """
    global _cloudinary_folders_created
    if folder_path in _cloudinary_folders_created:
        return  # already created this server lifetime

    cloud_name  = os.getenv("CLOUDINARY_CLOUD_NAME", "")
    api_key     = os.getenv("CLOUDINARY_API_KEY", "")
    api_secret  = os.getenv("CLOUDINARY_API_SECRET", "")
    if not cloud_name:
        return

    # Build all parent paths so nested folders are created top-down
    # e.g. "dodesk/tenants/1/avatars" → ["dodesk", "dodesk/tenants", ...]
    parts = folder_path.split("/")
    paths_to_create = ["/".join(parts[:i+1]) for i in range(len(parts))]

    creds = base64.b64encode(f"{api_key}:{api_secret}".encode()).decode()
    headers = {
        "Authorization": f"Basic {creds}",
        "Content-Type": "application/json",
    }

    for path in paths_to_create:
        if path in _cloudinary_folders_created:
            continue
        url = f"https://api.cloudinary.com/v1_1/{cloud_name}/folders/{urllib.parse.quote(path, safe='/')}"
        try:
            req = urllib.request.Request(url, data=b"{}", headers=headers, method="POST")
            with urllib.request.urlopen(req) as resp:
                resp.read()
            _cloudinary_folders_created.add(path)
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if e.code == 409 or "already exists" in body.lower():
                # Folder already exists — that's fine
                _cloudinary_folders_created.add(path)
            else:
                print(f"⚠️ Cloudinary folder create failed for '{path}': {e.code} {body[:200]}")
        except Exception as e:
            print(f"⚠️ Cloudinary folder create error for '{path}': {e}")

def upload_to_cloudinary(file_bytes: bytes, public_id: str, folder: str = "dodesk",
                         filename: str = "file") -> str:
    """Upload a file to Cloudinary as authenticated (private) and return the stored public_id.

    Handles both Cloudinary account modes:
    - Dynamic folder mode: use asset_folder for path, public_id = filename only
    - Fixed folder mode: use public_id = full/path/filename

    Returns the public_id exactly as Cloudinary stored it (from the API response).
    """
    cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "")
    if not cloud_name:
        raise HTTPException(status_code=500, detail="Cloudinary is not configured.")
    _configure_cloudinary()

    if folder:
        _ensure_cloudinary_folder(folder)

    resource_type = _detect_resource_type(filename or public_id)

    # full_public_id = folder/filename — used in fixed folder mode
    full_public_id = f"{folder}/{public_id}" if folder else public_id
    # filename_only — used in dynamic folder mode (no slashes in public_id)
    filename_only = public_id

    import io

    # Try dynamic folder mode first: asset_folder = full path, public_id = filename only
    try:
        result = cloudinary.uploader.upload(
            io.BytesIO(file_bytes),
            public_id=filename_only,     # just the filename — NO path separators
            asset_folder=folder,         # full folder path goes here
            resource_type=resource_type,
            type="authenticated",
            overwrite=True,
            use_filename=False,
            unique_filename=False,
            invalidate=True,
        )
        pid = result.get("public_id") or full_public_id
        # Cloudinary in dynamic mode stores the full path in public_id response
        print(f"✅ Cloudinary upload (dynamic folder): {pid}")
        return pid
    except Exception as e:
        err = str(e)
        print(f"⚠️ Cloudinary dynamic upload failed ({err[:80]}), trying fixed folder mode...")
        # Fallback: fixed folder mode — public_id = full path including folder
        try:
            file_bytes_io = io.BytesIO(file_bytes)
            result = cloudinary.uploader.upload(
                file_bytes_io,
                public_id=full_public_id,   # full path as public_id
                resource_type=resource_type,
                type="authenticated",
                overwrite=True,
                use_filename=False,
                unique_filename=False,
            )
            pid = result.get("public_id") or full_public_id
            print(f"✅ Cloudinary upload (fixed folder): {pid}")
            return pid
        except Exception as e2:
            raise HTTPException(status_code=500, detail=f"Cloudinary upload failed: {str(e2)}")

def get_signed_url(public_id: str, resource_type: str = "image",
                   expires_in_seconds: int = 3600) -> str:
    """Generate a time-limited signed URL for a private Cloudinary asset."""
    if not public_id:
        return ""
    _configure_cloudinary()
    import time as _time
    import cloudinary.utils as _cu
    expires_at = int(_time.time()) + expires_in_seconds
    ext = os.path.splitext(public_id)[1].lower().lstrip(".")
    # Auto-detect resource type if not specified
    if resource_type == "auto" or resource_type == "image":
        resource_type = "image" if ext in {"png","jpg","jpeg","gif","webp","svg"} else "raw"
    url, _ = _cu.cloudinary_url(
        public_id,
        resource_type=resource_type,
        type="authenticated",
        sign_url=True,
        expires_at=expires_at,
        secure=True,
        **({"format": ext} if ext and ext in {"png","jpg","jpeg","gif","webp"} else {}),
    )
    return url


import time as _time_module
from email.mime.text import MIMEText
from apscheduler.schedulers.background import BackgroundScheduler

from fastapi import FastAPI, Depends, HTTPException, status, Query, Header, UploadFile, File, Request, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from jose import JWTError, jwt
from pydantic import BaseModel
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Boolean, Enum as SAEnum, ForeignKey, Text, Date, Float, UniqueConstraint
import sqlalchemy as sa
from sqlalchemy.orm import declarative_base, sessionmaker, Session, relationship, backref
from sqlalchemy.sql import func as sa_func

# Rate limiting
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# Load .env file
from dotenv import load_dotenv
load_dotenv()

# =============================================================================
# DATABASE SETUP
# =============================================================================

SQLALCHEMY_DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./test.db")

# Neon PgBouncer: use POOLED_DATABASE_URL if set (port 6543), otherwise use direct connection
# To enable: Neon dashboard → Connection Details → Pooled connection → copy URL → set as POOLED_DATABASE_URL on Render
POOLED_DATABASE_URL = os.getenv("POOLED_DATABASE_URL", SQLALCHEMY_DATABASE_URL)

# Fix URL schemes
for _url_attr in ["SQLALCHEMY_DATABASE_URL", "POOLED_DATABASE_URL"]:
    _val = locals()[_url_attr]
    if _val.startswith("postgres://"):
        locals()[_url_attr] = _val.replace("postgres://", "postgresql://", 1)

if SQLALCHEMY_DATABASE_URL.startswith("postgres://"):
    SQLALCHEMY_DATABASE_URL = SQLALCHEMY_DATABASE_URL.replace("postgres://", "postgresql://", 1)
if POOLED_DATABASE_URL.startswith("postgres://"):
    POOLED_DATABASE_URL = POOLED_DATABASE_URL.replace("postgres://", "postgresql://", 1)

# SQLite needs check_same_thread, PostgreSQL does not
if SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
    engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
else:
    # Detect if using Neon's PgBouncer pooled connection (hostname contains -pooler)
    _is_pooled = "-pooler." in POOLED_DATABASE_URL

    engine = create_engine(
        POOLED_DATABASE_URL,
        # PgBouncer transaction mode: disable pre-ping (it uses prepared statements)
        pool_pre_ping=not _is_pooled,
        pool_recycle=300,
        # Neon PgBouncer: keep pool small — pooler manages the actual Postgres connections
        pool_size=3 if _is_pooled else 5,
        max_overflow=7 if _is_pooled else 10,
        pool_timeout=30,
        connect_args={
            "connect_timeout": 10,
            # sslmode is already in the Neon connection URL — don't pass here
            # Neon PgBouncer rejects extra startup parameters
            "keepalives": 1,
            "keepalives_idle": 30,
            "keepalives_interval": 10,
            "keepalives_count": 5,
        },
    )
    print(f"✅ DB engine: {'PgBouncer pooled' if _is_pooled else 'direct'} connection")
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# =============================================================================
# SUBSCRIPTION PLANS
# =============================================================================

PLAN_LIMITS = {
    # ── Free ─────────────────────────────────────────────────────────────────
    "free": {
        "label": "Free", "max_agents": 1, "max_assets": 0,
        "ai_chatbot_conversations": 0, "storage_gb_per_agent": 1,
        "trial_days": 14, "trial_max_agents": 3,
        "ticketing": True, "knowledge_base": True, "service_catalog": False,
        "asset_tracking": False, "branding": False, "basic_sla": True,
        "multiple_sla": False, "workflow_automation": False,
        "change_management": False, "problem_management": False, "release_management": False,
        "ai_chatbot": False, "custom_analytics": False, "mfa": False,
        "sso": False, "approval_workflows": False, "audit_log": False, "sandbox": False,
        "price_monthly": 0, "price_annual": 0, "price_per_extra_seat": 0,
        "sla": True, "max_users": 1, "max_tenants": 1, "grace_users": 0,
    },
    # ── Essentials ───────────────────────────────────────────────────────────
    # $15/agent/month · $153/agent/year (15% off)
    "essentials": {
        "label": "Essentials", "max_agents": None, "max_assets": 250,
        "ai_chatbot_conversations": 0, "storage_gb_per_agent": 2,
        "trial_days": 14, "trial_max_agents": 3,
        "ticketing": True, "knowledge_base": True, "service_catalog": True,
        "asset_tracking": True, "branding": True, "basic_sla": True,
        "multiple_sla": False, "workflow_automation": False,
        "change_management": False, "problem_management": False, "release_management": False,
        "ai_chatbot": False, "custom_analytics": False, "mfa": False,
        "sso": False, "approval_workflows": False, "audit_log": False, "sandbox": False,
        "price_monthly": 15, "price_annual": 153, "price_per_extra_seat": 0,
        "sla": True, "max_users": None, "max_tenants": 1, "grace_users": 0,
    },
    # ── Business ─────────────────────────────────────────────────────────────
    # $35/agent/month · $357/agent/year (15% off)
    "business": {
        "label": "Business", "max_agents": None, "max_assets": 1000,
        "ai_chatbot_conversations": 0, "storage_gb_per_agent": 10,
        "trial_days": 14, "trial_max_agents": 3,
        "ticketing": True, "knowledge_base": True, "service_catalog": True,
        "asset_tracking": True, "branding": True, "basic_sla": True,
        "multiple_sla": True, "workflow_automation": True,
        "change_management": False, "problem_management": False, "release_management": False,
        "ai_chatbot": False, "custom_analytics": True, "mfa": True,
        "sso": False, "approval_workflows": True, "audit_log": True, "sandbox": False,
        "price_monthly": 35, "price_annual": 357, "price_per_extra_seat": 0,
        "sla": True, "max_users": None, "max_tenants": 1, "grace_users": 0,
    },
    # ── Pro (Advanced) ────────────────────────────────────────────────────────
    # $65/agent/month · $663/agent/year (15% off)
    "pro": {
        "label": "Pro", "max_agents": None, "max_assets": 5000,
        "ai_chatbot_conversations": 500, "storage_gb_per_agent": 25,
        "trial_days": 14, "trial_max_agents": 3,
        "ticketing": True, "knowledge_base": True, "service_catalog": True,
        "asset_tracking": True, "branding": True, "basic_sla": True,
        "multiple_sla": True, "workflow_automation": True,
        "change_management": True, "problem_management": True, "release_management": True,
        "ai_chatbot": True, "custom_analytics": True, "mfa": True,
        "sso": False, "approval_workflows": True, "audit_log": True, "sandbox": False,
        "price_monthly": 65, "price_annual": 663, "price_per_extra_seat": 0,
        "sla": True, "max_users": None, "max_tenants": 1, "grace_users": 0,
    },
    # ── Enterprise ────────────────────────────────────────────────────────────
    "enterprise": {
        "label": "Enterprise", "max_agents": None, "max_assets": None,
        "ai_chatbot_conversations": None, "storage_gb_per_agent": None,
        "trial_days": None, "trial_max_agents": None,
        "ticketing": True, "knowledge_base": True, "service_catalog": True,
        "asset_tracking": True, "branding": True, "basic_sla": True,
        "multiple_sla": True, "workflow_automation": True,
        "change_management": True, "problem_management": True, "release_management": True,
        "ai_chatbot": True, "custom_analytics": True, "mfa": True,
        "sso": True, "approval_workflows": True, "audit_log": True, "sandbox": True,
        "price_monthly": None, "price_annual": None, "price_per_extra_seat": 0,
        "sla": True, "max_users": None, "max_tenants": None, "grace_users": 0,
    },
}


def _sql_safe_search(term: str) -> str:
    """Escape LIKE special characters in user search input.
    % and _ are wildcards in SQL LIKE/ILIKE — without escaping, a search for
    "50% off" would match any string, and "_test" would match any single char + "test".
    SQLAlchemy's ilike() parameterizes values (no SQL injection), but these
    characters still act as wildcards unless escaped.
    """
    if not term:
        return ""
    return term.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def get_plan_limits(plan: str) -> dict:
    return PLAN_LIMITS.get(plan, PLAN_LIMITS["free"])

def plan_requires(feature: str, tenant, detail: str | None = None):
    """Raise 403 if the tenant's plan doesn't include the given feature.
    Usage: plan_requires('change_management', tenant)
    """
    limits = get_plan_limits(tenant.plan if tenant else "free")
    if not limits.get(feature, False):
        plan_label = limits.get("label", tenant.plan if tenant else "free")
        msg = detail or f"This feature is not available on the {plan_label} plan. Please upgrade to access it."
        raise HTTPException(status_code=403, detail=msg)



def check_tenant_limit(db: Session, admin: "User"):
    """Raise HTTPException if the admin's tenant has reached the plan's max_tenants.
    Only applies to regular admins — super_admin can always create tenants."""
    if str(admin.role) in ("super_admin", "platform_admin"):
        return  # super_admin is never limited

    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        return
    limits = get_plan_limits(tenant.plan)
    max_tenants = limits.get("max_tenants")
    if max_tenants is None:
        return  # unlimited (Enterprise)

    # Count tenants this admin has created (or just count all tenants — since each company
    # should have exactly 1, this effectively prevents any additional tenant creation)
    owned = db.query(Tenant).filter(Tenant.id == admin.tenant_id).count()
    if owned >= max_tenants:
        plan_label = limits["label"]
        raise HTTPException(
            status_code=403,
            detail=f"Your {plan_label} plan is limited to {max_tenants} tenant{'s' if max_tenants != 1 else ''}. "
                   f"Each company should have its own DodoDesk subscription. "
                   f"Contact us about Enterprise if you manage multiple client organisations."
        )



def get_trial_status(tenant: "Tenant") -> dict:
    """Compute trial status for a tenant.
    A tenant is on trial if:
    - billing_status == 'trialing' AND
    - their plan has trial_days defined AND
    - they are within the trial window since created_at
    Once they subscribe (billing_status = 'active'), trial is over.
    """
    # If already paid/active subscription, not on trial
    billing_status = getattr(tenant, "billing_status", None)
    if billing_status and billing_status not in ("trialing", None):
        return {"on_trial": False, "trial_days_remaining": None, "trial_expired": False, "trial_plan": None}

    limits = get_plan_limits(tenant.plan)
    trial_days = limits.get("trial_days")
    if not trial_days or not tenant.created_at:
        return {"on_trial": False, "trial_days_remaining": None, "trial_expired": False, "trial_plan": None}

    elapsed = datetime.utcnow() - tenant.created_at
    remaining = trial_days - elapsed.days
    return {
        "on_trial": True,
        "trial_days_remaining": max(remaining, 0),
        "trial_expired": remaining <= 0,
        "trial_plan": tenant.plan,       # which plan they're trialling
        "trial_plan_label": limits.get("label", tenant.plan),
    }


def check_user_limit(db: Session, tenant_id: int, additional: int = 1, role: "UserRole | str | None" = None):
    """Raise HTTPException if adding `additional` staff (agent/admin/super_admin) would exceed the plan limit.
    Employees (end-users raising tickets) don't count toward the limit.

    Billing model:
    - Trial: max 3 agents total (trial_max_agents)
    - Free: max 1 agent
    - Paid plans: unlimited agents — billing is per-agent, metered via Dodo Payments
    - On paid plans we allow adding agents freely and notify the billing system
    """
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        return

    # Employees never count toward seat limits
    role_value = role.value if isinstance(role, UserRole) else role
    if str(role_value) == "employee":
        return

    limits = get_plan_limits(tenant.plan)

    # Count current agents/admins
    current_count = db.query(User).filter(
        User.tenant_id == tenant_id,
        User.is_active == True,
        User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
    ).count()

    # ── Trial enforcement ──────────────────────────────────────────────────────
    trial = get_trial_status(tenant)
    if trial.get("on_trial"):
        trial_max = limits.get("trial_max_agents", 3)
        if trial_max and current_count + additional > trial_max:
            raise HTTPException(
                status_code=403,
                detail=f"Trial limit reached. You can add up to {trial_max} agents/admins during the 14-day trial. "
                       f"Upgrade to a paid plan to add more."
            )
        return  # within trial limit — allow

    # ── Free plan enforcement ──────────────────────────────────────────────────
    if tenant.billing_status in (None, "trial_expired", "cancelled") or tenant.plan == "free":
        max_users = limits.get("max_users") or limits.get("max_agents")
        if max_users and current_count + additional > max_users:
            raise HTTPException(
                status_code=403,
                detail=f"The Free plan supports {max_users} agent seat. "
                       f"Upgrade to Essentials or higher to add more agents."
            )
        return

    # ── Paid plan: auto-update seat count in Dodo Payments ────────────────────
    if tenant.billing_status == "active":
        new_count = current_count + additional
        # Call Dodo API to update subscription quantity
        success = _update_dodo_seat_count(tenant, new_count)
        if not success:
            # Payment failed — block user creation and show billing page
            portal_url = f"https://{'customer' if DODO_ENVIRONMENT == 'live_mode' else 'test.customer'}.dodopayments.com/login/{os.getenv('DODO_BUSINESS_ID', '')}"
            raise HTTPException(
                status_code=402,
                detail=f"Could not update your subscription for the additional seat. "
                       f"Please check your payment method at {portal_url} and try again."
            )
        return

    # ── Expired/unknown state — apply free plan limits ─────────────────────────
    raise HTTPException(
        status_code=403,
        detail="Your plan has expired. Please renew your subscription to add agents."
    )

# =============================================================================
# ENUMS
# =============================================================================

class UserRole(str, enum.Enum):
    READONLY       = "readonly"        # view only
    EMPLOYEE       = "employee"        # raises tickets only
    AGENT          = "agent"           # resolves tickets
    ADMIN          = "admin"           # manages one tenant
    SUPER_ADMIN    = "super_admin"     # MSP admin — manages their client tenants
    PLATFORM_ADMIN = "platform_admin"  # DodoDesk owner — sees ALL tenants, ALL data

class TicketStatus(str, enum.Enum):
    PENDING_APPROVAL    = "pending_approval"
    OPEN                = "open"
    IN_PROGRESS         = "in_progress"
    PENDING_USER        = "pending_user"      # waiting for requester's input/reply
    PENDING_VENDOR      = "pending_vendor"    # waiting for third-party/vendor
    RESOLVED            = "resolved"
    CLOSED              = "closed"

class TicketPriority(str, enum.Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"

class TicketType(str, enum.Enum):
    INCIDENT        = "incident"
    SERVICE_REQUEST = "service_request"

class AssetType(str, enum.Enum):
    HARDWARE    = "hardware"
    SOFTWARE    = "software"
    NETWORK     = "network"
    MOBILE      = "mobile"
    PERIPHERAL  = "peripheral"
    SAAS        = "saas"
    CLOUD       = "cloud"
    OTHER       = "other"

class AssetStatus(str, enum.Enum):
    AVAILABLE   = "available"
    ASSIGNED    = "assigned"
    MAINTENANCE = "maintenance"
    RETIRED     = "retired"
    DISPOSED    = "disposed"
    LOST        = "lost"
    STOLEN      = "stolen"

class ChangeType(str, enum.Enum):
    NORMAL    = "normal"      # Standard ITIL change requiring CAB approval
    STANDARD  = "standard"    # Pre-approved, low-risk, routine change
    EMERGENCY = "emergency"   # Urgent, bypasses normal CAB cycle

class ChangeRisk(str, enum.Enum):
    LOW    = "low"
    MEDIUM = "medium"
    HIGH   = "high"
    CRITICAL = "critical"

class ChangeStatus(str, enum.Enum):
    DRAFT            = "draft"
    PENDING_APPROVAL = "pending_approval"
    IN_REVIEW        = "in_review"
    APPROVED         = "approved"
    SCHEDULED        = "scheduled"
    IN_PROGRESS      = "in_progress"
    IMPLEMENTED      = "implemented"
    REJECTED         = "rejected"
    CANCELLED        = "cancelled"
    FAILED           = "failed"

class Permission(str, enum.Enum):
    VIEW_ALL_TICKETS = "view_all_tickets"
    CREATE_TICKETS = "create_tickets"
    EDIT_TICKETS = "edit_tickets"
    DELETE_TICKETS = "delete_tickets"
    MANAGE_ASSETS = "manage_assets"
    MANAGE_USERS = "manage_users"
    MANAGE_KB = "manage_kb"
    VIEW_REPORTS = "view_reports"
    MANAGE_CANNED = "manage_canned"
    CREATE_CHANGES = "create_changes"
    APPROVE_CHANGES = "approve_changes"
    MANAGE_CATALOG = "manage_catalog"
    MANAGE_TENANT = "manage_tenant"

# =============================================================================
# MODELS
# =============================================================================

class CustomRole(Base):
    __tablename__ = "custom_roles"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    permissions = Column(Text, nullable=False)  # JSON list of Permission values
    is_default = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())

    tenant = relationship("Tenant", back_populates="custom_roles")

class Tenant(Base):
    __tablename__ = "tenants"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    slug = Column(String, unique=True, index=True, nullable=False)
    logo_url = Column(String, nullable=True)
    primary_color = Column(String, default="#4f46e5")
    accent_color = Column(String, default="#818cf8")
    company_tagline = Column(String, nullable=True)
    support_email = Column(String, nullable=True)
    custom_css = Column(Text, nullable=True)
    favicon_url = Column(String, nullable=True)
    is_active = Column(Boolean, default=True)
    plan = Column(String, default="free")  # free | pro | enterprise
    # Billing (Paddle)
    dodo_customer_id = Column(String, nullable=True)       # was paddle_customer_id
    dodo_subscription_id = Column(String, nullable=True)   # was paddle_subscription_id
    billing_status = Column(String, nullable=True)  # active | past_due | canceled | paused
    plan_renews_at = Column(DateTime, nullable=True)
    # Security settings
    mfa_enabled = Column(Boolean, default=False)       # MFA available for voluntary enrollment
    mfa_required = Column(Boolean, default=False)      # MFA mandatory for all users
    sso_enabled = Column(Boolean, default=False)
    sso_provider = Column(String, default="google")
    sso_client_id = Column(String, nullable=True)
    sso_client_secret = Column(String, nullable=True)
    sso_domain        = Column(String, nullable=True)   # allowed email domain e.g. company.com
    sso_tenant_id     = Column(String, nullable=True)   # Azure tenant ID / IdP SSO URL
    sso_sso_url       = Column(String, nullable=True)   # IdP Single Sign-On URL
    saml_cert         = Column(Text,   nullable=True)   # IdP X.509 certificate (PEM)
    ip_whitelist      = Column(Text,   nullable=True)   # JSON array of allowed CIDRs
    scheduled_reports = Column(Text,   nullable=True)   # JSON config for scheduled reports
    billing_notes     = Column(Text,   nullable=True)   # JSON flags e.g. {"warned_7d": "..."}
    onboarding_emails = Column(Text,   nullable=True)   # JSON tracking onboarding email sends
    created_at = Column(DateTime, server_default=sa_func.now())

    users = relationship("User", back_populates="tenant")
    tickets = relationship("Ticket", back_populates="tenant")
    assets = relationship("Asset", back_populates="tenant")
    kb_articles = relationship("KBArticle", back_populates="tenant")
    change_requests = relationship("ChangeRequest", back_populates="tenant")
    service_catalog_items = relationship("ServiceCatalogItem", back_populates="tenant")
    custom_roles = relationship("CustomRole", back_populates="tenant")

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, unique=True, index=True, nullable=False)
    hashed_password = Column(String, nullable=False)
    full_name = Column(String, nullable=False)
    role = Column(String, default="employee")
    custom_role_id = Column(Integer, ForeignKey("custom_roles.id"), nullable=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    is_active = Column(Boolean, default=True)
    language = Column(String, default='en')
    theme = Column(String, default='light')
    profile_photo = Column(String, nullable=True)
    job_title = Column(String, nullable=True)
    department = Column(String, nullable=True)
    employee_id = Column(String, nullable=True)  # custom employee ID set by admin
    country = Column(String, nullable=True)  # country name
    failed_login_attempts = Column(Integer, default=0)
    locked_until = Column(DateTime, nullable=True)
    status_changed_at = Column(DateTime, nullable=True)  # last time is_active was toggled
    current_session_id = Column(String, nullable=True)  # for single-session enforcement
    pending_email = Column(String, nullable=True)            # new email awaiting confirmation
    email_change_token = Column(String, nullable=True)       # token sent to new email
    email_change_expires_at = Column(DateTime, nullable=True)
    mfa_enabled = Column(Boolean, default=False)
    mfa_secret = Column(String, nullable=True)
    mfa_backup_codes = Column(Text, nullable=True)  # JSON array of unused backup codes
    email_verified = Column(Boolean, default=False)  # must verify email before tenant is activated
    password_reset_token = Column(String, nullable=True)
    password_reset_expires_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())
    # New profile fields
    phone = Column(String, nullable=True)
    timezone = Column(String, default="UTC")
    availability = Column(String, default="online")    # online | busy | away | offline
    notification_prefs = Column(Text, nullable=True)   # JSON: per-event toggles

    tenant = relationship("Tenant", back_populates="users")
    custom_role = relationship("CustomRole")

class SignupVerification(Base):
    """Stores pending email verification tokens for self-serve signup."""
    __tablename__ = "signup_verifications"
    id = Column(Integer, primary_key=True, index=True)
    token = Column(String, unique=True, index=True, nullable=False)
    email = Column(String, nullable=False)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    plan = Column(String, default="free")  # plan they signed up for (determines post-verify redirect)
    expires_at = Column(DateTime, nullable=False)
    used = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())


class Ticket(Base):
    __tablename__ = "tickets"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    ticket_type = Column(String, default="incident")
    title = Column(String, nullable=False)
    description = Column(String, nullable=False)
    category = Column(String, nullable=True)
    priority = Column(String, default="medium")
    status = Column(String, default="open")
    requester_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    group_id = Column(Integer, ForeignKey("groups.id", use_alter=True, name="fk_ticket_group"), nullable=True)
    asset_id = Column(Integer, ForeignKey("assets.id"), nullable=True)
    sla_response_deadline = Column(DateTime, nullable=True)
    sla_resolution_deadline = Column(DateTime, nullable=True)
    sla_breach_notified_at = Column(DateTime, nullable=True)
    escalated_at = Column(DateTime, nullable=True)
    first_response_at = Column(DateTime, nullable=True)  # when first agent reply was posted
    tags = Column(Text, nullable=True)  # JSON array of tag strings e.g. ["vpn","network"]
    merged_into_id = Column(Integer, nullable=True)  # if merged, points to primary ticket id
    resolution_note = Column(Text, nullable=True)    # what was done to resolve the ticket
    resolved_at = Column(DateTime, nullable=True)    # when it was resolved
    resolution_kb_article_id = Column(Integer, ForeignKey("kb_articles.id"), nullable=True)  # linked KB article
    csat_token = Column(String, unique=True, nullable=True)
    csat_rating = Column(Integer, nullable=True)
    csat_comment = Column(Text, nullable=True)
    due_date = Column(DateTime, nullable=True)           # manual due date set by agent
    custom_fields_data = Column(Text, nullable=True)     # JSON: {field_key: value, ...}
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, onupdate=sa_func.now())

    tenant = relationship("Tenant", back_populates="tickets")
    requester = relationship("User", foreign_keys=[requester_id])
    assigned_to = relationship("User", foreign_keys=[assigned_to_id])
    asset = relationship("Asset", back_populates="tickets", foreign_keys=[asset_id])
    comments = relationship("Comment", back_populates="ticket", order_by="Comment.created_at")
    attachments = relationship("Attachment", back_populates="ticket", order_by="Attachment.uploaded_at")

class Comment(Base):
    __tablename__ = "comments"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    author_id = Column(Integer, ForeignKey("users.id"), nullable=True)  # nullable for system comments
    body = Column(String, nullable=False)
    is_internal = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())

    ticket = relationship("Ticket", back_populates="comments")
    author = relationship("User", foreign_keys=[author_id])

# ── Custom ticket fields ──────────────────────────────────────────────────────
class CustomField(Base):
    """Admin-defined extra fields for tickets, per tenant."""
    __tablename__ = "custom_fields"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)           # e.g. "Customer PO Number"
    field_key = Column(String, nullable=False)       # e.g. "customer_po_number"
    field_type = Column(String, default="text")      # text | number | date | dropdown | checkbox
    options = Column(Text, nullable=True)            # JSON list of options for dropdown
    is_required = Column(Boolean, default=False)
    applies_to = Column(String, default="all")       # all | incident | service_request | change | asset | kb_article
    sort_order = Column(Integer, default=0)
    created_at = Column(DateTime, server_default=sa_func.now())

# ── Macros ───────────────────────────────────────────────────────────────────
class Macro(Base):
    """One-click multi-action sequences for agents."""
    __tablename__ = "macros"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    description = Column(String, nullable=True)
    actions = Column(Text, nullable=False)           # JSON list of actions
    is_shared = Column(Boolean, default=True)        # shared=all agents, False=creator only
    created_by_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    run_count = Column(Integer, default=0)
    created_at = Column(DateTime, server_default=sa_func.now())
    created_by = relationship("User", foreign_keys=[created_by_id])

# ── Saved ticket views ────────────────────────────────────────────────────────
class TicketView(Base):
    """Saved filter views per agent or shared across team."""
    __tablename__ = "ticket_views"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    created_by_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    name = Column(String, nullable=False)
    filters = Column(Text, nullable=False)           # JSON: {status, priority, assigned, category, tag, ...}
    is_shared = Column(Boolean, default=False)       # False = personal, True = shared with team
    sort_order = Column(Integer, default=0)
    created_at = Column(DateTime, server_default=sa_func.now())
    created_by = relationship("User", foreign_keys=[created_by_id])

# ── Ticket tasks ──────────────────────────────────────────────────────────────
class TicketTask(Base):
    """Sub-tasks / checklist items on a ticket."""
    __tablename__ = "ticket_tasks"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    title = Column(String, nullable=False)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    due_date = Column(DateTime, nullable=True)
    is_done = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())
    assigned_to = relationship("User", foreign_keys=[assigned_to_id])

# ── Ticket templates ──────────────────────────────────────────────────────────
class TicketTemplate(Base):
    """Pre-filled ticket forms for common request types."""
    __tablename__ = "ticket_templates"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)            # e.g. "VPN Access Request"
    ticket_type = Column(String, default="incident") # incident | service_request only (changes use /changes module)
    title = Column(String, nullable=True)
    description = Column(Text, nullable=True)
    category = Column(String, nullable=True)
    priority = Column(String, default="medium")
    tags = Column(Text, nullable=True)               # JSON array
    created_at = Column(DateTime, server_default=sa_func.now())

# ── Problem tickets ───────────────────────────────────────────────────────────
class ProblemLink(Base):
    """Links multiple incident tickets to a root-cause problem ticket."""
    __tablename__ = "problem_links"
    id = Column(Integer, primary_key=True, index=True)
    problem_ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    incident_ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)

class KBArticle(Base):
    __tablename__ = "kb_articles"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    title = Column(String, nullable=False)
    content = Column(Text, nullable=False)
    category = Column(String, nullable=True)
    folder = Column(String, nullable=True)           # sub-category / folder within category
    author_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    status = Column(String, default="draft", nullable=False)  # draft | published
    version = Column(Integer, default=1, nullable=False)
    view_count = Column(Integer, default=0, nullable=False)
    helpful_count = Column(Integer, default=0, nullable=False)      # 👍 count
    not_helpful_count = Column(Integer, default=0, nullable=False)  # 👎 count
    tags = Column(Text, nullable=True)               # JSON array of tag strings
    visibility = Column(String, default="all")       # all | agents_only | employees_only
    review_date = Column(DateTime, nullable=True)    # flag for review after this date
    sort_order = Column(Integer, default=0)
    custom_fields_data = Column(Text, nullable=True)      # JSON: {field_key: value, ...}
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, onupdate=sa_func.now())

    tenant = relationship("Tenant", back_populates="kb_articles")
    author = relationship("User")
    versions = relationship("KBVersion", back_populates="article", cascade="all, delete-orphan", order_by="KBVersion.version_number.desc()")

class KBVersion(Base):
    """Snapshot of a KB article at each save."""
    __tablename__ = "kb_versions"
    id = Column(Integer, primary_key=True, index=True)
    article_id = Column(Integer, ForeignKey("kb_articles.id"), nullable=False)
    version_number = Column(Integer, nullable=False)
    title = Column(String, nullable=False)
    content = Column(Text, nullable=False)
    category = Column(String, nullable=True)
    status = Column(String, nullable=True)
    change_note = Column(String, nullable=True)  # optional note about what changed
    edited_by_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())
    article = relationship("KBArticle", back_populates="versions")
    edited_by = relationship("User", foreign_keys=[edited_by_id])

class Asset(Base):
    __tablename__ = "assets"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    type = Column(String, nullable=False)  # stored as lowercase varchar, not enum
    model = Column(String, nullable=True)                 # e.g. "Dell Latitude 5420" — picked from admin-managed list per type
    serial_number = Column(String, unique=True, nullable=True)
    status = Column(String, default="available")  # stored as lowercase varchar
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    purchase_date = Column(DateTime, nullable=True)
    license_key = Column(String, nullable=True)
    vendor = Column(String, nullable=True)
    expiry_date = Column(Date, nullable=True)
    notes = Column(Text, nullable=True)
    # New fields
    location = Column(String, nullable=True)              # room, building, site
    purchase_cost = Column(Float, nullable=True)          # for depreciation
    warranty_expiry = Column(Date, nullable=True)         # warranty end date
    contract_number = Column(String, nullable=True)       # PO / contract ref
    quantity = Column(Integer, default=1)                 # for consumables
    seats_total = Column(Integer, nullable=True)          # software: total seats
    seats_used = Column(Integer, default=0)               # software: seats in use
    maintenance_date = Column(DateTime, nullable=True)    # next planned maintenance
    parent_asset_id = Column(Integer, ForeignKey("assets.id"), nullable=True)  # asset hierarchy
    tag_number = Column(String, nullable=True)            # asset tag / barcode
    custom_fields_data = Column(Text, nullable=True)      # JSON: {field_key: value, ...}
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, onupdate=sa_func.now())

    tenant = relationship("Tenant", back_populates="assets")
    assigned_to = relationship("User", foreign_keys=[assigned_to_id])
    tickets = relationship("Ticket", back_populates="asset", foreign_keys=[Ticket.asset_id])
    children = relationship("Asset", foreign_keys=[parent_asset_id], backref=backref("parent", remote_side="Asset.id"))

class AssetModelOption(Base):
    """Admin-managed list of model/manufacturer options shown in the asset creation
    dropdown, scoped per asset type (e.g. type=hardware → Dell Latitude 5420, HP EliteBook...)."""
    __tablename__ = "asset_model_options"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    asset_type = Column(String, nullable=False)           # stored as lowercase string e.g. "hardware"
    label = Column(String, nullable=False)                # e.g. "Dell Latitude 5420"
    sort_order = Column(Integer, default=0)
    created_at = Column(DateTime, server_default=sa_func.now())



class AssetHistory(Base):
    """Tracks every assignment change for an asset."""
    __tablename__ = "asset_history"
    id = Column(Integer, primary_key=True, index=True)
    asset_id = Column(Integer, ForeignKey("assets.id"), nullable=False)
    action = Column(String, nullable=False)  # "assigned", "unassigned", "status_changed"
    from_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    to_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    note = Column(String, nullable=True)
    changed_by_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    changed_at = Column(DateTime, server_default=sa_func.now())
    from_user = relationship("User", foreign_keys=[from_user_id])
    to_user = relationship("User", foreign_keys=[to_user_id])
    changed_by = relationship("User", foreign_keys=[changed_by_id])

class TimeEntry(Base):
    """Agent logs time spent on a ticket."""
    __tablename__ = "time_entries"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    agent_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    minutes = Column(Integer, nullable=False)  # time spent in minutes
    note = Column(String, nullable=True)        # what was done
    logged_at = Column(DateTime, server_default=sa_func.now())
    agent = relationship("User")

class TicketLink(Base):
    """Parent-child relationship between tickets."""
    __tablename__ = "ticket_links"
    id = Column(Integer, primary_key=True, index=True)
    parent_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    child_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)

class Group(Base):
    """Agent groups — tickets can be assigned to a group."""
    __tablename__ = "groups"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    description = Column(String, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())
    members = relationship("GroupMember", back_populates="group", cascade="all, delete-orphan")

class GroupMember(Base):
    __tablename__ = "group_members"
    id = Column(Integer, primary_key=True, index=True)
    group_id = Column(Integer, ForeignKey("groups.id"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    group = relationship("Group", back_populates="members")
    user = relationship("User")

class AutomationRule(Base):
    """If/then automation rules — run on ticket events or on a schedule."""
    __tablename__ = "automation_rules"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    description = Column(String, nullable=True)
    is_active = Column(Boolean, default=True)
    trigger = Column(String, nullable=False)   # on_create | on_update | on_status_change | time_based
    # Conditions stored as JSON: [{"field": "priority", "operator": "is", "value": "high"}]
    conditions = Column(Text, nullable=True)
    # Actions stored as JSON: [{"action": "assign_to", "value": "12"}]
    actions = Column(Text, nullable=False)
    run_count = Column(Integer, default=0)
    last_run_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())

class AdminTenantAccess(Base):
    """Super admin can grant an admin access to manage multiple tenants."""
    __tablename__ = "admin_tenant_access"
    id = Column(Integer, primary_key=True, index=True)
    admin_user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    granted_by_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    granted_at = Column(DateTime, server_default=sa_func.now())
    admin_user = relationship("User", foreign_keys=[admin_user_id])
    tenant = relationship("Tenant")

class CannedResponse(Base):
    __tablename__ = "canned_responses"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    title = Column(String, nullable=False)
    content = Column(Text, nullable=False)
    category = Column(String, nullable=True)          # folder / category
    author_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    visibility = Column(String, default="all")        # all | personal | group
    group_id = Column(Integer, ForeignKey("groups.id"), nullable=True)
    use_count = Column(Integer, default=0)
    sort_order = Column(Integer, default=0)
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, onupdate=sa_func.now())

    author = relationship("User")
    group = relationship("Group", foreign_keys=[group_id])

class Attachment(Base):
    __tablename__ = "attachments"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    filename = Column(String, nullable=False)
    stored_filename = Column(String, nullable=False)   # legacy local disk name OR Cloudinary public_id
    url = Column(String, nullable=True)                # Cloudinary secure_url (None = legacy local file)
    content_type = Column(String, nullable=True)
    size = Column(Integer, nullable=False)
    uploaded_at = Column(DateTime, server_default=sa_func.now())

    ticket = relationship("Ticket", back_populates="attachments")

class ChangeRequest(Base):
    __tablename__ = "change_requests"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    title = Column(String, nullable=False)
    description = Column(Text, nullable=False)
    change_type = Column(String, default="normal")          # normal | standard | emergency
    risk_level = Column(String, default="medium")
    risk_score = Column(Integer, nullable=True)             # 1-25 calculated from impact x likelihood
    status = Column(String, default="draft")
    requester_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    owner_id = Column(Integer, ForeignKey("users.id"), nullable=True)      # change owner (separate from requester)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    planned_date = Column(Date, nullable=True)
    start_date = Column(DateTime, nullable=True)            # implementation start
    end_date = Column(DateTime, nullable=True)              # implementation end
    impact = Column(Text, nullable=True)                    # who/what is affected
    rollback_plan = Column(Text, nullable=True)             # what to do if change fails
    test_plan = Column(Text, nullable=True)                 # how to verify success
    cab_members = Column(Text, nullable=True)               # JSON list of user_ids for CAB
    linked_ticket_ids = Column(Text, nullable=True)         # JSON list of ticket IDs
    linked_asset_ids = Column(Text, nullable=True)          # JSON list of asset IDs
    post_review_notes = Column(Text, nullable=True)         # post-implementation review
    post_review_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, onupdate=sa_func.now())

    tenant = relationship("Tenant", back_populates="change_requests")
    requester = relationship("User", foreign_keys=[requester_id])
    owner = relationship("User", foreign_keys=[owner_id])
    assigned_to = relationship("User", foreign_keys=[assigned_to_id])

# ── Change tasks ──────────────────────────────────────────────────────────────
class ChangeTask(Base):
    """Sub-tasks / checklist items on a change request."""
    __tablename__ = "change_tasks"
    id = Column(Integer, primary_key=True, index=True)
    change_id = Column(Integer, ForeignKey("change_requests.id"), nullable=False)
    title = Column(String, nullable=False)
    assigned_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    is_done = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())
    assigned_to = relationship("User", foreign_keys=[assigned_to_id])

# ── Change comments ───────────────────────────────────────────────────────────
class ChangeComment(Base):
    """Comments / discussion on a change request."""
    __tablename__ = "change_comments"
    id = Column(Integer, primary_key=True, index=True)
    change_id = Column(Integer, ForeignKey("change_requests.id"), nullable=False)
    author_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    body = Column(Text, nullable=False)
    is_internal = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())
    author = relationship("User", foreign_keys=[author_id])

class ServiceCatalogItem(Base):
    __tablename__ = "service_catalog_items"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    category = Column(String, nullable=True)
    estimated_cost = Column(Float, nullable=True)
    delivery_time_days = Column(Integer, nullable=True)
    approval_required = Column(Boolean, default=True)
    is_active = Column(Boolean, default=True)
    # Pre-fill fields (merged from TicketTemplate)
    ticket_title = Column(String, nullable=True)
    ticket_description = Column(Text, nullable=True)
    ticket_type = Column(String, default="service_request")
    priority = Column(String, default="medium")
    is_onboarding = Column(Boolean, default=False)      # triggers bulk ticket creation
    onboarding_tasks = Column(Text, nullable=True)       # JSON array of tasks
    is_featured = Column(Boolean, default=False)
    # New features
    sort_order = Column(Integer, default=0)
    icon = Column(String, nullable=True)
    request_form_fields = Column(Text, nullable=True)
    visibility = Column(String, default="all")
    sla_hours = Column(Integer, nullable=True)
    request_count = Column(Integer, default=0)
    fulfillment_checklist = Column(Text, nullable=True)
    approval_workflow_id = Column(Integer, ForeignKey("approval_workflows.id"), nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())

    tenant = relationship("Tenant", back_populates="service_catalog_items")

class EmailConfig(Base):
    __tablename__ = "email_configs"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), unique=True, nullable=False)
    smtp_host = Column(String, default="")
    smtp_port = Column(Integer, default=587)
    smtp_user = Column(String, default="")
    smtp_pass = Column(String, default="")
    smtp_from = Column(String, default="noreply@itsm.local")
    reply_to  = Column(String, default="")
    slack_webhook_url  = Column(String, default="")
    teams_webhook_url  = Column(String, default="")
    email_signature = Column(Text, default="")
    email_footer = Column(Text, default="")
    updated_at = Column(DateTime, onupdate=sa_func.now())

class EscalationRule(Base):
    __tablename__ = "escalation_rules"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    priority = Column(String, nullable=True)        # if None, applies to all priorities
    idle_hours = Column(Integer, nullable=False)    # hours without update before escalating
    escalate_to_id = Column(Integer, ForeignKey("users.id"), nullable=True)  # specific agent
    escalate_to_role = Column(String, nullable=True)  # or any agent/admin
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, server_default=sa_func.now())

class SLAConfig(Base):
    __tablename__ = "sla_configs"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), unique=True, nullable=False)
    low_response = Column(Integer, default=8)
    low_resolution = Column(Integer, default=72)
    medium_response = Column(Integer, default=4)
    medium_resolution = Column(Integer, default=48)
    high_response = Column(Integer, default=2)
    high_resolution = Column(Integer, default=24)
    critical_response = Column(Integer, default=1)
    critical_resolution = Column(Integer, default=8)
    updated_at = Column(DateTime, onupdate=sa_func.now())

class BusinessHoursConfig(Base):
    __tablename__ = "business_hours_configs"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), unique=True, nullable=False)
    enabled = Column(Boolean, default=False)
    start_hour = Column(Integer, default=9)   # 9 AM
    end_hour = Column(Integer, default=17)    # 5 PM
    # Working days: comma-separated 0=Mon,1=Tue,...,6=Sun
    working_days = Column(String, default="0,1,2,3,4")  # Mon-Fri
    timezone = Column(String, default="UTC")
    updated_at = Column(DateTime, onupdate=sa_func.now())

# ── AI Chatbot models (Enterprise plan) ──────────────────────────────────

class ChatSession(Base):
    __tablename__ = "chat_sessions"
    id         = Column(Integer, primary_key=True, index=True)
    tenant_id  = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    user_id    = Column(Integer, ForeignKey("users.id"), nullable=False)
    title      = Column(String, default="New conversation")
    created_at = Column(DateTime, server_default=sa_func.now())
    updated_at = Column(DateTime, server_default=sa_func.now(), onupdate=sa_func.now())
    messages   = relationship("ChatMessage", back_populates="session", cascade="all, delete-orphan")

class ChatMessage(Base):
    __tablename__ = "chat_messages"
    id         = Column(Integer, primary_key=True, index=True)
    session_id = Column(Integer, ForeignKey("chat_sessions.id"), nullable=False)
    role       = Column(String, nullable=False)   # "user" | "assistant"
    content    = Column(Text, nullable=False)
    tool_calls = Column(Text, nullable=True)       # JSON summary of tools used
    created_at = Column(DateTime, server_default=sa_func.now())
    session    = relationship("ChatSession", back_populates="messages")

class Notification(Base):
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    type = Column(String, nullable=False)
    title = Column(String, nullable=False)
    body = Column(String, nullable=False)
    link = Column(String, nullable=True)
    is_read = Column(Boolean, default=False)
    created_at = Column(DateTime, server_default=sa_func.now())

class SystemAuditLog(Base):
    """Platform-wide audit log for admin actions: user management, settings, plan changes, etc."""
    __tablename__ = "system_audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=True)
    actor_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    actor_email = Column(String, nullable=True)  # stored in case user is later deleted
    action = Column(String, nullable=False)       # e.g. "user.created", "plan.changed", "branding.updated"
    target_type = Column(String, nullable=True)   # e.g. "user", "tenant", "workflow"
    target_id = Column(String, nullable=True)     # ID of the affected object
    target_label = Column(String, nullable=True)  # human-readable e.g. "jane@acme.com"
    old_value = Column(String, nullable=True)
    new_value = Column(String, nullable=True)
    ip_address = Column(String, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())


class TicketAuditLog(Base):
    __tablename__ = "ticket_audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    actor_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    action = Column(String, nullable=False)
    field = Column(String, nullable=True)
    old_value = Column(String, nullable=True)
    new_value = Column(String, nullable=True)
    note = Column(String, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())

class ApprovalWorkflow(Base):
    __tablename__ = "approval_workflows"
    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    name = Column(String, nullable=False)
    category = Column(String, nullable=True)       # matches ticket category e.g. "Hardware"
    ticket_type = Column(String, default="service_request")
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, server_default=sa_func.now())
    steps = relationship("ApprovalStep", back_populates="workflow", order_by="ApprovalStep.step_order")

class ApprovalStep(Base):
    __tablename__ = "approval_steps"
    id = Column(Integer, primary_key=True, index=True)
    workflow_id = Column(Integer, ForeignKey("approval_workflows.id"), nullable=False)
    step_order = Column(Integer, nullable=False)   # 1, 2, 3...
    name = Column(String, nullable=False)          # e.g. "Line Manager Approval"
    approver_id = Column(Integer, ForeignKey("users.id"), nullable=True)   # specific user
    approver_role = Column(String, nullable=True)  # or any user with this role
    workflow = relationship("ApprovalWorkflow", back_populates="steps")

class TicketApproval(Base):
    __tablename__ = "ticket_approvals"
    id = Column(Integer, primary_key=True, index=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    workflow_id = Column(Integer, ForeignKey("approval_workflows.id"), nullable=False)
    step_id = Column(Integer, ForeignKey("approval_steps.id"), nullable=False)
    step_order = Column(Integer, nullable=False)
    step_name = Column(String, nullable=False)
    approver_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    approver_role = Column(String, nullable=True)
    status = Column(String, default="pending")     # pending, approved, rejected, skipped
    comment = Column(Text, nullable=True)
    decided_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, server_default=sa_func.now())

class TicketWatcher(Base):
    __tablename__ = "ticket_watchers"
    id         = Column(Integer, primary_key=True, index=True)
    ticket_id  = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    user_id    = Column(Integer, ForeignKey("users.id"), nullable=False)
    tenant_id  = Column(Integer, ForeignKey("tenants.id"), nullable=False)
    created_at = Column(DateTime, server_default=sa_func.now())
    __table_args__ = (UniqueConstraint("ticket_id", "user_id", name="uq_ticket_watcher"),)

# =============================================================================

class TicketCreate(BaseModel):
    model_config = {"extra": "ignore"}   # silently ignore unknown fields from frontend
    title: str
    description: str
    category: str | None = None          # optional — some tenants don't use categories
    priority: TicketPriority = "medium"
    ticket_type: TicketType = "incident"
    on_behalf_of_id: int | None = None
    tags: list[str] = []
    group_id: int | None = None
    due_date: datetime | None = None
    custom_fields_data: dict | None = None
    template_id: int | None = None
    asset_id: int | None = None          # link ticket to an asset at creation time
    impact: str | None = None            # optional impact field
    urgency: str | None = None           # optional urgency field
    assigned_to_id: int | None = None    # explicit agent assignment (overrides round-robin)

class TicketUpdate(BaseModel):
    status: TicketStatus | None = None
    assigned_to_id: int | None = None
    priority: TicketPriority | None = None
    category: str | None = None
    title: str | None = None
    description: str | None = None
    tags: list[str] | None = None
    group_id: int | None = None
    resolution_note: str | None = None
    resolution_kb_article_id: int | None = None
    due_date: datetime | None = None
    custom_fields_data: dict | None = None

class TicketOut(BaseModel):
    model_config = {"extra": "ignore", "from_attributes": True}
    id: int
    ticket_type: str = "incident"
    title: str
    description: str | None = None
    category: str | None = None
    priority: str = "medium"
    status: str = "open"
    requester_id: int | None = None
    requester_name: str = ""
    assigned_to_id: int | None = None
    asset_id: int | None = None
    sla_response_deadline: datetime | None = None
    sla_resolution_deadline: datetime | None = None
    sla_status: str | None = None
    created_at: datetime | None = None
    watchers: list[dict] = []

class CommentCreate(BaseModel):
    body: str
    is_internal: bool = False  # True = private note visible only to agents/admins

class CommentOut(BaseModel):
    id: int
    ticket_id: int
    author_id: int
    author_name: str
    body: str
    is_internal: bool = False
    created_at: datetime

    class Config:
        from_attributes = True

class KBArticleCreate(BaseModel):
    title: str
    content: str
    category: str
    folder: str | None = None
    status: str = "draft"
    tags: list[str] = []
    visibility: str = "all"
    review_date: datetime | None = None
    custom_fields_data: dict | None = None

class KBArticleUpdate(BaseModel):
    model_config = {"extra": "ignore"}
    title: str | None = None
    content: str | None = None
    category: str | None = None
    folder: str | None = None
    status: str | None = None
    change_note: str | None = None
    tags: list[str] | None = None
    visibility: str | None = None
    review_date: datetime | None = None
    sort_order: int | None = None
    custom_fields_data: dict | None = None



class KBArticleOut(BaseModel):
    id: int
    title: str
    content: str
    category: str | None = None
    folder: str | None = None
    author_id: int | None = None
    author_name: str = "Unknown"
    status: str = "published"
    version: int = 1
    view_count: int = 0
    helpful_count: int = 0
    not_helpful_count: int = 0
    tags: list[str] = []
    visibility: str = "all"
    review_date: datetime | None = None
    sort_order: int = 0
    custom_fields_data: dict | None = None
    created_at: datetime | None = None
    updated_at: datetime | None = None

    class Config:
        from_attributes = True

class AssetCreate(BaseModel):
    name: str
    type: str = "hardware"
    model: str | None = None
    serial_number: str | None = None
    status: AssetStatus = "available"
    assigned_to_id: int | None = None
    purchase_date: datetime | None = None
    license_key: str | None = None
    vendor: str | None = None
    expiry_date: date | None = None
    notes: str | None = None
    location: str | None = None
    purchase_cost: float | None = None
    warranty_expiry: date | None = None
    contract_number: str | None = None
    quantity: int = 1
    seats_total: int | None = None
    maintenance_date: datetime | None = None
    parent_asset_id: int | None = None
    tag_number: str | None = None
    custom_fields_data: dict | None = None

class AssetUpdate(BaseModel):
    name: str | None = None
    type: AssetType | None = None
    model: str | None = None
    serial_number: str | None = None
    status: AssetStatus | None = None
    assigned_to_id: int | None = None
    purchase_date: datetime | None = None
    license_key: str | None = None
    vendor: str | None = None
    expiry_date: date | None = None
    notes: str | None = None
    location: str | None = None
    purchase_cost: float | None = None
    warranty_expiry: date | None = None
    contract_number: str | None = None
    quantity: int | None = None
    seats_total: int | None = None
    seats_used: int | None = None
    maintenance_date: datetime | None = None
    parent_asset_id: int | None = None
    tag_number: str | None = None
    custom_fields_data: dict | None = None

class AssetOut(BaseModel):
    id: int
    name: str
    type: str = "hardware"
    model: str | None = None
    serial_number: str | None
    status: str = "available"
    assigned_to_id: int | None
    assigned_to_name: str | None = None
    purchase_date: datetime | None
    license_key: str | None = None
    vendor: str | None = None
    expiry_date: date | None = None
    notes: str | None
    location: str | None = None
    purchase_cost: float | None = None
    warranty_expiry: date | None = None
    contract_number: str | None = None
    quantity: int = 1
    seats_total: int | None = None
    seats_used: int = 0
    maintenance_date: datetime | None = None
    parent_asset_id: int | None = None
    tag_number: str | None = None
    ticket_count: int = 0
    created_at: datetime
    updated_at: datetime | None

    class Config:
        from_attributes = True

class LinkAssetRequest(BaseModel):
    asset_id: int | None = None

class UserOut(BaseModel):
    id: int
    email: str
    full_name: str
    role: str = "employee"
    is_active: bool
    language: str = "en"
    theme: str = "light"
    profile_photo: str | None = None
    job_title: str | None = None
    department: str | None = None
    employee_id: str | None = None
    country: str | None = None
    tenant_id: int | None = None
    created_at: datetime

    class Config:
        from_attributes = True

class UserCreate(BaseModel):
    email: str
    password: str
    full_name: str
    role: UserRole = UserRole.EMPLOYEE
    job_title: str | None = None
    department: str | None = None
    employee_id: str | None = None
    tenant_id: int | None = None

class UserInvite(BaseModel):
    email: str
    full_name: str
    role: UserRole = UserRole.EMPLOYEE
    job_title: str | None = None
    department: str | None = None

class SignupRequest(BaseModel):
    company_name: str
    full_name: str
    email: str
    password: str
    plan: str = "free"  # "free" or "pro" — Enterprise is not self-serve

class UserUpdate(BaseModel):
    email: str | None = None
    full_name: str | None = None
    role: UserRole | None = None
    is_active: bool | None = None
    password: str | None = None
    job_title: str | None = None
    department: str | None = None
    employee_id: str | None = None
    tenant_id: int | None = None

class UserProfileUpdate(BaseModel):
    full_name: str | None = None
    email: str | None = None
    country: str | None = None
    language: str | None = None
    theme: str | None = None
    job_title: str | None = None
    department: str | None = None
    phone: str | None = None
    timezone: str | None = None
    availability: str | None = None

class PasswordUpdate(BaseModel):
    current_password: str
    new_password: str

class CannedResponseCreate(BaseModel):
    title: str
    content: str
    category: str | None = None
    visibility: str = "all"   # all | personal | group
    group_id: int | None = None
    sort_order: int = 0

class CannedResponseUpdate(BaseModel):
    title: str | None = None
    content: str | None = None
    category: str | None = None
    visibility: str | None = None
    group_id: int | None = None
    sort_order: int | None = None

class CannedResponseOut(BaseModel):
    id: int
    title: str
    content: str
    category: str | None
    author_id: int
    author_name: str
    visibility: str = "all"
    group_id: int | None = None
    use_count: int = 0
    sort_order: int = 0
    created_at: datetime
    updated_at: datetime | None

    class Config:
        from_attributes = True

class AttachmentOut(BaseModel):
    id: int
    ticket_id: int
    filename: str
    url: str | None = None
    content_type: str | None
    size: int
    uploaded_at: datetime

    class Config:
        from_attributes = True

class EmailTicketRequest(BaseModel):
    from_email: str
    subject: str
    body: str

class ChangeCreate(BaseModel):
    title: str
    description: str
    change_type: str = "normal"
    risk_level: str = "medium"
    risk_score: int | None = None
    planned_date: date | None = None
    start_date: datetime | None = None
    end_date: datetime | None = None
    impact: str | None = None
    rollback_plan: str | None = None
    test_plan: str | None = None
    owner_id: int | None = None
    assigned_to_id: int | None = None
    cab_members: list[int] = []
    linked_ticket_ids: list[int] = []
    linked_asset_ids: list[int] = []

class ChangeUpdate(BaseModel):
    title: str | None = None
    description: str | None = None
    change_type: str | None = None
    risk_level: ChangeRisk | None = None
    risk_score: int | None = None
    status: ChangeStatus | None = None
    planned_date: date | None = None
    start_date: datetime | None = None
    end_date: datetime | None = None
    impact: str | None = None
    rollback_plan: str | None = None
    test_plan: str | None = None
    owner_id: int | None = None
    assigned_to_id: int | None = None
    cab_members: list[int] | None = None
    linked_ticket_ids: list[int] | None = None
    linked_asset_ids: list[int] | None = None
    post_review_notes: str | None = None

class ChangeOut(BaseModel):
    model_config = {"extra": "ignore", "from_attributes": True}
    id: int
    title: str
    description: str | None = None
    change_type: str = "normal"
    risk_level: str = "medium"
    risk_score: int | None = None
    status: str = "draft"
    requester_id: int | None = None
    requester_name: str = ""
    owner_id: int | None = None
    owner_name: str = ""
    assigned_to_id: int | None = None
    assigned_to_name: str = ""
    planned_date: date | None = None
    start_date: datetime | None = None
    end_date: datetime | None = None
    impact: str | None = None
    rollback_plan: str | None = None
    test_plan: str | None = None
    cab_members: list[int] = []
    linked_ticket_ids: list[int] = []
    linked_asset_ids: list[int] = []
    post_review_notes: str | None = None
    post_review_at: datetime | None = None
    created_at: datetime | None = None
    updated_at: datetime | None = None

# ---------- New schemas ----------

class TenantOut(BaseModel):
    id: int
    name: str
    slug: str
    logo_url: str | None
    primary_color: str
    accent_color: str = "#818cf8"
    company_tagline: str | None = None
    support_email: str | None = None
    is_active: bool
    created_at: datetime

    class Config:
        from_attributes = True

class ServiceCatalogItemCreate(BaseModel):
    name: str
    description: str | None = None
    category: str | None = None
    estimated_cost: float | None = None
    delivery_time_days: int | None = None
    approval_required: bool = True
    is_active: bool = True
    is_featured: bool = False

class ServiceCatalogItemOut(BaseModel):
    id: int
    tenant_id: int
    name: str
    description: str | None
    category: str | None
    estimated_cost: float | None
    delivery_time_days: int | None
    approval_required: bool
    is_active: bool
    is_featured: bool = False
    created_at: datetime

    class Config:
        from_attributes = True

class CustomRoleCreate(BaseModel):
    name: str
    permissions: list[Permission]

class CustomRoleOut(BaseModel):
    id: int
    tenant_id: int
    name: str
    permissions: str
    is_default: bool
    created_at: datetime

    class Config:
        from_attributes = True

# CSAT schemas
class CSATSubmit(BaseModel):
    rating: int
    comment: str | None = None

class CSATStats(BaseModel):
    average: float | None
    count: int
    distribution: dict

# =============================================================================
# AUTH UTILITIES
# =============================================================================

# Password hashing uses the bcrypt library directly (not passlib), because
# passlib 1.7.x's bcrypt backend self-test ("detect_wrap_bug") is broken on
# bcrypt>=4.x / Python 3.14, raising on the very first hash/verify call.
import bcrypt as _bcrypt_lib

SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY environment variable is not set")

ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "60"))
EMAIL_API_KEY = os.getenv("EMAIL_API_KEY", "dev-email-key")

MAX_FAILED_ATTEMPTS = int(os.getenv("MAX_FAILED_ATTEMPTS", "5"))
LOCKOUT_DURATION_MINUTES = int(os.getenv("LOCKOUT_DURATION_MINUTES", "15"))

# In-memory rate limiter
from collections import defaultdict
import time as _time

_login_ip_attempts = defaultdict(list)
RATE_LIMIT_WINDOW = 60  # seconds
RATE_LIMIT_MAX = int(os.getenv("RATE_LIMIT_MAX", "5"))

def check_ip_rate_limit(ip: str) -> bool:
    now = _time.time()
    _login_ip_attempts[ip] = [t for t in _login_ip_attempts[ip] if now - t < RATE_LIMIT_WINDOW]
    if len(_login_ip_attempts[ip]) >= RATE_LIMIT_MAX:
        return False
    _login_ip_attempts[ip].append(now)
    return True


PASSWORD_MIN_LENGTH = int(os.getenv("PASSWORD_MIN_LENGTH", "8"))

def validate_password_strength(password: str):
    if len(password) < PASSWORD_MIN_LENGTH:
        raise HTTPException(status_code=400, detail=f"Password must be at least {PASSWORD_MIN_LENGTH} characters")
    if not re.search(r"[A-Z]", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one uppercase letter")
    if not re.search(r"[a-z]", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one lowercase letter")
    if not re.search(r"\d", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one digit")
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one special character")

def verify_password(plain, hashed):
    if not hashed:
        return False
    plain_bytes = plain.encode("utf-8")[:72]  # bcrypt max input length
    try:
        return _bcrypt_lib.checkpw(plain_bytes, hashed.encode("utf-8"))
    except (ValueError, TypeError):
        return False

def get_password_hash(password):
    password_bytes = password.encode("utf-8")[:72]  # bcrypt max input length
    return _bcrypt_lib.hashpw(password_bytes, _bcrypt_lib.gensalt()).decode("utf-8")

# =============================================================================
# TOTP (RFC 6238) — implemented with stdlib only, no extra dependencies
# =============================================================================
import secrets
import secrets as _secrets

def generate_totp_secret() -> str:
    """Generate a base32 secret for TOTP enrollment (16 chars = 80 bits)."""
    return base64.b32encode(_secrets.token_bytes(10)).decode("utf-8")

def _totp_code(secret: str, for_time: int, digits: int = 6, period: int = 30) -> str:
    key = base64.b32decode(secret.upper() + "=" * ((8 - len(secret) % 8) % 8))
    counter = int(for_time // period)
    msg = struct.pack(">Q", counter)
    h = hmac_lib.new(key, msg, hashlib.sha1).digest()
    offset = h[-1] & 0x0F
    code = (struct.unpack(">I", h[offset:offset + 4])[0] & 0x7FFFFFFF) % (10 ** digits)
    return str(code).zfill(digits)

def verify_totp(secret: str, code: str, window: int = 1) -> bool:
    """Verify a TOTP code, allowing +/- `window` periods for clock drift."""
    if not secret or not code:
        return False
    code = code.strip().replace(" ", "")
    now = int(_time_module.time())
    for offset in range(-window, window + 1):
        if _totp_code(secret, now + offset * 30) == code:
            return True
    return False

def totp_provisioning_uri(secret: str, email: str, issuer: str = "DodoDesk") -> str:
    """Build a RFC-compliant otpauth:// URI for QR code generation.
    Both the label and issuer must be percent-encoded for authenticator apps
    (Google Authenticator, Authy, Microsoft Authenticator) to parse correctly.
    Format: otpauth://totp/{issuer}:{account}?secret=X&issuer=X&algorithm=SHA1&digits=6&period=30
    """
    # URL-encode each component individually
    encoded_issuer  = urllib.parse.quote(issuer, safe='')
    encoded_account = urllib.parse.quote(email, safe='')
    label = f"{encoded_issuer}:{encoded_account}"
    params = urllib.parse.urlencode({
        "secret": secret,
        "issuer": issuer,        # issuer param value — raw, urlencode handles it
        "algorithm": "SHA1",
        "digits": 6,
        "period": 30,
    })
    return f"otpauth://totp/{label}?{params}"

def generate_backup_codes(count: int = 8) -> list[str]:
    """Generate human-friendly backup codes like 'ABCD-1234'."""
    codes = []
    for _ in range(count):
        part1 = _secrets.token_hex(2).upper()
        part2 = _secrets.token_hex(2).upper()
        codes.append(f"{part1}-{part2}")
    return codes


def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def create_access_token_with_expiry(data: dict, minutes: int):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=minutes)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def decode_access_token(token: str):
    return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])

# =============================================================================
# EMAIL / NOTIFICATIONS
# =============================================================================

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "DodoDesk <noreply@dodobay.com>")
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")  # preferred over SMTP on Render
# Canonical verified from address for Resend — must match verified domain
RESEND_FROM = os.getenv("RESEND_FROM", SMTP_FROM or "DodoDesk <noreply@dodobay.com>")

# Webhook integrations — fallback globals (per-tenant config overrides these)
SLACK_WEBHOOK_URL  = os.getenv("SLACK_WEBHOOK_URL", "")
TEAMS_WEBHOOK_URL  = os.getenv("TEAMS_WEBHOOK_URL", "")

# =============================================================================
# PADDLE BILLING CONFIG
# =============================================================================

# ── Dodo Payments (replacing Paddle) ─────────────────────────────────────────
DODO_API_KEY            = os.getenv("DODO_PAYMENTS_API_KEY", "")
DODO_WEBHOOK_SECRET     = os.getenv("DODO_PAYMENTS_WEBHOOK_SECRET", "")
DODO_API_BASE           = os.getenv("DODO_API_BASE", "https://api.dodopayments.com")  # same URL for test and live
DODO_ENVIRONMENT        = os.getenv("DODO_ENVIRONMENT", "live_mode")  # "test_mode" or "live_mode"

# Product IDs per plan and billing interval
DODO_PRODUCTS = {
    "essentials": {
        "month": os.getenv("DODO_PRICE_ESSENTIALS_MONTHLY", "pdt_0NiG2gfkljCxtSpqXgNQg"),
        "year":  os.getenv("DODO_PRICE_ESSENTIALS_YEARLY",  "pdt_0NiJzTbLZfDW7v4NdDuBF"),
    },
    "business": {
        "month": os.getenv("DODO_PRICE_BUSINESS_MONTHLY", "pdt_0NiK0WaafQE5ilthVt2Vx"),
        "year":  os.getenv("DODO_PRICE_BUSINESS_YEARLY",  "pdt_0NiK1V6gVbN9vocSYhBrc"),
    },
    "pro": {
        "month": os.getenv("DODO_PRICE_PRO_MONTHLY", "pdt_0NiK2AH3V5xLke1ZT0LmP"),
        "year":  os.getenv("DODO_PRICE_PRO_YEARLY",  "pdt_0NiK4FyOdHWijrzrDsMfo"),
    },
}

# Seat add-on IDs — used for per-agent billing via changePlan API
DODO_ADDONS = {
    "essentials": {
        "month": os.getenv("DODO_ADDON_ESSENTIALS_MONTHLY", "adn_0NikooenMHyl2SAs7qTaI"),
        "year":  os.getenv("DODO_ADDON_ESSENTIALS_YEARLY",  "adn_0NikonJrbx1OWYYzaFsIM"),
    },
    "business": {
        "month": os.getenv("DODO_ADDON_BUSINESS_MONTHLY", "adn_0NikomFT4Nu1GA0r3K10V"),
        "year":  os.getenv("DODO_ADDON_BUSINESS_YEARLY",  "adn_0NikoiN1NbrrCLfnIa8wR"),
    },
    "pro": {
        "month": os.getenv("DODO_ADDON_PRO_MONTHLY", "adn_0NikojnarxvEMyCVVSqXu"),
        "year":  os.getenv("DODO_ADDON_PRO_YEARLY",  "adn_0Nikol56JxNhgYa2Wg3pg"),
    },
}

# Log product IDs at startup so you can verify correct IDs are loaded
print(f"📦 Dodo Products loaded ({DODO_ENVIRONMENT}):")
for plan, intervals in DODO_PRODUCTS.items():
    for interval, pid in intervals.items():
        print(f"   {plan}/{interval}: {pid}")

def _update_dodo_seat_count(tenant: "Tenant", new_agent_count: int) -> bool:
    """Call Dodo Payments API to update subscription addon quantity to match agent count.
    Uses seat add-ons (adn_*) so the base subscription stays the same and only
    the per-seat add-on quantity changes — Dodo prorates automatically.
    Returns True if successful, False if payment failed.
    """
    if not tenant.dodo_subscription_id:
        print(f"⚠️ Seat update skipped — tenant {tenant.id} has no dodo_subscription_id")
        return True
    if tenant.billing_status != "active":
        print(f"⚠️ Seat update skipped — billing_status={tenant.billing_status}")
        return True

    api_key  = DODO_API_KEY
    base_url = "https://live.dodopayments.com" if DODO_ENVIRONMENT == "live_mode" else "https://test.dodopayments.com"

    # Detect billing interval
    plan     = (tenant.plan or "essentials").lower()
    interval = "month"
    if tenant.plan_renews_at:
        from datetime import datetime as _dt
        try:
            days = (tenant.plan_renews_at - _dt.utcnow()).days
            if days > 60:
                interval = "year"
        except Exception:
            pass

    # Get addon ID for this plan + interval
    addon_id = DODO_ADDONS.get(plan, {}).get(interval)
    product_id = DODO_PRODUCTS.get(plan, {}).get(interval)

    if not addon_id or not product_id:
        print(f"⚠️ Seat update: no addon_id for plan={plan} interval={interval}")
        return True

    import httpx
    try:
        resp = httpx.post(
            f"{base_url}/subscriptions/{tenant.dodo_subscription_id}/change-plan",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "product_id": product_id,
                "quantity": new_agent_count,             # seats = agent count on base product
                "proration_billing_mode": "prorated_immediately",
                "on_payment_failure": "prevent_change",
            },
            timeout=15.0
        )
        if resp.status_code in (200, 201, 202):
            print(f"✅ Dodo seat update: tenant {tenant.id} ({tenant.name}) → {new_agent_count} seats [{plan}/{interval}]")
            return True
        else:
            print(f"❌ Dodo seat update failed: {resp.status_code} {resp.text[:300]}")
            return False
    except Exception as e:
        print(f"❌ Dodo seat update error: {e}")
        return False


# Anthropic AI chatbot (Enterprise plan)
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
ANTHROPIC_MODEL   = "claude-sonnet-4-6"


def get_email_config(db: Session, tenant_id: int) -> dict:
    """Get email config from DB, falling back to env vars."""
    cfg = db.query(EmailConfig).filter(EmailConfig.tenant_id == tenant_id).first()
    if cfg and cfg.smtp_host:
        return {
            "smtp_host": cfg.smtp_host, "smtp_port": cfg.smtp_port,
            "smtp_user": cfg.smtp_user, "smtp_pass": cfg.smtp_pass,
            "smtp_from": cfg.smtp_from, "reply_to": cfg.reply_to or "",
            "slack_webhook_url": cfg.slack_webhook_url or "",
            "teams_webhook_url": cfg.teams_webhook_url or "",
        }
    return {
        "smtp_host": SMTP_HOST, "smtp_port": SMTP_PORT,
        "smtp_user": SMTP_USER, "smtp_pass": SMTP_PASS,
        "smtp_from": SMTP_FROM, "reply_to": cfg.reply_to if cfg else "",
        "slack_webhook_url": os.getenv("SLACK_WEBHOOK_URL", ""),
        "teams_webhook_url": os.getenv("TEAMS_WEBHOOK_URL", ""),
    }

def build_html_email(subject: str, body_text: str, company_name: str = "DodoDesk", primary_color: str = "#4f46e5", cta_url: str = None, cta_label: str = None, logo_url: str = None) -> str:
    """Build a branded HTML email."""
    # Convert plain text body to HTML paragraphs
    paragraphs = ""
    for line in body_text.strip().split('\n'):
        line = line.strip()
        if line:
            paragraphs += f"<p style='margin:0 0 12px 0;color:#374151;font-size:15px;line-height:1.6;'>{line}</p>"

    cta_html = ""
    if cta_url and cta_label:
        cta_html = f"""
        <div style='text-align:center;margin:28px 0;'>
          <a href='{cta_url}' style='background:{primary_color};color:#ffffff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;font-size:15px;display:inline-block;'>
            {cta_label}
          </a>
        </div>"""

    # Header — white background, logo centered, company name below (industry standard)
    if logo_url:
        header_content = f"""
            <div style='text-align:center;padding:32px 36px 24px;background:#ffffff;border-radius:12px 12px 0 0;'>
              <img src='{logo_url}' alt='{company_name}'
                   style='height:56px;width:auto;object-fit:contain;display:block;margin:0 auto 12px auto;' />
              <p style='margin:0;font-size:22px;font-weight:800;color:#111827;letter-spacing:-0.5px;'>{company_name}</p>
            </div>
            <div style='height:4px;background:{primary_color};'></div>"""
    else:
        header_content = f"""
            <div style='text-align:center;padding:32px 36px 24px;background:#ffffff;border-radius:12px 12px 0 0;'>
              <p style='margin:0;font-size:26px;font-weight:800;color:#111827;letter-spacing:-0.5px;'>{company_name}</p>
            </div>
            <div style='height:4px;background:{primary_color};'></div>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'></head>
<body style='margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif;'>
  <table width='100%' cellpadding='0' cellspacing='0' style='background:#f3f4f6;padding:40px 20px;'>
    <tr><td align='center'>
      <table width='600' cellpadding='0' cellspacing='0' style='max-width:600px;width:100%;border-radius:12px;box-shadow:0 1px 3px rgba(0,0,0,0.1);'>
        <!-- Header: white bg, centered logo + name, brand colour strip -->
        <tr>
          <td style='border-radius:12px 12px 0 0;overflow:hidden;'>
            {header_content}
          </td>
        </tr>
        <!-- Body -->
        <tr>
          <td style='background:#ffffff;padding:36px;border-left:1px solid #e5e7eb;border-right:1px solid #e5e7eb;'>
            <h2 style='margin:0 0 20px 0;color:#111827;font-size:20px;font-weight:600;'>{subject}</h2>
            {paragraphs}
            {cta_html}
          </td>
        </tr>
        <!-- Footer -->
        <tr>
          <td style='background:#f9fafb;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;padding:20px 36px;text-align:center;'>
            <p style='margin:0;color:#9ca3af;font-size:12px;'>This email was sent by DodoDesk.</p>
            <p style='margin:6px 0 0 0;color:#9ca3af;font-size:12px;'>If you did not expect this email, please ignore it.</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""

# Email translations — key phrases in French
EMAIL_TRANSLATIONS = {
    "fr": {
        "view_ticket": "Voir le ticket →",
        "view_now": "Voir maintenant →",
        "subscribe": "S'abonner",
        "go_to_dashboard": "Accéder au tableau de bord →",
        "sla_breach_subject": "⚠ Breach SLA : Ticket #{id} — {title}",
        "sla_breach_cta": "Voir le ticket maintenant →",
        "assigned_subject": "Ticket assigné : #{id} — {title}",
        "assigned_cta": "Voir le ticket →",
        "comment_subject": "Nouveau commentaire : Ticket #{id} — {title}",
        "comment_cta": "Voir le commentaire →",
        "trial_7d_subject": "⏳ Votre essai DodoDesk {plan} se termine dans 7 jours",
        "trial_1d_subject": "🚨 Votre essai DodoDesk {plan} se termine DEMAIN",
        "trial_expired_subject": "Votre essai DodoDesk {plan} est terminé",
        "csat_subject": "Comment s'est passée votre expérience ? — Ticket #{id}",
        "csat_cta": "Donner mon avis →",
        "welcome_subject": "Bienvenue sur DodoDesk, {name} 👋 — commençons",
        "welcome_cta": "Accéder à votre tableau de bord →",
    }
}

def get_user_language(db, email: str) -> str:
    """Look up the language preference for a user by email. Defaults to en."""
    if not db or not email:
        return 'en'
    try:
        from sqlalchemy import text as _t
        row = db.execute(_t("SELECT language FROM users WHERE email=:e LIMIT 1"), {"e": email}).fetchone()
        return (row[0] or 'en') if row else 'en'
    except Exception:
        return 'en'

def translate_email(key: str, lang: str, **kwargs) -> str:
    """Get translated email string, falling back to English key if not found."""
    if lang == 'en' or lang not in EMAIL_TRANSLATIONS:
        return key
    translated = EMAIL_TRANSLATIONS[lang].get(key, key)
    try:
        return translated.format(**kwargs)
    except Exception:
        return translated

def send_email(to: str, subject: str, body: str, cfg: dict = None, cta_url: str = None, cta_label: str = None, db=None, tenant_id: int = None, lang: str = None) -> bool:
    """Send email via Resend API (preferred) or SMTP fallback.
    Returns True if sent, False if all methods failed.
    Never silently swallows failures — always logs the outcome.
    lang: if provided, overrides auto-detection from recipient's profile.
    """
    # Auto-detect language from recipient's profile if not provided
    if not lang and db:
        lang = get_user_language(db, to)
    lang = lang or 'en'
    
    from_addr = (cfg or {}).get("smtp_from") or SMTP_FROM or "DodoDesk <noreply@dodobay.com>"
    reply_to  = (cfg or {}).get("reply_to") or ""

    # Use tenant branding if available
    company_name  = os.getenv("PLATFORM_NAME", "DodoDesk")
    primary_color = os.getenv("PLATFORM_PRIMARY_COLOR", "#059669")
    logo_url      = os.getenv("PLATFORM_LOGO_URL", None)
    support_email = None

    if db and tenant_id:
        try:
            tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
            if tenant:
                company_name  = tenant.name or company_name
                primary_color = tenant.primary_color or primary_color
                logo_url      = tenant.logo_url or logo_url
                support_email = tenant.support_email or None
        except Exception:
            pass
    elif db and cfg and cfg.get("tenant_id"):
        try:
            tenant = db.query(Tenant).filter(Tenant.id == cfg["tenant_id"]).first()
            if tenant:
                company_name  = tenant.name or company_name
                primary_color = tenant.primary_color or primary_color
                logo_url      = tenant.logo_url or logo_url
                support_email = tenant.support_email or None
        except Exception:
            pass

    # Set From display name
    platform_name = os.getenv("PLATFORM_NAME", "DodoDesk")
    if company_name and company_name.lower() not in (platform_name.lower(), "dododesk", "dodo desk"):
        # Client tenant — show their name with DodoDesk attribution
        resend_from_name = f"{company_name} (via DodoDesk)"
    else:
        # DodoDesk itself — just show DodoDesk
        resend_from_name = platform_name
    resend_from_addr = f"{resend_from_name} <{RESEND_FROM.split('<')[-1].rstrip('>') if '<' in RESEND_FROM else 'noreply@dodobay.com'}>"

    # Set Reply-To to tenant's support email if configured
    if support_email and not reply_to:
        reply_to = support_email

    html_body = build_html_email(subject, body, company_name, primary_color, cta_url, cta_label, logo_url)

    # ── Attempt 1: Resend API ─────────────────────────────────────────────
    resend_key = (cfg or {}).get("resend_api_key") or RESEND_API_KEY
    if resend_key:
        import json as _j, http.client as _hc, ssl as _ssl
        candidates = list(dict.fromkeys([resend_from_addr, RESEND_FROM, "DodoDesk <onboarding@resend.dev>"]))
        for from_addr_try in candidates:
            try:
                print(f"\U0001f4e7 Resend: to={to} from={from_addr_try}")
                payload = _j.dumps({
                    "from": from_addr_try, "to": [to],
                    "subject": subject, "html": html_body, "text": body,
                    **({ "reply_to": [reply_to] } if reply_to else {}),
                }).encode()
                ctx  = _ssl.create_default_context()
                conn = _hc.HTTPSConnection("api.resend.com", port=443, timeout=20, context=ctx)
                conn.request("POST", "/emails", body=payload, headers={
                    "Authorization": f"Bearer {resend_key}",
                    "Content-Type": "application/json",
                })
                resp      = conn.getresponse()
                resp_body = resp.read().decode()
                conn.close()
                if resp.status in (200, 201):
                    result = _j.loads(resp_body)
                    print(f"\u2705 Email sent via Resend to {to} id={result.get('id')}")
                    return True
                print(f"\u26a0\ufe0f Resend {resp.status} ({from_addr_try}): {resp_body[:300]}")
            except Exception as e:
                print(f"\u26a0\ufe0f Resend error ({from_addr_try}): {type(e).__name__}: {e}")
        print(f"\u26a0\ufe0f All Resend attempts failed for {to}, trying SMTP...")

    # ── Attempt 2: SMTP ───────────────────────────────────────────────────
    host     = (cfg or {}).get("smtp_host") or SMTP_HOST
    port     = int((cfg or {}).get("smtp_port") or SMTP_PORT or 587)
    user     = (cfg or {}).get("smtp_user") or SMTP_USER
    password = (cfg or {}).get("smtp_pass") or SMTP_PASS

    if not host:
        print(f"\u26a0\ufe0f No email provider configured. Email NOT sent to {to}.")
        print(f"--- Unsent email ---\nTo: {to}\nSubject: {subject}\n{body}\n---")
        return False

    print(f"\U0001f4e7 SMTP: to={to} host={host}:{port}")
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText as _MIMEText
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = from_addr
    msg["To"]      = to
    if reply_to:
        msg["Reply-To"] = reply_to
    msg.attach(_MIMEText(body, "plain"))
    msg.attach(_MIMEText(html_body, "html"))
    try:
        import ssl as _ssl2
        if port == 465:
            ctx = _ssl2.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=30) as server:
                if user: server.login(user, password)
                server.sendmail(from_addr, [to], msg.as_string())
        else:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.ehlo(); server.starttls(); server.ehlo()
                if user: server.login(user, password)
                server.sendmail(from_addr, [to], msg.as_string())
        print(f"\u2705 Email sent via SMTP to {to}")
        return True
    except Exception as e:
        print(f"\u274c SMTP failed for {to}: {type(e).__name__}: {e}")
        return False

def send_email_background(to: str, subject: str, body: str, cta_url: str = None, cta_label: str = None):
    """Non-daemon thread email — survives request completion on Render.
    Use for all critical transactional emails (verification, password reset, invite).
    """
    import threading
    def _run():
        ok = send_email(to, subject, body, cta_url=cta_url, cta_label=cta_label)
        if not ok:
            print(f"\u274c Background email FAILED to {to}: {subject}")
    threading.Thread(target=_run, daemon=False).start()


def send_notification(message: str, cfg: dict = None):
    """Send notification to Slack and/or Teams webhooks.
    Slack expects: {"text": "..."}
    Teams expects: {"type": "message", "attachments": [...]} (Adaptive Card) or legacy {"text": "..."}
    We send the correct format for each.
    """
    slack_url = (cfg or {}).get("slack_webhook_url") or SLACK_WEBHOOK_URL
    teams_url = (cfg or {}).get("teams_webhook_url") or TEAMS_WEBHOOK_URL

    slack_payload = json.dumps({"text": message}).encode("utf-8")

    # Teams payload — supports both legacy MessageCard and new Workflows webhooks
    # New Workflows webhook (post-April 2026) accepts simple text format
    teams_payload = json.dumps({
        "type": "message",
        "attachments": [{
            "contentType": "application/vnd.microsoft.card.adaptive",
            "content": {
                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                "type": "AdaptiveCard",
                "version": "1.2",
                "body": [{
                    "type": "TextBlock",
                    "text": message.replace("<", "&lt;").replace(">", "&gt;"),
                    "wrap": True,
                    "size": "Small",
                    "fontType": "Monospace"
                }]
            }
        }]
    }).encode("utf-8")

    # Also prepare legacy MessageCard format as fallback
    teams_legacy_payload = json.dumps({
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": "059669",
        "summary": "DodoDesk Notification",
        "text": message.replace("*", "**")
    }).encode("utf-8")

    tasks = [
        (slack_url, "Slack", slack_payload, None),
        (teams_url, "Teams", teams_payload, teams_legacy_payload),
    ]
    for url, name, payload, fallback_payload in tasks:
        if not url:
            continue
        try:
            req = urllib.request.Request(
                url, data=payload,
                headers={"Content-Type": "application/json"},
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                if resp.status not in (200, 202, 204):
                    # Try legacy format as fallback for Teams
                    if fallback_payload:
                        req2 = urllib.request.Request(
                            url, data=fallback_payload,
                            headers={"Content-Type": "application/json"},
                            method="POST"
                        )
                        with urllib.request.urlopen(req2, timeout=10) as resp2:
                            if resp2.status in (200, 202, 204):
                                print(f"✅ {name} notification sent (legacy format)")
                            else:
                                print(f"⚠ {name} notification failed: {resp2.status}")
                    else:
                        print(f"⚠ {name} notification failed: {resp.status}")
                else:
                    print(f"✅ {name} notification sent")
        except Exception as e:
            print(f"❌ Failed to send {name} notification: {e}")

def trigger_approval_workflow(db: Session, ticket: "Ticket"):
    """
    Check if an approval workflow matches this ticket's category/type.
    If yes, create TicketApproval records and notify the first approver.
    """
    workflow = db.query(ApprovalWorkflow).filter(
        ApprovalWorkflow.tenant_id == ticket.tenant_id,
        ApprovalWorkflow.is_active == True,
        ApprovalWorkflow.ticket_type == str(ticket.ticket_type),
    ).filter(
        (ApprovalWorkflow.category == None) |
        (ApprovalWorkflow.category == ticket.category)
    ).first()

    if not workflow or not workflow.steps:
        return None

    # Create approval records for each step
    for step in workflow.steps:
        approval = TicketApproval(
            ticket_id=ticket.id,
            workflow_id=workflow.id,
            step_id=step.id,
            step_order=step.step_order,
            step_name=step.name,
            approver_id=step.approver_id,
            approver_role=step.approver_role,
            status="pending" if step.step_order == 1 else "waiting",
        )
        db.add(approval)
    db.flush()

    # Notify first approver
    first_step = workflow.steps[0]
    if first_step.approver_id:
        create_notification(db, first_step.approver_id, ticket.tenant_id,
            "approval_required",
            f"✅ Approval required: {ticket.title}",
            f'Step 1 of {len(workflow.steps)}: {first_step.name}',
            f"/tickets/{ticket.id}")
    elif first_step.approver_role:
        approvers = db.query(User).filter(
            User.tenant_id == ticket.tenant_id,
            User.role == first_step.approver_role,
            User.is_active == True
        ).all()
        for approver in approvers:
            create_notification(db, approver.id, ticket.tenant_id,
                "approval_required",
                f"✅ Approval required: {ticket.title}",
                f'Step 1 of {len(workflow.steps)}: {first_step.name}',
                f"/tickets/{ticket.id}")
    return workflow

def create_notification(db: Session, user_id: int, tenant_id: int, type: str, title: str, body: str, link: str = None):
    """Create an in-app notification for a user — respects user notification preferences."""
    # Map notification type to preference key
    _TYPE_PREF_MAP = {
        'ticket_assigned':   'ticket_assigned',
        'ticket_commented':  'ticket_commented',
        'ticket_status':     'ticket_status_changed',
        'sla_breach':        'ticket_sla_breach',
        'mention':           'ticket_mentioned',
        'approval_required': 'change_approved',
        'approval_approved': 'change_approved',
        'approval_rejected': 'change_rejected',
    }
    pref_key = _TYPE_PREF_MAP.get(type)
    if pref_key:
        try:
            target_user = db.query(User).filter(User.id == user_id).first()
            if target_user and target_user.notification_prefs:
                prefs = json.loads(target_user.notification_prefs)
                if not prefs.get(pref_key, True):
                    return  # User has disabled this notification
        except Exception:
            pass  # On error, default to sending
    notif = Notification(user_id=user_id, tenant_id=tenant_id, type=type, title=title, body=body, link=link)
    db.add(notif)
    db.commit()

def log_ticket_event(db: Session, ticket_id: int, tenant_id: int, actor_id: int,
                     action: str, field: str = None, old_value: str = None,
                     new_value: str = None, note: str = None):
    """Append an audit log entry for a ticket."""
    entry = TicketAuditLog(
        ticket_id=ticket_id, tenant_id=tenant_id, actor_id=actor_id,
        action=action, field=field, old_value=old_value, new_value=new_value, note=note
    )
    db.add(entry)
    # Don't commit here — caller commits

def log_system_event(db: Session, actor: "User", action: str,
                     target_type: str = None, target_id: str = None,
                     target_label: str = None, old_value: str = None,
                     new_value: str = None, ip_address: str = None):
    """Append a system-level audit log entry (user management, settings, plan changes, etc.)."""
    entry = SystemAuditLog(
        tenant_id=actor.tenant_id if actor else None,
        actor_id=actor.id if actor else None,
        actor_email=actor.email if actor else None,
        action=action,
        target_type=target_type,
        target_id=str(target_id) if target_id is not None else None,
        target_label=target_label,
        old_value=str(old_value) if old_value is not None else None,
        new_value=str(new_value) if new_value is not None else None,
        ip_address=ip_address,
    )
    db.add(entry)
    # Don't commit here — caller commits

# =============================================================================
# SLA RULES
# =============================================================================

SLA_RULES = {
    "low":      {"response": 8,  "resolution": 72},
    "medium":   {"response": 4,  "resolution": 48},
    "high":     {"response": 2,  "resolution": 24},
    "critical": {"response": 1,  "resolution": 8},
}

def get_sla_rules(db: Session, tenant_id: int) -> dict:
    """Get SLA rules from DB, falling back to hardcoded defaults."""
    cfg = db.query(SLAConfig).filter(SLAConfig.tenant_id == tenant_id).first()
    if cfg:
        return {
            "low":      {"response": cfg.low_response,      "resolution": cfg.low_resolution},
            "medium":   {"response": cfg.medium_response,   "resolution": cfg.medium_resolution},
            "high":     {"response": cfg.high_response,     "resolution": cfg.high_resolution},
            "critical": {"response": cfg.critical_response, "resolution": cfg.critical_resolution},
        }
    return SLA_RULES

def get_business_hours_config(db: Session, tenant_id: int) -> dict:
    """Get business hours config for a tenant."""
    cfg = db.query(BusinessHoursConfig).filter(BusinessHoursConfig.tenant_id == tenant_id).first()
    if cfg and cfg.enabled:
        return {
            "enabled": True,
            "start_hour": cfg.start_hour,
            "end_hour": cfg.end_hour,
            "working_days": [int(d) for d in cfg.working_days.split(",")],
        }
    return {"enabled": False}

def add_business_hours(start: datetime, hours: int, bh: dict) -> datetime:
    """
    Add `hours` of business time to `start`, skipping non-business hours and weekends.
    bh = {"start_hour": 9, "end_hour": 17, "working_days": [0,1,2,3,4]}
    """
    if not bh.get("enabled"):
        return start + timedelta(hours=hours)

    start_h = bh["start_hour"]
    end_h = bh["end_hour"]
    working_days = bh["working_days"]
    hours_per_day = end_h - start_h
    current = start
    remaining = hours

    # If starting outside business hours, advance to next business start
    def next_business_start(dt):
        # If before start of day
        if dt.weekday() not in working_days:
            dt = dt.replace(hour=start_h, minute=0, second=0, microsecond=0)
            dt += timedelta(days=1)
            while dt.weekday() not in working_days:
                dt += timedelta(days=1)
            return dt
        if dt.hour < start_h:
            return dt.replace(hour=start_h, minute=0, second=0, microsecond=0)
        if dt.hour >= end_h:
            dt = dt.replace(hour=start_h, minute=0, second=0, microsecond=0) + timedelta(days=1)
            while dt.weekday() not in working_days:
                dt += timedelta(days=1)
            return dt
        return dt

    current = next_business_start(current)

    while remaining > 0:
        # Hours left in current business day
        day_end = current.replace(hour=end_h, minute=0, second=0, microsecond=0)
        hours_today = (day_end - current).total_seconds() / 3600

        if remaining <= hours_today:
            current += timedelta(hours=remaining)
            remaining = 0
        else:
            remaining -= hours_today
            # Move to next business day start
            current = current.replace(hour=start_h, minute=0, second=0, microsecond=0) + timedelta(days=1)
            while current.weekday() not in working_days:
                current += timedelta(days=1)

    return current

def compute_sla_deadlines(priority: str, created_at: datetime, db: Session = None, tenant_id: int = None):
    if db and tenant_id:
        rules = get_sla_rules(db, tenant_id).get(priority, {"response": 4, "resolution": 48})
        bh = get_business_hours_config(db, tenant_id)
    else:
        rules = SLA_RULES.get(priority, {"response": 4, "resolution": 48})
        bh = {"enabled": False}

    response_deadline = add_business_hours(created_at, rules["response"], bh)
    resolution_deadline = add_business_hours(created_at, rules["resolution"], bh)
    return response_deadline, resolution_deadline

def compute_sla_status(ticket: Ticket) -> str:
    if ticket.status in ["resolved", "closed"]:
        return "ok"
    now = datetime.utcnow()
    if ticket.sla_resolution_deadline and now > ticket.sla_resolution_deadline:
        return "overdue"
    if ticket.sla_response_deadline and now > ticket.sla_response_deadline:
        return "warning"
    return "ok"

# =============================================================================
# SEED FUNCTION
# =============================================================================

def seed():
    db = SessionLocal()
    # Skip seeding if ANY users exist — database is already set up
    if db.query(User).count() > 0:
        print("✅ Database already seeded — skipping.")
        db.close()
        return
    # Skip if tenant exists with custom logo or color
    existing_tenant = db.query(Tenant).first()
    if existing_tenant and (existing_tenant.logo_url or existing_tenant.primary_color != "#4f46e5"):
        print("✅ Tenant already customised — skipping seed.")
        db.close()
        return

    # Default tenant — only create if doesn't exist
    existing = db.query(Tenant).filter(Tenant.slug == "default").first()
    if not existing:
        tenant = Tenant(name="My Company", slug="default", logo_url=None, primary_color="#4f46e5")
        db.add(tenant)
        db.commit()
        db.refresh(tenant)
        tenant_id = tenant.id
    else:
        tenant_id = existing.id

    # Custom roles
    if not db.query(CustomRole).first():
        admin_role = CustomRole(tenant_id=tenant_id, name="Admin",
                                permissions=json.dumps([p.value for p in Permission]),
                                is_default=True)
        agent_role = CustomRole(tenant_id=tenant_id, name="Agent",
                                permissions=json.dumps([
                                    Permission.VIEW_ALL_TICKETS.value,
                                    Permission.EDIT_TICKETS.value,
                                    Permission.MANAGE_KB.value,
                                    Permission.VIEW_REPORTS.value,
                                    Permission.MANAGE_CANNED.value,
                                    Permission.CREATE_CHANGES.value,
                                    Permission.APPROVE_CHANGES.value,
                                    Permission.MANAGE_ASSETS.value
                                ]))
        readonly_agent_role = CustomRole(tenant_id=tenant_id, name="Read‑only Agent",
                                         permissions=json.dumps([
                                             Permission.VIEW_ALL_TICKETS.value,
                                             Permission.VIEW_REPORTS.value
                                         ]))
        db.add_all([admin_role, agent_role, readonly_agent_role])
        db.commit()
        db.refresh(admin_role)
        db.refresh(agent_role)
        db.refresh(readonly_agent_role)
        admin_role_id = admin_role.id
        agent_role_id = agent_role.id
        readonly_agent_role_id = readonly_agent_role.id
    else:
        roles = db.query(CustomRole).all()
        admin_role_id = next((r.id for r in roles if r.name == "Admin"), None)
        agent_role_id = next((r.id for r in roles if r.name == "Agent"), None)

    # Users
    # Users — only create if they don't exist, never update existing
    seed_admin_email = os.getenv("SEED_ADMIN_EMAIL", "admin@example.com")
    seed_admin_pass  = os.getenv("SEED_ADMIN_PASSWORD", "Admin1234")
    seed_agent_email = os.getenv("SEED_AGENT_EMAIL", "agent@example.com")
    seed_agent_pass  = os.getenv("SEED_AGENT_PASSWORD", "Agent1234")
    seed_emp_email   = os.getenv("SEED_EMPLOYEE_EMAIL", "employee@example.com")
    seed_emp_pass    = os.getenv("SEED_EMPLOYEE_PASSWORD", "Emp1234")

    if not db.query(User).filter(User.email == seed_admin_email).first():
        db.add(User(email=seed_admin_email,
                    hashed_password=get_password_hash(seed_admin_pass),
                    full_name="Admin User",
                    role=UserRole.ADMIN,
                    custom_role_id=admin_role_id,
                    tenant_id=tenant_id))
    if not db.query(User).filter(User.email == seed_emp_email).first():
        db.add(User(email=seed_emp_email,
                    hashed_password=get_password_hash(seed_emp_pass),
                    full_name="Alice Employee",
                    role=UserRole.EMPLOYEE,
                    tenant_id=tenant_id))
    if not db.query(User).filter(User.email == seed_agent_email).first():
        db.add(User(email=seed_agent_email,
                    hashed_password=get_password_hash(seed_agent_pass),
                    full_name="Bob Agent",
                    role=UserRole.AGENT,
                    custom_role_id=agent_role_id,
                    tenant_id=tenant_id))
    # KB
    if not db.query(KBArticle).first():
        db.add(KBArticle(title="How to reset your password",
                         content="1. Go to the login page.\n2. Click 'Forgot password'.\n3. Follow the instructions sent to your email.",
                         category="Account", author_id=2, tenant_id=tenant_id))
        db.add(KBArticle(title="Printer troubleshooting",
                         content="If the printer is offline:\n- Check the power cable.\n- Restart the printer.\n- Ensure it's connected to the network.",
                         category="Hardware", author_id=2, tenant_id=tenant_id))
    # Assets
    if not db.query(Asset).first():
        db.add(Asset(name="Dell Laptop #1", type="hardware", serial_number="SN-001",
                     status="available", notes="15 inch, i7", tenant_id=tenant_id))
        db.add(Asset(name="Microsoft Office License", type="software", serial_number="LIC-001",
                     status="assigned", assigned_to_id=1,
                     license_key="XXXX-XXXX-XXXX", vendor="Microsoft",
                     expiry_date=date.today() + timedelta(days=10), tenant_id=tenant_id))
    # Tickets
    if not db.query(Ticket).first():
        now = datetime.utcnow()
        created_incident = now - timedelta(hours=3)
        resp, reso = compute_sla_deadlines("high", created_incident)
        db.add(Ticket(ticket_type="incident", title="VPN connection issue",
                     description="Unable to connect to VPN from home office.",
                     category="Network", priority="high",
                     status="open", requester_id=1,
                     sla_response_deadline=resp, sla_resolution_deadline=reso,
                     created_at=created_incident, tenant_id=tenant_id))
        created_request = now - timedelta(days=1)
        resp2, reso2 = compute_sla_deadlines("medium", created_request)
        db.add(Ticket(ticket_type="service_request",
                     title="New laptop request",
                     description="I need a developer-grade laptop with 32GB RAM.",
                     category="Hardware", priority="medium",
                     status="pending_approval", requester_id=1,
                     sla_response_deadline=resp2, sla_resolution_deadline=reso2,
                     created_at=created_request, tenant_id=tenant_id))
    # Canned
    if not db.query(CannedResponse).first():
        db.add(CannedResponse(title="Printer offline check",
                              content="Please restart the printer by turning it off and on again.",
                              category="Hardware", author_id=2))
        db.add(CannedResponse(title="Password reset instructions",
                              content="Please visit the forgot password page.",
                              category="Account", author_id=2))
    # Change
    if not db.query(ChangeRequest).first():
        db.add(ChangeRequest(title="Server maintenance reboot",
                             description="Planned reboot of the application server.",
                             risk_level="medium",
                             status="pending_approval",
                             requester_id=1,
                             planned_date=date.today() + timedelta(days=3),
                             tenant_id=tenant_id))
    # Service catalog items
    if not db.query(ServiceCatalogItem).first():
        db.add(ServiceCatalogItem(tenant_id=tenant_id, name="New Laptop",
                                  description="Standard developer laptop (16GB RAM, 512GB SSD)",
                                  category="Hardware", estimated_cost=1500.0,
                                  delivery_time_days=5, approval_required=True))
        db.add(ServiceCatalogItem(tenant_id=tenant_id, name="VPN Access",
                                  description="VPN access for remote workers",
                                  category="Software", approval_required=False))
    db.commit()
    db.close()
    print("Seed data created (if not already present).")

# =============================================================================
# FASTAPI APP & LIFESPAN
# =============================================================================

UPLOAD_DIR = "uploads"
AVATAR_DIR = os.path.join(UPLOAD_DIR, "avatars")

# =============================================================================
# AUTOMATION ENGINE
# =============================================================================

def _evaluate_condition(ticket: "Ticket", cond: dict) -> bool:
    """Evaluate a single condition against a ticket."""
    field = cond.get("field", "")
    operator = cond.get("operator", "is")
    value = str(cond.get("value", "")).lower().strip()

    ticket_val = ""
    if field == "priority":
        ticket_val = str(ticket.priority) if ticket.priority else ""
    elif field == "status":
        ticket_val = (str(ticket.status) if hasattr(ticket.status, "value") else str(ticket.status)) if ticket.status else ""
    elif field == "ticket_type":
        ticket_val = str(ticket.ticket_type) if ticket.ticket_type else ""
    elif field == "category":
        ticket_val = (ticket.category or "").lower()
    elif field == "tag":
        tags = json.loads(ticket.tags) if ticket.tags else []
        if operator == "contains":
            return value in [t.lower() for t in tags]
        return value in [t.lower() for t in tags]
    elif field == "assigned_to":
        ticket_val = str(ticket.assigned_to_id or "")
    elif field == "group_id":
        ticket_val = str(ticket.group_id or "")
    else:
        return True  # unknown field — skip

    ticket_val = ticket_val.lower()

    if operator == "is":
        return ticket_val == value
    elif operator == "is_not":
        return ticket_val != value
    elif operator == "contains":
        return value in ticket_val
    elif operator == "is_empty":
        return not ticket_val
    elif operator == "is_not_empty":
        return bool(ticket_val)
    return True

def _execute_action(ticket: "Ticket", action_def: dict, db: "Session", tenant_id: int) -> None:
    """Execute a single automation action on a ticket."""
    action = action_def.get("action", "")
    value = action_def.get("value", "")

    if action == "assign_to" and value:
        ticket.assigned_to_id = int(value)
    elif action == "assign_round_robin":
        rr = _round_robin_assign(ticket.tenant_id, ticket.group_id, db)
        if rr:
            ticket.assigned_to_id = rr
    elif action == "assign_to_group" and value:
        ticket.group_id = int(value)
    elif action == "set_priority" and value:
        try:
            ticket.priority = str(value).lower()
        except ValueError:
            pass
    elif action == "set_status" and value:
        try:
            ticket.status = str(value).lower()
            if str(ticket.status) == "resolved":
                ticket.resolved_at = ticket.resolved_at or datetime.utcnow()
        except ValueError:
            pass
    elif action == "add_tag" and value:
        existing = json.loads(ticket.tags) if ticket.tags else []
        if value not in existing:
            existing.append(value)
            ticket.tags = json.dumps(existing)
    elif action == "add_comment" and value:
        comment = Comment(ticket_id=ticket.id, author_id=None, body=f"🤖 Automation: {value}", is_internal=True)
        db.add(comment)
    elif action == "close_ticket":
        ticket.status = "closed"
        ticket.resolved_at = ticket.resolved_at or datetime.utcnow()

def run_automation_rules(ticket: "Ticket", trigger: str, db: "Session") -> int:
    """
    Evaluate all active automation rules for a tenant against a ticket.
    Returns count of rules that fired.
    """
    try:
        rules = db.query(AutomationRule).filter(
            AutomationRule.tenant_id == ticket.tenant_id,
            AutomationRule.is_active == True,
            AutomationRule.trigger == trigger
        ).all()
        fired = 0
        for rule in rules:
            try:
                conditions = json.loads(rule.conditions) if rule.conditions else []
                actions = json.loads(rule.actions) if rule.actions else []
                # ALL conditions must pass (AND logic)
                if all(_evaluate_condition(ticket, c) for c in conditions):
                    for action_def in actions:
                        _execute_action(ticket, action_def, db, ticket.tenant_id)
                    rule.run_count = (rule.run_count or 0) + 1
                    rule.last_run_at = datetime.utcnow()
                    fired += 1
            except Exception as e:
                print(f"⚠️ Automation rule {rule.id} error: {e}")
        return fired
    except Exception as e:
        print(f"⚠️ run_automation_rules error: {e}")
        return 0

def check_time_based_automations():
    """Runs every 30 minutes. Executes time_based automation rules."""
    try:
        db = SessionLocal()
        rules = db.query(AutomationRule).filter(
            AutomationRule.is_active == True,
            AutomationRule.trigger == "time_based"
        ).all()
        for rule in rules:
            try:
                conditions = json.loads(rule.conditions) if rule.conditions else []
                actions = json.loads(rule.actions) if rule.actions else []
                # For time-based: conditions include hours_since_update, hours_since_created
                query = db.query(Ticket).filter(Ticket.tenant_id == rule.tenant_id,
                                                Ticket.status.notin_(["resolved", "closed"]))
                for cond in conditions:
                    if cond.get("field") == "hours_since_update":
                        cutoff = datetime.utcnow() - timedelta(hours=int(cond.get("value", 24)))
                        query = query.filter(Ticket.updated_at < cutoff)
                    elif cond.get("field") == "hours_since_created":
                        cutoff = datetime.utcnow() - timedelta(hours=int(cond.get("value", 48)))
                        query = query.filter(Ticket.created_at < cutoff)
                    elif cond.get("field") == "priority":
                        try:
                            query = query.filter(Ticket.priority == str(cond.get("value","")).lower())
                        except ValueError:
                            pass
                tickets = query.all()
                for ticket in tickets:
                    for action_def in actions:
                        _execute_action(ticket, action_def, db, rule.tenant_id)
                    rule.run_count = (rule.run_count or 0) + 1
                    rule.last_run_at = datetime.utcnow()
                db.commit()
            except Exception as e:
                print(f"⚠️ Time automation rule {rule.id}: {e}")
    except Exception as e:
        print(f"⚠️ check_time_based_automations: {e}")
    finally:
        try: db.close()
        except: pass

def auto_close_tickets():
    """Runs every hour.
    Auto-closes tickets that are pending_user for 10+ days with no reply.
    Sends a warning comment at day 7 (3 days before close).
    """
    try:
        db = SessionLocal()
        now = datetime.utcnow()
        warning_cutoff = now - timedelta(days=7)
        close_cutoff   = now - timedelta(days=10)

        # Find tickets pending user reply
        pending_tickets = db.query(Ticket).filter(
            Ticket.status == 'pending_user',
            Ticket.updated_at < warning_cutoff,
        ).all()

        for ticket in pending_tickets:
            age_days = (now - ticket.updated_at).days

            if age_days >= 10:
                # Auto-close
                ticket.status = "closed"
                ticket.updated_at = now
                db.add(Comment(
                    ticket_id=ticket.id,
                    author_id=None,
                    body="🔒 This ticket has been automatically closed after 10 days with no response from the requester. If you still need assistance, please open a new ticket.",
                    is_internal=False,
                ))
                print(f"✅ Auto-closed ticket {ticket.id} (pending_user {age_days} days)")

            elif age_days >= 7:
                # Warning — only send once (check if warning already sent)
                already_warned = db.query(Comment).filter(
                    Comment.ticket_id == ticket.id,
                    Comment.body.like("%will be automatically closed in 3 days%"),
                ).first()
                if not already_warned:
                    db.add(Comment(
                        ticket_id=ticket.id,
                        author_id=None,
                        body="⚠️ We are still waiting for your response on this ticket. If we do not hear back within 3 days, this ticket will be automatically closed. Please reply to keep it open.",
                        is_internal=False,
                    ))
                    print(f"✅ Sent auto-close warning for ticket {ticket.id}")

        db.commit()
    except Exception as e:
        print(f"⚠️ auto_close_tickets: {e}")
    finally:
        try: db.close()
        except: pass

# =============================================================================

def _dispatch_scheduled_reports():
    """Runs every hour. Checks all tenants for scheduled reports due to be sent."""
    from datetime import datetime as _dt
    now = _dt.utcnow()
    current_hour = now.strftime("%H:00")
    current_day  = now.strftime("%A").lower()   # e.g. "monday"
    current_date = now.day                       # day of month for monthly

    db = next(get_db())
    try:
        tenants = db.query(Tenant).filter(Tenant.is_active == True).all()
        for tenant in tenants:
            raw = getattr(tenant, "scheduled_reports", None)
            if not raw:
                continue
            try:
                config = json.loads(raw)
            except Exception:
                continue
            if not config.get("enabled"):
                continue
            if not config.get("recipients"):
                continue

            freq = config.get("frequency", "weekly")
            sched_time = config.get("time", "08:00")[:5]  # HH:MM

            if sched_time != current_hour:
                continue

            should_send = False
            if freq == "daily":
                should_send = True
            elif freq == "weekly":
                should_send = (current_day == config.get("day", "monday").lower())
            elif freq == "monthly":
                should_send = (current_date == int(config.get("day_of_month", 1)))

            if should_send:
                try:
                    _send_scheduled_report(tenant.id)
                except Exception as e:
                    print(f"⚠️ Scheduled report dispatch error tenant {tenant.id}: {e}")
    except Exception as e:
        print(f"⚠️ _dispatch_scheduled_reports error: {e}")
    finally:
        db.close()


def send_trial_expiry_warnings():
    """Runs every 12 hours. Sends warning at 7 days and 1 day.
    Uses billing_notes to track sent warnings — prevents duplicates on restarts."""
    try:
        db = SessionLocal()
        now = datetime.utcnow()
        trial_tenants = db.query(Tenant).filter(
            Tenant.is_active == True,
            Tenant.plan.in_(["free", "essentials", "business", "pro"]),
        ).all()

        for tenant in trial_tenants:
            trial = get_trial_status(tenant)
            if not trial.get("on_trial"):
                continue
            days_left = trial.get("trial_days_remaining", 0)
            plan_label = trial.get("trial_plan_label", tenant.plan)
            admin = db.query(User).filter(
                User.tenant_id == tenant.id,
                User.role.in_(["admin", "super_admin", "platform_admin"]),
                User.is_active == True,
            ).first()
            if not admin:
                continue

            # Track sent warnings to avoid duplicates across Render restarts
            try:
                sent_flags = json.loads(tenant.billing_notes or "{}")
            except Exception:
                sent_flags = {}

            upgrade_url = f"{FRONTEND_URL}/settings?tab=billing"

            if 6 <= days_left <= 8 and not sent_flags.get("warned_7d"):
                send_email_background(
                    to=admin.email,
                    subject=f"\u23f3 Your DodoDesk {plan_label} trial ends in 7 days",
                    body=(f"Hi {admin.full_name},\n\nYour DodoDesk {plan_label} trial for {tenant.name} ends in {days_left} days.\n\nSubscribe now to keep all your {plan_label} features.\n\n— The DodoDesk Team"),
                    cta_url=upgrade_url, cta_label=f"Subscribe to {plan_label}",
                )
                sent_flags["warned_7d"] = now.isoformat()
                tenant.billing_notes = json.dumps(sent_flags)
                db.commit()
                print(f"\U0001f4e7 Trial 7-day warning sent: {admin.email} ({tenant.name})")

            elif 0 < days_left <= 2 and not sent_flags.get("warned_1d"):
                send_email_background(
                    to=admin.email,
                    subject=f"\U0001f6a8 Your DodoDesk {plan_label} trial ends TOMORROW",
                    body=(f"Hi {admin.full_name},\n\nYour DodoDesk {plan_label} trial for {tenant.name} ends tomorrow.\n\nSubscribe today to avoid losing access.\n\n— The DodoDesk Team"),
                    cta_url=upgrade_url, cta_label="Subscribe Now",
                )
                sent_flags["warned_1d"] = now.isoformat()
                tenant.billing_notes = json.dumps(sent_flags)
                db.commit()
                print(f"\U0001f4e7 Trial 1-day warning sent: {admin.email} ({tenant.name})")

            elif trial.get("trial_expired") and days_left <= 0 and not sent_flags.get("expired"):
                if tenant.plan != "free":
                    tenant.plan = "free"
                    tenant.billing_status = "trial_expired"
                sent_flags["expired"] = now.isoformat()
                tenant.billing_notes = json.dumps(sent_flags)
                db.commit()
                send_email_background(
                    to=admin.email,
                    subject=f"Your DodoDesk {plan_label} trial has ended",
                    body=(f"Hi {admin.full_name},\n\nYour 14-day trial has ended. Your account is now on the Free plan. Your data is safe — subscribe anytime to restore access.\n\n— The DodoDesk Team"),
                    cta_url=upgrade_url, cta_label="Restore Full Access",
                )
                print(f"\u2b07\ufe0f Trial expired: {tenant.name}")

        db.close()
    except Exception as e:
        print(f"\u26a0\ufe0f send_trial_expiry_warnings error: {e}")


def check_sla_breaches():
    """
    Runs every 5 minutes. Finds tickets that:
    - Are open or in_progress
    - Have breached their resolution deadline
    - Haven't been notified in the last 4 hours (to avoid spam)
    Sends in-app notification + email + Slack/Teams to assigned agent and all admins.
    """
    db = SessionLocal()
    try:
        now = datetime.utcnow()
        notify_cooldown = now - timedelta(hours=4)

        breached = db.query(Ticket).filter(
            Ticket.status.in_(["open","in_progress"]),
            Ticket.sla_resolution_deadline < now,
            (Ticket.sla_breach_notified_at == None) |
            (Ticket.sla_breach_notified_at < notify_cooldown)
        ).all()

        for ticket in breached:
            cfg = get_email_config(db, ticket.tenant_id)
            priority_str = str(ticket.priority).capitalize()
            deadline_str = ticket.sla_resolution_deadline.strftime("%Y-%m-%d %H:%M UTC")
            ticket_url = f"{FRONTEND_URL}/tickets/{ticket.id}"
            notified_ids = set()

            # Notify assigned agent (if any)
            if ticket.assigned_to_id:
                agent = db.query(User).filter(User.id == ticket.assigned_to_id).first()
                if agent:
                    create_notification(db, agent.id, ticket.tenant_id,
                        "sla_breach",
                        f"⚠ SLA Breached — {ticket.title}",
                        f"Ticket #{ticket.id} has exceeded its resolution SLA. Immediate attention required.",
                        f"/tickets/{ticket.id}")
                    _lang = get_user_language(db, agent.email)
                    if _lang == 'fr':
                        _subj = f"⚠ Breach SLA : Ticket #{ticket.id} — {ticket.title}"
                        _body = f"Bonjour {agent.full_name},\n\nLe ticket #{ticket.id} « {ticket.title} » a dépassé son délai de résolution SLA.\nPriorité : {priority_str}\nDate limite : {deadline_str}\n\nVeuillez traiter ce ticket immédiatement."
                        _cta = "Voir le ticket maintenant →"
                    else:
                        _subj = f"⚠ SLA Breach: Ticket #{ticket.id} — {ticket.title}"
                        _body = f"Hi {agent.full_name},\n\nTicket #{ticket.id} \"{ticket.title}\" has breached its SLA resolution deadline.\nPriority: {priority_str}\nDeadline was: {deadline_str}\n\nPlease action this ticket immediately."
                        _cta = "View Ticket Now →"
                    if _user_wants_notif(db, agent.id, 'email_sla_breach'):
                        print(f"📧 SLA breach email to {agent.email}")
                        send_email(agent.email, _subj, _body, cfg,
                        cta_url=ticket_url, cta_label=_cta,
                        db=None, tenant_id=ticket.tenant_id, lang=_lang)
                    notified_ids.add(agent.id)

            # Notify all admins/super_admins/platform_admins
            admins = db.query(User).filter(
                User.tenant_id == ticket.tenant_id,
                User.role.in_(["admin", "super_admin", "platform_admin"]),
                User.is_active == True
            ).all()
            for admin in admins:
                create_notification(db, admin.id, ticket.tenant_id,
                    "sla_breach",
                    f"⚠ SLA Breached — {ticket.title}",
                    f"Ticket #{ticket.id} has exceeded its resolution SLA.",
                    f"/tickets/{ticket.id}")
                if admin.id not in notified_ids:
                    assigned_name = "Unassigned"
                    if ticket.assigned_to_id:
                        try:
                            au = db.query(User).filter(User.id == ticket.assigned_to_id).first()
                            assigned_name = au.full_name if au else "Unassigned"
                        except Exception:
                            pass
                    send_email(admin.email,
                        f"⚠ SLA Breach: Ticket #{ticket.id} — {ticket.title}",
                        f"Hi {admin.full_name},\n\n"
                        f"Ticket #{ticket.id} \"{ticket.title}\" has breached its SLA resolution deadline.\n"
                        f"Priority: {priority_str}\n"
                        f"Deadline was: {deadline_str}\n"
                        f"Assigned to: {assigned_name}\n\n"
                        f"Please ensure this ticket is actioned immediately.",
                        cfg,
                        cta_url=ticket_url,
                        cta_label="View Ticket →",
                        db=None, tenant_id=ticket.tenant_id)
                    notified_ids.add(admin.id)

            # Slack/Teams alert
            send_notification(
                f"⚠ *SLA Breach*: Ticket #{ticket.id} \"{ticket.title}\" "
                f"(Priority: {priority_str}) has exceeded its resolution deadline.",
                cfg
            )

            ticket.sla_breach_notified_at = now
            db.commit()

        if breached:
            print(f"✅ SLA breach check: notified {len(breached)} ticket(s)")

    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"❌ SLA breach check error: {e}")
    finally:
        db.close()

def check_escalations():
    """
    Runs every 10 minutes. Finds open/in-progress tickets that have been
    idle (no updates) for longer than the escalation rule threshold,
    and reassigns/notifies accordingly.
    """
    db = SessionLocal()
    try:
        now = datetime.utcnow()
        rules = db.query(EscalationRule).filter(EscalationRule.is_active == True).all()

        for rule in rules:
            idle_cutoff = now - timedelta(hours=rule.idle_hours)
            escalation_cooldown = now - timedelta(hours=rule.idle_hours)

            query = db.query(Ticket).filter(
                Ticket.tenant_id == rule.tenant_id,
                Ticket.status.in_(['open','in_progress']),
                Ticket.updated_at < idle_cutoff,
                (Ticket.escalated_at == None) | (Ticket.escalated_at < escalation_cooldown)
            )
            if rule.priority:
                query = query.filter(Ticket.priority == str(rule.priority).lower())

            tickets = query.all()

            for ticket in tickets:
                old_assignee_id = ticket.assigned_to_id
                new_assignee = None

                # Escalate to specific agent
                if rule.escalate_to_id:
                    new_assignee = db.query(User).filter(User.id == rule.escalate_to_id).first()

                # Escalate to any available agent/admin (least loaded)
                elif rule.escalate_to_role:
                    new_assignee = db.query(User).filter(
                        User.tenant_id == rule.tenant_id,
                        User.role == rule.escalate_to_role,
                        User.is_active == True,
                        User.id != old_assignee_id
                    ).first()

                if new_assignee:
                    ticket.assigned_to_id = new_assignee.id
                    ticket.escalated_at = now
                    log_ticket_event(db, ticket.id, ticket.tenant_id, new_assignee.id,
                                     action="assigned",
                                     field="assigned_to",
                                     old_value=db.query(User).filter(User.id == old_assignee_id).first().full_name if old_assignee_id else "Unassigned",
                                     new_value=new_assignee.full_name,
                                     note=f"Auto-escalated by rule: {rule.name}")

                    # Notify new assignee
                    create_notification(db, new_assignee.id, ticket.tenant_id,
                        "ticket_assigned",
                        f"🔺 Escalated to you: Ticket #{ticket.id}",
                        f'"{ticket.title}" has been escalated to you after {rule.idle_hours}h of inactivity.',
                        f"/tickets/{ticket.id}")

                    # Email new assignee
                    cfg = get_email_config(db, ticket.tenant_id)
                    _el = get_user_language(db, new_assignee.email)
                    if _el == 'fr':
                        _es = f"🔺 Ticket escaladé #{ticket.id} : {ticket.title}"
                        _eb = f"Bonjour {new_assignee.full_name},\n\nLe ticket #{ticket.id} « {ticket.title} » vous a été escaladé après {rule.idle_hours}h d'inactivité.\nPriorité : {str(ticket.priority)}"
                        _ec = "Voir le ticket escaladé →"
                    else:
                        _es = f"🔺 Escalated Ticket #{ticket.id}: {ticket.title}"
                        _eb = f'Hi {new_assignee.full_name},\n\nTicket #{ticket.id} "{ticket.title}" has been escalated to you after {rule.idle_hours} hours of inactivity.\nPriority: {str(ticket.priority)}'
                        _ec = "View Escalated Ticket →"
                    send_email(new_assignee.email, _es, _eb, cfg,
                        cta_url=f"{FRONTEND_URL}/tickets/{ticket.id}",
                        cta_label=_ec, db=None, tenant_id=ticket.tenant_id, lang=_el)

                    db.commit()
                    print(f"✅ Escalated ticket #{ticket.id} to {new_assignee.full_name} (rule: {rule.name})")

    except Exception as e:
        print(f"❌ Escalation check error: {e}")
    finally:
        db.close()

def run_migrations():
    """Add any missing columns to existing tables (lightweight migration for SQLite/PostgreSQL)."""
    from sqlalchemy import inspect, text
    inspector = inspect(engine)

    # Add new enum value to PostgreSQL userrole enum type if missing (no-op on SQLite)
    # Note: SQLAlchemy's SAEnum stores the Python enum NAME (e.g. 'SUPER_ADMIN'), not .value
    if not SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
        for enum_name in ("userrole", "UserRole"):
            for value in ("SUPER_ADMIN", "super_admin"):
                try:
                    with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
                        conn.execute(text(f"ALTER TYPE {enum_name} ADD VALUE IF NOT EXISTS '{value}'"))
                        print(f"✅ Migration: ensured '{value}' exists in {enum_name} enum")
                except Exception as e:
                    print(f"⚠️ Migration skipped for {enum_name}.{value}: {e}")

    try:
        existing_columns = {col['name'] for col in inspector.get_columns('users')}
    except Exception:
        return  # table doesn't exist yet — create_all will handle it

    migrations = {
        'status_changed_at': 'TIMESTAMP',
        'current_session_id': 'VARCHAR',
        'pending_email': 'VARCHAR',
        'email_change_token': 'VARCHAR',
        'email_change_expires_at': 'TIMESTAMP',
        'mfa_enabled': 'BOOLEAN DEFAULT FALSE',
        'mfa_secret': 'VARCHAR',
        'mfa_backup_codes': 'TEXT',
        'email_verified': 'BOOLEAN DEFAULT FALSE',
        'password_reset_token': 'VARCHAR',
        'password_reset_expires_at': 'TIMESTAMP',
        'employee_id': 'VARCHAR',
        'country': 'VARCHAR',
    }

    # CRITICAL: Ensure all User model columns exist before any request is served.
    # If any column is missing, every authenticated endpoint returns 500.
    # Run this synchronously at startup, not deferred.
    try:
        with engine.connect() as conn:
            existing_user_cols = {col['name'] for col in inspector.get_columns('users')}
            critical_cols = {
                'current_session_id': 'VARCHAR',
                'pending_email': 'VARCHAR',
                'email_change_token': 'VARCHAR',
                'email_change_expires_at': 'TIMESTAMP',
                'mfa_enabled': 'BOOLEAN DEFAULT FALSE',
                'mfa_secret': 'VARCHAR',
                'mfa_backup_codes': 'TEXT',
                'email_verified': 'BOOLEAN DEFAULT FALSE',
                'password_reset_token': 'VARCHAR',
                'password_reset_expires_at': 'TIMESTAMP',
                'employee_id': 'VARCHAR',
                'country': 'VARCHAR',
                'status_changed_at': 'TIMESTAMP',
            }
            for col, defn in critical_cols.items():
                if col not in existing_user_cols:
                    conn.execute(text(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ CRITICAL migration: users.{col} added")
                    existing_user_cols.add(col)
            print(f"✅ User columns verified ({len(existing_user_cols)} total)")
    except Exception as e:
        print(f"⚠️ CRITICAL user column migration error: {e}")

    # Ensure email_configs columns exist
    try:
        with engine.connect() as conn:
            ec_cols = {col['name'] for col in inspector.get_columns('email_configs')}
            for col, defn in [
                ('email_signature',   "TEXT DEFAULT ''"),
                ('email_footer',      "TEXT DEFAULT ''"),
                ('slack_webhook_url', "VARCHAR DEFAULT ''"),
                ('teams_webhook_url', "VARCHAR DEFAULT ''"),
            ]:
                if col not in ec_cols:
                    conn.execute(text(f"ALTER TABLE email_configs ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ CRITICAL migration: email_configs.{col} added")
    except Exception as e:
        print(f"⚠️ email_configs migration error: {e}")

    # Ensure tenants columns exist
    try:
        with engine.connect() as conn:
            t_cols = {col['name'] for col in inspector.get_columns('tenants')}
            for col, defn in [
                ('dodo_customer_id',     'VARCHAR'),
                ('dodo_subscription_id', 'VARCHAR'),
                ('billing_status',       'VARCHAR'),
                ('plan_renews_at',       'TIMESTAMP'),
            ]:
                if col not in t_cols:
                    conn.execute(text(f"ALTER TABLE tenants ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ CRITICAL migration: tenants.{col} added")
    except Exception as e:
        print(f"⚠️ tenants migration error: {e}")


    try:
        with engine.connect() as conn:
            result = conn.execute(text(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name='users' AND column_name='current_session_id'"
            )).fetchone()
            if result:
                print("✅ Single-session enforcement: current_session_id column confirmed")
            else:
                # Column missing — add it now
                conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS current_session_id VARCHAR"))
                conn.commit()
                print("✅ Single-session enforcement: current_session_id column ADDED (was missing)")
    except Exception as e:
        print(f"⚠️ current_session_id check failed: {e}")

    # Run the user column migrations
    try:
        with engine.connect() as conn:
            existing_cols = {col['name'] for col in inspector.get_columns('users')}
            for col, defn in migrations.items():
                if col not in existing_cols:
                    conn.execute(text(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ Migration: users.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: user columns: {e}")

    # Add 'readonly' value to userrole enum if not already present
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TYPE userrole ADD VALUE IF NOT EXISTS 'readonly'"))
            conn.commit()
            print("✅ Migration: userrole enum updated with 'readonly'")
        except Exception as e:
            print(f"⚠️ userrole enum migration: {e}")

    # Ticket column migrations
    try:
        with engine.connect() as conn:
            ticket_cols = {col['name'] for col in inspector.get_columns('tickets')}
            ticket_migrations = {
                'first_response_at': 'TIMESTAMP',
                'tags': 'TEXT',
                'merged_into_id': 'INTEGER',
                'sla_breach_notified_at': 'TIMESTAMP',
                'escalated_at': 'TIMESTAMP',
                'resolution_note': 'TEXT',
                'resolved_at': 'TIMESTAMP',
                'resolution_kb_article_id': 'INTEGER',
            }
            for col_name, col_type in ticket_migrations.items():
                if col_name not in ticket_cols:
                    try:
                        conn.execute(text(f'ALTER TABLE tickets ADD COLUMN {col_name} {col_type}'))
                        conn.commit()
                        print(f"✅ Migration: added column tickets.{col_name}")
                    except Exception as e:
                        print(f"⚠️ Migration skipped for tickets.{col_name}: {e}")
    except Exception as e:
        print(f"⚠️ Ticket column migration failed: {e}")

    with engine.connect() as conn:
        for col_name, col_type in migrations.items():
            if col_name not in existing_columns:
                try:
                    conn.execute(text(f'ALTER TABLE users ADD COLUMN {col_name} {col_type}'))
                    conn.commit()
                    print(f"✅ Migration: added column users.{col_name}")
                except Exception as e:
                    print(f"⚠️ Migration skipped for users.{col_name}: {e}")

    # Create signup_verifications table if it doesn't exist
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS signup_verifications (
                    id SERIAL PRIMARY KEY,
                    token VARCHAR UNIQUE NOT NULL,
                    email VARCHAR NOT NULL,
                    tenant_id INTEGER REFERENCES tenants(id),
                    user_id INTEGER REFERENCES users(id),
                    plan VARCHAR DEFAULT 'free',
                    expires_at TIMESTAMP NOT NULL,
                    used BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: signup_verifications table ready")
    except Exception as e:
        print(f"⚠️ Migration: signup_verifications: {e}")

    # Attachments — add url column for Cloudinary storage
    try:
        with engine.connect() as conn:
            att_cols = {col['name'] for col in inspector.get_columns('attachments')}
            if 'url' not in att_cols:
                conn.execute(text("ALTER TABLE attachments ADD COLUMN url VARCHAR"))
                conn.commit()
                print("✅ Migration: attachments.url added")
    except Exception as e:
        print(f"⚠️ Migration: attachments.url: {e}")

    # Ticket new columns (due_date, custom_fields_data)
    try:
        with engine.connect() as conn:
            t_cols = {col['name'] for col in inspector.get_columns('tickets')}
            for col, defn in [('due_date', 'TIMESTAMP'), ('custom_fields_data', 'TEXT')]:
                if col not in t_cols:
                    conn.execute(text(f'ALTER TABLE tickets ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: tickets.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: tickets columns: {e}")

    # Custom fields table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS custom_fields (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    name VARCHAR NOT NULL,
                    field_key VARCHAR NOT NULL,
                    field_type VARCHAR DEFAULT 'text',
                    options TEXT,
                    is_required BOOLEAN DEFAULT FALSE,
                    applies_to VARCHAR DEFAULT 'all',
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: custom_fields table ready")
    except Exception as e:
        print(f"⚠️ Migration: custom_fields: {e}")

    # Macros table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS macros (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    name VARCHAR NOT NULL,
                    description VARCHAR,
                    actions TEXT NOT NULL,
                    is_shared BOOLEAN DEFAULT TRUE,
                    created_by_id INTEGER REFERENCES users(id),
                    run_count INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: macros table ready")
    except Exception as e:
        print(f"⚠️ Migration: macros: {e}")

    # Ticket views table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS ticket_views (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    created_by_id INTEGER NOT NULL REFERENCES users(id),
                    name VARCHAR NOT NULL,
                    filters TEXT NOT NULL,
                    is_shared BOOLEAN DEFAULT FALSE,
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: ticket_views table ready")
    except Exception as e:
        print(f"⚠️ Migration: ticket_views: {e}")

    # Ticket tasks table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS ticket_tasks (
                    id SERIAL PRIMARY KEY,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    title VARCHAR NOT NULL,
                    assigned_to_id INTEGER REFERENCES users(id),
                    due_date TIMESTAMP,
                    is_done BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: ticket_tasks table ready")
    except Exception as e:
        print(f"⚠️ Migration: ticket_tasks: {e}")

    # Ticket templates table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS ticket_templates (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    name VARCHAR NOT NULL,
                    ticket_type VARCHAR DEFAULT 'incident',
                    title VARCHAR,
                    description TEXT,
                    category VARCHAR,
                    priority VARCHAR DEFAULT 'medium',
                    tags TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: ticket_templates table ready")
    except Exception as e:
        print(f"⚠️ Migration: ticket_templates: {e}")

    # User new profile columns
    try:
        with engine.connect() as conn:
            user_cols = {col['name'] for col in inspector.get_columns('users')}
            for col, defn in [
                ('phone',               'VARCHAR'),
                ('timezone',            "VARCHAR DEFAULT 'UTC'"),
                ('availability',        "VARCHAR DEFAULT 'online'"),
                ('notification_prefs',  'TEXT'),
            ]:
                if col not in user_cols:
                    conn.execute(text(f'ALTER TABLE users ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: users.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: users columns: {e}")

    # Tenant new branding columns
    try:
        with engine.connect() as conn:
            tenant_cols = {col['name'] for col in inspector.get_columns('tenants')}
            for col, defn in [
                ('custom_css',   'TEXT'),
                ('favicon_url',  'VARCHAR'),
            ]:
                if col not in tenant_cols:
                    conn.execute(text(f'ALTER TABLE tenants ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: tenants.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: tenants columns: {e}")

    # EmailConfig new columns
    try:
        with engine.connect() as conn:
            ec_cols = {col['name'] for col in inspector.get_columns('email_configs')}
            for col, defn in [
                ('email_signature',    "TEXT DEFAULT ''"),
                ('email_footer',       "TEXT DEFAULT ''"),
                ('slack_webhook_url',  "VARCHAR DEFAULT ''"),
                ('teams_webhook_url',  "VARCHAR DEFAULT ''"),
            ]:
                if col not in ec_cols:
                    conn.execute(text(f'ALTER TABLE email_configs ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: email_configs.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: email_configs columns: {e}")

    # One-time fix: clean corrupted logo_url values where the API base URL was
    # accidentally prepended to a Cloudinary URL, producing broken URLs like:
    # "https://dodo-desk-api.onrender.comhttps//res.cloudinary.com/..."
    # Also fixes partial Cloudinary public IDs (e.g. "tenant_26_logo") that are
    # missing the full https://res.cloudinary.com/... prefix.
    try:
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "")
        with engine.connect() as conn:
            rows = conn.execute(text(
                "SELECT id, logo_url FROM tenants WHERE logo_url IS NOT NULL"
            )).fetchall()
            fixed = 0
            for row in rows:
                url = row[1]
                if not url:
                    continue
                clean_url = None
                # Fix 1: double-URL corruption
                if 'cloudinary.com' in url and not url.startswith('https://res.cloudinary.com'):
                    idx = url.find('https://res.cloudinary.com')
                    if idx == -1:
                        idx = url.find('http://res.cloudinary.com')
                    if idx > 0:
                        clean_url = url[idx:]
                # Fix 2: partial public ID (no http/https prefix)
                elif cloud_name and not url.startswith('http') and not url.startswith('/'):
                    clean_url = f"https://res.cloudinary.com/{cloud_name}/image/upload/{url}"
                # Fix 3: relative /logos/ path — can't fix without file, skip
                if clean_url:
                    conn.execute(text("UPDATE tenants SET logo_url = :url WHERE id = :id"),
                                 {"url": clean_url, "id": row[0]})
                    fixed += 1
                    print(f"✅ Migration: fixed logo_url for tenant {row[0]}: {url[:40]} → {clean_url[:60]}")
            if fixed:
                conn.commit()
                print(f"✅ Migration: fixed {fixed} logo_url(s)")
            else:
                print("✅ Migration: no broken logo_url values found")
    except Exception as e:
        print(f"⚠️ Migration: logo_url cleanup: {e}")

    # One-time backfill: normalise KB article and Catalog item categories
    # to match the shared TICKET_CATEGORIES list. Blank or non-matching
    # values are set to "Other" so the new Category Focus report groups cleanly.
    try:
        VALID_CATEGORIES = {
            "Hardware", "Software", "Network", "Account", "Email",
            "Security", "Printer", "Mobile Device", "Cloud Services",
            "Telephony", "Other"
        }
        with engine.connect() as conn:
            # KB articles
            kb_rows = conn.execute(text(
                "SELECT id, category FROM kb_articles WHERE category IS NULL OR category = '' "
                "OR category NOT IN :valid"
            ), {"valid": tuple(VALID_CATEGORIES)}).fetchall()
            for row in kb_rows:
                conn.execute(text("UPDATE kb_articles SET category = 'Other' WHERE id = :id"), {"id": row[0]})
            if kb_rows:
                conn.commit()
                print(f"✅ Migration: backfilled {len(kb_rows)} kb_articles.category → 'Other'")

            # Catalog items
            cat_rows = conn.execute(text(
                "SELECT id, category FROM service_catalog_items WHERE category IS NULL OR category = '' "
                "OR category NOT IN :valid"
            ), {"valid": tuple(VALID_CATEGORIES)}).fetchall()
            for row in cat_rows:
                conn.execute(text("UPDATE service_catalog_items SET category = 'Other' WHERE id = :id"), {"id": row[0]})
            if cat_rows:
                conn.commit()
                print(f"✅ Migration: backfilled {len(cat_rows)} service_catalog_items.category → 'Other'")
    except Exception as e:
        print(f"⚠️ Migration: category backfill: {e}")

    # Canned response new columns
    try:
        with engine.connect() as conn:
            cr_cols = {col['name'] for col in inspector.get_columns('canned_responses')}
            for col, defn in [
                ('tenant_id',  'INTEGER'),
                ('visibility', "VARCHAR DEFAULT 'all'"),
                ('group_id',   'INTEGER'),
                ('use_count',  'INTEGER DEFAULT 0'),
                ('sort_order', 'INTEGER DEFAULT 0'),
            ]:
                if col not in cr_cols:
                    conn.execute(text(f'ALTER TABLE canned_responses ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: canned_responses.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: canned_responses columns: {e}")

    # Change request new columns
    try:
        with engine.connect() as conn:
            chg_cols = {col['name'] for col in inspector.get_columns('change_requests')}
            for col, defn in [
                ('change_type',         "VARCHAR DEFAULT 'normal'"),
                ('risk_score',          'INTEGER'),
                ('owner_id',            'INTEGER'),
                ('start_date',          'TIMESTAMP'),
                ('end_date',            'TIMESTAMP'),
                ('impact',              'TEXT'),
                ('rollback_plan',       'TEXT'),
                ('test_plan',           'TEXT'),
                ('cab_members',         'TEXT'),
                ('linked_ticket_ids',   'TEXT'),
                ('linked_asset_ids',    'TEXT'),
                ('post_review_notes',   'TEXT'),
                ('post_review_at',      'TIMESTAMP'),
            ]:
                if col not in chg_cols:
                    conn.execute(text(f'ALTER TABLE change_requests ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: change_requests.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: change_requests columns: {e}")

    # Change request enum types — Postgres native enums don't auto-update when the
    # Python enum gains new values, so we must explicitly ALTER TYPE ... ADD VALUE.
    # SQLAlchemy's SAEnum stores the Python enum MEMBER NAME (uppercase, e.g. "DRAFT"),
    # not its .value (lowercase "draft") — so the Postgres enum values must be uppercase
    # to match what SQLAlchemy actually sends on INSERT.
    # Each ADD VALUE must run in its own auto-commit connection (cannot be inside
    # a multi-statement transaction in Postgres).
    try:
        with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
            for status_value in ["DRAFT", "IN_REVIEW", "SCHEDULED", "IN_PROGRESS", "CANCELLED", "FAILED"]:
                try:
                    conn.execute(text(f"ALTER TYPE changestatus ADD VALUE IF NOT EXISTS '{status_value}'"))
                    print(f"✅ Migration: changestatus enum value '{status_value}' ensured")
                except Exception as inner_e:
                    print(f"⚠️ Migration: changestatus value '{status_value}': {inner_e}")
    except Exception as e:
        print(f"⚠️ Migration: changestatus enum type: {e}")

    try:
        with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
            try:
                conn.execute(text("ALTER TYPE changerisk ADD VALUE IF NOT EXISTS 'CRITICAL'"))
                print("✅ Migration: changerisk enum value 'CRITICAL' ensured")
            except Exception as inner_e:
                print(f"⚠️ Migration: changerisk value 'CRITICAL': {inner_e}")
    except Exception as e:
        print(f"⚠️ Migration: changerisk enum type: {e}")

    # Backfill any existing changes still on the old default 'pending_approval'
    # status with no explicit submission yet — leave as-is, only new rows default to draft now.

    # Change tasks table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS change_tasks (
                    id SERIAL PRIMARY KEY,
                    change_id INTEGER NOT NULL REFERENCES change_requests(id) ON DELETE CASCADE,
                    title VARCHAR NOT NULL,
                    assigned_to_id INTEGER REFERENCES users(id),
                    is_done BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: change_tasks table ready")
    except Exception as e:
        print(f"⚠️ Migration: change_tasks: {e}")

    # Change comments table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS change_comments (
                    id SERIAL PRIMARY KEY,
                    change_id INTEGER NOT NULL REFERENCES change_requests(id) ON DELETE CASCADE,
                    author_id INTEGER NOT NULL REFERENCES users(id),
                    body TEXT NOT NULL,
                    is_internal BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: change_comments table ready")
    except Exception as e:
        print(f"⚠️ Migration: change_comments: {e}")

    # Asset new columns
    try:
        with engine.connect() as conn:
            asset_cols = {col['name'] for col in inspector.get_columns('assets')}
            for col, defn in [
                ('model', 'VARCHAR'),
                ('location', 'VARCHAR'),
                ('purchase_cost', 'FLOAT'),
                ('warranty_expiry', 'DATE'),
                ('contract_number', 'VARCHAR'),
                ('quantity', 'INTEGER DEFAULT 1'),
                ('seats_total', 'INTEGER'),
                ('seats_used', 'INTEGER DEFAULT 0'),
                ('maintenance_date', 'TIMESTAMP'),
                ('parent_asset_id', 'INTEGER'),
                ('tag_number', 'VARCHAR'),
                ('custom_fields_data', 'TEXT'),
            ]:
                if col not in asset_cols:
                    conn.execute(text(f'ALTER TABLE assets ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: assets.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: assets columns: {e}")

    # Asset model options table — admin-managed dropdown per asset type
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS asset_model_options (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
                    asset_type VARCHAR NOT NULL,
                    label VARCHAR NOT NULL,
                    sort_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text(
                "CREATE INDEX IF NOT EXISTS idx_amo_tenant_type "
                "ON asset_model_options(tenant_id, asset_type)"
            ))
            print("✅ Migration: asset_model_options table ready")
    except Exception as e:
        print(f"⚠️ Migration: asset_model_options table: {e}")

    # Convert ALL enum columns to lowercase VARCHAR — permanent fix for SAEnum case mismatch
    enum_conversions = [
        # (table, column, default_value)
        ('tickets',  'status',      'open'),
        ('tickets',  'priority',    'medium'),
        ('tickets',  'ticket_type', 'incident'),
        ('users',    'role',        'employee'),
        ('changes',  'status',      'draft'),
        ('changes',  'risk_level',  'medium'),
        ('changes',  'change_type', 'normal'),
        ('change_requests',  'status',      'draft'),
        ('change_requests',  'risk_level',  'medium'),
        ('change_requests',  'change_type', 'normal'),
        ('assets',   'type',        'hardware'),
        ('assets',   'status',      'available'),
    ]
    try:
        with engine.begin() as conn:
            for table, col, default in enum_conversions:
                try:
                    col_type = conn.execute(text(
                        f"SELECT data_type FROM information_schema.columns "
                        f"WHERE table_name='{table}' AND column_name='{col}'"
                    )).scalar()
                    if col_type and col_type != 'character varying':
                        conn.execute(text(
                            f"ALTER TABLE {table} ALTER COLUMN {col} "
                            f"TYPE VARCHAR USING lower({col}::text)"
                        ))
                        print(f"✅ Migration: {table}.{col} → VARCHAR (lowercase)")
                except Exception as e:
                    print(f"⚠️ Migration: {table}.{col} conversion: {e}")
    except Exception as e:
        print(f"⚠️ Enum→VARCHAR migration error: {e}")

    # Ensure assets table exists (fallback if Base.metadata.create_all missed it)
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS assets (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
                    name VARCHAR NOT NULL,
                    type VARCHAR NOT NULL DEFAULT 'hardware',
                    model VARCHAR,
                    serial_number VARCHAR,
                    status VARCHAR NOT NULL DEFAULT 'available',
                    assigned_to_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    purchase_date DATE,
                    purchase_cost NUMERIC(10,2),
                    vendor VARCHAR,
                    license_key VARCHAR,
                    expiry_date DATE,
                    warranty_expiry DATE,
                    maintenance_date DATE,
                    contract_number VARCHAR,
                    location VARCHAR,
                    notes TEXT,
                    tag_number VARCHAR,
                    quantity INTEGER DEFAULT 1,
                    seats_total INTEGER,
                    seats_used INTEGER DEFAULT 0,
                    parent_asset_id INTEGER REFERENCES assets(id) ON DELETE SET NULL,
                    custom_fields_data TEXT,
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text("CREATE INDEX IF NOT EXISTS idx_assets_tenant ON assets(tenant_id)"))
            print("✅ Migration: assets table ready")
    except Exception as e:
        print(f"⚠️ Migration: assets table: {e}")

    # Convert assets.type and assets.status from enum to lowercase VARCHAR
    try:
        with engine.begin() as conn:
            type_col = conn.execute(text(
                "SELECT data_type FROM information_schema.columns "
                "WHERE table_name='assets' AND column_name='type'"
            )).scalar()
            if type_col and type_col != 'character varying':
                conn.execute(text(
                    "ALTER TABLE assets ALTER COLUMN type TYPE VARCHAR USING lower(type::text)"
                ))
                print("✅ Migration: assets.type converted to VARCHAR (lowercase)")
            status_col = conn.execute(text(
                "SELECT data_type FROM information_schema.columns "
                "WHERE table_name='assets' AND column_name='status'"
            )).scalar()
            if status_col and status_col != 'character varying':
                conn.execute(text(
                    "ALTER TABLE assets ALTER COLUMN status TYPE VARCHAR USING lower(status::text)"
                ))
                print("✅ Migration: assets.status converted to VARCHAR (lowercase)")
    except Exception as e:
        print(f"⚠️ assets enum→varchar migration: {e}")

    # Convert asset_model_options.asset_type from enum to varchar (permanent fix for case mismatch)
    try:
        with engine.begin() as conn:
            col_type = conn.execute(text(
                "SELECT data_type FROM information_schema.columns "
                "WHERE table_name='asset_model_options' AND column_name='asset_type'"
            )).scalar()
            if col_type and col_type != 'character varying':
                # Convert enum column to varchar, normalising to lowercase
                conn.execute(text(
                    "ALTER TABLE asset_model_options "
                    "ALTER COLUMN asset_type TYPE VARCHAR USING lower(asset_type::text)"
                ))
                print("✅ Migration: asset_model_options.asset_type converted from enum to VARCHAR")
            else:
                print("✅ Migration: asset_model_options.asset_type already VARCHAR")
    except Exception as e:
        print(f"⚠️ asset_model_options varchar migration: {e}")

    # Seed sensible defaults for any tenant that has none yet
    try:
        with engine.begin() as seed_conn:
            DEFAULT_MODEL_OPTIONS = {
                "hardware":   ["Dell Latitude 5420", "Dell OptiPlex 7090", "HP EliteBook 840",
                               "HP ProBook 450", "Lenovo ThinkPad T14", "Lenovo ThinkCentre M70q",
                               "Apple MacBook Pro 14\"", "Apple MacBook Air M2", "Apple iMac 24\""],
                "software":   ["Microsoft Office 365", "Adobe Creative Cloud", "Windows 11 Pro",
                               "AutoCAD", "Slack", "Zoom"],
                "network":    ["Cisco Catalyst 2960", "Ubiquiti UniFi Switch", "TP-Link Switch",
                               "Fortinet FortiGate", "Netgear Router"],
                "mobile":     ["Apple iPhone 14", "Apple iPhone 15", "Samsung Galaxy S23",
                               "Samsung Galaxy A54", "Google Pixel 8"],
                "peripheral": ["Dell UltraSharp Monitor", "HP LaserJet Printer", "Logitech MX Keys",
                               "Logitech MX Master Mouse", "Jabra Headset"],
                "saas":       ["Salesforce", "HubSpot", "Google Workspace", "DodoDesk", "Notion"],
                "cloud":      ["AWS EC2 Instance", "Azure VM", "Google Cloud Compute", "DigitalOcean Droplet"],
                "other":      ["Other / Custom"],
            }
            tenant_ids = [row[0] for row in seed_conn.execute(text("SELECT id FROM tenants")).fetchall()]

            # Get enum values
            try:
                enum_vals = [r[0] for r in seed_conn.execute(text(
                    "SELECT enumlabel FROM pg_enum JOIN pg_type ON pg_type.oid = pg_enum.enumtypid "
                    "WHERE pg_type.typname = 'assettype'"
                )).fetchall()]
                enum_map = {v.lower(): v for v in enum_vals}
            except Exception:
                enum_map = {}

        for tid in tenant_ids:
            # Use a fresh connection per tenant to avoid aborted transaction bleed
            try:
                with engine.begin() as tconn:
                    existing = tconn.execute(text(
                        "SELECT COUNT(*) FROM asset_model_options WHERE tenant_id = :tid"
                    ), {"tid": tid}).scalar()
                    if existing == 0:
                        seeded = 0
                        for asset_type, labels in DEFAULT_MODEL_OPTIONS.items():
                            db_val = enum_map.get(asset_type.lower(), asset_type)
                            for i, label in enumerate(labels):
                                try:
                                    tconn.execute(text(
                                        "INSERT INTO asset_model_options (tenant_id, asset_type, label, sort_order) "
                                        "VALUES (:tid, :atype::assettype, :label, :sort)"
                                    ), {"tid": tid, "atype": db_val, "label": label, "sort": i})
                                    seeded += 1
                                except Exception:
                                    pass
                        if seeded > 0:
                            print(f"✅ Migration: seeded {seeded} asset model options for tenant {tid}")
            except Exception as e:
                print(f"⚠️ Migration: asset_model_options tenant {tid}: {e}")
    except Exception as e:
        print(f"⚠️ Migration: asset_model_options: {e}")

    # Service catalog new columns
    try:
        with engine.connect() as conn:
            cat_cols = {col['name'] for col in inspector.get_columns('service_catalog_items')}
            for col, defn in [
                ('sort_order', 'INTEGER DEFAULT 0'),
                ('icon', 'VARCHAR'),
                ('request_form_fields', 'TEXT'),
                ('visibility', "VARCHAR DEFAULT 'all'"),
                ('sla_hours', 'INTEGER'),
                ('request_count', 'INTEGER DEFAULT 0'),
                ('fulfillment_checklist', 'TEXT'),
                ('approval_workflow_id', 'INTEGER'),
            ]:
                if col not in cat_cols:
                    conn.execute(text(f'ALTER TABLE service_catalog_items ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: service_catalog_items.{col} added")
    except Exception as e:
        print(f"⚠️ Migration: service_catalog_items: {e}")

    # Problem links table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS problem_links (
                    id SERIAL PRIMARY KEY,
                    problem_ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    incident_ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE
                )
            """))
            conn.commit()
            print("✅ Migration: problem_links table ready")
    except Exception as e:
        print(f"⚠️ Migration: problem_links: {e}")

    # email_configs reply_to column
    try:
        with engine.connect() as conn:
            cols = {col['name'] for col in inspector.get_columns('email_configs')}
            if 'reply_to' not in cols:
                conn.execute(text("ALTER TABLE email_configs ADD COLUMN reply_to VARCHAR DEFAULT ''"))
                conn.commit()
                print("✅ Migration: email_configs.reply_to added")
    except Exception as e:
        print(f"⚠️ Migration: email_configs.reply_to: {e}")

    # KB article new columns (status, version, view_count + new features)
    try:
        with engine.connect() as conn:
            kb_cols = {col['name'] for col in inspector.get_columns('kb_articles')}
            for col, defn in [
                ('status', "VARCHAR DEFAULT 'published'"),
                ('version', 'INTEGER DEFAULT 1'),
                ('view_count', 'INTEGER DEFAULT 0'),
                ('helpful_count', 'INTEGER DEFAULT 0'),
                ('not_helpful_count', 'INTEGER DEFAULT 0'),
                ('tags', 'TEXT'),
                ('folder', 'VARCHAR'),
                ('visibility', "VARCHAR DEFAULT 'all'"),
                ('review_date', 'TIMESTAMP'),
                ('sort_order', 'INTEGER DEFAULT 0'),
                ('custom_fields_data', 'TEXT'),
            ]:
                if col not in kb_cols:
                    conn.execute(text(f'ALTER TABLE kb_articles ADD COLUMN {col} {defn}'))
                    conn.commit()
                    print(f"✅ Migration: added kb_articles.{col}")
    except Exception as e:
        print(f"⚠️ Migration: kb_articles columns: {e}")

    # KB versions table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS kb_versions (
                    id SERIAL PRIMARY KEY,
                    article_id INTEGER NOT NULL REFERENCES kb_articles(id) ON DELETE CASCADE,
                    version_number INTEGER NOT NULL,
                    title VARCHAR NOT NULL,
                    content TEXT NOT NULL,
                    category VARCHAR,
                    status VARCHAR,
                    change_note VARCHAR,
                    edited_by_id INTEGER REFERENCES users(id),
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: kb_versions table ready")
    except Exception as e:
        print(f"⚠️ Migration: kb_versions: {e}")

    # Automation rules table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS automation_rules (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
                    name VARCHAR NOT NULL,
                    description VARCHAR,
                    is_active BOOLEAN DEFAULT TRUE,
                    trigger VARCHAR NOT NULL,
                    conditions TEXT,
                    actions TEXT NOT NULL,
                    run_count INTEGER DEFAULT 0,
                    last_run_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: automation_rules table ready")
    except Exception as e:
        print(f"⚠️ Migration: automation_rules: {e}")

    # Admin multi-tenant access table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS admin_tenant_access (
                    id SERIAL PRIMARY KEY,
                    admin_user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
                    granted_by_id INTEGER REFERENCES users(id),
                    granted_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(admin_user_id, tenant_id)
                )
            """))
            conn.commit()
            print("✅ Migration: admin_tenant_access table ready")
    except Exception as e:
        print(f"⚠️ Migration: admin_tenant_access: {e}")

    # Asset history table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS asset_history (
                    id SERIAL PRIMARY KEY,
                    asset_id INTEGER NOT NULL REFERENCES assets(id) ON DELETE CASCADE,
                    action VARCHAR NOT NULL,
                    from_user_id INTEGER REFERENCES users(id),
                    to_user_id INTEGER REFERENCES users(id),
                    note VARCHAR,
                    changed_by_id INTEGER REFERENCES users(id),
                    changed_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: asset_history table ready")
    except Exception as e:
        print(f"⚠️ Migration: asset_history: {e}")

    # Time entries table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS time_entries (
                    id SERIAL PRIMARY KEY,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    agent_id INTEGER NOT NULL REFERENCES users(id),
                    minutes INTEGER NOT NULL,
                    note VARCHAR,
                    logged_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: time_entries table ready")
    except Exception as e:
        print(f"⚠️ Migration: time_entries: {e}")

    # Ticket links table (parent-child)
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS ticket_links (
                    id SERIAL PRIMARY KEY,
                    parent_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    child_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    UNIQUE(parent_id, child_id)
                )
            """))
            conn.commit()
            print("✅ Migration: ticket_links table ready")
    except Exception as e:
        print(f"⚠️ Migration: ticket_links: {e}")

    # Groups and group members tables
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS groups (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    name VARCHAR NOT NULL,
                    description VARCHAR,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS group_members (
                    id SERIAL PRIMARY KEY,
                    group_id INTEGER NOT NULL REFERENCES groups(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE
                )
            """))
            # Add group_id to tickets after groups table exists
            try:
                conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS group_id INTEGER REFERENCES groups(id)"))
            except Exception:
                pass
            conn.commit()
            print("✅ Migration: groups tables ready")
    except Exception as e:
        print(f"⚠️ Migration: groups: {e}")

    # System audit log table
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS system_audit_logs (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER REFERENCES tenants(id),
                    actor_id INTEGER REFERENCES users(id),
                    actor_email VARCHAR,
                    action VARCHAR NOT NULL,
                    target_type VARCHAR,
                    target_id VARCHAR,
                    target_label VARCHAR,
                    old_value VARCHAR,
                    new_value VARCHAR,
                    ip_address VARCHAR,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()
            print("✅ Migration: system_audit_logs table ready")
    except Exception as e:
        print(f"⚠️ Migration: system_audit_logs: {e}")

    # Comments — is_internal (private notes)
    try:
        comment_columns = {col['name'] for col in inspector.get_columns('comments')}
        if 'is_internal' not in comment_columns:
            with engine.connect() as conn:
                conn.execute(text('ALTER TABLE comments ADD COLUMN is_internal BOOLEAN DEFAULT FALSE'))
                conn.commit()
                print("✅ Migration: added column comments.is_internal")
    except Exception as e:
        print(f"⚠️ Migration: comments.is_internal: {e}")

    # AI chatbot tables
    try:
        existing_tables = inspector.get_table_names()
        if "chat_sessions" not in existing_tables:
            with engine.connect() as conn:
                conn.execute(text("""CREATE TABLE chat_sessions (
                    id SERIAL PRIMARY KEY,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id),
                    user_id   INTEGER NOT NULL REFERENCES users(id),
                    title     VARCHAR DEFAULT 'New conversation',
                    created_at TIMESTAMP DEFAULT NOW(),
                    updated_at TIMESTAMP DEFAULT NOW()
                )"""))
                conn.commit()
                print("✅ Migration: chat_sessions table created")
        if "chat_messages" not in existing_tables:
            with engine.connect() as conn:
                conn.execute(text("""CREATE TABLE chat_messages (
                    id         SERIAL PRIMARY KEY,
                    session_id INTEGER NOT NULL REFERENCES chat_sessions(id) ON DELETE CASCADE,
                    role       VARCHAR NOT NULL,
                    content    TEXT    NOT NULL,
                    tool_calls TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )"""))
                conn.commit()
                print("✅ Migration: chat_messages table created")
    except Exception as e:
        print(f"⚠️ Migration: chat tables: {e}")

    # Ticket watchers
    try:
        existing_tables = inspector.get_table_names()
        if "ticket_watchers" not in existing_tables:
            with engine.connect() as conn:
                conn.execute(text("""CREATE TABLE ticket_watchers (
                    id SERIAL PRIMARY KEY,
                    ticket_id INTEGER NOT NULL REFERENCES tickets(id) ON DELETE CASCADE,
                    user_id   INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    tenant_id INTEGER NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
                    created_at TIMESTAMP DEFAULT NOW(),
                    CONSTRAINT uq_ticket_watcher UNIQUE (ticket_id, user_id)
                )"""))
                conn.commit()
                print("✅ Migration: ticket_watchers table created")
    except Exception as e:
        print(f"⚠️ Migration: ticket_watchers: {e}")

    # Service catalog items — is_featured
    try:
        sc_columns = {col['name'] for col in inspector.get_columns('service_catalog_items')}
        if 'is_featured' not in sc_columns:
            with engine.connect() as conn:
                conn.execute(text('ALTER TABLE service_catalog_items ADD COLUMN is_featured BOOLEAN DEFAULT FALSE'))
                conn.commit()
                print("✅ Migration: added column service_catalog_items.is_featured")
    except Exception as e:
        print(f"⚠️ Migration skipped for service_catalog_items.is_featured: {e}")

    # Tenants — security config columns
    try:
        tenant_columns = {col['name'] for col in inspector.get_columns('tenants')}
        tenant_migrations = {
            'plan': "VARCHAR DEFAULT 'free'",
            'dodo_customer_id': 'VARCHAR',
            'dodo_subscription_id': 'VARCHAR',
            'billing_status': 'VARCHAR',
            'plan_renews_at': 'TIMESTAMP',
            'mfa_enabled': 'BOOLEAN DEFAULT FALSE',
            'mfa_required': 'BOOLEAN DEFAULT FALSE',
            'sso_enabled': 'BOOLEAN DEFAULT FALSE',
            'sso_provider': "VARCHAR DEFAULT 'google'",
            'sso_client_id': 'VARCHAR',
            'sso_client_secret': 'VARCHAR',
            'sso_domain': 'VARCHAR',
            'sso_sso_url': 'VARCHAR',
            'saml_cert': 'TEXT',
            'ip_whitelist': 'TEXT',
            'scheduled_reports': 'TEXT',
            'billing_notes': 'TEXT',
            'onboarding_emails': 'TEXT',
            'sso_tenant_id': 'VARCHAR',
        }
        with engine.connect() as conn:
            for col_name, col_type in tenant_migrations.items():
                if col_name not in tenant_columns:
                    try:
                        conn.execute(text(f'ALTER TABLE tenants ADD COLUMN {col_name} {col_type}'))
                        conn.commit()
                        print(f"✅ Migration: added column tenants.{col_name}")
                    except Exception as e:
                        print(f"⚠️ Migration skipped for tenants.{col_name}: {e}")
    except Exception as e:
        print(f"⚠️ Tenant migration check failed: {e}")



# =============================================================================
# ONBOARDING EMAIL SEQUENCE — Day 0, 3, 7, 10
# =============================================================================
def _send_onboarding_sequence():
    """Runs every 6 hours. Sends onboarding emails at Day 0/3/7/10 after signup.
    Uses tenant.onboarding_emails JSON to track which emails have been sent."""
    try:
        db = SessionLocal()
        now = datetime.utcnow()
        FRONTEND = os.getenv("FRONTEND_URL", "https://dododesk.dodobay.com")

        tenants = db.query(Tenant).filter(Tenant.is_active == True).all()

        for tenant in tenants:
            try:
                flags = json.loads(tenant.onboarding_emails or "{}") if tenant.onboarding_emails else {}
            except Exception:
                flags = {}

            admin = db.query(User).filter(
                User.tenant_id == tenant.id,
                User.role.in_(["admin", "super_admin", "platform_admin"]),
                User.is_active == True,
            ).order_by(User.id).first()
            if not admin or not tenant.created_at:
                continue

            days_since = (now - tenant.created_at).days
            name = (admin.full_name or "").split()[0] or "there"
            plan_label = (tenant.plan or "essentials").capitalize()
            days_left = max(0, 14 - days_since)
            ticket_count = db.query(Ticket).filter(Ticket.tenant_id == tenant.id).count()
            user_count = db.query(User).filter(User.tenant_id == tenant.id, User.is_active == True).count()

            # Day 0 — Welcome email (send on day 0 or 1)
            if days_since <= 1 and not flags.get("day0"):
                send_email(
                    to=admin.email,
                    cfg={}, db=None, tenant_id=tenant.id,
                    subject=f"Welcome to DodoDesk, {name} 👋 — let's get you started",
                    body=(
                        f"Hi {name},\n\n"
                        f"Welcome to DodoDesk! Your {plan_label} trial is now active — you have 14 days to explore everything, no credit card needed.\n\n"
                        f"Here are 3 things to do in your first 10 minutes:\n\n"
                        f"1. 🎫  Create your first ticket — click 'New Ticket' in the sidebar and log a real IT issue\n"
                        f"2. 👥  Invite your team — go to Users → Add User to bring in your agents and employees\n"
                        f"3. ⚙️  Set up your branding — add your logo and company colours under Settings → Profile\n\n"
                        f"Your {plan_label} trial includes full access to all features. We're here if you need anything — just reply to this email."
                    ),
                    cta_url=f"{FRONTEND}/",
                    cta_label="Go to your dashboard →",
                )
                flags["day0"] = now.isoformat()
                tenant.onboarding_emails = json.dumps(flags)
                db.commit()
                print(f"📧 Onboarding Day 0 → {admin.email} ({tenant.name})")

            # Day 3 — First ticket check
            elif days_since >= 3 and not flags.get("day3"):
                if ticket_count == 0:
                    subj = f"{name}, have you created your first ticket yet?"
                    body = (
                        f"Hi {name},\n\n"
                        f"You signed up for DodoDesk 3 days ago — we wanted to check in.\n\n"
                        f"Creating your first ticket takes under 60 seconds:\n\n"
                        f"1. Click 'New Ticket' in the sidebar\n"
                        f"2. Choose Incident or Service Request\n"
                        f"3. Fill in the title and description\n"
                        f"4. Hit Submit\n\n"
                        f"Your agents get notified automatically and can start working on it right away.\n\n"
                        f"You have {days_left} days left on your trial — plenty of time to see DodoDesk in action."
                    )
                    cta = "Create your first ticket →"
                    url = f"{FRONTEND}/create-ticket"
                else:
                    subj = f"{name}, great start — {ticket_count} ticket{'s' if ticket_count > 1 else ''} already!"
                    body = (
                        f"Hi {name},\n\n"
                        f"You're off to a great start with {ticket_count} ticket{'s' if ticket_count > 1 else ''} already in DodoDesk.\n\n"
                        f"A few more things worth exploring:\n\n"
                        f"• 📚  Knowledge Base — document solutions so your team can self-serve\n"
                        f"• ⚡  Automation Rules — auto-assign tickets by category or priority\n"
                        f"• 📊  Reports — see how your team is performing in real time\n\n"
                        f"You have {days_left} days left on your {plan_label} trial."
                    )
                    cta = "Explore your dashboard →"
                    url = f"{FRONTEND}/"
                send_email(to=admin.email, subject=subj, body=body, cfg={}, cta_url=url, cta_label=cta, db=None, tenant_id=tenant.id)
                flags["day3"] = now.isoformat()
                tenant.onboarding_emails = json.dumps(flags)
                db.commit()
                print(f"📧 Onboarding Day 3 → {admin.email} ({tenant.name})")

            # Day 7 — Invite your team
            elif days_since >= 7 and not flags.get("day7"):
                if user_count <= 1:
                    body = (
                        f"Hi {name},\n\n"
                        f"One week with DodoDesk — how's it going?\n\n"
                        f"We noticed you haven't invited your team yet. DodoDesk is much more powerful with your whole IT team on board:\n\n"
                        f"• Agents can be assigned tickets and manage their own queue\n"
                        f"• Employees get a self-service portal to raise requests without calling IT\n"
                        f"• Managers get live visibility across the entire team\n\n"
                        f"Inviting your team takes 2 minutes — go to Users → Add User.\n\n"
                        f"You have {days_left} days left on your trial."
                    )
                    cta = "Invite your team now →"
                    url = f"{FRONTEND}/admin/users"
                else:
                    body = (
                        f"Hi {name},\n\n"
                        f"One week in and {user_count} team members already on board — brilliant!\n\n"
                        f"Here are some advanced features worth exploring this week:\n\n"
                        f"• 🔄  Change Management — log and approve IT changes with full CAB workflows\n"
                        f"• 🖥️  Asset Management — track all hardware and software across your organisation\n"
                        f"• 🔗  SSO Integration — connect Google Workspace or Microsoft Entra in Settings → Security\n"
                        f"• 📋  SLA Policies — set response and resolution targets per priority level\n\n"
                        f"You have {days_left} days left on your {plan_label} trial."
                    )
                    cta = "Explore advanced features →"
                    url = f"{FRONTEND}/settings"
                send_email(
                    to=admin.email,
                    cfg={}, db=None, tenant_id=tenant.id,
                    subject=f"One week with DodoDesk, {name} — here's what to try next",
                    body=body, cta_url=url, cta_label=cta,
                )
                flags["day7"] = now.isoformat()
                tenant.onboarding_emails = json.dumps(flags)
                db.commit()
                print(f"📧 Onboarding Day 7 → {admin.email} ({tenant.name})")

            # Day 10 — Trial ending soon
            elif days_since >= 10 and not flags.get("day10"):
                send_email(
                    to=admin.email,
                    cfg={}, db=None, tenant_id=tenant.id,
                    subject=f"⏳ {name}, your DodoDesk trial ends in {days_left} day{'s' if days_left != 1 else ''}",
                    body=(
                        f"Hi {name},\n\n"
                        f"Your DodoDesk {plan_label} trial ends in {days_left} day{'s' if days_left != 1 else ''}.\n\n"
                        f"Here's what you've built so far:\n"
                        f"• {ticket_count} ticket{'s' if ticket_count != 1 else ''} logged\n"
                        f"• {user_count} team member{'s' if user_count != 1 else ''} on board\n\n"
                        f"When your trial ends, your account moves to the Free plan (1 agent only). "
                        f"Subscribe now to keep everything running without interruption.\n\n"
                        f"DodoDesk {plan_label} starts at just $15 per agent per month — less than a coffee a day per agent.\n\n"
                        f"Any questions? Just reply to this email — we read everything."
                    ),
                    cta_url=f"{FRONTEND}/settings?tab=billing",
                    cta_label=f"Subscribe to {plan_label} →",
                )
                flags["day10"] = now.isoformat()
                tenant.onboarding_emails = json.dumps(flags)
                db.commit()
                print(f"📧 Onboarding Day 10 → {admin.email} ({tenant.name})")

        db.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"⚠️ _send_onboarding_sequence error: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    import threading
    os.makedirs(AVATAR_DIR, exist_ok=True)

    # ── Connect to DB with retry (Neon can be slow on cold start) ─────────────
    import time as _startup_time
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        try:
            Base.metadata.create_all(bind=engine)
            print(f"✅ Database connected (attempt {attempt})")
            break
        except Exception as e:
            print(f"⚠️ DB connection attempt {attempt}/{max_retries} failed: {type(e).__name__}: {str(e)[:100]}")
            if attempt < max_retries:
                wait = attempt * 3  # 3s, 6s, 9s, 12s
                print(f"   Retrying in {wait}s...")
                _startup_time.sleep(wait)
            else:
                print("❌ Could not connect to DB after all retries — starting anyway, requests may fail until DB recovers")

    # ── STEP 1: Critical column migrations — run SYNCHRONOUSLY ────────────────
    # These must complete before any request is served.
    # Missing columns cause 500 on every authenticated endpoint.
    try:
        from sqlalchemy import inspect as _inspect, text as _text
        _inspector = _inspect(engine)
        with engine.connect() as conn:
            # User table — all columns the model expects
            u_cols = {c['name'] for c in _inspector.get_columns('users')}
            for col, defn in {
                'current_session_id': 'VARCHAR',
                'pending_email': 'VARCHAR',
                'email_change_token': 'VARCHAR',
                'email_change_expires_at': 'TIMESTAMP',
                'mfa_enabled': 'BOOLEAN DEFAULT FALSE',
                'mfa_secret': 'VARCHAR',
                'mfa_backup_codes': 'TEXT',
                'email_verified': 'BOOLEAN DEFAULT FALSE',
                'password_reset_token': 'VARCHAR',
                'password_reset_expires_at': 'TIMESTAMP',
                'employee_id': 'VARCHAR',
                'country': 'VARCHAR',
                'status_changed_at': 'TIMESTAMP',
            }.items():
                if col not in u_cols:
                    conn.execute(_text(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ Critical: users.{col} added")
            # email_configs
            ec_cols = {c['name'] for c in _inspector.get_columns('email_configs')}
            for col, defn in {
                'email_signature': "TEXT DEFAULT ''",
                'email_footer': "TEXT DEFAULT ''",
                'slack_webhook_url': "VARCHAR DEFAULT ''",
                'teams_webhook_url': "VARCHAR DEFAULT ''",
            }.items():
                if col not in ec_cols:
                    conn.execute(_text(f"ALTER TABLE email_configs ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ Critical: email_configs.{col} added")
            # tenants
            t_cols = {c['name'] for c in _inspector.get_columns('tenants')}
            for col, defn in {
                'dodo_customer_id': 'VARCHAR',
                'dodo_subscription_id': 'VARCHAR',
                'billing_status': 'VARCHAR',
                'plan_renews_at': 'TIMESTAMP',
            }.items():
                if col not in t_cols:
                    conn.execute(_text(f"ALTER TABLE tenants ADD COLUMN IF NOT EXISTS {col} {defn}"))
                    conn.commit()
                    print(f"✅ Critical: tenants.{col} added")
        print("✅ Critical column check complete — all required columns present")
    except Exception as e:
        print(f"⚠️ Critical column migration error: {e}")

    # ── STEP 2: Remaining migrations — run in background thread ───────────────
    def _run_migrations_safe():
        try:
            run_migrations()
            # Drop legacy Paddle columns if they still exist
            try:
                from sqlalchemy import text as _t2, inspect as _ins2
                _insp = _ins2(engine)
                tenant_cols = {c['name'] for c in _insp.get_columns('tenants')}
                with engine.begin() as _conn:
                    for col in ['paddle_customer_id', 'paddle_subscription_id']:
                        if col in tenant_cols:
                            _conn.execute(_t2(f"ALTER TABLE tenants DROP COLUMN IF EXISTS {col}"))
                            print(f"✅ Dropped legacy column: tenants.{col}")
            except Exception as e:
                print(f"⚠️ Paddle column cleanup skipped: {e}")
        except Exception as e:
            print(f"⚠️ Migration error (non-fatal): {e}")

    mig_thread = threading.Thread(target=_run_migrations_safe, daemon=False)
    mig_thread.start()

    # Log webhook and billing config status
    print(f"📦 Dodo Environment: {DODO_ENVIRONMENT}")
    print(f"📦 Dodo API Key: {'✅ set' if DODO_API_KEY else '❌ MISSING'}")
    print(f"📦 Dodo Webhook Secret: {'✅ set' if DODO_WEBHOOK_SECRET else '❌ MISSING — webhooks will not verify'}")
    print(f"📦 Dodo Business ID: {os.getenv('DODO_BUSINESS_ID', '❌ MISSING')}")

    seed()


    # Start schedulers after a short delay so the server is live first
    scheduler = BackgroundScheduler()
    scheduler.add_job(check_sla_breaches, 'interval', minutes=5, id='sla_breach_check',
                      next_run_time=datetime.utcnow() + timedelta(seconds=60))
    scheduler.add_job(check_escalations, 'interval', minutes=10, id='escalation_check',
                      next_run_time=datetime.utcnow() + timedelta(seconds=90))
    scheduler.add_job(check_time_based_automations, 'interval', minutes=30, id='automation_time_check',
                      next_run_time=datetime.utcnow() + timedelta(seconds=120))
    scheduler.add_job(auto_close_tickets, 'interval', hours=1, id='auto_close_check',
                      next_run_time=datetime.utcnow() + timedelta(seconds=150))
    scheduler.add_job(send_trial_expiry_warnings, 'interval', hours=12, id='trial_expiry_warnings',
                      next_run_time=datetime.utcnow() + timedelta(seconds=180))
    scheduler.add_job(_dispatch_scheduled_reports, 'interval', hours=1, id='scheduled_reports',
                      next_run_time=datetime.utcnow() + timedelta(seconds=210))
    scheduler.add_job(_send_onboarding_sequence, 'interval', hours=6, id='onboarding_emails',
                      next_run_time=datetime.utcnow() + timedelta(seconds=240))
    scheduler.start()
    print("✅ SLA breach + escalation + automation + auto-close + trial warning + scheduled reports + onboarding emails schedulers started")

    yield

    scheduler.shutdown()
    print("SLA breach scheduler stopped")

app = FastAPI(lifespan=lifespan)

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")
API_URL      = os.getenv("API_URL", "https://dodo-desk-api.onrender.com")

# =============================================================================
# DEPENDENCIES
# =============================================================================

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

def get_db():
    db = SessionLocal()
    try:
        yield db
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass  # SSL already closed — ignore rollback failure
        raise
    finally:
        try:
            db.close()
        except Exception:
            pass  # SSL already closed — ignore close failure

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = decode_access_token(token)
        email = payload.get("sub")
        session_id = payload.get("sid")
        if email is None or payload.get("mfa_pending"):
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    try:
        user = db.query(User).filter(User.email == email).first()
    except Exception as e:
        print(f"⚠️ get_current_user DB error: {e}")
        raise HTTPException(status_code=503, detail="Service temporarily unavailable — please retry in a moment.")

    if user is None:
        raise credentials_exception
    if not user.is_active:
        raise HTTPException(status_code=403, detail="User account is disabled")

    # Single-session enforcement — use getattr to handle missing column gracefully
    try:
        current_sid = getattr(user, "current_session_id", None)
        if current_sid is None:
            # Column exists but no session recorded yet — skip (first login after migration)
            pass
        elif not session_id:
            # Old token with no sid — reject if DB has a session ID
            print(f"⚠️ Session rejected: old token (no sid) for {user.email}")
            raise HTTPException(
                status_code=401,
                detail="Your session has expired. Please log in again."
            )
        elif session_id != current_sid:
            # Token sid doesn't match DB — logged in elsewhere
            print(f"⚠️ Session rejected: sid mismatch for {user.email} (token={session_id[:8]}... db={current_sid[:8]}...)")
            raise HTTPException(
                status_code=401,
                detail="You have been logged out because your account was signed in from another device or browser."
            )
    except HTTPException:
        raise
    except Exception as e:
        print(f"⚠️ Session check error for {user.email}: {e}")
        pass  # column not yet migrated — skip enforcement

    return user


# =============================================================================
# SAML SSO — Single Sign-On via SAML 2.0
# Supports: Google Workspace, Okta, Azure AD, Auth0, and any SAML 2.0 IdP
# =============================================================================

def _get_saml_settings(tenant: "Tenant") -> dict:
    """Build python3-saml settings dict from tenant's SSO configuration."""
    api_url = API_URL.rstrip("/")
    return {
        "strict": True,
        "debug": False,
        "sp": {
            "entityId": f"{api_url}/auth/sso/metadata/{tenant.slug}",
            "assertionConsumerService": {
                "url": f"{api_url}/auth/sso/callback/{tenant.slug}",
                "binding": "urn:oasis:names:tc:SAML:2.0:bindings:HTTP-POST",
            },
            "singleLogoutService": {
                "url": f"{api_url}/auth/sso/logout/{tenant.slug}",
                "binding": "urn:oasis:names:tc:SAML:2.0:bindings:HTTP-Redirect",
            },
            "NameIDFormat": "urn:oasis:names:tc:SAML:1.1:nameid-format:emailAddress",
            "x509cert": "",
            "privateKey": "",
        },
        "idp": {
            "entityId": tenant.sso_client_id or "",
            "singleSignOnService": {
                "url": tenant.sso_sso_url or getattr(tenant, "sso_tenant_id", "") or "",
                "binding": "urn:oasis:names:tc:SAML:2.0:bindings:HTTP-Redirect",
            },
            "singleLogoutService": {
                "url": "",
                "binding": "urn:oasis:names:tc:SAML:2.0:bindings:HTTP-Redirect",
            },
            "x509cert": getattr(tenant, "saml_cert", "") or tenant.sso_domain or "",
        },
        "security": {
            "nameIdEncrypted": False,
            "authnRequestsSigned": False,
            "logoutRequestSigned": False,
            "logoutResponseSigned": False,
            "signMetadata": False,
            "wantMessagesSigned": False,
            "wantAssertionsSigned": True,
            "wantNameId": True,
            "wantAttributeStatement": False,
            "requestedAuthnContext": False,
            "signatureAlgorithm": "http://www.w3.org/2001/04/xmldsig-more#rsa-sha256",
            "digestAlgorithm": "http://www.w3.org/2001/04/xmlenc#sha256",
        }
    }


@app.get("/auth/sso/login/{tenant_slug}")
def sso_login(tenant_slug: str, db: Session = Depends(get_db)):
    """Initiate SAML SSO login — redirects user to their IdP."""
    tenant = db.query(Tenant).filter(
        Tenant.slug == tenant_slug,
        Tenant.is_active == True
    ).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Organisation not found")
    if not tenant.sso_enabled:
        raise HTTPException(status_code=400, detail="SSO is not enabled for this organisation")
    if not tenant.sso_client_id:
        raise HTTPException(status_code=400, detail="SSO is not configured. Please contact your administrator.")

    try:
        from onelogin.saml2.auth import OneLogin_Saml2_Auth
        from onelogin.saml2.utils import OneLogin_Saml2_Utils
        saml_settings = _get_saml_settings(tenant)

        # Build auth object without a real request (we just need the redirect URL)
        req = {
            "https": "on",
            "http_host": API_URL.replace("https://", "").replace("http://", ""),
            "script_name": f"/auth/sso/login/{tenant_slug}",
            "server_port": "443",
            "get_data": {},
            "post_data": {},
        }
        auth = OneLogin_Saml2_Auth(req, saml_settings)
        redirect_url = auth.login()
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=redirect_url)
    except ImportError:
        raise HTTPException(status_code=500, detail="SAML library not installed. Run: pip install python3-saml")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SSO login error: {str(e)}")


@app.post("/auth/sso/callback/{tenant_slug}")
async def sso_callback(tenant_slug: str, request: Request, db: Session = Depends(get_db)):
    """Receive SAML Response from IdP, validate it, and issue a DodoDesk JWT."""
    tenant = db.query(Tenant).filter(
        Tenant.slug == tenant_slug,
        Tenant.is_active == True
    ).first()
    if not tenant:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=org_not_found")

    try:
        from onelogin.saml2.auth import OneLogin_Saml2_Auth
        form_data = await request.form()
        saml_response = form_data.get("SAMLResponse", "")

        req = {
            "https": "on",
            "http_host": API_URL.replace("https://", "").replace("http://", ""),
            "script_name": f"/auth/sso/callback/{tenant_slug}",
            "server_port": "443",
            "get_data": dict(request.query_params),
            "post_data": {"SAMLResponse": saml_response},
        }

        saml_settings = _get_saml_settings(tenant)
        auth = OneLogin_Saml2_Auth(req, saml_settings)
        auth.process_response()
        errors = auth.get_errors()

        if errors:
            error_msg = auth.get_last_error_reason() or ", ".join(errors)
            print(f"❌ SAML errors for {tenant_slug}: {error_msg}")
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=saml_failed&detail={error_msg[:100]}")

        if not auth.is_authenticated():
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=not_authenticated")

        # Extract user attributes from SAML assertion
        attrs      = auth.get_attributes()
        name_id    = auth.get_nameid()  # usually the email
        email      = (attrs.get("email", [None])[0] or
                      attrs.get("http://schemas.xmlsoap.org/ws/2005/05/identity/claims/emailaddress", [None])[0] or
                      name_id or "")
        first_name = (attrs.get("firstName", [None])[0] or
                      attrs.get("http://schemas.xmlsoap.org/ws/2005/05/identity/claims/givenname", [None])[0] or "")
        last_name  = (attrs.get("lastName", [None])[0] or
                      attrs.get("http://schemas.xmlsoap.org/ws/2005/05/identity/claims/surname", [None])[0] or "")
        full_name  = f"{first_name} {last_name}".strip() or email.split("@")[0]

        if not email:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=no_email")

        # Restrict to configured SSO domain if set
        if tenant.sso_domain and not email.endswith(f"@{tenant.sso_domain.lstrip('@')}"):
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=domain_mismatch")

        # Find or create user
        user = db.query(User).filter(
            User.email == email.lower(),
            User.tenant_id == tenant.id
        ).first()

        if not user:
            # Auto-provision user on first SSO login
            user = User(
                tenant_id=tenant.id,
                email=email.lower(),
                full_name=full_name,
                hashed_password=get_password_hash(os.urandom(32).hex()),  # random unusable password
                role=UserRole.EMPLOYEE,
                is_active=True,
                email_verified=True,
            )
            db.add(user)
            db.commit()
            db.refresh(user)
            log_system_event(db, user, "user.sso_provisioned",
                             target_type="user", target_id=user.id, target_label=email)
            db.commit()
            print(f"✅ SSO: auto-provisioned user {email} for tenant {tenant.name}")
        elif not user.is_active:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=account_disabled")

        # Update name if changed in IdP
        if full_name and full_name != user.full_name:
            user.full_name = full_name
            db.commit()

        # Issue JWT — single session enforcement
        import uuid as _uuid
        session_id = str(_uuid.uuid4())
        user.current_session_id = session_id
        db.commit()

        access_token = create_access_token({"sub": user.email, "sid": session_id})
        log_system_event(db, user, "user.sso_login",
                         target_type="user", target_id=user.id, target_label=email)
        db.commit()

        print(f"✅ SSO login: {email} → tenant {tenant.name}")

        # Redirect to frontend with token in URL fragment (never in query string)
        from fastapi.responses import RedirectResponse
        return RedirectResponse(
            url=f"{FRONTEND_URL}/sso-callback#token={access_token}&email={email}"
        )

    except Exception as e:
        import traceback; traceback.print_exc()
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=sso_error")


@app.get("/auth/sso/metadata/{tenant_slug}")
def sso_metadata(tenant_slug: str, db: Session = Depends(get_db)):
    """Return SP metadata XML — paste this into your IdP when configuring DodoDesk."""
    tenant = db.query(Tenant).filter(Tenant.slug == tenant_slug).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Organisation not found")

    try:
        from onelogin.saml2.settings import OneLogin_Saml2_Settings
        from onelogin.saml2.errors import OneLogin_Saml2_Error
        saml_settings = _get_saml_settings(tenant)
        settings_obj  = OneLogin_Saml2_Settings(settings=saml_settings, sp_validation_only=True)
        metadata      = settings_obj.get_sp_metadata()
        from fastapi.responses import Response
        return Response(content=metadata, media_type="application/xml")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not generate metadata: {str(e)}")


@app.get("/auth/sso/check/{email_or_slug}")
def sso_check(email_or_slug: str, db: Session = Depends(get_db)):
    """Check if SSO is enabled for a given email domain or tenant slug.
    Used by the login page to show SSO option automatically.
    """
    # Try by email domain
    if "@" in email_or_slug:
        domain = email_or_slug.split("@")[1].lower()
        tenant = db.query(Tenant).filter(
            Tenant.sso_enabled == True,
            Tenant.sso_domain == domain,
            Tenant.is_active == True
        ).first()
    else:
        # Try by slug
        tenant = db.query(Tenant).filter(
            Tenant.slug == email_or_slug,
            Tenant.sso_enabled == True,
            Tenant.is_active == True
        ).first()

    if tenant:
        provider = tenant.sso_provider or "saml"
        # OAuth providers use /auth/oauth/login, SAML uses /auth/sso/login
        oauth_providers = {"google", "microsoft", "okta"}
        login_url = (f"{API_URL}/auth/oauth/login/{tenant.slug}"
                     if provider in oauth_providers
                     else f"{API_URL}/auth/sso/login/{tenant.slug}")
        return {
            "sso_enabled":  True,
            "tenant_slug":  tenant.slug,
            "tenant_name":  tenant.name,
            "sso_provider": provider,
            "login_url":    login_url,
        }
    return {"sso_enabled": False}
# =============================================================================

OAUTH_PROVIDERS = {
    "google": {
        "auth_url":   "https://accounts.google.com/o/oauth2/v2/auth",
        "token_url":  "https://oauth2.googleapis.com/token",
        "userinfo_url": "https://www.googleapis.com/oauth2/v3/userinfo",
        "scope":      "openid email profile",
    },
    "microsoft": {
        # tenant_id is substituted at runtime from tenant config
        "auth_url":   "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/authorize",
        "token_url":  "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        "userinfo_url": "https://graph.microsoft.com/oidc/userinfo",
        "scope":      "openid email profile",
    },
    "okta": {
        # domain is substituted at runtime e.g. your-org.okta.com
        "auth_url":   "https://{domain}/oauth2/v1/authorize",
        "token_url":  "https://{domain}/oauth2/v1/token",
        "userinfo_url": "https://{domain}/oauth2/v1/userinfo",
        "scope":      "openid email profile",
    },
}


def _get_oauth_redirect_uri(tenant_slug: str) -> str:
    return f"{API_URL}/auth/oauth/callback/{tenant_slug}"


@app.get("/auth/oauth/login/{tenant_slug}")
def oauth_login(tenant_slug: str, db: Session = Depends(get_db)):
    """Initiate OAuth 2.0 / OIDC login for Google, Microsoft, or Okta."""
    tenant = db.query(Tenant).filter(
        Tenant.slug == tenant_slug,
        Tenant.is_active == True
    ).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Organisation not found")
    if not tenant.sso_enabled:
        raise HTTPException(status_code=400, detail="SSO is not enabled for this organisation")

    provider = (tenant.sso_provider or "google").lower()
    if provider not in OAUTH_PROVIDERS:
        # Fall back to SAML
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"{API_URL}/auth/sso/login/{tenant_slug}")

    provider_cfg = OAUTH_PROVIDERS[provider]
    client_id = tenant.sso_client_id or ""
    if not client_id:
        raise HTTPException(status_code=400, detail="OAuth Client ID not configured. Please set it in Settings → Security.")

    # Build auth URL with provider-specific substitutions
    auth_url = provider_cfg["auth_url"]
    if provider == "microsoft":
        tenant_id = tenant.sso_tenant_id or "common"
        auth_url = auth_url.replace("{tenant_id}", tenant_id)
    elif provider == "okta":
        domain = tenant.sso_domain or tenant.sso_tenant_id or ""
        if not domain:
            raise HTTPException(status_code=400, detail="Okta domain not configured (e.g. your-org.okta.com)")
        auth_url = auth_url.replace("{domain}", domain)

    import urllib.parse, secrets
    state = secrets.token_urlsafe(32)
    params = {
        "client_id":     client_id,
        "response_type": "code",
        "redirect_uri":  _get_oauth_redirect_uri(tenant_slug),
        "scope":         provider_cfg["scope"],
        "state":         f"{tenant_slug}:{state}",
        "access_type":   "online",
    }
    if provider == "microsoft":
        params["response_mode"] = "query"

    redirect = f"{auth_url}?{urllib.parse.urlencode(params)}"
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=redirect)


@app.get("/auth/oauth/callback/{tenant_slug}")
async def oauth_callback(
    tenant_slug: str,
    code: str = "",
    state: str = "",
    error: str = "",
    db: Session = Depends(get_db)
):
    """Handle OAuth 2.0 callback — exchange code for token, get user info, issue JWT."""
    from fastapi.responses import RedirectResponse

    if error:
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=oauth_{error}")

    tenant = db.query(Tenant).filter(
        Tenant.slug == tenant_slug,
        Tenant.is_active == True
    ).first()
    if not tenant:
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=org_not_found")

    provider = (tenant.sso_provider or "google").lower()
    if provider not in OAUTH_PROVIDERS:
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=unsupported_provider")

    provider_cfg  = OAUTH_PROVIDERS[provider]
    client_id     = tenant.sso_client_id or ""
    client_secret = tenant.sso_client_secret or ""
    token_url     = provider_cfg["token_url"]
    userinfo_url  = provider_cfg["userinfo_url"]

    if provider == "microsoft":
        tid = tenant.sso_tenant_id or "common"
        token_url    = token_url.replace("{tenant_id}", tid)
        userinfo_url = userinfo_url  # graph endpoint is fixed
    elif provider == "okta":
        domain = tenant.sso_domain or tenant.sso_tenant_id or ""
        token_url    = token_url.replace("{domain}", domain)
        userinfo_url = userinfo_url.replace("{domain}", domain)

    try:
        import httpx as _httpx

        # Step 1: Exchange authorization code for tokens
        token_resp = _httpx.post(token_url, data={
            "grant_type":   "authorization_code",
            "code":          code,
            "redirect_uri":  _get_oauth_redirect_uri(tenant_slug),
            "client_id":     client_id,
            "client_secret": client_secret,
        }, timeout=15.0)

        if token_resp.status_code != 200:
            print(f"❌ OAuth token exchange failed: {token_resp.status_code} {token_resp.text[:200]}")
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=token_exchange_failed")

        tokens      = token_resp.json()
        access_token = tokens.get("access_token", "")

        # Step 2: Get user info
        userinfo_resp = _httpx.get(userinfo_url, headers={
            "Authorization": f"Bearer {access_token}"
        }, timeout=10.0)

        if userinfo_resp.status_code != 200:
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=userinfo_failed")

        user_data = userinfo_resp.json()

        # Normalise across providers
        email = (user_data.get("email") or
                 user_data.get("mail") or
                 user_data.get("preferred_username") or "").lower()
        first = user_data.get("given_name") or user_data.get("givenName") or ""
        last  = user_data.get("family_name") or user_data.get("surname") or ""
        full_name = f"{first} {last}".strip() or email.split("@")[0]

        if not email:
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=no_email")

        # Domain restriction
        if tenant.sso_domain and not email.endswith(f"@{tenant.sso_domain.lstrip('@')}"):
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=domain_mismatch")

        # Find or auto-provision user
        db_user = db.query(User).filter(
            User.email == email,
            User.tenant_id == tenant.id
        ).first()

        if not db_user:
            import os as _os
            db_user = User(
                tenant_id=tenant.id,
                email=email,
                full_name=full_name,
                hashed_password=get_password_hash(_os.urandom(32).hex()),
                role=UserRole.EMPLOYEE,
                is_active=True,
                email_verified=True,
            )
            db.add(db_user)
            db.commit()
            db.refresh(db_user)
            log_system_event(db, db_user, "user.oauth_provisioned",
                             target_type="user", target_id=db_user.id, target_label=email)
            db.commit()
            print(f"✅ OAuth SSO: auto-provisioned {email} via {provider}")
        elif not db_user.is_active:
            return RedirectResponse(url=f"{FRONTEND_URL}/login?error=account_disabled")

        # Update name if changed
        if full_name and full_name != db_user.full_name:
            db_user.full_name = full_name
            db.commit()

        # Issue JWT
        import uuid as _uuid2
        session_id = str(_uuid2.uuid4())
        db_user.current_session_id = session_id
        db.commit()

        jwt_token = create_access_token({"sub": db_user.email, "sid": session_id})
        log_system_event(db, db_user, "user.oauth_login",
                         target_type="user", target_id=db_user.id, target_label=email)
        db.commit()
        print(f"✅ OAuth login: {email} via {provider} → tenant {tenant.name}")

        return RedirectResponse(
            url=f"{FRONTEND_URL}/sso-callback#token={jwt_token}&email={email}"
        )

    except Exception as e:
        import traceback; traceback.print_exc()
        return RedirectResponse(url=f"{FRONTEND_URL}/login?error=oauth_error")


@app.get("/auth/oauth/providers")
def list_oauth_providers():
    """Return supported OAuth providers for the frontend."""
    return {
        "providers": [
            {"key": "google",    "label": "Google Workspace",         "icon": "google"},
            {"key": "microsoft", "label": "Microsoft Entra ID",       "icon": "microsoft"},
            {"key": "okta",      "label": "Okta",                     "icon": "okta"},
            {"key": "saml",      "label": "Generic SAML 2.0",         "icon": "saml"},
        ]
    }




ALLOWED_ORIGIN = os.getenv("ALLOWED_ORIGIN", "http://localhost:5173")
# Support multiple comma-separated origins e.g. "https://app.vercel.app,http://localhost:5173"
_allowed_origins = list(set(
    [o.strip() for o in ALLOWED_ORIGIN.split(",") if o.strip()]
    + ["http://localhost:5173", "http://localhost:3000"]
))
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure CORS headers are present even on 500 errors
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest
from starlette.responses import Response as StarletteResponse

class CORSOnErrorMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: StarletteRequest, call_next):
        origin = request.headers.get("origin", "")
        is_allowed = origin in _allowed_origins

        # Handle OPTIONS preflight directly — don't pass to app
        if request.method == "OPTIONS":
            from starlette.responses import Response as _R
            r = _R(status_code=204)
            if is_allowed:
                r.headers["Access-Control-Allow-Origin"]      = origin
                r.headers["Access-Control-Allow-Credentials"] = "true"
                r.headers["Access-Control-Allow-Methods"]     = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
                r.headers["Access-Control-Allow-Headers"]     = "Content-Type, Authorization, X-Requested-With"
                r.headers["Access-Control-Max-Age"]           = "86400"
            return r

        try:
            response = await call_next(request)
        except Exception:
            from starlette.responses import JSONResponse
            response = JSONResponse({"detail": "Internal server error"}, status_code=500)

        if is_allowed:
            response.headers["Access-Control-Allow-Origin"]      = origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
        return response

app.add_middleware(CORSOnErrorMiddleware)



def get_current_admin_user(current_user: User = Depends(get_current_user)):
    role = current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)
    if role not in ('admin', 'super_admin', 'platform_admin'):
        raise HTTPException(status_code=403, detail="Only admins can perform this action")
    return current_user

def has_permission(user: User, permission: Permission) -> bool:
    role = user.role.value if hasattr(user.role, 'value') else str(user.role)
    if role in ('admin', 'super_admin', 'platform_admin'):
        return True
    if user.custom_role:
        permissions = json.loads(user.custom_role.permissions)
        return permission.value in permissions
    # Readonly — can view tickets/assets/kb/reports but cannot create or edit anything
    if str(user.role) == "readonly":
        return permission in [
            Permission.VIEW_ALL_TICKETS,
            Permission.VIEW_REPORTS,
        ]
    # Legacy fallback
    if str(user.role) == "agent":
        return permission in [
            Permission.VIEW_ALL_TICKETS,
            Permission.EDIT_TICKETS,
            Permission.CREATE_TICKETS,
            Permission.MANAGE_ASSETS,
            Permission.MANAGE_KB,
            Permission.VIEW_REPORTS,
            Permission.MANAGE_CANNED,
            Permission.CREATE_CHANGES,
            Permission.APPROVE_CHANGES
        ]
    if str(user.role) == "employee":
        return permission in [
            Permission.CREATE_TICKETS,
            Permission.CREATE_CHANGES
        ]
    return False

def apply_filters(query, ticket_type: str | None, start_date: date | None, end_date: date | None):
    if ticket_type and ticket_type != 'change':
        query = query.filter(Ticket.ticket_type == ticket_type.lower().strip())
    if start_date:
        query = query.filter(Ticket.created_at >= datetime(start_date.year, start_date.month, start_date.day))
    if end_date:
        end_dt = datetime(end_date.year, end_date.month, end_date.day) + timedelta(days=1)
        query = query.filter(Ticket.created_at < end_dt)
    return query

# =============================================================================
# ENDPOINTS
# =============================================================================

# ---------- Authentication ----------
@app.post("/auth/forgot-password")
def forgot_password(data: dict, db: Session = Depends(get_db)):
    from sqlalchemy import text as _text
    email = data.get("email", "").lower().strip()
    # Allow locked or inactive users to reset password — account locked ≠ permanently deleted
    # We look for any user with this email (active or locked) so they can regain access
    user = db.query(User).filter(User.email == email).first()
    if not user:
        return {"ok": True, "message": "If that email exists, a reset link has been sent."}
    # Don't allow reset for invited-but-not-yet-activated users — they should use the invite link
    if not user.is_active and user.password_reset_token and user.password_reset_token.startswith("invite_"):
        return {"ok": True, "message": "If that email exists, a reset link has been sent."}

    token    = uuid.uuid4().hex
    reset_val = f"reset_{token}"
    expires_at = datetime.utcnow() + timedelta(hours=1)

    # Store token with expiry — use raw SQL for reliability
    try:
        with db.bind.connect() as conn:
            conn.execute(
                _text("UPDATE users SET password_reset_token = :tok, password_reset_expires_at = :exp WHERE id = :uid"),
                {"tok": reset_val, "uid": user.id, "exp": expires_at}
            )
            conn.commit()
        print(f"✅ Reset token stored for {user.email}, expires {expires_at}")
    except Exception as e:
        print(f"❌ Failed to store reset token: {e}")
        raise HTTPException(status_code=500, detail="Could not generate reset token.")

    reset_url = f"{FRONTEND_URL}/reset-password?token={token}"

    # Send via Resend in background thread — returns immediately
    import threading
    _email = user.email
    _name  = user.full_name
    _url   = reset_url
    _key   = RESEND_API_KEY
    _from  = RESEND_FROM
    # Fetch super admin branding for email
    try:
        _super = db.query(Tenant).filter(Tenant.id == 1).first()
        _logo  = _super.logo_url if _super else None
        _color = _super.primary_color if _super else "#4f46e5"
        _cname = _super.name if _super else "DodoDesk"
    except Exception:
        _logo = None; _color = "#4f46e5"; _cname = "DodoDesk"

    send_email_background(
        to=_email,
        subject="Reset your DodoDesk password",
        body=(
            f"Hi {_name},\n\n"
            f"You requested a password reset. Click the link below to set a new password:\n\n"
            f"{_url}\n\n"
            f"This link expires in 1 hour. If you didn't request this, you can safely ignore this email."
        ),
        cta_url=_url,
        cta_label="Reset My Password",
    )
    return {"ok": True, "message": "If that email exists, a reset link has been sent."}

@app.post("/auth/reset-password")
def reset_password(data: dict, db: Session = Depends(get_db)):
    import traceback
    from sqlalchemy import text as _text
    token        = data.get("token", "")
    new_password = data.get("new_password", "")
    print(f"🔑 reset_password called token_len={len(token)} pw_len={len(new_password)}")

    if not token or not new_password:
        raise HTTPException(status_code=400, detail="Token and new password are required")

    # Accept either a forgot-password reset token or an invite token —
    # both are stored the same way, just with a different prefix so we know
    # whether to also activate the account (invites) or leave status untouched (resets).
    reset_val = f"reset_{token}"
    invite_val = f"invite_{token}"

    try:
        # Step 1 — ensure column exists
        try:
            with db.bind.connect() as conn:
                conn.execute(_text("ALTER TABLE users ADD COLUMN IF NOT EXISTS password_reset_token VARCHAR"))
                conn.commit()
        except Exception as e:
            print(f"⚠️ ALTER TABLE skipped: {e}")

        # Step 2 — look up token (try reset first, then invite) and check expiry
        is_invite = False
        with db.bind.connect() as conn:
            result = conn.execute(
                _text("SELECT id, password_reset_expires_at FROM users WHERE password_reset_token = :tok"),
                {"tok": reset_val}
            ).fetchone()
            if not result:
                result = conn.execute(
                    _text("SELECT id, password_reset_expires_at FROM users WHERE password_reset_token = :tok"),
                    {"tok": invite_val}
                ).fetchone()
                if result:
                    is_invite = True
        print(f"🔍 Token lookup result: {result} (invite={is_invite})")

        if not result:
            raise HTTPException(status_code=400, detail="Invalid or expired link. Please request a new one.")

        user_id = result[0]
        expires_at = result[1]

        # Check expiry
        if expires_at and datetime.utcnow() > expires_at:
            with db.bind.connect() as conn:
                conn.execute(_text("UPDATE users SET password_reset_token = NULL, password_reset_expires_at = NULL WHERE id = :uid"), {"uid": user_id})
                conn.commit()
            detail = "This invite link has expired. Please ask your admin to resend it." if is_invite else "This reset link has expired. Please request a new password reset."
            raise HTTPException(status_code=400, detail=detail)

        # Step 3 — validate and hash
        validate_password_strength(new_password)
        hashed = get_password_hash(new_password[:72])

        # Step 4 — update password, clear token, unlock if locked
        with db.bind.connect() as conn:
            if is_invite:
                conn.execute(
                    _text("UPDATE users SET hashed_password = :pw, password_reset_token = NULL, password_reset_expires_at = NULL, is_active = TRUE, email_verified = TRUE WHERE id = :uid"),
                    {"pw": hashed, "uid": user_id}
                )
            else:
                # Also clear any lock — if someone is locked out and resets password, they should regain access
                conn.execute(
                    _text("UPDATE users SET hashed_password = :pw, password_reset_token = NULL, password_reset_expires_at = NULL, is_active = TRUE, locked_until = NULL, failed_login_attempts = 0 WHERE id = :uid"),
                    {"pw": hashed, "uid": user_id}
                )
            conn.commit()

        print(f"✅ Password set successful for user_id={user_id} (invite={is_invite})")
        message = "Account activated! You can now log in." if is_invite else "Password reset successfully. You can now log in."
        return {"ok": True, "message": message}

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ reset_password error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Reset failed: {str(e)}")

# =============================================================================
# SELF-SERVE SIGNUP
# =============================================================================

def slugify_company_name(name: str) -> str:
    """Convert a company name into a URL-safe slug, e.g. 'Acme Corp!' -> 'acme-corp'."""
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    return slug or "company"

def generate_unique_slug(db: Session, base_slug: str) -> str:
    """Append a numeric suffix if the slug is already taken."""
    slug = base_slug
    suffix = 1
    while db.query(Tenant).filter(Tenant.slug == slug).first():
        suffix += 1
        slug = f"{base_slug}-{suffix}"
    return slug

@app.get("/signup/verify")
def verify_signup(token: str, db: Session = Depends(get_db)):
    """Verify a signup email token, activate the tenant + admin user, and return a login token."""
    try:
        payload = decode_access_token(token)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or expired verification link.")

    if not payload.get("signup_verify"):
        raise HTTPException(status_code=400, detail="Invalid verification token.")

    tenant = db.query(Tenant).filter(Tenant.id == payload.get("tenant_id")).first()
    user = db.query(User).filter(User.id == payload.get("user_id")).first()
    if not tenant or not user:
        raise HTTPException(status_code=404, detail="Account not found.")

    if not tenant.is_active or not user.is_active:
        tenant.is_active = True
        user.is_active = True
        db.commit()

    # Issue a normal login session token so the frontend can log them straight in
    session_id = str(uuid.uuid4())
    user.current_session_id = session_id
    db.commit()
    access_token = create_access_token({"sub": user.email, "sid": session_id})

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "plan_selected": payload.get("plan", "free"),
        "tenant_slug": tenant.slug,
    }

# =============================================================================
# SELF-SERVE SIGNUP
# =============================================================================

def slugify(name: str) -> str:
    """Convert company name to a URL-safe slug: 'Acme Corp' -> 'acme-corp'."""
    import re
    slug = name.lower().strip()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s_-]+', '-', slug)
    slug = slug.strip('-')
    return slug or "tenant"

def unique_slug(db: Session, base: str) -> str:
    """Append a number if the slug is already taken: 'acme-corp', 'acme-corp-2', etc."""
    slug = base[:40]  # keep reasonable length
    existing = db.query(Tenant).filter(Tenant.slug == slug).first()
    if not existing:
        return slug
    counter = 2
    while True:
        candidate = f"{slug[:37]}-{counter}"
        if not db.query(Tenant).filter(Tenant.slug == candidate).first():
            return candidate
        counter += 1

def generate_verification_token() -> str:
    import secrets
    return secrets.token_urlsafe(32)

@app.post("/auth/signup")
@limiter.limit("5/hour")
def signup(request: Request, data: dict, db: Session = Depends(get_db)):
    """Self-serve signup. Plan can be: essentials, business, pro, free.
    Tenant starts on selected plan as 14-day trial. Drops to free if no payment after trial.
    """
    company_name = (data.get("company_name") or "").strip()
    full_name    = (data.get("full_name") or "").strip()
    email        = (data.get("email") or "").strip().lower()
    password     = (data.get("password") or "").strip()
    plan         = (data.get("plan") or "essentials").strip().lower()

    valid_plans = ("free", "essentials", "business", "pro")
    if plan not in valid_plans:
        plan = "essentials"

    if not company_name or not full_name or not email or not password:
        raise HTTPException(status_code=400, detail="All fields are required.")

    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        if existing_user.is_active:
            raise HTTPException(status_code=400, detail="An account with this email already exists. Please log in or use a different email.")
        if existing_user.password_reset_token and existing_user.password_reset_token.startswith("invite_"):
            raise HTTPException(status_code=400, detail="This email address has already been invited to DodoDesk. Check your inbox for the invitation link.")
        old_tenant = db.query(Tenant).filter(Tenant.id == existing_user.tenant_id, Tenant.is_active == False).first()
        db.query(SignupVerification).filter(SignupVerification.user_id == existing_user.id).delete()
        db.delete(existing_user)
        if old_tenant:
            db.delete(old_tenant)
        db.commit()

    validate_password_strength(password)
    base_slug = slugify(company_name)
    slug = unique_slug(db, base_slug)

    try:
        super_tenant = db.query(Tenant).filter(Tenant.id == 1).first()
        default_color  = super_tenant.primary_color if super_tenant and super_tenant.primary_color else "#4f46e5"
        default_accent = super_tenant.accent_color if super_tenant and super_tenant.accent_color else "#818cf8"

        # Create tenant on selected plan as 14-day trial
        tenant = Tenant(
            name=company_name, slug=slug,
            is_active=False,
            plan=plan,
            billing_status="trialing",
            primary_color=default_color,
            accent_color=default_accent,
        )
        db.add(tenant)
        db.flush()

        admin_user = User(
            tenant_id=tenant.id, email=email,
            hashed_password=get_password_hash(password),
            full_name=full_name, role=UserRole.ADMIN,
            is_active=False, email_verified=False,
        )
        db.add(admin_user)
        db.flush()

        token = generate_verification_token()
        verification = SignupVerification(
            token=token, email=email,
            tenant_id=tenant.id, user_id=admin_user.id,
            plan=plan, expires_at=datetime.utcnow() + timedelta(hours=24),
        )
        db.add(verification)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"❌ Signup DB error: {e}")
        raise HTTPException(status_code=500, detail=f"Account creation failed: {str(e)}")

    # Send verification email in background — capture plain values only, no DB objects
    frontend_url  = os.getenv("FRONTEND_URL", "https://dodo-desk-pied.vercel.app")
    verify_url    = f"{frontend_url}/verify-email?token={token}"
    _to           = str(email)
    _full_name    = str(full_name)
    _company_name = str(company_name)
    _verify_url   = str(verify_url)

    send_email_background(
        to=_to,
        subject="Verify your DodoDesk account",
        body=f"Hi {_full_name},\n\nWelcome to DodoDesk! Please verify your email address to activate your account for {_company_name}.\n\nThis link expires in 24 hours.",
        cta_url=_verify_url,
        cta_label="Verify Email",
    )

    return {
        "message": "Account created! Please check your email to verify your address before logging in.",
        "email": email,
    }


@app.get("/auth/verify-email")
def verify_email(token: str, db: Session = Depends(get_db)):
    """Verifies an email token and activates the tenant + admin user."""
    try:
        verification = db.query(SignupVerification).filter(
            SignupVerification.token == token,
            SignupVerification.used == False,
        ).first()

        if not verification:
            raise HTTPException(status_code=400, detail="Verification link is invalid or has already been used.")
        if verification.expires_at < datetime.utcnow():
            raise HTTPException(status_code=400, detail="Verification link has expired. Please sign up again.")

        tenant = db.query(Tenant).filter(Tenant.id == verification.tenant_id).first()
        user   = db.query(User).filter(User.id == verification.user_id).first()

        if not tenant or not user:
            raise HTTPException(status_code=400, detail="Account data not found. Please sign up again.")

        tenant.is_active   = True
        user.is_active     = True
        user.email_verified = True
        verification.used  = True

        session_id = str(uuid.uuid4())
        user.current_session_id = session_id
        db.commit()

        access_token = create_access_token({"sub": user.email, "sid": session_id})

        return {
            "access_token": access_token,
            "token_type": "bearer",
            "plan_selected": verification.plan,
            "tenant_slug": tenant.slug,
            "message": "Email verified! Your account is now active.",
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        db.rollback()
        print(f"❌ verify_email error: {e}")
        raise HTTPException(status_code=500, detail=f"Verification failed due to a server error. Please try again or contact support. ({type(e).__name__}: {str(e)[:200]})")


@app.post("/auth/resend-verification")
def resend_verification(data: dict, db: Session = Depends(get_db)):
    """Resends the verification email for a pending unverified signup."""
    email = (data.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required.")

    user = db.query(User).filter(User.email == email, User.email_verified == False).first()
    if not user:
        # Don't reveal if email exists or is already verified
        return {"message": "If this email has a pending verification, a new link has been sent."}

    # Invalidate old tokens
    db.query(SignupVerification).filter(
        SignupVerification.email == email,
        SignupVerification.used == False,
    ).update({"used": True})
    db.flush()

    # Issue new token
    token = generate_verification_token()
    verification = SignupVerification(
        token=token,
        email=email,
        tenant_id=user.tenant_id,
        user_id=user.id,
        plan=db.query(SignupVerification).filter(
            SignupVerification.user_id == user.id
        ).order_by(SignupVerification.id.desc()).first().plan if db.query(SignupVerification).filter(
            SignupVerification.user_id == user.id
        ).first() else "free",
        expires_at=datetime.utcnow() + timedelta(hours=24),
    )
    db.add(verification)
    db.commit()

    frontend_url   = os.getenv("FRONTEND_URL", "https://dodo-desk-pied.vercel.app")
    verify_url     = f"{frontend_url}/verify-email?token={token}"
    _to            = str(email)
    _full_name     = str(user.full_name)   # extract before session closes
    _verify_url    = str(verify_url)

    def _send_resend():
        print(f"📧 [thread] Resending verification email to {_to}, RESEND_API_KEY set={bool(RESEND_API_KEY)}")
        try:
            send_email(
                to=_to,
                subject="Verify your DodoDesk account (new link)",
                body=f"Hi {_full_name},\n\nHere's a new verification link for your DodoDesk account. The previous link has been invalidated.\n\nThis link expires in 24 hours.",
                cta_url=_verify_url,
                cta_label="Verify Email",
            )
        except Exception as e:
            print(f"⚠️ Failed to resend verification email: {e}")

    import threading
    threading.Thread(target=_send_resend, daemon=True).start()

    return {"message": "If this email has a pending verification, a new link has been sent."}


@app.post("/auth/login")
def login(
    request: Request,
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    # IP-based rate limiting
    client_ip = request.client.host if request.client else "unknown"
    if not check_ip_rate_limit(client_ip):
        raise HTTPException(status_code=429, detail="Too many login attempts. Please wait 1 minute before trying again.")

    user = db.query(User).filter(User.email == form_data.username).first()

    # Check if account is locked
    if user and user.locked_until and user.locked_until > datetime.utcnow():
        raise HTTPException(status_code=423, detail="Account locked due to too many failed attempts. Please contact your administrator.")

    if not user or not verify_password(form_data.password, user.hashed_password):
        if user:
            user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
            if user.failed_login_attempts >= MAX_FAILED_ATTEMPTS:
                user.locked_until = datetime.utcnow() + timedelta(days=3650)
                user.failed_login_attempts = 0
                db.commit()
                raise HTTPException(status_code=423, detail=f"Account locked after {MAX_FAILED_ATTEMPTS} failed attempts. Please contact your administrator.")
            db.commit()
            remaining = MAX_FAILED_ATTEMPTS - user.failed_login_attempts
            raise HTTPException(status_code=401, detail=f"Invalid credentials. {remaining} attempt(s) remaining before account lockout.")
        raise HTTPException(status_code=401, detail="Invalid credentials.")

    # Successful login — reset counters
    user.failed_login_attempts = 0
    user.locked_until = None
    if not user.is_active:
        db.commit()
        if not user.email_verified:
            raise HTTPException(status_code=403, detail="Please verify your email address before logging in. Check your inbox for the verification link.")
        raise HTTPException(status_code=403, detail="User account is disabled.")
    db.commit()

    # MFA check — if enabled, return a short-lived MFA challenge token instead of full access
    if user.mfa_enabled:
        mfa_token = create_access_token_with_expiry(
            data={"sub": user.email, "tenant_id": user.tenant_id, "mfa_pending": True},
            minutes=5
        )
        return {"mfa_required": True, "mfa_token": mfa_token}

    # If tenant requires MFA but this user hasn't set it up yet, allow login but flag it
    tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first()
    mfa_setup_required = bool(tenant and tenant.mfa_required and not user.mfa_enabled)

    # Single-session enforcement — generate new session ID, invalidating any previous session
    import uuid as _uuid
    session_id = str(_uuid.uuid4())
    user.current_session_id = session_id
    log_system_event(db, user, "user.login",
                     target_type="user", target_id=user.id, target_label=user.email,
                     ip_address=request.client.host if request.client else None)
    db.commit()

    access_token = create_access_token(data={"sub": user.email, "tenant_id": user.tenant_id, "sid": session_id})
    return {"access_token": access_token, "token_type": "bearer", "mfa_setup_required": mfa_setup_required}

@app.post("/auth/login/mfa")
def login_mfa_verify(data: dict, db: Session = Depends(get_db)):
    """Step 2 of login when MFA is enabled. Accepts the mfa_token from /auth/login plus a 6-digit code or backup code."""
    mfa_token = data.get("mfa_token", "")
    code = data.get("code", "")
    try:
        payload = decode_access_token(mfa_token)
        if not payload.get("mfa_pending"):
            raise HTTPException(status_code=401, detail="Invalid MFA session.")
        email = payload.get("sub")
    except JWTError:
        raise HTTPException(status_code=401, detail="MFA session expired. Please log in again.")

    user = db.query(User).filter(User.email == email).first()
    if not user or not user.mfa_enabled:
        raise HTTPException(status_code=401, detail="Invalid MFA session.")

    # Try TOTP code first
    valid = verify_totp(user.mfa_secret, code)

    # Fall back to backup codes
    if not valid:
        backup_codes = json.loads(user.mfa_backup_codes or "[]")
        normalized = code.strip().upper()
        if normalized in backup_codes:
            valid = True
            backup_codes.remove(normalized)
            user.mfa_backup_codes = json.dumps(backup_codes)

    if not valid:
        raise HTTPException(status_code=401, detail="Invalid authentication code.")

    # Single-session enforcement
    import uuid as _uuid
    session_id = str(_uuid.uuid4())
    user.current_session_id = session_id
    db.commit()

    access_token = create_access_token(data={"sub": user.email, "tenant_id": user.tenant_id, "sid": session_id})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/")
def list_users(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """List all active users in the tenant. Accessible to agents and admins."""
    if current_user.role not in ['agent', 'admin', 'super_admin', 'platform_admin']:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    users = db.query(User).filter(
        User.tenant_id == current_user.tenant_id,
        User.is_active == True
    ).all()
    return [{
        "id": u.id,
        "email": u.email,
        "full_name": u.full_name,
        "role": str(u.role) if hasattr(u.role, 'value') else str(u.role),
        "is_active": u.is_active,
        "job_title": u.job_title,
        "department": u.department,
        "profile_photo": u.profile_photo,
        "availability": u.availability or "online",
        "created_at": str(u.created_at) if u.created_at else None,
    } for u in users]

@app.get("/users/me")
def read_users_me(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    role = (current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role)) if hasattr(current_user.role, 'value') else str(current_user.role)

    # Super admin gets all features unlocked regardless of plan
    if role in ('super_admin', 'platform_admin'):
        limits = get_plan_limits('enterprise')
    else:
        limits = get_plan_limits(tenant.plan if tenant else 'free')

    return {
        "id": current_user.id,
        "email": current_user.email,
        "full_name": current_user.full_name,
        "role": role,
        "is_active": current_user.is_active,
        "language": current_user.language or "en",
        "theme": current_user.theme or "light",
        "profile_photo": current_user.profile_photo,
        "created_at": current_user.created_at,
        "plan_limits": limits,
        "branding": {
            "company_name": tenant.name if tenant else "ITSM Portal",
            "company_tagline": tenant.company_tagline if tenant else None,
            "primary_color": tenant.primary_color if tenant else "#4f46e5",
            "accent_color": tenant.accent_color if tenant else "#818cf8",
            "logo_url": tenant.logo_url if tenant else None,
            "support_email": tenant.support_email if tenant else None,
            "plan_limits": limits,
        } if tenant else None,
    }

# ---------- Tickets (tenant‑scoped + permissions + QUICK FILTERS + CSAT) ----------
def _round_robin_assign(tenant_id: int, group_id: int | None, db) -> int | None:
    """Round-robin auto-assignment.
    Finds the active agent (or agent in the specified group) who was assigned
    a ticket least recently — ensuring even distribution across the team.
    Returns the user_id to assign to, or None if no agents available.
    """
    # Base query — active agents/admins in this tenant
    agent_query = db.query(User).filter(
        User.tenant_id == tenant_id,
        User.is_active == True,
        User.role.in_(['agent', 'admin']),
    )

    if group_id:
        # Restrict to agents in the specified group
        group_member_ids = db.query(GroupMember.user_id).filter(
            GroupMember.group_id == group_id
        ).subquery()
        agent_query = agent_query.filter(User.id.in_(group_member_ids))

    agents = agent_query.all()
    if not agents:
        return None

    agent_ids = [a.id for a in agents]

    # Find last assignment time for each agent
    from sqlalchemy import func as _func
    last_assignments = db.query(
        Ticket.assigned_to_id,
        _func.max(Ticket.created_at).label("last_assigned_at")
    ).filter(
        Ticket.tenant_id == tenant_id,
        Ticket.assigned_to_id.in_(agent_ids),
    ).group_by(Ticket.assigned_to_id).all()

    # Build a map of agent_id → last assigned time
    last_map = {row.assigned_to_id: row.last_assigned_at for row in last_assignments}

    # Sort agents: those never assigned first (None → earliest), then by oldest assignment
    agents_sorted = sorted(
        agents,
        key=lambda a: last_map.get(a.id) or datetime.min
    )

    selected = agents_sorted[0]
    print(f"✅ Round-robin assigned ticket to {selected.full_name} (id={selected.id})")
    return selected.id


def _ticket_tenant_filter(query, ticket_id: int, current_user):
    """Filter ticket by ID, allowing platform_admin to access any tenant."""
    role = str(current_user.role)
    if role in ("platform_admin", "super_admin"):
        return query.filter(Ticket.id == ticket_id)
    return query.filter(Ticket.id == ticket_id, Ticket.tenant_id == current_user.tenant_id)


@app.post("/tickets/", response_model=TicketOut)
def create_ticket(ticket: TicketCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.CREATE_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if tenant:
        trial = get_trial_status(tenant)
        if trial["trial_expired"]:
            raise HTTPException(
                status_code=403,
                detail="Your 14-day free trial has ended. Please upgrade to the Pro plan to continue creating tickets."
            )

    # Determine requester — agents/admins can log on behalf of another user
    requester_id = current_user.id
    if ticket.on_behalf_of_id and has_permission(current_user, Permission.EDIT_TICKETS):
        behalf_user = db.query(User).filter(
            User.id == ticket.on_behalf_of_id,
            User.tenant_id == current_user.tenant_id
        ).first()
        if behalf_user:
            requester_id = behalf_user.id

    # If template_id provided, pre-fill from template
    if ticket.template_id:
        tmpl = db.query(TicketTemplate).filter(
            TicketTemplate.id == ticket.template_id,
            TicketTemplate.tenant_id == current_user.tenant_id
        ).first()
        if tmpl:
            if not ticket.title and tmpl.title:
                ticket.title = tmpl.title
            if not ticket.description and tmpl.description:
                ticket.description = tmpl.description
            if not ticket.category and tmpl.category:
                ticket.category = tmpl.category

    now = datetime.utcnow()
    initial_status = "pending_approval" if str(ticket.ticket_type) == "service_request" else "open"
    try:
        resp, reso = compute_sla_deadlines(str(ticket.priority), now, db, current_user.tenant_id)
    except Exception as e:
        print(f"⚠️ SLA deadline error: {e} — priority={ticket.priority}")
        resp, reso = None, None
    db_ticket = Ticket(
        tenant_id=current_user.tenant_id,
        ticket_type=ticket.ticket_type,
        title=ticket.title,
        description=ticket.description,
        category=ticket.category,
        priority=ticket.priority,
        requester_id=requester_id,
        status=initial_status,
        sla_response_deadline=resp,
        sla_resolution_deadline=reso,
        tags=json.dumps(ticket.tags) if ticket.tags else None,
        group_id=ticket.group_id,
        due_date=ticket.due_date,
        custom_fields_data=json.dumps(ticket.custom_fields_data) if ticket.custom_fields_data else None,
        asset_id=ticket.asset_id,        # link to asset if provided
        created_at=now
    )

    # Auto-assign via round-robin if no agent specified
    if not getattr(ticket, 'assigned_to_id', None):
        rr_agent = _round_robin_assign(current_user.tenant_id, ticket.group_id, db)
        if rr_agent:
            db_ticket.assigned_to_id = rr_agent
    else:
        db_ticket.assigned_to_id = ticket.assigned_to_id

    db.add(db_ticket)
    try:
        db.commit()
        db.refresh(db_ticket)
    except Exception as e:
        import traceback; traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ticket save failed: {type(e).__name__}: {str(e)[:300]}")

    requester = db.query(User).filter(User.id == requester_id).first()
    on_behalf_note = f" (logged by {current_user.full_name} on behalf of {requester.full_name})" if requester_id != current_user.id else ""

    # Post-save actions — all wrapped so they never block the success response
    try:
        notif_cfg = get_email_config(db, current_user.tenant_id)
        ticket_ref = f"{'INC' if db_str(ticket.ticket_type) == 'incident' else 'REQ'}{db_ticket.id:06d}"
        send_notification(
            f"🆕 *New ticket: {ticket_ref}*\n"
            f"*{db_ticket.title}*\n"
            f"From: {requester.full_name if requester else current_user.full_name}{on_behalf_note}\n"
            f"Priority: {db_str(ticket.priority).capitalize()}\n"
            f"<{FRONTEND_URL}/tickets/{db_ticket.id}|View ticket>",
            notif_cfg
        )
    except Exception as e:
        print(f"⚠️ Slack/Teams notification failed (ticket still created): {e}")

    try:
        log_ticket_event(db, db_ticket.id, current_user.tenant_id, current_user.id,
                         action="created",
                         note=f'Ticket "{db_ticket.title}" created{on_behalf_note}.')
        if db_str(ticket.ticket_type) == 'service_request':
            trigger_approval_workflow(db, db_ticket)
        db.commit()
    except Exception as e:
        print(f"⚠️ Audit log / approval workflow error (ticket still created): {e}")

    try:
        if requester and requester.email:
            ticket_id_fmt = f"{'INC' if db_str(ticket.ticket_type) == 'incident' else 'REQ'}{db_ticket.id:06d}"
            _cfg_tc = get_email_config(db, current_user.tenant_id)
            _lang_tc = get_user_language(db, requester.email)
            _priority_tc = str(ticket.priority).capitalize()
            _status_tc = str(initial_status).replace('_', ' ').capitalize() if initial_status else 'Open'
            _title_tc = db_ticket.title
            _name_tc = requester.full_name
            _email_tc = requester.email
            _url_tc = f"{FRONTEND_URL}/tickets/{db_ticket.id}"
            _tid_tc = current_user.tenant_id
            if _lang_tc == 'fr':
                _subj_tc = f"✅ Ticket {ticket_id_fmt} créé : {_title_tc}"
                _body_tc = (f"Bonjour {_name_tc},\n\n"
                            f"Votre ticket a bien été créé et notre équipe vous répondra dans les plus brefs délais.\n\n"
                            f"Ticket : {ticket_id_fmt}\n"
                            f"Titre : {_title_tc}\n"
                            f"Priorité : {_priority_tc}\n\n"
                            f"Merci.")
                _cta_tc = "Voir votre ticket →"
            else:
                _subj_tc = f"✅ Ticket {ticket_id_fmt} created: {_title_tc}"
                _body_tc = (f"Hi {_name_tc},\n\n"
                            f"Your ticket has been successfully created and our team will get back to you shortly.\n\n"
                            f"Ticket: {ticket_id_fmt}\n"
                            f"Title: {_title_tc}\n"
                            f"Priority: {_priority_tc}\n\n"
                            f"Thank you.")
                _cta_tc = "View Your Ticket →"
            import threading as _th_tc
            def _send_ticket_created(_e=_email_tc,_s=_subj_tc,_b=_body_tc,_cfg=_cfg_tc,_u=_url_tc,_cta=_cta_tc,_t=_tid_tc,_l=_lang_tc):
                send_email(_e, _s, _b, _cfg, _u, _cta, None, _t, _l)
            _th_tc.Thread(target=_send_ticket_created, daemon=True).start()
            print(f"📧 Ticket created email sent to {_email_tc} lang={_lang_tc}")
    except Exception as e:
        print(f"⚠️ Ticket created email failed (ticket still created): {e}")

    try:
        run_automation_rules(db_ticket, "on_create", db)
        db.commit()
    except Exception as e:
        print(f"⚠️ on_create automation error (ticket still created): {e}")

    return _ticket_to_out(db_ticket, db)

@app.get("/tickets/")
def list_tickets(
    search: str | None = Query(None),
    assigned: str | None = Query(None),
    assigned_to_id: int | None = Query(None),
    status: str | None = Query(None),
    priority: str | None = Query(None),
    category: str | None = Query(None),
    ticket_type: str | None = Query(None),
    tag: str | None = Query(None),
    group_id: int | None = Query(None),
    asset_id: int | None = Query(None),
    resolved_after: str | None = Query(None, description="Resolved tickets updated after this ISO datetime"),
    updated_after: str | None = Query(None, description="Tickets updated after this ISO datetime"),
    due_date_from: str | None = Query(None, description="Tickets with due_date >= this ISO datetime"),
    due_date_to: str | None = Query(None, description="Tickets with due_date < this ISO datetime"),
    sort_by: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=200),
    tenant_id: int | None = Query(None),  # MSP filter by specific client tenant
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    role = current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)

    # Platform admin (DodoDesk owner) — sees ALL tickets across ALL tenants
    if role == 'platform_admin':
        if tenant_id:
            query = db.query(Ticket).filter(Ticket.tenant_id == tenant_id)
        else:
            query = db.query(Ticket)  # no tenant filter — all tickets

    # Super admin (MSP) — sees tickets from their managed client tenants
    elif role == 'super_admin':
        from sqlalchemy import text as _t2
        access_rows = db.execute(_t2(
            "SELECT DISTINCT tenant_id FROM admin_tenant_access WHERE admin_user_id=:uid "
            "UNION SELECT :own_tid"
        ), {"uid": current_user.id, "own_tid": current_user.tenant_id}).fetchall()
        accessible_tenant_ids = [r[0] for r in access_rows]

        if tenant_id and tenant_id in accessible_tenant_ids:
            query = db.query(Ticket).filter(Ticket.tenant_id == tenant_id)
        elif len(accessible_tenant_ids) > 1:
            query = db.query(Ticket).filter(Ticket.tenant_id.in_(accessible_tenant_ids))
        else:
            query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)

    # Regular admin/agent/employee — own tenant only
    else:
        query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)

    if not has_permission(current_user, Permission.VIEW_ALL_TICKETS):
        query = query.filter(Ticket.requester_id == current_user.id)

    if assigned == "me":
        query = query.filter(Ticket.assigned_to_id == current_user.id)
    elif assigned == "unassigned":
        query = query.filter(Ticket.assigned_to_id == None)

    if assigned_to_id:
        query = query.filter(Ticket.assigned_to_id == assigned_to_id)

    if resolved_after:
        try:
            after_dt = datetime.fromisoformat(resolved_after)
            query = query.filter(Ticket.updated_at >= after_dt)
        except ValueError:
            pass

    if updated_after:
        try:
            query = query.filter(Ticket.updated_at >= datetime.fromisoformat(updated_after))
        except ValueError:
            pass

    if due_date_from:
        try:
            query = query.filter(Ticket.due_date >= datetime.fromisoformat(due_date_from))
        except ValueError:
            pass

    if due_date_to:
        try:
            query = query.filter(Ticket.due_date < datetime.fromisoformat(due_date_to))
        except ValueError:
            pass

    if status:
        if status == "overdue":
            query = query.filter(
                Ticket.sla_resolution_deadline < datetime.utcnow(),
                Ticket.status.in_(['open','in_progress'])
            )
        elif status == "open":
            query = query.filter(Ticket.status.in_(['open','in_progress','pending_approval']))
        else:
            try:
                query = query.filter(Ticket.status == status.lower())
            except ValueError:
                pass

    if priority:
        try:
            query = query.filter(Ticket.priority == priority.lower())
        except ValueError:
            pass

    if category:
        query = query.filter(Ticket.category.ilike(f"%{_sql_safe_search(category)}%"))

    if ticket_type:
        try:
            query = query.filter(Ticket.ticket_type == ticket_type.lower())
        except ValueError:
            pass

    if tag:
        query = query.filter(Ticket.tags.ilike(f'%"{_sql_safe_search(tag)}"%'))

    if group_id:
        query = query.filter(Ticket.group_id == group_id)
    if asset_id:
        query = query.filter(Ticket.asset_id == asset_id)

    if search:
        term = f"%{search}%"
        try:
            ticket_id = int(search)
            query = query.filter(
                (Ticket.id == ticket_id) |
                (Ticket.title.ilike(term)) |
                (Ticket.description.ilike(term))
            )
        except ValueError:
            query = query.filter(
                (Ticket.title.ilike(term)) |
                (Ticket.description.ilike(term))
            )

    total = query.count()

    # Sorting
    if sort_by == "priority":
        from sqlalchemy import case
        priority_order = case(
            (Ticket.priority == 'critical', 0),
            (Ticket.priority == 'high', 1),
            (Ticket.priority == "medium", 2),
            (Ticket.priority == "low", 3),
            else_=4
        )
        query = query.order_by(priority_order, Ticket.created_at.desc())
    elif sort_by == "sla":
        query = query.order_by(Ticket.sla_resolution_deadline.asc().nullslast(), Ticket.created_at.desc())
    else:
        query = query.order_by(Ticket.created_at.desc())

    tickets = query.offset(skip).limit(limit).all()
    return {"items": [_ticket_to_out(t, db) for t in tickets], "total": total, "skip": skip, "limit": limit}

@app.get("/tickets/{ticket_id}", response_model=TicketOut)
def get_ticket(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not has_permission(current_user, Permission.VIEW_ALL_TICKETS) and ticket.requester_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return _ticket_to_out(ticket, db)

@app.patch("/tickets/{ticket_id}", response_model=TicketOut)
def update_ticket(ticket_id: int, update: TicketUpdate,
                  background_tasks: BackgroundTasks,
                  current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    update_data = update.model_dump(exclude_unset=True)
    old_status = (str(ticket.status) if hasattr(ticket.status, "value") else str(ticket.status)) if ticket.status else None
    old_assigned = ticket.assigned_to_id
    if "status" in update_data:
        new_status = update_data["status"]
        # Extract clean string value from enum or string
        _new_status_str = str(new_status).split(".")[-1].lower()
        print(f"🔄 Ticket {ticket_id} status change: {old_status} → {_new_status_str} by user {current_user.id}")
        ticket.status = new_status
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="status_changed", field="status",
                         old_value=old_status,
                         new_value=new_status.value if hasattr(new_status, 'value') else str(new_status))
        # Set resolved_at timestamp when resolved
        if _new_status_str == "resolved":
            ticket.resolved_at = ticket.resolved_at or datetime.utcnow()
        elif _new_status_str in ("open", "in_progress"):
            ticket.resolved_at = None  # clear if reopened via status change
            ticket.csat_token = None   # allow new CSAT email on next resolve
        try:
            # --- CSAT trigger on RESOLVED ---
            if _new_status_str == "resolved":
                requester = db.query(User).filter(User.id == ticket.requester_id).first()
                if requester:
                    _lang2 = get_user_language(db, requester.email)
                    _email = requester.email
                    _name  = requester.full_name
                    _title = ticket.title
                    cfg_csat = get_email_config(db, ticket.tenant_id)
                    _tid_csat = ticket.tenant_id
                    # Generate CSAT token if not already set
                    if not ticket.csat_token:
                        ticket.csat_token = uuid.uuid4().hex
                    survey_url = f"{FRONTEND_URL}/csat/{ticket.csat_token}"
                    _url = survey_url
                    if _lang2 == 'fr':
                        _csat_subj = f"✅ Votre ticket a été résolu : {_title}"
                        _csat_body = (f"Bonjour {_name},\n\nVotre ticket « {_title} » a été résolu.\n"
                                      f"Nous aimerions connaître votre avis sur notre service.")
                        _csat_cta = "Donner mon avis →"
                    else:
                        _csat_subj = f"✅ Your ticket has been resolved: {_title}"
                        _csat_body = (f"Hi {_name},\n\nYour ticket \"{_title}\" has been resolved.\n"
                                      f"Please take a moment to rate our service.")
                        _csat_cta = "Rate our service →"
                    print(f"📧 CSAT email to {_email} for ticket {ticket.id} lang={_lang2}")
                    import threading as _th
                    def _send_csat_email(_e=_email,_s=_csat_subj,_b=_csat_body,_cfg=cfg_csat,_u=_url,_cta=_csat_cta,_t=_tid_csat,_l=_lang2):
                        send_email(_e, _s, _b, _cfg, _u, _cta, None, _t, _l)
                    _th.Thread(target=_send_csat_email, daemon=True).start()

            # --- Status change notification for ALL other statuses ---
            elif _new_status_str in ["open","in_progress","closed","pending_user","pending_approval"]:
                requester = db.query(User).filter(User.id == ticket.requester_id).first()
                if requester and requester.id != current_user.id:
                    _lang = get_user_language(db, requester.email)
                    prefix = "INC" if str(ticket.ticket_type) == 'incident' else "REQ"
                    ticket_ref = f"{prefix}-{ticket.id:04d}"
                    _url = f"{FRONTEND_URL}/tickets/{ticket.id}"
                    new_st = _new_status_str
                    if _lang == 'fr':
                        status_labels_fr = {
                            "open":               "🔓 Ouvert",
                            "in_progress":        "🔄 En cours",
                            "closed":             "🔒 Fermé",
                            "pending_user":       "⏳ En attente de l'utilisateur",
                            "pending_approval":   "⏳ En attente d'approbation",
                        }
                        status_label = status_labels_fr.get(new_st, new_st)
                        _subj = f"[{ticket_ref}] Statut mis à jour : {status_label}"
                        _body = (f"Bonjour {requester.full_name},\n\n"
                                 f"Le statut de votre ticket a été mis à jour.\n\n"
                                 f"Ticket : {ticket_ref}\n"
                                 f"Titre : {ticket.title}\n"
                                 f"Nouveau statut : {status_label}\n"
                                 f"Mis à jour par : {current_user.full_name}\n\n"
                                 f"Merci.")
                        _cta = "Voir le ticket →"
                    else:
                        status_labels_en = {
                            "open":               "🔓 Open",
                            "in_progress":        "🔄 In Progress",
                            "closed":             "🔒 Closed",
                            "pending_user":       "⏳ Pending User",
                            "pending_approval":   "⏳ Pending Approval",
                        }
                        status_label = status_labels_en.get(new_st, new_st)
                        _subj = f"[{ticket_ref}] Status updated: {status_label}"
                        _body = (f"Hi {requester.full_name},\n\n"
                                 f"The status of your ticket has been updated.\n\n"
                                 f"Ticket: {ticket_ref}\n"
                                 f"Title: {ticket.title}\n"
                                 f"New Status: {status_label}\n"
                                 f"Updated by: {current_user.full_name}\n\n"
                                 f"Thank you.")
                        _cta = "View Ticket →"
                    cfg_st = get_email_config(db, ticket.tenant_id)
                    _tid_st = ticket.tenant_id
                    _req_email = requester.email
                    print(f"📧 Status email to {_req_email} lang={_lang} — {ticket_ref} → {status_label}")
                    import threading as _th3
                    def _send_st(_s=_subj, _b=_body, _c=cfg_st, _u=_url, _ct=_cta, _t=_tid_st, _l=_lang, _e=_req_email):
                        send_email(_e, _s, _b, _c, _u, _ct, None, _t, _l)
                    _th3.Thread(target=_send_st, daemon=True).start()
            # --- end status emails ---
        except Exception as _email_err:
            import traceback
            print(f"⚠️ Status notification email failed: {_email_err}")
            traceback.print_exc()

    if "assigned_to_id" in update_data:
        new_assigned = update_data["assigned_to_id"]
        ticket.assigned_to_id = new_assigned
        old_name = db.query(User).filter(User.id == old_assigned).first()
        new_name = db.query(User).filter(User.id == new_assigned).first()
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="assigned", field="assigned_to",
                         old_value=old_name.full_name if old_name else "Unassigned",
                         new_value=new_name.full_name if new_name else "Unassigned")
    if "priority" in update_data:
        old_priority = str(ticket.priority) if ticket.priority else None
        new_priority = update_data["priority"]
        ticket.priority = new_priority
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="priority_changed", field="priority",
                         old_value=old_priority,
                         new_value=new_priority.value if hasattr(new_priority, 'value') else str(new_priority))
    if "category" in update_data and update_data["category"]:
        old_category = ticket.category
        ticket.category = update_data["category"]
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="category_changed", field="category",
                         old_value=old_category, new_value=update_data["category"])
    if "title" in update_data and update_data["title"]:
        old_title = ticket.title
        ticket.title = update_data["title"]
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="title_changed", field="title",
                         old_value=old_title, new_value=update_data["title"])
    if "description" in update_data and update_data["description"]:
        ticket.description = update_data["description"]
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="description_updated", field="description")
    if "tags" in update_data:
        old_tags = json.loads(ticket.tags) if ticket.tags else []
        new_tags = update_data["tags"] or []
        ticket.tags = json.dumps(new_tags)
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="tags_updated", field="tags",
                         old_value=",".join(old_tags), new_value=",".join(new_tags))
    if "group_id" in update_data:
        ticket.group_id = update_data["group_id"]
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="group_assigned", field="group_id",
                         new_value=str(update_data["group_id"]) if update_data["group_id"] else "unassigned")
    if "resolution_note" in update_data and update_data["resolution_note"] is not None:
        ticket.resolution_note = update_data["resolution_note"]
        log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                         action="resolution_added", note="Resolution note updated")
    if "resolution_kb_article_id" in update_data:
        ticket.resolution_kb_article_id = update_data["resolution_kb_article_id"]
        if update_data["resolution_kb_article_id"]:
            log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                             action="kb_linked", note=f"KB article #{update_data['resolution_kb_article_id']} linked as resolution")
    if "due_date" in update_data:
        ticket.due_date = update_data["due_date"]
    if "custom_fields_data" in update_data and update_data["custom_fields_data"] is not None:
        ticket.custom_fields_data = json.dumps(update_data["custom_fields_data"])
    db.commit()
    db.refresh(ticket)
    # Run on_update and on_status_change automation rules
    try:
        run_automation_rules(ticket, "on_update", db)
        if "status" in update_data:
            run_automation_rules(ticket, "on_status_change", db)
        db.commit()
    except Exception as e:
        print(f"⚠️ on_update automation error: {e}")

    # Slack/Teams notification — only for specific status changes that matter to users
    if "status" in update_data:
        try:
            new_status = update_data["status"]
            status_val = new_status.value if hasattr(new_status, "value") else str(new_status)
            ticket_ref = f"{'INC' if str(ticket.ticket_type) == 'incident' else 'REQ'}{ticket.id:06d}"
            ticket_url = f"{FRONTEND_URL}/tickets/{ticket.id}"
            notif_cfg = get_email_config(db, current_user.tenant_id)
            requester = db.query(User).filter(User.id == ticket.requester_id).first()
            requester_name = requester.full_name if requester else "User"

            msg = None
            if status_val == "resolved":
                msg = (
                    f"✅ *{ticket_ref} Resolved*\n"
                    f"*{ticket.title}*\n"
                    f"Resolved by: {current_user.full_name}\n"
                    f"User: {requester_name}\n"
                    f"<{ticket_url}|View ticket>"
                )
            elif status_val == "closed":
                msg = (
                    f"🔒 *{ticket_ref} Closed*\n"
                    f"*{ticket.title}*\n"
                    f"Closed by: {current_user.full_name}\n"
                    f"<{ticket_url}|View ticket>"
                )
            elif status_val == "pending_user":
                msg = (
                    f"⏳ *{ticket_ref} — Waiting for user response*\n"
                    f"*{ticket.title}*\n"
                    f"Agent {current_user.full_name} is waiting for a reply from {requester_name}\n"
                    f"<{ticket_url}|View ticket>"
                )
            elif status_val == "in_progress":
                msg = (
                    f"🔧 *{ticket_ref} In Progress*\n"
                    f"*{ticket.title}*\n"
                    f"Assigned to: {current_user.full_name}\n"
                    f"<{ticket_url}|View ticket>"
                )

            if msg:
                send_notification(msg, notif_cfg)
        except Exception as e:
            print(f"⚠️ Slack/Teams status notification failed: {e}")

    # Notify watchers on status change (in background to avoid blocking)
    if "status" in update_data:
        status_label = update_data["status"].value if hasattr(update_data["status"], "value") else str(update_data["status"])
        import threading
        threading.Thread(
            target=_notify_watchers,
            args=(ticket, f"Status changed to {status_label}", current_user, db),
            kwargs={"exclude_user_id": current_user.id},
            daemon=True
        ).start()
    return _ticket_to_out(ticket, db)

@app.patch("/tickets/{ticket_id}/link-asset")
def link_asset(ticket_id: int, link: LinkAssetRequest,
               current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    ticket.asset_id = link.asset_id
    db.commit()
    db.refresh(ticket)
    try:
        return _ticket_to_out(ticket, db)
    except Exception:
        return {"ok": True, "asset_id": link.asset_id}

def _ticket_to_out(ticket: Ticket, db: Session = None) -> dict:
    requester = ticket.requester
    assigned = ticket.assigned_to if ticket.assigned_to_id else None
    watchers = []
    tenant_name = None
    if db:
        watcher_rows = db.query(TicketWatcher, User).join(
            User, TicketWatcher.user_id == User.id
        ).filter(TicketWatcher.ticket_id == ticket.id).all()
        watchers = [{"user_id": w.user_id, "full_name": u.full_name, "email": u.email}
                    for w, u in watcher_rows]
        try:
            t = db.query(Tenant).filter(Tenant.id == ticket.tenant_id).first()
            tenant_name = t.name if t else None
        except Exception:
            pass
    return {
        "id": ticket.id,
        "tenant_id": ticket.tenant_id,
        "tenant_name": tenant_name,
        "ticket_type": ticket.ticket_type,
        "title": ticket.title,
        "description": ticket.description,
        "category": ticket.category,
        "priority": ticket.priority,
        "status": ticket.status,
        "requester_id": ticket.requester_id,
        "requester_name": requester.full_name if requester else "Unknown",
        "assigned_to_id": ticket.assigned_to_id,
        "assigned_to_name": assigned.full_name if assigned else None,
        "assigned_to_availability": (assigned.availability or "online") if assigned else None,
        "asset_id": ticket.asset_id,
        "sla_response_deadline": ticket.sla_response_deadline,
        "sla_resolution_deadline": ticket.sla_resolution_deadline,
        "sla_status": compute_sla_status(ticket),
        "first_response_at": ticket.first_response_at,
        "tags": json.loads(ticket.tags) if ticket.tags else [],
        "merged_into_id": ticket.merged_into_id,
        "group_id": ticket.group_id,
        "resolution_note": ticket.resolution_note,
        "resolved_at": ticket.resolved_at,
        "resolution_kb_article_id": ticket.resolution_kb_article_id,
        "created_at": ticket.created_at,
        "watchers": watchers,
    }

# =============================================================================
# COLLISION DETECTION — track who is currently viewing a ticket
# =============================================================================
_ticket_viewers = {}  # in-memory presence store: { ticket_id: { user_id: {...} } }

@app.post("/tickets/{ticket_id}/presence")
def update_presence(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Called every 15s by the frontend to register/refresh presence on a ticket."""
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket_id not in _ticket_viewers:
        _ticket_viewers[ticket_id] = {}
    _ticket_viewers[ticket_id][current_user.id] = {
        "user_id": current_user.id,
        "full_name": current_user.full_name,
        "last_seen": datetime.utcnow().isoformat(),
    }
    cutoff = datetime.utcnow().timestamp() - 30
    others = [
        v for uid, v in _ticket_viewers.get(ticket_id, {}).items()
        if uid != current_user.id and
        datetime.fromisoformat(v["last_seen"]).timestamp() > cutoff
    ]
    return {"viewers": others}

@app.delete("/tickets/{ticket_id}/presence")
def remove_presence(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Called when agent leaves the ticket page."""
    if ticket_id in _ticket_viewers:
        _ticket_viewers[ticket_id].pop(current_user.id, None)
    return {"ok": True}

@app.post("/tickets/{ticket_id}/merge")
def merge_ticket(ticket_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Merge ticket_id INTO primary_ticket_id. Moves comments/attachments, closes duplicate."""
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    primary_id = data.get("primary_ticket_id")
    if not primary_id:
        raise HTTPException(status_code=400, detail="primary_ticket_id is required")
    if primary_id == ticket_id:
        raise HTTPException(status_code=400, detail="Cannot merge a ticket into itself")

    duplicate = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    primary = db.query(Ticket).filter(Ticket.id == primary_id, Ticket.tenant_id == current_user.tenant_id).first()
    if not duplicate or not primary:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if duplicate.merged_into_id:
        raise HTTPException(status_code=400, detail="This ticket has already been merged")

    # Move all comments to primary ticket
    db.query(Comment).filter(Comment.ticket_id == ticket_id).update({"ticket_id": primary_id})
    # Move attachments
    db.query(Attachment).filter(Attachment.ticket_id == ticket_id).update({"ticket_id": primary_id})

    # Add a system note on both tickets
    merge_note = Comment(
        ticket_id=primary_id,
        author_id=current_user.id,
        body=f"🔀 Ticket #{ticket_id} was merged into this ticket by {current_user.full_name}.",
        is_internal=True
    )
    db.add(merge_note)

    # Close the duplicate and mark as merged
    duplicate.status = "closed"
    duplicate.merged_into_id = primary_id
    log_ticket_event(db, ticket_id, duplicate.tenant_id, current_user.id,
                     action="merged", note=f"Merged into #{primary_id}")
    log_ticket_event(db, primary_id, primary.tenant_id, current_user.id,
                     action="merge_received", note=f"Received merge from #{ticket_id}")
    db.commit()
    return {"ok": True, "primary_id": primary_id, "merged_id": ticket_id}

# =============================================================================
# TIME TRACKING
# =============================================================================

@app.get("/tickets/{ticket_id}/time-entries")
def list_time_entries(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    entries = db.query(TimeEntry).filter(TimeEntry.ticket_id == ticket_id).order_by(TimeEntry.logged_at.desc()).all()
    total_minutes = sum(e.minutes for e in entries)
    return {
        "entries": [{
            "id": e.id,
            "agent_name": e.agent.full_name if e.agent else "Unknown",
            "agent_id": e.agent_id,
            "minutes": e.minutes,
            "hours": round(e.minutes / 60, 2),
            "note": e.note,
            "logged_at": e.logged_at,
        } for e in entries],
        "total_minutes": total_minutes,
        "total_hours": round(total_minutes / 60, 2),
    }

@app.post("/tickets/{ticket_id}/time-entries")
def log_time(ticket_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    minutes = data.get("minutes")
    if not minutes or int(minutes) <= 0:
        raise HTTPException(status_code=400, detail="Minutes must be a positive number")
    entry = TimeEntry(
        ticket_id=ticket_id,
        agent_id=current_user.id,
        minutes=int(minutes),
        note=data.get("note", "").strip() or None,
    )
    db.add(entry)
    log_ticket_event(db, ticket_id, ticket.tenant_id, current_user.id,
                     action="time_logged", note=f"{minutes}min logged by {current_user.full_name}")
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "minutes": entry.minutes, "note": entry.note, "logged_at": entry.logged_at}

@app.delete("/tickets/{ticket_id}/time-entries/{entry_id}")
def delete_time_entry(ticket_id: int, entry_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Verify the ticket belongs to the current user's tenant first
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    entry = db.query(TimeEntry).filter(TimeEntry.id == entry_id, TimeEntry.ticket_id == ticket_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.agent_id != current_user.id and not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Can only delete your own time entries")
    db.delete(entry)
    db.commit()
    return {"ok": True}

# =============================================================================
# PARENT-CHILD TICKET LINKING
# =============================================================================

@app.get("/tickets/{ticket_id}/links")
def get_ticket_links(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    # Children of this ticket
    children = db.query(TicketLink).filter(TicketLink.parent_id == ticket_id).all()
    # Parent of this ticket
    parent_link = db.query(TicketLink).filter(TicketLink.child_id == ticket_id).first()

    def ticket_summary(t_id):
        t = db.query(Ticket).filter(Ticket.id == t_id).first()
        if not t: return None
        return {"id": t.id, "title": t.title, "status": str(t.status) if t.status else "", "ticket_type": str(t.ticket_type) if t.ticket_type else ""}

    return {
        "parent": ticket_summary(parent_link.parent_id) if parent_link else None,
        "children": [ticket_summary(c.child_id) for c in children if ticket_summary(c.child_id)],
    }

@app.post("/tickets/{ticket_id}/links")
def link_ticket(ticket_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    child_id = data.get("child_id")
    if not child_id:
        raise HTTPException(status_code=400, detail="child_id is required")
    if int(child_id) == ticket_id:
        raise HTTPException(status_code=400, detail="A ticket cannot be its own child")
    # Verify both tickets belong to this tenant
    parent = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    child = db.query(Ticket).filter(Ticket.id == child_id, Ticket.tenant_id == current_user.tenant_id).first()
    if not parent or not child:
        raise HTTPException(status_code=404, detail="Ticket not found")
    # Check not already linked
    existing = db.query(TicketLink).filter(TicketLink.parent_id == ticket_id, TicketLink.child_id == child_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Already linked")
    link = TicketLink(parent_id=ticket_id, child_id=int(child_id), tenant_id=current_user.tenant_id)
    db.add(link)
    log_ticket_event(db, ticket_id, current_user.tenant_id, current_user.id,
                     action="child_linked", note=f"Linked child ticket #{child_id}")
    db.commit()
    return {"ok": True, "parent_id": ticket_id, "child_id": child_id}

@app.delete("/tickets/{ticket_id}/links/{child_id}")
def unlink_ticket(ticket_id: int, child_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    link = db.query(TicketLink).filter(TicketLink.parent_id == ticket_id, TicketLink.child_id == child_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    db.delete(link)
    db.commit()
    return {"ok": True}


@app.post("/tickets/{ticket_id}/reopen")
def reopen_ticket(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Re-open a resolved or closed ticket. Agents/admins only."""
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket.status not in ["resolved", "closed"]:
        raise HTTPException(status_code=400, detail="Only resolved or closed tickets can be reopened.")
    old_status = (str(ticket.status) if hasattr(ticket.status, "value") else str(ticket.status))
    ticket.status = "open"
    ticket.csat_token = None  # Reset CSAT so it can be re-sent on next resolution
    log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                     action="status_changed", field="status",
                     old_value=old_status, new_value="open")
    db.commit()
    db.refresh(ticket)
    # Notify requester
    requester = db.query(User).filter(User.id == ticket.requester_id).first()
    if requester and requester.id != current_user.id:
        _rl2 = get_user_language(db, requester.email)
        if _rl2 == 'fr':
            _rs2 = f"Ticket rouvert : {ticket.title}"
            _rb2 = f"Bonjour {requester.full_name},\n\nVotre ticket \"{ticket.title}\" a été rouvert et est en cours de traitement."
        else:
            _rs2 = f"Ticket reopened: {ticket.title}"
            _rb2 = f"Hi {requester.full_name},\n\nYour ticket \"{ticket.title}\" has been reopened and is being worked on again."
        send_email(requester.email, _rs2, _rb2, cta_url=f"{FRONTEND_URL}/tickets/{ticket.id}", cta_label="View ticket →" if _rl2 != 'fr' else "Voir le ticket →", db=None, tenant_id=ticket.tenant_id, lang=_rl2)
    return _ticket_to_out(ticket, db)

# ---------- Ticket Watchers ----------

def _notify_watchers(ticket: Ticket, event: str, actor: User, db: Session, exclude_user_id: int = None):
    """Send email notifications to all watchers of a ticket."""
    watchers = db.query(TicketWatcher).filter(TicketWatcher.ticket_id == ticket.id).all()
    prefix = "INC" if str(ticket.ticket_type) == 'incident' else "REQ"
    ticket_ref = f"{prefix}-{ticket.id:04d}"
    for w in watchers:
        if w.user_id == exclude_user_id:
            continue
        watcher_user = db.query(User).filter(User.id == w.user_id).first()
        if watcher_user:
            _wl = get_user_language(db, watcher_user.email)
            if _wl == 'fr':
                _ws = f"[Observation] {ticket_ref} : {ticket.title} — {event}"
                _wb = (f"Bonjour {watcher_user.full_name},\n\n"
                       f"Mise à jour d'un ticket que vous observez :\n\n"
                       f"Ticket : {ticket_ref} — {ticket.title}\n"
                       f"Mise à jour : {event}\n"
                       f"Par : {actor.full_name}\n\n")
            else:
                _ws = f"[Watching] {ticket_ref}: {ticket.title} — {event}"
                _wb = (f"Hi {watcher_user.full_name},\n\n"
                       f"An update on a ticket you're watching:\n\n"
                       f"Ticket: {ticket_ref} — {ticket.title}\n"
                       f"Update: {event}\n"
                       f"By: {actor.full_name}\n\n")
            send_email(
                watcher_user.email, _ws, _wb,
                f"To stop watching this ticket, open it and click 'Unwatch'."
            )

@app.get("/tickets/{ticket_id}/watchers")
def get_watchers(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all watchers for a ticket."""
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    rows = db.query(TicketWatcher, User).join(User, TicketWatcher.user_id == User.id).filter(
        TicketWatcher.ticket_id == ticket_id
    ).all()
    return [{"user_id": w.user_id, "full_name": u.full_name, "email": u.email} for w, u in rows]

@app.post("/tickets/{ticket_id}/watch")
def watch_ticket(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Add current user as a watcher."""
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    existing = db.query(TicketWatcher).filter(
        TicketWatcher.ticket_id == ticket_id, TicketWatcher.user_id == current_user.id
    ).first()
    if existing:
        return {"ok": True, "watching": True, "message": "Already watching"}
    db.add(TicketWatcher(ticket_id=ticket_id, user_id=current_user.id, tenant_id=current_user.tenant_id))
    db.commit()
    return {"ok": True, "watching": True, "message": f"You are now watching ticket {ticket_id}"}

@app.delete("/tickets/{ticket_id}/watch")
def unwatch_ticket(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Remove current user as a watcher."""
    watcher = db.query(TicketWatcher).filter(
        TicketWatcher.ticket_id == ticket_id, TicketWatcher.user_id == current_user.id
    ).first()
    if watcher:
        db.delete(watcher)
        db.commit()
    return {"ok": True, "watching": False}

@app.post("/tickets/{ticket_id}/watchers/add")
def add_watcher(ticket_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Agent/admin adds another user as a watcher."""
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    user_id = data.get("user_id")
    user = db.query(User).filter(User.id == user_id, User.tenant_id == current_user.tenant_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    existing = db.query(TicketWatcher).filter(
        TicketWatcher.ticket_id == ticket_id, TicketWatcher.user_id == user_id
    ).first()
    if not existing:
        db.add(TicketWatcher(ticket_id=ticket_id, user_id=user_id, tenant_id=current_user.tenant_id))
        db.commit()
    return {"ok": True, "message": f"{user.full_name} is now watching this ticket"}

@app.delete("/tickets/{ticket_id}/watchers/{user_id}")
def remove_watcher(ticket_id: int, user_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Agent/admin removes a watcher."""
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    watcher = db.query(TicketWatcher).filter(
        TicketWatcher.ticket_id == ticket_id, TicketWatcher.user_id == user_id,
        TicketWatcher.tenant_id == current_user.tenant_id
    ).first()
    if watcher:
        db.delete(watcher)
        db.commit()
    return {"ok": True}

# ---------- Approval workflow ----------
@app.post("/tickets/{ticket_id}/approve", response_model=TicketOut)
def approve_ticket(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket.status != "pending_approval":
        raise HTTPException(status_code=400, detail="Ticket is not in pending approval status")
    ticket.status = "open"
    log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                     action="approved", field="status",
                     old_value="pending_approval", new_value="open")
    db.commit()
    db.refresh(ticket)
    requester = db.query(User).filter(User.id == ticket.requester_id).first()
    if requester:
        _rl = get_user_language(db, requester.email)
        if _rl == 'fr':
            _rs = f"Votre demande a été approuvée : #{ticket.id} {ticket.title}"
            _rb = f"Votre demande de service a été approuvée et est en cours de traitement."
            _rc = "Voir le ticket →"
        else:
            _rs = f"Your request has been approved: #{ticket.id} {ticket.title}"
            _rb = f"Your service request has been approved and is now being processed."
            _rc = "View ticket →"
        send_email(requester.email, _rs, _rb, cta_url=f"{FRONTEND_URL}/tickets/{ticket.id}", cta_label=_rc, db=None, tenant_id=ticket.tenant_id, lang=_rl)
    return _ticket_to_out(ticket, db)

@app.post("/tickets/{ticket_id}/reject", response_model=TicketOut)
def reject_ticket(ticket_id: int, comment: CommentCreate,
                  current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket.status != "pending_approval":
        raise HTTPException(status_code=400, detail="Ticket is not in pending approval status")
    ticket.status = "closed"
    db_comment = Comment(ticket_id=ticket_id, author_id=current_user.id, body=comment.body)
    db.add(db_comment)
    log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                     action="rejected", field="status",
                     old_value="pending_approval", new_value="closed",
                     note=comment.body)
    db.commit()
    db.refresh(ticket)
    requester = db.query(User).filter(User.id == ticket.requester_id).first()
    if requester:
        _rl = get_user_language(db, requester.email)
        if _rl == 'fr':
            _rs = f"Votre demande a été rejetée : #{ticket.id} {ticket.title}"
            _rb = f"Votre demande de service a été rejetée.\nRaison : {comment.body}"
            _rc = "Voir le ticket →"
        else:
            _rs = f"Your request has been rejected: #{ticket.id} {ticket.title}"
            _rb = f"Your service request has been rejected.\nReason: {comment.body}"
            _rc = "View ticket →"
        send_email(requester.email, _rs, _rb, cta_url=f"{FRONTEND_URL}/tickets/{ticket.id}", cta_label=_rc, db=None, tenant_id=ticket.tenant_id, lang=_rl)
    return _ticket_to_out(ticket, db)

# ---------- Comments ----------
@app.post("/tickets/{ticket_id}/comments", response_model=CommentOut)
def add_comment(ticket_id: int, comment: CommentCreate,
                current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not has_permission(current_user, Permission.EDIT_TICKETS) and ticket.requester_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    # Only agents/admins can post internal notes
    is_internal = comment.is_internal and has_permission(current_user, Permission.EDIT_TICKETS)

    db_comment = Comment(ticket_id=ticket_id, author_id=current_user.id, body=comment.body, is_internal=is_internal)
    db.add(db_comment)

    # Track first response time — set when an agent/admin posts the first non-internal reply
    if not is_internal and has_permission(current_user, Permission.EDIT_TICKETS):
        if not ticket.first_response_at and ticket.requester_id != current_user.id:
            ticket.first_response_at = datetime.utcnow()
            log_ticket_event(db, ticket_id, ticket.tenant_id, current_user.id,
                             action="first_response", note=f"First response by {current_user.full_name}")

    db.commit()
    db.refresh(db_comment)
    log_ticket_event(db, ticket_id, ticket.tenant_id, current_user.id,
                     action="internal_note_added" if is_internal else "comment_added",
                     note=f'{comment.body[:120]}{"..." if len(comment.body) > 120 else ""}')
    db.commit()

    # Process @mentions — notify mentioned agents
    if is_internal and "@" in comment.body:
        process_mentions(comment.body, ticket_id, current_user.tenant_id, current_user, db)
        db.commit()

    # Don't send email/notification for internal notes — they're agent-only
    if not is_internal:
        if str(current_user.role) in ["agent", "admin", "super_admin", "platform_admin"] and ticket.requester_id != current_user.id:
            requester = db.query(User).filter(User.id == ticket.requester_id).first()
            if requester:
                _lang = get_user_language(db, requester.email)
                if _lang == 'fr':
                    _subj = f"Nouveau commentaire sur le ticket #{ticket.id} : {ticket.title}"
                    _body = f"L'agent {current_user.full_name} a répondu :\n\n{comment.body}"
                    _cta = "Voir le commentaire →"
                else:
                    _subj = f"New reply on ticket #{ticket.id}: {ticket.title}"
                    _body = f"Agent {current_user.full_name} replied:\n\n{comment.body}"
                    _cta = "View ticket →"
                _cfg_c = get_email_config(db, ticket.tenant_id)
                _tid_c = ticket.tenant_id
                _url_c = f"{FRONTEND_URL}/tickets/{ticket.id}"
                _email_c = requester.email
                import threading as _th_c
                def _send_comment(_e=_email_c,_s=_subj,_b=_body,_cfg=_cfg_c,_u=_url_c,_cta=_cta,_t=_tid_c,_l=_lang):
                    send_email(_e, _s, _b, _cfg, _u, _cta, None, _t, _l)
                _th_c.Thread(target=_send_comment, daemon=True).start()
                print(f"📧 Comment email to {_email_c} lang={_lang}")
        comment_cfg = get_email_config(db, current_user.tenant_id)
        send_notification(
            f"💬 New comment on ticket #{ticket.id} *{ticket.title}*\n"
            f"By: {current_user.full_name}\n"
            f"Comment: {comment.body[:100]}{'...' if len(comment.body) > 100 else ''}\n"
            f"View: {FRONTEND_URL}/tickets/{ticket.id}",
            comment_cfg
        )
        # Notify watchers in background
        import threading
        threading.Thread(
            target=_notify_watchers,
            args=(ticket, f"New comment by {current_user.full_name}: {comment.body[:80]}{'...' if len(comment.body) > 80 else ''}", current_user, db),
            kwargs={"exclude_user_id": current_user.id},
            daemon=True
        ).start()
    return {"id": db_comment.id, "ticket_id": db_comment.ticket_id, "author_id": db_comment.author_id,
            "author_name": current_user.full_name, "body": db_comment.body,
            "is_internal": db_comment.is_internal, "created_at": db_comment.created_at}

@app.get("/tickets/{ticket_id}/comments", response_model=list[CommentOut])
def list_comments(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not has_permission(current_user, Permission.VIEW_ALL_TICKETS) and ticket.requester_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    comments = db.query(Comment).filter(Comment.ticket_id == ticket_id).all()
    is_agent_or_admin = has_permission(current_user, Permission.EDIT_TICKETS)
    result = []
    for c in comments:
        # Requesters (employees) cannot see internal notes
        if c.is_internal and not is_agent_or_admin:
            continue
        author = db.query(User).filter(User.id == c.author_id).first()
        result.append({"id": c.id, "ticket_id": c.ticket_id, "author_id": c.author_id,
                       "author_name": author.full_name if author else "Unknown",
                       "body": c.body, "is_internal": c.is_internal,
                       "created_at": c.created_at})
    return result

@app.get("/tickets/{ticket_id}/audit-log")
def get_audit_log(ticket_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not has_permission(current_user, Permission.VIEW_ALL_TICKETS) and ticket.requester_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    entries = db.query(TicketAuditLog).filter(
        TicketAuditLog.ticket_id == ticket_id
    ).order_by(TicketAuditLog.created_at.asc()).all()
    result = []
    for e in entries:
        actor = db.query(User).filter(User.id == e.actor_id).first()
        result.append({
            "id": e.id,
            "action": e.action,
            "field": e.field,
            "old_value": e.old_value,
            "new_value": e.new_value,
            "note": e.note,
            "actor_name": actor.full_name if actor else "Unknown",
            "created_at": e.created_at,
        })
    return result

@app.get("/admin/audit-log")
def get_system_audit_log(
    search: str | None = Query(None),
    action: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """System-wide audit log for the tenant. Shows all admin actions."""
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("audit_log", tenant, "Full audit log is available on the Business plan and above. Please upgrade.")
    query = db.query(SystemAuditLog).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    )
    if action:
        query = query.filter(SystemAuditLog.action.ilike(f"{_sql_safe_search(action)}%"))
    if search:
        s = f"%{_sql_safe_search(search)}%"
        query = query.filter(
            SystemAuditLog.action.ilike(s) |
            SystemAuditLog.target_label.ilike(s) |
            SystemAuditLog.actor_email.ilike(s)
        )
    if start_date:
        try:
            query = query.filter(SystemAuditLog.created_at >= datetime.fromisoformat(start_date))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(SystemAuditLog.created_at <= datetime.fromisoformat(end_date))
        except Exception:
            pass

    total = query.count()
    logs  = query.order_by(SystemAuditLog.created_at.desc()).offset(skip).limit(limit).all()

    # Build by_category counts for the frontend
    all_logs = db.query(SystemAuditLog.action).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    ).all()
    by_category: dict = {}
    for (act,) in all_logs:
        if act:
            cat = act.split('.')[0]
            by_category[cat] = by_category.get(cat, 0) + 1

    # Resolve actor names
    actor_ids = {l.actor_id for l in logs if l.actor_id}
    actor_map  = {}
    if actor_ids:
        actors = db.query(User).filter(User.id.in_(actor_ids)).all()
        actor_map = {u.id: u.full_name for u in actors}

    items = [
        {
            "id": l.id,
            "action": l.action or "",
            "target_type": getattr(l, "target_type", "") or "",
            "target_id": getattr(l, "target_id", None),
            "target_label": getattr(l, "target_label", "") or "",
            "old_value": getattr(l, "old_value", "") or "",
            "new_value": getattr(l, "new_value", "") or "",
            "actor_id": l.actor_id,
            "actor_name": actor_map.get(l.actor_id, ""),
            "actor_email": getattr(l, "actor_email", "") or "",
            "created_at": str(l.created_at)[:19] if l.created_at else "",
        }
        for l in logs
    ]
    return {"items": items, "total": total, "by_category": by_category}

@app.get("/admin/audit-log/export/csv")
def export_audit_log_csv(
    search: str | None = Query(None),
    action: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Export the full audit log as CSV."""
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("audit_log", tenant, "Audit log export is available on the Business plan and above.")
    query = db.query(SystemAuditLog).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    )
    if action:
        query = query.filter(SystemAuditLog.action.ilike(f"{_sql_safe_search(action)}%"))
    if search:
        s = f"%{_sql_safe_search(search)}%"
        query = query.filter(
            SystemAuditLog.action.ilike(s) |
            SystemAuditLog.actor_email.ilike(s)
        )
    logs = query.order_by(SystemAuditLog.created_at.desc()).limit(10000).all()

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "Action", "Target Type", "Target", "Actor Email", "Old Value", "New Value"])
    for l in logs:
        writer.writerow([
            str(l.created_at)[:19] if l.created_at else "",
            l.action or "",
            getattr(l, "target_type", "") or "",
            getattr(l, "target_label", "") or "",
            getattr(l, "actor_email", "") or "",
            getattr(l, "old_value", "") or "",
            getattr(l, "new_value", "") or "",
        ])
    output.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=\"audit_log.csv\""}
    )


@app.get("/kb/articles/")
def search_kb_articles(search: str | None = Query(None), skip: int = Query(0, ge=0),
                       limit: int = Query(20, ge=1, le=200), status: str | None = Query(None),
                       category: str | None = Query(None), folder: str | None = Query(None),
                       tag: str | None = Query(None), needs_review: bool = Query(False),
                       db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(KBArticle).filter(KBArticle.tenant_id == current_user.tenant_id)
    # Employees only see published + visible articles
    if not has_permission(current_user, Permission.MANAGE_KB):
        query = query.filter(KBArticle.status == "published")
        query = query.filter(KBArticle.visibility.in_(["all", "employees_only"]))
    else:
        if status:
            query = query.filter(KBArticle.status == status)
        # agents_only articles visible to agents/admins
        query = query.filter(KBArticle.visibility.in_(["all", "agents_only"]))
    if search:
        term = f"%{search}%"
        query = query.filter(KBArticle.title.ilike(term) | KBArticle.content.ilike(term) | KBArticle.tags.ilike(term))
    if category:
        query = query.filter(KBArticle.category == category)
    if folder:
        query = query.filter(KBArticle.folder == folder)
    if tag:
        query = query.filter(KBArticle.tags.ilike(f'%"{_sql_safe_search(tag)}"%'))
    if needs_review:
        query = query.filter(KBArticle.review_date.isnot(None), KBArticle.review_date < datetime.utcnow())
    total = query.count()
    articles = query.order_by(KBArticle.sort_order, KBArticle.updated_at.desc()).offset(skip).limit(limit).all()
    result = []
    for art in articles:
        author = db.query(User).filter(User.id == art.author_id).first()
        result.append({
            "id": art.id, "title": art.title, "content": art.content,
            "category": art.category, "folder": art.folder,
            "author_id": art.author_id, "author_name": author.full_name if author else "Unknown",
            "status": art.status or "published", "version": art.version or 1,
            "view_count": art.view_count or 0,
            "helpful_count": art.helpful_count or 0,
            "not_helpful_count": art.not_helpful_count or 0,
            "tags": json.loads(art.tags) if art.tags else [],
            "visibility": art.visibility or "all",
            "review_date": art.review_date,
            "sort_order": art.sort_order or 0,
            "created_at": art.created_at, "updated_at": art.updated_at,
        })
    return {"items": result, "total": total, "skip": skip, "limit": limit}

@app.get("/kb/articles/{article_id}/versions")
def get_kb_versions(article_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Full version history. Agents/admins only."""
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    article = db.query(KBArticle).filter(KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    versions = db.query(KBVersion).filter(KBVersion.article_id == article_id).order_by(KBVersion.version_number.desc()).all()
    return [{"id": v.id, "version_number": v.version_number, "title": v.title,
             "content": v.content, "category": v.category, "status": v.status,
             "change_note": v.change_note,
             "edited_by": v.edited_by.full_name if v.edited_by else "Unknown",
             "created_at": v.created_at} for v in versions]

@app.post("/kb/articles/{article_id}/restore/{version_id}")
def restore_kb_version(article_id: int, version_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Restore a previous version as current content."""
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    article = db.query(KBArticle).filter(KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    version = db.query(KBVersion).filter(KBVersion.id == version_id, KBVersion.article_id == article_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")
    new_ver_num = (article.version or 1) + 1
    db.add(KBVersion(article_id=article_id, version_number=new_ver_num,
                     title=article.title, content=article.content, category=article.category,
                     status=article.status, change_note=f"Restored from v{version.version_number}",
                     edited_by_id=current_user.id))
    article.title = version.title
    article.content = version.content
    article.category = version.category
    article.version = new_ver_num
    article.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "restored_from_version": version.version_number, "new_version": new_ver_num}

@app.get("/kb/articles/{article_id}")
def get_kb_article(article_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    article = db.query(KBArticle).filter(KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    # Only count views for employees (not agents editing)
    if not has_permission(current_user, Permission.MANAGE_KB):
        article.view_count = (article.view_count or 0) + 1
        db.commit()
    author = db.query(User).filter(User.id == article.author_id).first()
    return {"id": article.id, "title": article.title, "content": article.content,
            "category": article.category, "folder": article.folder,
            "author_id": article.author_id, "author_name": author.full_name if author else "Unknown",
            "status": article.status or "published", "version": article.version or 1,
            "view_count": article.view_count or 0,
            "helpful_count": article.helpful_count or 0,
            "not_helpful_count": article.not_helpful_count or 0,
            "tags": json.loads(article.tags) if article.tags else [],
            "visibility": article.visibility or "all",
            "review_date": article.review_date,
            "sort_order": article.sort_order or 0,
            "custom_fields_data": json.loads(article.custom_fields_data) if article.custom_fields_data else {},
            "created_at": article.created_at, "updated_at": article.updated_at}

@app.post("/kb/articles/")
def create_kb_article(article: KBArticleCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_article = KBArticle(
        tenant_id=current_user.tenant_id,
        title=article.title, content=article.content, category=article.category,
        folder=article.folder, author_id=current_user.id, status=article.status, version=1,
        tags=json.dumps(article.tags) if article.tags else None,
        visibility=article.visibility or "all",
        review_date=article.review_date,
        custom_fields_data=json.dumps(article.custom_fields_data) if article.custom_fields_data else None,
    )
    db.add(db_article)
    db.flush()
    db.add(KBVersion(article_id=db_article.id, version_number=1, title=article.title,
                     content=article.content, category=article.category, status=article.status,
                     change_note="Initial version", edited_by_id=current_user.id))
    db.commit()
    db.refresh(db_article)
    return {"id": db_article.id, "title": db_article.title, "content": db_article.content,
            "category": db_article.category, "folder": db_article.folder,
            "author_id": db_article.author_id, "author_name": current_user.full_name,
            "status": db_article.status, "version": db_article.version,
            "view_count": db_article.view_count or 0,
            "helpful_count": 0, "not_helpful_count": 0,
            "tags": article.tags or [], "visibility": db_article.visibility or "all",
            "review_date": db_article.review_date, "sort_order": 0,
            "custom_fields_data": article.custom_fields_data or {},
            "created_at": db_article.created_at, "updated_at": db_article.updated_at}

@app.post("/tickets/{ticket_id}/create-kb-article")
def create_kb_from_ticket(ticket_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Create a KB article pre-filled from ticket resolution note. Links it back to the ticket."""
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    title = data.get("title", ticket.title)
    content = data.get("content") or ticket.resolution_note or ""
    category = data.get("category", ticket.category or "General")
    if not content:
        raise HTTPException(status_code=400, detail="Resolution note is empty — add a resolution note before creating a KB article")
    article = KBArticle(tenant_id=current_user.tenant_id, title=title, content=content,
                        category=category, author_id=current_user.id, status="draft", version=1)
    db.add(article)
    db.flush()
    # Initial version snapshot
    db.add(KBVersion(article_id=article.id, version_number=1, title=title, content=content,
                     category=category, status="draft",
                     change_note="Created from ticket resolution", edited_by_id=current_user.id))
    db.add(article)
    db.flush()
    # Link the article back to the ticket
    ticket.resolution_kb_article_id = article.id
    log_ticket_event(db, ticket_id, ticket.tenant_id, current_user.id,
                     action="kb_created", note=f"KB article created from resolution: {title}")
    db.commit()
    db.refresh(article)
    return {"id": article.id, "title": article.title, "category": article.category, "created_at": article.created_at}

@app.put("/kb/articles/{article_id}")
async def update_kb_article(article_id: int, request: Request,
                      current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    raw_body = await request.json()
    # Convert empty strings to None so Pydantic accepts optional fields
    cleaned = {k: (None if v == "" else v) for k, v in raw_body.items()}
    article = KBArticleUpdate(**cleaned)
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_article = db.query(KBArticle).filter(KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id).first()
    if not db_article:
        raise HTTPException(status_code=404, detail="Article not found")
    if article.category is not None and not article.category.strip():
        raise HTTPException(status_code=422, detail="Category is required")
    update_data = article.model_dump(exclude_unset=True)
    change_note = update_data.pop("change_note", None)
    new_version = (db_article.version or 1) + 1
    db.add(KBVersion(
        article_id=article_id, version_number=new_version,
        title=update_data.get("title", db_article.title),
        content=update_data.get("content", db_article.content),
        category=update_data.get("category", db_article.category),
        status=update_data.get("status", db_article.status),
        change_note=change_note, edited_by_id=current_user.id
    ))
    for field in ["title", "content", "category", "folder", "status", "visibility", "review_date", "sort_order"]:
        if field in update_data:
            setattr(db_article, field, update_data[field])
    if "tags" in update_data:
        db_article.tags = json.dumps(update_data["tags"]) if update_data["tags"] else None
    if "custom_fields_data" in update_data:
        db_article.custom_fields_data = json.dumps(update_data["custom_fields_data"]) if update_data["custom_fields_data"] else None
    db_article.version = new_version
    db_article.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_article)
    author = db.query(User).filter(User.id == db_article.author_id).first()
    return {"id": db_article.id, "title": db_article.title, "content": db_article.content,
            "category": db_article.category, "folder": db_article.folder,
            "author_id": db_article.author_id, "author_name": author.full_name if author else "Unknown",
            "status": db_article.status, "version": db_article.version,
            "view_count": db_article.view_count or 0,
            "helpful_count": db_article.helpful_count or 0,
            "not_helpful_count": db_article.not_helpful_count or 0,
            "tags": json.loads(db_article.tags) if db_article.tags else [],
            "visibility": db_article.visibility or "all",
            "review_date": db_article.review_date, "sort_order": db_article.sort_order or 0,
            "custom_fields_data": json.loads(db_article.custom_fields_data) if db_article.custom_fields_data else {},
            "created_at": db_article.created_at, "updated_at": db_article.updated_at}

@app.delete("/kb/articles/{article_id}")
def delete_kb_article(article_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_article = db.query(KBArticle).filter(KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id).first()
    if not db_article:
        raise HTTPException(status_code=404, detail="Article not found")
    db.delete(db_article)
    db.commit()
    return {"detail": "Article deleted"}

@app.post("/kb/articles/{article_id}/feedback")
def submit_kb_feedback(article_id: int, data: dict,
                       current_user: User = Depends(get_current_user),
                       db: Session = Depends(get_db)):
    """Submit 👍/👎 feedback on a KB article."""
    article = db.query(KBArticle).filter(
        KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id
    ).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    helpful = data.get("helpful")  # True = 👍, False = 👎
    if helpful is True:
        article.helpful_count = (article.helpful_count or 0) + 1
    elif helpful is False:
        article.not_helpful_count = (article.not_helpful_count or 0) + 1
    db.commit()
    return {"helpful_count": article.helpful_count, "not_helpful_count": article.not_helpful_count}

@app.get("/kb/categories")
def get_kb_categories(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all distinct categories and folders for KB navigation."""
    query = db.query(KBArticle.category, KBArticle.folder).filter(
        KBArticle.tenant_id == current_user.tenant_id,
        KBArticle.status == "published"
    )
    if not has_permission(current_user, Permission.MANAGE_KB):
        query = query.filter(KBArticle.visibility.in_(["all", "employees_only"]))
    rows = query.distinct().all()
    structure = {}
    for cat, folder in rows:
        cat = cat or "General"
        if cat not in structure:
            structure[cat] = []
        if folder and folder not in structure[cat]:
            structure[cat].append(folder)
    return [{"category": cat, "folders": folders} for cat, folders in sorted(structure.items())]

@app.get("/kb/articles/{article_id}/related")
def get_related_articles(article_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get related articles based on same category/tags."""
    article = db.query(KBArticle).filter(
        KBArticle.id == article_id, KBArticle.tenant_id == current_user.tenant_id
    ).first()
    if not article:
        return []
    query = db.query(KBArticle).filter(
        KBArticle.tenant_id == current_user.tenant_id,
        KBArticle.id != article_id,
        KBArticle.status == "published",
        KBArticle.category == article.category
    ).limit(5)
    related = query.all()
    return [{"id": a.id, "title": a.title, "category": a.category,
             "view_count": a.view_count or 0} for a in related]

@app.get("/kb/insights")
def get_kb_insights(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """KB insights for agents/admins — most viewed, least helpful, needs review."""
    if not has_permission(current_user, Permission.MANAGE_KB):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    articles = db.query(KBArticle).filter(
        KBArticle.tenant_id == current_user.tenant_id,
        KBArticle.status == "published"
    ).all()
    most_viewed = sorted(articles, key=lambda a: a.view_count or 0, reverse=True)[:5]
    least_helpful = [a for a in articles if (a.not_helpful_count or 0) > 0]
    least_helpful = sorted(least_helpful, key=lambda a: (a.not_helpful_count or 0), reverse=True)[:5]
    needs_review_all = [a for a in articles if a.review_date and a.review_date < datetime.utcnow()]
    needs_review = needs_review_all[:5]
    def fmt(a):
        return {"id": a.id, "title": a.title, "category": a.category,
                "view_count": a.view_count or 0,
                "helpful_count": a.helpful_count or 0,
                "not_helpful_count": a.not_helpful_count or 0,
                "review_date": a.review_date}
    return {
        "most_viewed": [fmt(a) for a in most_viewed],
        "least_helpful": [fmt(a) for a in least_helpful],
        "needs_review": [fmt(a) for a in needs_review],
        "needs_review_count": len(needs_review_all),
        "total_articles": len(articles),
        "total_views": sum(a.view_count or 0 for a in articles),
    }

# ---------- Asset Management (tenant‑scoped + permissions) ----------
def _asset_to_out(a, db):
    assigned = db.query(User).filter(User.id == a.assigned_to_id).first() if a.assigned_to_id else None
    ticket_count = db.query(Ticket).filter(Ticket.asset_id == a.id).count()
    try:
        asset_type = str(a.type) if hasattr(a.type, 'value') else str(a.type).lower() if a.type else None
    except Exception:
        asset_type = str(a.type) if a.type else None
    try:
        asset_status = str(a.status) if hasattr(a.status, 'value') else str(a.status) if a.status else None
    except Exception:
        asset_status = str(a.status) if a.status else None
    return {
        "id": a.id, "name": a.name, "type": asset_type, "model": a.model, "serial_number": a.serial_number,
        "status": asset_status, "assigned_to_id": a.assigned_to_id,
        "assigned_to_name": assigned.full_name if assigned else None,
        "purchase_date": str(a.purchase_date) if a.purchase_date else None,
        "license_key": a.license_key,
        "vendor": a.vendor,
        "expiry_date": str(a.expiry_date) if a.expiry_date else None,
        "notes": a.notes,
        "location": a.location, "purchase_cost": float(a.purchase_cost) if a.purchase_cost else None,
        "warranty_expiry": str(a.warranty_expiry) if a.warranty_expiry else None,
        "contract_number": a.contract_number,
        "quantity": a.quantity or 1, "seats_total": a.seats_total,
        "seats_used": a.seats_used or 0,
        "maintenance_date": str(a.maintenance_date) if a.maintenance_date else None,
        "parent_asset_id": a.parent_asset_id, "tag_number": a.tag_number,
        "custom_fields_data": json.loads(a.custom_fields_data) if a.custom_fields_data else {},
        "ticket_count": ticket_count,
        "created_at": str(a.created_at) if a.created_at else None,
        "updated_at": str(a.updated_at) if a.updated_at else None,
    }

@app.get("/assets/")
def list_assets(search: str | None = Query(None), skip: int = Query(0, ge=0),
                limit: int = Query(20, ge=1, le=200),
                asset_type: str | None = Query(None),
                status: str | None = Query(None),
                location: str | None = Query(None),
                expiring_soon: bool = Query(False),
                days: int = Query(90),
                db: Session = Depends(get_db),
                current_user: User = Depends(get_current_user)):
    try:
        query = db.query(Asset).filter(Asset.tenant_id == current_user.tenant_id)
        if expiring_soon:
            cutoff = datetime.utcnow().date() + timedelta(days=days)
            today = datetime.utcnow().date()
            query = query.filter(
                ((Asset.expiry_date != None) & (Asset.expiry_date >= today) & (Asset.expiry_date <= cutoff)) |
                ((Asset.warranty_expiry != None) & (Asset.warranty_expiry >= today) & (Asset.warranty_expiry <= cutoff))
            )
        if search:
            term = f"%{search}%"
            query = query.filter(
                Asset.name.ilike(term) | Asset.serial_number.ilike(term) |
                Asset.vendor.ilike(term) | Asset.tag_number.ilike(term) |
                Asset.location.ilike(term)
            )
        if asset_type:
            # Use text comparison to avoid enum case mismatch
            from sqlalchemy import text as _t, cast, String
            query = query.filter(cast(Asset.type, String).ilike(asset_type))
        if status:
            from sqlalchemy import cast, String
            query = query.filter(cast(Asset.status, String).ilike(status))
        if location:
            query = query.filter(Asset.location.ilike(f"%{_sql_safe_search(location)}%"))
        total = query.count()
        assets = query.order_by(Asset.name).offset(skip).limit(limit).all()
        items = []
        for a in assets:
            try:
                items.append(_asset_to_out(a, db))
            except Exception as e:
                print(f"⚠️ _asset_to_out error for asset {a.id}: {e}")
                # Return minimal safe dict
                items.append({"id": a.id, "name": a.name or "Unknown",
                               "type": str(a.type) if a.type else None,
                               "status": str(a.status) if a.status else None})
        return {"items": items, "total": total, "skip": skip, "limit": limit}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Assets fetch error: {str(e)[:300]}")

@app.get("/assets/expiring")
def expiring_assets(days: int = Query(30), db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    """Returns assets whose license OR warranty expires within the given window."""
    try:
        today = date.today()
        deadline = today + timedelta(days=days)
        from sqlalchemy import or_, and_
        assets = db.query(Asset).filter(
            Asset.tenant_id == current_user.tenant_id,
            or_(
                and_(Asset.expiry_date.isnot(None), Asset.expiry_date > today, Asset.expiry_date <= deadline),
                and_(Asset.warranty_expiry.isnot(None), Asset.warranty_expiry > today, Asset.warranty_expiry <= deadline),
            )
        ).order_by(sa_func.coalesce(Asset.expiry_date, Asset.warranty_expiry)).all()
        return [_asset_to_out(a, db) for a in assets]
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Expiring assets error: {str(e)[:300]}")

@app.get("/assets/{asset_id}")
def get_asset(asset_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    asset = db.query(Asset).filter(Asset.id == asset_id, Asset.tenant_id == current_user.tenant_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return _asset_to_out(asset, db)

@app.get("/asset-model-options/")
def list_asset_model_options(asset_type: str | None = Query(None),
                              db: Session = Depends(get_db),
                              current_user: User = Depends(get_current_user)):
    """Returns admin-managed model/manufacturer options, optionally filtered to one asset type."""
    from sqlalchemy import text as _t
    try:
        if asset_type:
            rows = db.execute(_t(
                "SELECT id, asset_type::text, label, sort_order FROM asset_model_options "
                "WHERE tenant_id = :tid AND lower(asset_type::text) = :atype "
                "ORDER BY sort_order, label"
            ), {"tid": current_user.tenant_id, "atype": asset_type.lower()}).fetchall()
        else:
            rows = db.execute(_t(
                "SELECT id, asset_type::text, label, sort_order FROM asset_model_options "
                "WHERE tenant_id = :tid ORDER BY asset_type, sort_order, label"
            ), {"tid": current_user.tenant_id}).fetchall()
        return [{"id": r[0], "asset_type": r[1].lower() if r[1] else r[1],
                 "label": r[2], "sort_order": r[3]} for r in rows]
    except Exception as e:
        print(f"⚠️ list_asset_model_options error: {e}")
        return []

@app.post("/asset-model-options/")
def create_asset_model_option(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        if str(current_user.role) not in ("agent", "admin", "super_admin", "platform_admin"):
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        label = (data.get("label") or "").strip()
        if not label:
            raise HTTPException(status_code=422, detail="Label is required")
        asset_type_str = (data.get("asset_type") or "").lower().strip()
        valid_types = [e.value for e in AssetType]
        if asset_type_str not in valid_types:
            raise HTTPException(status_code=422, detail=f"Invalid asset_type '{asset_type_str}'.")
        sort_order = int(data.get("sort_order") or 0)
        from sqlalchemy import text as _t
        result = db.execute(_t(
            "INSERT INTO asset_model_options (tenant_id, asset_type, label, sort_order) "
            "VALUES (:tid, :atype, :label, :sort) RETURNING id"
        ), {"tid": current_user.tenant_id, "atype": asset_type_str, "label": label, "sort": sort_order})
        row = result.fetchone()
        db.commit()
        return {"id": row[0], "asset_type": asset_type_str, "label": label, "sort_order": sort_order}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Could not create model option: {str(e)[:200]}")

@app.delete("/asset-model-options/{option_id}")
def delete_asset_model_option(option_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_SETTINGS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    option = db.query(AssetModelOption).filter(
        AssetModelOption.id == option_id, AssetModelOption.tenant_id == current_user.tenant_id
    ).first()
    if not option:
        raise HTTPException(status_code=404, detail="Option not found")
    db.delete(option)
    db.commit()
    return {"ok": True}

@app.post("/assets/", response_model=AssetOut)
def create_asset(asset: AssetCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    plan_requires("asset_tracking", tenant, "Asset tracking is not available on the Free plan. Upgrade to Starter or higher.")
    # Check asset limit for the plan
    limits = get_plan_limits(tenant.plan if tenant else "free")
    max_assets = limits.get("max_assets")
    if max_assets is not None:
        current_count = db.query(Asset).filter(Asset.tenant_id == current_user.tenant_id).count()
        if current_count >= max_assets:
            raise HTTPException(status_code=403, detail=f"You've reached the {limits.get('label')} plan limit of {max_assets} assets. Upgrade to add more.")
    asset_data = asset.dict()
    if asset_data.get("custom_fields_data"):
        asset_data["custom_fields_data"] = json.dumps(asset_data["custom_fields_data"])
    else:
        asset_data["custom_fields_data"] = None
    db_asset = Asset(tenant_id=current_user.tenant_id, **asset_data)
    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)
    # Log the asset's creation as the first lifecycle event
    cost_note = f" — cost {db_asset.purchase_cost}" if db_asset.purchase_cost else ""
    vendor_note = f" from {db_asset.vendor}" if db_asset.vendor else ""
    db.add(AssetHistory(asset_id=db_asset.id, action="purchased",
        note=f"Asset created{vendor_note}{cost_note}".strip() or None,
        changed_by_id=current_user.id))
    # If assigned at creation time, log that too
    if db_asset.assigned_to_id:
        db.add(AssetHistory(asset_id=db_asset.id, action="assigned",
            from_user_id=None, to_user_id=db_asset.assigned_to_id,
            changed_by_id=current_user.id))
    db.commit()
    assigned = db.query(User).filter(User.id == db_asset.assigned_to_id).first()
    # Audit log
    log_system_event(db, current_user, "asset.created",
                     target_type="asset", target_id=db_asset.id,
                     target_label=f"{db_asset.name} ({db_asset.type})")
    db.commit()
    return {
        "id": db_asset.id, "name": db_asset.name, "type": db_asset.type, "model": db_asset.model, "serial_number": db_asset.serial_number,
        "tag_number": db_asset.tag_number,
        "status": db_asset.status, "assigned_to_id": db_asset.assigned_to_id,
        "assigned_to_name": assigned.full_name if assigned else None,
        "purchase_date": db_asset.purchase_date, "purchase_cost": db_asset.purchase_cost,
        "location": db_asset.location,
        "license_key": db_asset.license_key, "vendor": db_asset.vendor, "expiry_date": db_asset.expiry_date,
        "warranty_expiry": db_asset.warranty_expiry,
        "notes": db_asset.notes,
        "created_at": db_asset.created_at, "updated_at": db_asset.updated_at
    }

@app.patch("/assets/{asset_id}", response_model=AssetOut)
def update_asset(asset_id: int, asset_update: AssetUpdate,
                 current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_asset = db.query(Asset).filter(Asset.id == asset_id, Asset.tenant_id == current_user.tenant_id).first()
    if not db_asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    update_data = asset_update.model_dump(exclude_unset=True)
    # Track assignment changes
    if "assigned_to_id" in update_data and update_data["assigned_to_id"] != db_asset.assigned_to_id:
        db.add(AssetHistory(asset_id=asset_id,
            action="assigned" if update_data["assigned_to_id"] else "unassigned",
            from_user_id=db_asset.assigned_to_id, to_user_id=update_data["assigned_to_id"],
            changed_by_id=current_user.id))
    if "status" in update_data and update_data["status"] != db_asset.status:
        db.add(AssetHistory(asset_id=asset_id, action="status_changed",
            note=f"{db_assestr(t.status) if db_asset.status else '?'} → {update_data['status'].value if hasattr(update_data['status'], 'value') else update_data['status']}",
            changed_by_id=current_user.id))
    if "location" in update_data and (update_data["location"] or None) != (db_asset.location or None):
        old_loc = db_asset.location or "Unspecified"
        new_loc = update_data["location"] or "Unspecified"
        db.add(AssetHistory(asset_id=asset_id, action="location_changed",
            note=f"{old_loc} → {new_loc}",
            changed_by_id=current_user.id))
    for field, value in update_data.items():
        if field == "custom_fields_data":
            setattr(db_asset, field, json.dumps(value) if value else None)
        else:
            setattr(db_asset, field, value)
    db.commit()
    db.refresh(db_asset)
    # Audit log
    changed_fields = ", ".join(update_data.keys())
    log_system_event(db, current_user, "asset.updated",
                     target_type="asset", target_id=asset_id,
                     target_label=db_asset.name,
                     new_value=changed_fields)
    db.commit()
    return _asset_to_out(db_asset, db)

@app.get("/assets/{asset_id}/history")
def get_asset_history(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == asset_id, Asset.tenant_id == current_user.tenant_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    history = db.query(AssetHistory).filter(AssetHistory.asset_id == asset_id).order_by(AssetHistory.changed_at.desc()).all()
    return [{
        "id": h.id,
        "action": h.action,
        "from_user": h.from_user.full_name if h.from_user else None,
        "to_user": h.to_user.full_name if h.to_user else None,
        "note": h.note,
        "changed_by": h.changed_by.full_name if h.changed_by else None,
        "changed_at": h.changed_at,
    } for h in history]

@app.delete("/assets/{asset_id}")
def delete_asset(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    from sqlalchemy import text as _t
    row = db.execute(_t(
        "SELECT id FROM assets WHERE id=:id AND tenant_id=:tid"
    ), {"id": asset_id, "tid": current_user.tenant_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Asset not found")
    try:
        # Get asset name for audit log before deleting
        asset_row = db.execute(_t(
            "SELECT name, type FROM assets WHERE id=:id"
        ), {"id": asset_id}).fetchone()
        asset_name = asset_row[0] if asset_row else f"Asset #{asset_id}"
        asset_type = asset_row[1] if asset_row else "unknown"

        # Delete related records first to avoid FK violations
        db.execute(_t("DELETE FROM asset_history WHERE asset_id=:id"), {"id": asset_id})
        db.execute(_t("UPDATE tickets SET asset_id=NULL WHERE asset_id=:id"), {"id": asset_id})
        db.execute(_t("DELETE FROM assets WHERE id=:id AND tenant_id=:tid"),
                   {"id": asset_id, "tid": current_user.tenant_id})

        # Audit log
        log_system_event(db, current_user, "asset.deleted",
                         target_type="asset", target_id=asset_id,
                         target_label=f"{asset_name} ({asset_type})")
        db.commit()
        return {"detail": "Asset deleted"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)[:200]}")

@app.get("/assets/insights/summary")
def asset_insights(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Asset insights dashboard — counts by type, status, expiry alerts."""
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    assets = db.query(Asset).filter(Asset.tenant_id == current_user.tenant_id).all()
    today = date.today()
    by_type = {}
    by_status = {}
    expiring_30 = 0
    expiring_90 = 0
    warranty_expiring = 0
    maintenance_due = 0
    total_cost = 0.0
    for a in assets:
        t = str(a.type) if a.type else "other"
        s = str(a.status) if a.status else "unknown"
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
        if a.expiry_date:
            days = (a.expiry_date - today).days
            if days <= 30: expiring_30 += 1
            elif days <= 90: expiring_90 += 1
        if a.warranty_expiry and (a.warranty_expiry - today).days <= 30:
            warranty_expiring += 1
        if a.maintenance_date and a.maintenance_date.date() <= today:
            maintenance_due += 1
        if a.purchase_cost:
            total_cost += a.purchase_cost
    return {
        "total": len(assets),
        "by_type": by_type,
        "by_status": by_status,
        "expiring_30_days": expiring_30,
        "expiring_90_days": expiring_90,
        "warranty_expiring_30_days": warranty_expiring,
        "maintenance_due": maintenance_due,
        "total_purchase_cost": round(total_cost, 2),
    }

@app.post("/assets/bulk-import")
def bulk_import_assets(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Bulk import assets from CSV rows. data = {rows: [{name, type, serial_number, ...}]}"""
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    rows = data.get("rows", [])
    created = 0
    errors = []
    for i, row in enumerate(rows):
        try:
            name = row.get("name", "").strip()
            if not name:
                errors.append(f"Row {i+1}: name is required")
                continue
            raw_type = row.get("type", "hardware").lower().strip()
            try:
                asset_type = raw_type
            except ValueError:
                asset_type = "hardware"
            db_asset = Asset(
                tenant_id=current_user.tenant_id,
                name=name, type=asset_type,
                serial_number=row.get("serial_number") or None,
                vendor=row.get("vendor") or None,
                location=row.get("location") or None,
                notes=row.get("notes") or None,
                tag_number=row.get("tag_number") or None,
                purchase_cost=float(row["purchase_cost"]) if row.get("purchase_cost") else None,
                status=str(row.get("status", "available")).lower(),
            )
            db.add(db_asset)
            created += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)}")
    db.commit()
    return {"created": created, "errors": errors}

@app.post("/assets/bulk-action")
def bulk_asset_action(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Bulk action on multiple assets. data = {asset_ids: [], action: 'retire'|'assign'|'maintenance', value: ...}"""
    if not has_permission(current_user, Permission.MANAGE_ASSETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    asset_ids = data.get("asset_ids", [])
    action = data.get("action")
    value = data.get("value")
    updated = 0
    for asset_id in asset_ids:
        asset = db.query(Asset).filter(Asset.id == asset_id, Asset.tenant_id == current_user.tenant_id).first()
        if not asset:
            continue
        if action == "retire":
            asset.status = "retired"
        elif action == "maintenance":
            asset.status = "maintenance"
        elif action == "available":
            asset.status = "available"
            asset.assigned_to_id = None
        elif action == "assign" and value:
            asset.assigned_to_id = int(value)
            asset.status = "assigned"
            db.add(AssetHistory(asset_id=asset_id, action="assigned",
                                to_user_id=int(value), changed_by_id=current_user.id))
        updated += 1
    db.commit()
    return {"updated": updated}

# =============================================================================
# ATTACHMENT ENDPOINTS
# =============================================================================

ALLOWED_EXTENSIONS = {".txt", ".pdf", ".png", ".jpg", ".jpeg", ".docx", ".xlsx", ".csv", ".zip", ".pptx"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

@app.post("/tickets/{ticket_id}/attachments", response_model=AttachmentOut)
def upload_attachment(
    ticket_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not has_permission(current_user, Permission.CREATE_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' is not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}")
    file_content = file.file.read()
    if len(file_content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File size exceeds 10 MB limit")

    # Sanitise the filename — strip non-ASCII and path traversal chars
    safe_name = "".join(c for c in file.filename if c.isalnum() or c in "._- ")[:100]
    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    file_url = None

    if CLOUDINARY_CLOUD_NAME:
        # Upload to tenant-scoped Cloudinary folder:
        # dodesk/tenants/{tenant_id}/tickets/{ticket_id}/{uuid_filename}
        folder = _cloudinary_folder(current_user.tenant_id, "tickets", ticket_id)
        try:
            file_url = upload_to_cloudinary(file_content, unique_name, folder=folder, filename=file.filename)
        except Exception as e:
            print(f"⚠️ Cloudinary upload failed, falling back to local: {e}")
            # Fall back to local disk if Cloudinary fails so uploads don't silently break
            file_url = None

    if not file_url:
        # Local disk fallback (ephemeral on Render — warn but don't block)
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        file_path = os.path.join(UPLOAD_DIR, unique_name)
        with open(file_path, "wb") as f:
            f.write(file_content)
        print(f"⚠️ File stored locally at {file_path} — this will be lost on redeploy. Configure Cloudinary to fix this.")

    db_attachment = Attachment(
        ticket_id=ticket_id,
        filename=file.filename,
        stored_filename=unique_name,
        url=file_url,              # None if using local fallback
        content_type=file.content_type,
        size=len(file_content)
    )
    db.add(db_attachment)
    db.commit()
    db.refresh(db_attachment)
    return db_attachment

@app.get("/tickets/{ticket_id}/attachments", response_model=list[AttachmentOut])
def list_attachments(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return db.query(Attachment).filter(Attachment.ticket_id == ticket_id).all()

@app.get("/attachments/{attachment_id}/download")
def download_attachment(attachment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    attachment = db.query(Attachment).filter(Attachment.id == attachment_id).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    # Tenant safety check — prevent cross-tenant access
    ticket = db.query(Ticket).filter(Ticket.id == attachment.ticket_id, Ticket.tenant_id == current_user.tenant_id).first()
    if not ticket:
        raise HTTPException(status_code=403, detail="Access denied")
    # If stored as Cloudinary public_id (authenticated type), generate signed URL
    if attachment.url and not attachment.url.startswith("http"):
        # It's a public_id stored after auth migration
        ext = os.path.splitext(attachment.filename)[1].lower()
        rtype = "image" if ext in {".png",".jpg",".jpeg",".gif",".webp",".svg"} else "raw"
        signed = get_signed_url(attachment.url, resource_type=rtype)
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=signed)
    # Legacy: stored URL (public or old Cloudinary URL)
    if attachment.url:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=attachment.url)
    # Legacy local file fallback
    file_path = os.path.join(UPLOAD_DIR, attachment.stored_filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found — it may have been lost during a server redeploy. Please re-upload.")
    return FileResponse(file_path, media_type=attachment.content_type or "application/octet-stream",
                        filename=attachment.filename)

# =============================================================================
# REPORTING ENDPOINTS (tenant‑scoped, permissions)
# =============================================================================

@app.get("/reports/my-clients")
def get_my_clients(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns the list of client tenants accessible to the current super_admin/platform_admin.
    Used by the Reports page to show a client selector for MSPs."""
    role = str(current_user.role)
    
    # platform_admin sees all tenants
    if role == 'platform_admin':
        tenants = db.query(Tenant).filter(Tenant.is_active == True).all()
        return [{"id": t.id, "name": t.name, "plan": t.plan} for t in tenants]
    
    # super_admin sees own tenant + assigned client tenants
    if role == 'super_admin':
        # Own tenant first
        own = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
        result = [{"id": own.id, "name": own.name + " (own)", "plan": own.plan}] if own else []
        
        # Assigned client tenants
        granted = db.query(AdminTenantAccess).filter(
            AdminTenantAccess.admin_user_id == current_user.id
        ).all()
        for g in granted:
            t = db.query(Tenant).filter(Tenant.id == g.tenant_id, Tenant.is_active == True).first()
            if t and t.id != current_user.tenant_id:
                result.append({"id": t.id, "name": t.name, "plan": t.plan})
        return result
    
    # Other roles — no client selector
    return []


@app.get("/reports/summary")
def report_summary(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    client_tenant_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    try:
        from sqlalchemy import text as _t
        # MSP: allow viewing client tenant reports
        if client_tenant_id and str(current_user.role) in ("super_admin", "platform_admin"):
            tid = client_tenant_id
        else:
            tid = current_user.tenant_id

        # Use raw SQL to avoid ORM enum deserialisation issues
        def count_q(where_extra=""):
            sql = f"SELECT COUNT(*) FROM tickets WHERE tenant_id=:tid {where_extra}"
            return db.execute(_t(sql), {"tid": tid}).scalar() or 0

        total        = count_q()
        open_count   = count_q("AND status IN ('open','in_progress','pending_approval')")
        overdue      = count_q(f"AND sla_resolution_deadline < NOW() AND status IN ('open','in_progress')")
        today_start  = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        resolved_today = db.execute(_t(
            "SELECT COUNT(*) FROM tickets WHERE tenant_id=:tid AND status='resolved' AND updated_at>=:ts"
        ), {"tid": tid, "ts": today_start}).scalar() or 0

        # Avg resolution hours
        try:
            rows = db.execute(_t(
                "SELECT created_at, updated_at FROM tickets "
                "WHERE tenant_id=:tid AND status='resolved' AND created_at IS NOT NULL AND updated_at IS NOT NULL"
            ), {"tid": tid}).fetchall()
            avg_resolution_hours = round(
                sum((r[1]-r[0]).total_seconds()/3600 for r in rows if r[1] and r[0]) / len(rows), 1
            ) if rows else 0
        except Exception:
            avg_resolution_hours = 0

        # Avg first response hours
        try:
            rows2 = db.execute(_t(
                "SELECT created_at, first_response_at FROM tickets "
                "WHERE tenant_id=:tid AND first_response_at IS NOT NULL AND created_at IS NOT NULL"
            ), {"tid": tid}).fetchall()
            avg_first_response_hours = round(
                sum((r[1]-r[0]).total_seconds()/3600 for r in rows2 if r[1] and r[0]) / len(rows2), 1
            ) if rows2 else 0
        except Exception:
            avg_first_response_hours = 0

        # Open changes — all non-completed statuses
        try:
            open_changes = db.execute(_t(
                "SELECT COUNT(*) FROM change_requests WHERE tenant_id=:tid "
                "AND status NOT IN ('completed','cancelled','failed')"
            ), {"tid": tid}).scalar() or 0
        except Exception:
            open_changes = 0

        return {
            "total": total, "open": open_count, "overdue": overdue,
            "resolved_today": resolved_today,
            "avg_resolution_hours": avg_resolution_hours,
            "avg_first_response_hours": avg_first_response_hours,
            "open_changes": open_changes,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Summary error: {str(e)[:200]}")

@app.get("/reports/sla-compliance")
def sla_compliance(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    base_query = db.query(Ticket).filter(Ticket.tenant_id == _eff_tid)
    base_query = apply_filters(base_query, ticket_type, start_date, end_date)
    resolved_total = base_query.filter(Ticket.status == 'resolved').count()
    if resolved_total == 0:
        return {"compliance_percent": 100.0, "total_resolved": 0}
    on_time = base_query.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at <= Ticket.sla_resolution_deadline
    ).count()
    compliance = round((on_time / resolved_total) * 100, 1)
    return {"compliance_percent": compliance, "total_resolved": resolved_total, "on_time": on_time}

@app.get("/reports/tickets-by-priority")
def tickets_by_priority(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    query = db.query(Ticket.priority, sa_func.count(Ticket.id)).filter(Ticket.tenant_id == _eff_tid)
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(Ticket.priority).all()
    return [{"priority": (p.value if hasattr(p, "value") else str(p)) if p else "unknown", "count": c} for p, c in results]

@app.get("/reports/tickets-by-status")
def tickets_by_status(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    query = db.query(Ticket.status, sa_func.count(Ticket.id)).filter(Ticket.tenant_id == _eff_tid)
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(Ticket.status).all()
    return [{"status": (s.value if hasattr(s, "value") else str(s)) if s else "unknown", "count": c} for s, c in results]

@app.get("/reports/tickets-created-daily")
def tickets_created_daily(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    if start_date and end_date:
        start = start_date
        end = end_date
    else:
        today = datetime.utcnow().date()
        start = today - timedelta(days=6)
        end = today
    days = []
    current = start
    while current <= end:
        day_start = datetime(current.year, current.month, current.day)
        day_end = day_start + timedelta(days=1)
        query = db.query(Ticket).filter(
            Ticket.tenant_id == _eff_tid,
            Ticket.created_at >= day_start,
            Ticket.created_at < day_end
        )
        query = apply_filters(query, ticket_type, None, None)
        count = query.count()
        days.append({"date": current.isoformat(), "count": count})
        current += timedelta(days=1)
    return days

@app.get("/reports/my-stats")
def my_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Personal stats for the current agent: assigned, due today, overdue, resolved this week."""
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end   = today_start + timedelta(days=1)
    week_start  = today_start - timedelta(days=today_start.weekday())
    now         = datetime.utcnow()
    base = db.query(Ticket).filter(
        Ticket.tenant_id == current_user.tenant_id,
        Ticket.assigned_to_id == current_user.id
    )
    assigned_open = base.filter(Ticket.status.in_(['open','in_progress','pending_approval'])).count()
    due_today = base.filter(
        Ticket.status.in_(['open','in_progress']),
        (
            (Ticket.due_date >= today_start) & (Ticket.due_date < today_end)
        ) | (
            (Ticket.sla_resolution_deadline >= today_start) & (Ticket.sla_resolution_deadline < today_end)
        )
    ).count()
    overdue_mine = base.filter(
        Ticket.sla_resolution_deadline < now,
        Ticket.status.in_(['open','in_progress'])
    ).count()
    resolved_week = base.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at >= week_start
    ).count()
    # Avg resolution time this week
    resolved_tix = base.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at >= week_start,
        Ticket.updated_at.isnot(None)
    ).with_entities(Ticket.created_at, Ticket.updated_at).all()
    avg_res = 0
    if resolved_tix:
        avg_res = round(sum((t.updated_at - t.created_at).total_seconds() / 3600 for t in resolved_tix if t.updated_at and t.created_at) / len(resolved_tix), 1)
    return {
        "assigned_open": assigned_open,
        "due_today": due_today,
        "overdue_mine": overdue_mine,
        "resolved_week": resolved_week,
        "avg_resolution_hours": avg_res,
    }

@app.get("/reports/agent-workload")
def agent_workload(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    agents = db.query(User).filter(User.tenant_id == _eff_tid, User.role == 'agent').all()
    if not agents:
        return []
    agent_ids = [a.id for a in agents]

    # Single query for all assigned/resolved counts, grouped by agent — replaces N×2 per-agent count() calls
    base_q = db.query(Ticket).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.assigned_to_id.in_(agent_ids)
    )
    base_q = apply_filters(base_q, ticket_type, start_date, end_date)
    tickets = base_q.with_entities(Ticket.id, Ticket.assigned_to_id, Ticket.status).all()

    assigned_counts = {}
    resolved_counts = {}
    ticket_ids_by_agent = {}
    for tid, agent_id, status in tickets:
        assigned_counts[agent_id] = assigned_counts.get(agent_id, 0) + 1
        ticket_ids_by_agent.setdefault(agent_id, []).append(tid)
        if status == "resolved":
            resolved_counts[agent_id] = resolved_counts.get(agent_id, 0) + 1

    # Single query for all time entries across all agents — replaces N separate queries
    all_ticket_ids = [tid for ids in ticket_ids_by_agent.values() for tid in ids]
    minutes_by_agent = {}
    if all_ticket_ids:
        time_rows = db.query(TimeEntry.agent_id, TimeEntry.minutes).filter(
            TimeEntry.agent_id.in_(agent_ids),
            TimeEntry.ticket_id.in_(all_ticket_ids)
        ).all()
        for agent_id, minutes in time_rows:
            minutes_by_agent[agent_id] = minutes_by_agent.get(agent_id, 0) + minutes

    result = []
    for agent in agents:
        result.append({
            "agent_name": agent.full_name,
            "assigned": assigned_counts.get(agent.id, 0),
            "resolved": resolved_counts.get(agent.id, 0),
            "total_hours": round(minutes_by_agent.get(agent.id, 0) / 60, 1),
        })
    return result

@app.get("/reports/changes-summary")
def changes_summary(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    client_tenant_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Summary stats for change requests."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    q = db.query(ChangeRequest).filter(ChangeRequest.tenant_id == _eff_tid)
    if start_date:
        q = q.filter(ChangeRequest.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        q = q.filter(ChangeRequest.created_at <= datetime.combine(end_date, datetime.max.time()))
    total = q.count()
    # By status
    by_status = {}
    for row in q.with_entities(ChangeRequest.status, sa_func.count()).group_by(ChangeRequest.status).all():
        by_status[str(row[0].value if hasattr(row[0],'value') else row[0]) if row[0] else 'unknown'] = row[1]
    # By risk
    by_risk = {}
    for row in q.with_entities(ChangeRequest.risk_level, sa_func.count()).group_by(ChangeRequest.risk_level).all():
        by_risk[str(row[0].value if hasattr(row[0],'value') else row[0]) if row[0] else 'unknown'] = row[1]
    # Daily trend (last 30 days)
    from sqlalchemy import cast, Date as SADate
    daily = []
    try:
        rows = (q.with_entities(cast(ChangeRequest.created_at, SADate).label('day'), sa_func.count())
                  .group_by(cast(ChangeRequest.created_at, SADate))
                  .order_by(cast(ChangeRequest.created_at, SADate)).all())
        daily = [{"date": str(r[0]), "count": r[1]} for r in rows]
    except Exception:
        pass
    # Open count
    open_statuses = ["pending_approval", "approved"]
    open_count = q.filter(ChangeRequest.status.in_(open_statuses)).count()
    implemented = by_status.get('implemented', 0)
    rejected    = by_status.get('rejected', 0)
    return {
        "total": total,
        "open": open_count,
        "implemented": implemented,
        "rejected": rejected,
        "by_status": by_status,
        "by_risk": by_risk,
        "daily": daily,
    }


@app.get("/reports/export/csv")
def export_csv(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Type", "Title", "Category", "Priority", "Status", "Requester", "Assigned To", "Created", "SLA Status", "Attachments"])

    if ticket_type == "change":
        # Export change requests
        query = db.query(ChangeRequest).filter(ChangeRequest.tenant_id == current_user.tenant_id)
        if start_date:
            query = query.filter(ChangeRequest.created_at >= datetime(start_date.year, start_date.month, start_date.day))
        if end_date:
            end_dt = datetime(end_date.year, end_date.month, end_date.day) + timedelta(days=1)
            query = query.filter(ChangeRequest.created_at < end_dt)
        changes = query.order_by(ChangeRequest.id).all()
        req_ids = {c.requester_id for c in changes if c.requester_id}
        req_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(req_ids)).all()} if req_ids else {}
        for c in changes:
            try:
                writer.writerow([
                    f"CHG-{c.id:04d}", "change_request", c.title or "",
                    getattr(c, 'category', '') or "",
                    str(c.risk_level) if c.risk_level else "",
                    str(c.status) if c.status else "",
                    req_map.get(c.requester_id, ""),
                    "", c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "", ""
                ])
            except Exception:
                continue
    else:
        # Export tickets (incidents and/or service requests)
        query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)
        query = apply_filters(query, ticket_type, start_date, end_date)
        tickets = query.order_by(Ticket.id).all()
        # Pre-load users to avoid lazy loading issues
        user_ids = set()
        for t in tickets:
            if t.requester_id: user_ids.add(t.requester_id)
            if t.assigned_to_id: user_ids.add(t.assigned_to_id)
        user_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}

        for t in tickets:
            if t.ticket_type == "incident":
                ticket_ref = f"INC-{t.id:04d}"
            elif t.ticket_type == "service_request":
                ticket_ref = f"REQ-{t.id:04d}"
            else:
                ticket_ref = f"CHG-{t.id:04d}"
            try:
                att_list = db.query(Attachment).filter(Attachment.ticket_id == t.id).all()
                att_urls = " | ".join([a.url or a.file_path or "" for a in att_list if (a.url or a.file_path)])
                writer.writerow([
                    ticket_ref,
                    str(t.ticket_type) if t.ticket_type else "",
                    t.title or "",
                    t.category or "",
                    str(t.priority) if t.priority else "",
                    str(t.status) if t.status else "",
                    user_map.get(t.requester_id, ""),
                    user_map.get(t.assigned_to_id, "Unassigned"),
                    t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                    compute_sla_status(t),
                    att_urls,
                ])
            except Exception:
                continue

    csv_content = output.getvalue()
    output.close()
    from fastapi.responses import Response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=dodesk_export.csv"}
    )

@app.get("/reports/tickets-by-category")
def tickets_by_category(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Category breakdown with volume, open count, and avg resolution time —
    used to identify which categories need the most operational focus."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id

    base = db.query(Ticket).filter(Ticket.tenant_id == _eff_tid)
    base = apply_filters(base, ticket_type, start_date, end_date)
    tickets = base.all()

    by_cat = {}
    for t in tickets:
        cat = t.category or "Uncategorised"
        if cat not in by_cat:
            by_cat[cat] = {"category": cat, "count": 0, "open": 0, "overdue": 0, "res_hours": [], "critical": 0}
        entry = by_cat[cat]
        entry["count"] += 1
        if t.status in ("open", "in_progress", "pending_approval"):
            entry["open"] += 1
            if t.sla_resolution_deadline and t.sla_resolution_deadline < datetime.utcnow():
                entry["overdue"] += 1
        if t.priority == "critical":
            entry["critical"] += 1
        if t.status == "resolved" and t.updated_at and t.created_at:
            entry["res_hours"].append((t.updated_at - t.created_at).total_seconds() / 3600)

    results = []
    for cat, e in by_cat.items():
        avg_res = round(sum(e["res_hours"]) / len(e["res_hours"]), 1) if e["res_hours"] else None
        # Focus score: weighted combination of volume, overdue count, and critical tickets
        # Higher score = needs more attention
        focus_score = e["count"] + (e["overdue"] * 3) + (e["critical"] * 2)
        results.append({
            "category": cat,
            "count": e["count"],
            "open": e["open"],
            "overdue": e["overdue"],
            "critical": e["critical"],
            "avg_resolution_hours": avg_res,
            "focus_score": focus_score,
        })

    results.sort(key=lambda r: r["focus_score"], reverse=True)
    return results

@app.get("/reports/resolution-time-trend")
def resolution_time_trend(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Average resolution time per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.updated_at, SADate).label("day"),
        sa_func.avg(
            sa_func.extract("epoch", Ticket.updated_at - Ticket.created_at) / 3600
        ).label("avg_hours")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.status == 'resolved',
        Ticket.updated_at.isnot(None)
    )
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(cast(Ticket.updated_at, SADate)).order_by(cast(Ticket.updated_at, SADate)).all()
    return [{"date": str(r.day), "avg_hours": round(float(r.avg_hours or 0), 1)} for r in results]

@app.get("/reports/first-response-trend")
def first_response_trend(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Average first response time per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.created_at, SADate).label("day"),
        sa_func.avg(
            sa_func.extract("epoch", Ticket.first_response_at - Ticket.created_at) / 3600
        ).label("avg_hours")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.first_response_at.isnot(None)
    )
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(cast(Ticket.created_at, SADate)).order_by(cast(Ticket.created_at, SADate)).all()
    return [{"date": str(r.day), "avg_hours": round(float(r.avg_hours or 0), 1)} for r in results]

@app.get("/reports/tickets-aging")
def tickets_aging(
    ticket_type: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Open tickets bucketed by age: <1d, 1-3d, 3-7d, 7-30d, >30d."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    now = datetime.utcnow()
    open_statuses = ["open", "in_progress", "pending_approval"]
    query = db.query(Ticket).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.status.in_(open_statuses)
    )
    if ticket_type:
        try: query = query.filter(Ticket.ticket_type == ticket_type.lower())
        except ValueError: pass
    tickets = query.all()
    buckets = {"<1 day": 0, "1-3 days": 0, "3-7 days": 0, "7-30 days": 0, ">30 days": 0}
    for t in tickets:
        if not t.created_at: continue
        age = (now - t.created_at).days
        if age < 1:   buckets["<1 day"] += 1
        elif age < 3:  buckets["1-3 days"] += 1
        elif age < 7:  buckets["3-7 days"] += 1
        elif age < 30: buckets["7-30 days"] += 1
        else:          buckets[">30 days"] += 1
    return [{"bucket": k, "count": v} for k, v in buckets.items()]

@app.get("/reports/csat-trend")
def csat_trend(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """CSAT average score per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.updated_at, SADate).label("day"),
        sa_func.avg(Ticket.csat_rating).label("avg_rating"),
        sa_func.count(Ticket.id).label("count")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    )
    if start_date:
        query = query.filter(Ticket.updated_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(Ticket.updated_at <= datetime.combine(end_date, datetime.max.time()))
    results = query.group_by(cast(Ticket.updated_at, SADate)).order_by(cast(Ticket.updated_at, SADate)).all()
    return [{"date": str(r.day), "avg_rating": round(float(r.avg_rating or 0), 2), "count": r.count} for r in results]

@app.get("/reports/kb-analytics")
def kb_analytics(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """KB article analytics — views, feedback, by category."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    articles = db.query(KBArticle).filter(
        KBArticle.tenant_id == _eff_tid,
        KBArticle.status == "published"
    ).all()
    total_views = sum(a.view_count or 0 for a in articles)
    total_helpful = sum(a.helpful_count or 0 for a in articles)
    total_not_helpful = sum(a.not_helpful_count or 0 for a in articles)
    by_category = {}
    for a in articles:
        cat = a.category or "General"
        if cat not in by_category:
            by_category[cat] = {"articles": 0, "views": 0}
        by_category[cat]["articles"] += 1
        by_category[cat]["views"] += a.view_count or 0
    most_viewed = sorted(articles, key=lambda a: a.view_count or 0, reverse=True)[:10]
    return {
        "total_articles": len(articles),
        "total_views": total_views,
        "total_helpful": total_helpful,
        "total_not_helpful": total_not_helpful,
        "satisfaction_rate": round(total_helpful / max(total_helpful + total_not_helpful, 1) * 100, 1),
        "by_category": [{"category": k, **v} for k, v in by_category.items()],
        "most_viewed": [{"id": a.id, "title": a.title, "category": a.category, "views": a.view_count or 0, "helpful": a.helpful_count or 0} for a in most_viewed],
    }

@app.get("/reports/asset-summary")
def asset_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Asset report — by type, status, expiry alerts."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    assets = db.query(Asset).filter(Asset.tenant_id == _eff_tid).all()
    today = date.today()
    by_type = {}
    by_status = {}
    for a in assets:
        t = str(a.type) if a.type else "other"
        s = str(a.status) if a.status else "unknown"
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
    expiring_30 = [a for a in assets if a.expiry_date and 0 <= (a.expiry_date - today).days <= 30]
    return {
        "total": len(assets),
        "by_type": [{"type": k, "count": v} for k, v in by_type.items()],
        "by_status": [{"status": k, "count": v} for k, v in by_status.items()],
        "expiring_30_days": len(expiring_30),
        "expiring_soon": [{"id": a.id, "name": a.name, "expiry_date": str(a.expiry_date)} for a in expiring_30[:10]],
        "total_cost": round(sum(a.purchase_cost or 0 for a in assets), 2),
    }

@app.get("/reports/export/excel")
def export_excel(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export tickets as Excel file."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["ID", "Type", "Title", "Category", "Priority", "Status", "Requester", "Assigned To", "Created", "SLA Deadline", "Resolution Time (hrs)", "Attachments"]
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)
    query = apply_filters(query, ticket_type, start_date, end_date)
    tickets = query.order_by(Ticket.id).all()
    req_ids = {t.requester_id for t in tickets if t.requester_id}
    asgn_ids = {t.assigned_to_id for t in tickets if t.assigned_to_id}
    all_ids = req_ids | asgn_ids
    user_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(all_ids)).all()} if all_ids else {}
    # Load attachments for all tickets
    ticket_ids = [t.id for t in tickets]
    attachments_map = {}
    if ticket_ids:
        all_attachments = db.query(Attachment).filter(Attachment.ticket_id.in_(ticket_ids)).all()
        for att in all_attachments:
            url = att.url or att.file_path or ""
            if url:
                attachments_map.setdefault(att.ticket_id, []).append(url)
    for row, t in enumerate(tickets, 2):
        prefix = {"incident": "INC", "service_request": "REQ", "change": "CHG"}.get(str(t.ticket_type) if t.ticket_type else "incident", "INC")
        res_hours = ""
        if t.status == "resolved" and t.updated_at and t.created_at:
            res_hours = round((t.updated_at - t.created_at).total_seconds() / 3600, 1)
        att_urls = " | ".join(attachments_map.get(t.id, []))
        ws.append([
            f"{prefix}{t.id:06d}",
            str(t.ticket_type) if t.ticket_type else "",
            t.title or "",
            t.category or "",
            str(t.priority) if t.priority else "",
            str(t.status) if t.status else "",
            user_map.get(t.requester_id, ""),
            user_map.get(t.assigned_to_id, ""),
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            str(t.sla_resolution_deadline.date()) if t.sla_resolution_deadline else "",
            res_hours,
            att_urls,
        ])
    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dodesk_tickets.xlsx"}
    )

# =============================================================================
# CHANGE MANAGEMENT (fixed permissions)
# =============================================================================

@app.post("/changes/")
def create_change(change: ChangeCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.CREATE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    plan_requires("change_management", tenant, "Change Management is available on the Pro plan and above. Please upgrade.")
    db_change = ChangeRequest(
        tenant_id=current_user.tenant_id,
        title=change.title,
        description=change.description,
        change_type=change.change_type or "normal",
        risk_level=change.risk_level,
        risk_score=change.risk_score,
        planned_date=change.planned_date,
        start_date=change.start_date,
        end_date=change.end_date,
        impact=change.impact,
        rollback_plan=change.rollback_plan,
        test_plan=change.test_plan,
        owner_id=change.owner_id,
        assigned_to_id=change.assigned_to_id,
        cab_members=json.dumps(change.cab_members) if change.cab_members else None,
        linked_ticket_ids=json.dumps(change.linked_ticket_ids) if change.linked_ticket_ids else None,
        linked_asset_ids=json.dumps(change.linked_asset_ids) if change.linked_asset_ids else None,
        requester_id=current_user.id,
        status="draft"
    )
    db.add(db_change)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create change: {str(e)}")
    db.refresh(db_change)
    return _change_to_out(db_change, db=db)

@app.get("/changes/")
def list_changes(skip: int = Query(0, ge=0), limit: int = Query(20, ge=1, le=200),
                 search: str = Query("", alias="search"),
                 status: str | None = Query(None),
                 change_type: str | None = Query(None),
                 risk_level: str | None = Query(None),
                 current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.APPROVE_CHANGES) and not has_permission(current_user, Permission.CREATE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    query = db.query(ChangeRequest).filter(ChangeRequest.tenant_id == current_user.tenant_id)
    if not has_permission(current_user, Permission.APPROVE_CHANGES):
        query = query.filter(ChangeRequest.requester_id == current_user.id)
    if search:
        from sqlalchemy import or_
        import re as _re
        search_term = f"%{search}%"
        id_match = _re.search(r'(\d+)', search)
        numeric_id = int(id_match.group(1)) if id_match else None
        conditions = [ChangeRequest.title.ilike(search_term), ChangeRequest.description.ilike(search_term)]
        if numeric_id: conditions.append(ChangeRequest.id == numeric_id)
        query = query.filter(or_(*conditions))
    if status:
        query = query.filter(ChangeRequest.status == status)
    if change_type:
        query = query.filter(ChangeRequest.change_type == change_type)
    if risk_level:
        query = query.filter(ChangeRequest.risk_level == risk_level)
    total = query.count()
    changes = query.order_by(ChangeRequest.created_at.desc()).offset(skip).limit(limit).all()
    req_ids = {c.requester_id for c in changes if c.requester_id}
    user_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(req_ids)).all()} if req_ids else {}
    return {"items": [_change_to_out(c, user_map) for c in changes], "total": total, "skip": skip, "limit": limit}

@app.get("/changes/calendar")
def get_change_calendar(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Returns all changes with dates for calendar view."""
    if not has_permission(current_user, Permission.APPROVE_CHANGES) and not has_permission(current_user, Permission.CREATE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    changes = db.query(ChangeRequest).filter(
        ChangeRequest.tenant_id == current_user.tenant_id,
        (ChangeRequest.planned_date != None) | (ChangeRequest.start_date != None)
    ).order_by(ChangeRequest.planned_date).all()
    return [{"id": c.id, "title": c.title, "change_type": str(c.change_type) if c.change_type else "normal",
             "risk_level": str(c.risk_level) if c.risk_level else "medium",
             "status": str(c.status) if c.status else "draft",
             "planned_date": c.planned_date.isoformat() if c.planned_date else None,
             "start_date": c.start_date.isoformat() if c.start_date else None,
             "end_date": c.end_date.isoformat() if c.end_date else None}
            for c in changes]

@app.get("/changes/{change_id}")
def get_change(change_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    from sqlalchemy import text as _t
    try:
        row = db.execute(_t(
            "SELECT id, title, description, change_type::text, risk_level::text, risk_score, "
            "status::text, requester_id, owner_id, assigned_to_id, "
            "planned_date, start_date, end_date, impact, rollback_plan, test_plan, "
            "cab_members, linked_ticket_ids, linked_asset_ids, "
            "post_review_notes, post_review_at, created_at, updated_at "
            "FROM change_requests WHERE id=:id AND tenant_id=:tid"
        ), {"id": change_id, "tid": current_user.tenant_id}).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Change not found")

        def safe_json(val):
            if not val: return []
            try: return json.loads(val)
            except: return []

        def get_name(uid):
            if not uid: return ""
            try:
                u = db.execute(_t("SELECT full_name FROM users WHERE id=:id"), {"id": uid}).fetchone()
                return u[0] if u else ""
            except: return ""

        return {
            "id": row[0], "title": row[1], "description": row[2],
            "change_type": str(row[3]).lower() if row[3] else "normal",
            "risk_level": str(row[4]).lower() if row[4] else "medium",
            "risk_score": row[5],
            "status": str(row[6]).lower() if row[6] else "draft",
            "requester_id": row[7], "requester_name": get_name(row[7]),
            "owner_id": row[8], "owner_name": get_name(row[8]),
            "assigned_to_id": row[9], "assigned_to_name": get_name(row[9]),
            "planned_date": str(row[10]) if row[10] else None,
            "start_date": str(row[11]) if row[11] else None,
            "end_date": str(row[12]) if row[12] else None,
            "impact": row[13], "rollback_plan": row[14], "test_plan": row[15],
            "cab_members": safe_json(row[16]),
            "linked_ticket_ids": safe_json(row[17]),
            "linked_asset_ids": safe_json(row[18]),
            "post_review_notes": row[19],
            "post_review_at": str(row[20]) if row[20] else None,
            "created_at": str(row[21]) if row[21] else None,
            "updated_at": str(row[22]) if row[22] else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Change fetch error: {str(e)[:200]}")

@app.patch("/changes/{change_id}")
def update_change(change_id: int, update: ChangeUpdate,
                  current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.CREATE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.requester_id != current_user.id and not has_permission(current_user, Permission.APPROVE_CHANGES):
        raise HTTPException(status_code=403, detail="Access denied")
    update_data = update.model_dump(exclude_unset=True)
    for json_field in ["cab_members", "linked_ticket_ids", "linked_asset_ids"]:
        if json_field in update_data:
            update_data[json_field] = json.dumps(update_data[json_field]) if update_data[json_field] is not None else None
    if "post_review_notes" in update_data and update_data["post_review_notes"]:
        change.post_review_at = datetime.utcnow()
    for field, value in update_data.items():
        setattr(change, field, value)
    db.commit()
    db.refresh(change)
    return _change_to_out(change, db=db)

@app.post("/changes/{change_id}/submit")
def submit_change_for_approval(change_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Move a Draft change to Pending Approval, notifying CAB members / approvers."""
    if not has_permission(current_user, Permission.CREATE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.requester_id != current_user.id and not has_permission(current_user, Permission.APPROVE_CHANGES):
        raise HTTPException(status_code=403, detail="Access denied")
    if change.status != "draft":
        raise HTTPException(status_code=400, detail="Only draft changes can be submitted for approval")
    # Standard changes are pre-approved by policy — skip CAB review and go straight to Approved
    if change.change_type == "standard":
        change.status = "approved"
    else:
        change.status = "pending_approval"
    db.commit()
    db.refresh(change)
    # Notify approvers (admins/super_admins) that a change needs review
    if change.status == "pending_approval":
        approvers = db.query(User).filter(
            User.tenant_id == current_user.tenant_id,
            User.role.in_(['admin', 'super_admin', 'platform_admin']),
            User.is_active == True
        ).all()
        for approver in approvers:
            _cl = get_user_language(db, approver.email)
            if _cl == 'fr':
                _cs = f"Changement en attente d'approbation : #{change.id} {change.title}"
                _cb = f"Une demande de changement nécessite votre révision.\n\nType : {change.change_type}\nRisque : {str(change.risk_level) if change.risk_level else 'n/a'}"
                _cc = "Voir le changement →"
            else:
                _cs = f"Change pending your approval: #{change.id} {change.title}"
                _cb = f"A change request needs your review.\n\nType: {change.change_type}\nRisk: {str(change.risk_level) if change.risk_level else 'n/a'}"
                _cc = "View change →"
            send_email(approver.email, _cs, _cb, cta_url=f"{FRONTEND_URL}/changes/{change.id}", cta_label=_cc, db=None, tenant_id=change.tenant_id, lang=_cl)
    return _change_to_out(change, db=db)

@app.post("/changes/{change_id}/approve")
def approve_change(change_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.APPROVE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.status != "pending_approval":
        raise HTTPException(status_code=400, detail="Change is not in pending approval status")
    change.status = "approved"
    db.commit()
    db.refresh(change)
    requester = db.query(User).filter(User.id == change.requester_id).first()
    if requester:
        _cl2 = get_user_language(db, requester.email)
        if _cl2 == 'fr':
            _cs2 = f"Changement approuvé : #{change.id} {change.title}"
            _cb2 = f"Votre demande de changement a été approuvée."
            _cc2 = "Voir le changement →"
        else:
            _cs2 = f"Change approved: #{change.id} {change.title}"
            _cb2 = f"Your change request has been approved."
            _cc2 = "View change →"
        send_email(requester.email, _cs2, _cb2, cta_url=f"{FRONTEND_URL}/changes/{change.id}", cta_label=_cc2, db=None, tenant_id=change.tenant_id, lang=_cl2)
    return _change_to_out(change)

@app.post("/changes/{change_id}/reject")
def reject_change(change_id: int, comment: CommentCreate,
                  current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.APPROVE_CHANGES):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change not found")
    if change.status != "pending_approval":
        raise HTTPException(status_code=400, detail="Change is not in pending approval status")
    change.status = "rejected"
    db.commit()
    db.refresh(change)
    requester = db.query(User).filter(User.id == change.requester_id).first()
    if requester:
        _cl3 = get_user_language(db, requester.email)
        if _cl3 == 'fr':
            _cs3 = f"Changement rejeté : #{change.id} {change.title}"
            _cb3 = f"Votre demande de changement a été rejetée.\nRaison : {comment.body}"
            _cc3 = "Voir le changement →"
        else:
            _cs3 = f"Change rejected: #{change.id} {change.title}"
            _cb3 = f"Your change request has been rejected.\nReason: {comment.body}"
            _cc3 = "View change →"
        send_email(requester.email, _cs3, _cb3, cta_url=f"{FRONTEND_URL}/changes/{change.id}", cta_label=_cc3, db=None, tenant_id=change.tenant_id, lang=_cl3)
    return _change_to_out(change)

def _change_to_out(change: ChangeRequest, user_map: dict = None, db=None) -> dict:
    if user_map is not None:
        requester_name = user_map.get(change.requester_id, "Unknown")
    else:
        try:
            requester_name = change.requester.full_name if change.requester else "Unknown"
        except Exception:
            requester_name = "Unknown"
    try:
        owner_name = change.owner.full_name if change.owner else ""
    except Exception:
        owner_name = ""
    try:
        assigned_name = change.assigned_to.full_name if change.assigned_to else ""
    except Exception:
        assigned_name = ""

    def _safe_json(val):
        if not val: return []
        try: return json.loads(val)
        except Exception: return []

    # Fetch CAB member names if db provided
    cab_ids = _safe_json(change.cab_members)
    cab_names = []
    if db and cab_ids:
        cab_users = db.query(User).filter(User.id.in_(cab_ids)).all()
        cab_names = [{"id": u.id, "name": u.full_name} for u in cab_users]

    return {
        "id": change.id,
        "title": change.title,
        "description": change.description,
        "change_type": str(change.change_type) if change.change_type else "normal",
        "risk_level": str(change.risk_level) if change.risk_level else "medium",
        "risk_score": change.risk_score,
        "status": str(change.status) if change.status else "draft",
        "requester_id": change.requester_id,
        "requester_name": requester_name,
        "owner_id": change.owner_id,
        "owner_name": owner_name,
        "assigned_to_id": change.assigned_to_id,
        "assigned_to_name": assigned_name,
        "planned_date": change.planned_date,
        "start_date": change.start_date,
        "end_date": change.end_date,
        "impact": change.impact,
        "rollback_plan": change.rollback_plan,
        "test_plan": change.test_plan,
        "cab_members": cab_ids,
        "cab_member_names": cab_names,
        "linked_ticket_ids": _safe_json(change.linked_ticket_ids),
        "linked_asset_ids": _safe_json(change.linked_asset_ids),
        "post_review_notes": change.post_review_notes,
        "post_review_at": change.post_review_at,
        "created_at": change.created_at,
        "updated_at": change.updated_at,
    }

# =============================================================================
# CHANGE TASKS
# =============================================================================

@app.get("/changes/{change_id}/tasks")
def list_change_tasks(change_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change: raise HTTPException(status_code=404, detail="Change not found")
    tasks = db.query(ChangeTask).filter(ChangeTask.change_id == change_id).order_by(ChangeTask.created_at).all()
    return [{"id": t.id, "title": t.title, "is_done": t.is_done,
             "assigned_to_id": t.assigned_to_id,
             "assigned_to_name": t.assigned_to.full_name if t.assigned_to else None,
             "created_at": t.created_at} for t in tasks]

@app.post("/changes/{change_id}/tasks")
def create_change_task(change_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change: raise HTTPException(status_code=404, detail="Change not found")
    task = ChangeTask(change_id=change_id, title=data.get("title", "New Task"),
                      assigned_to_id=data.get("assigned_to_id"))
    db.add(task); db.commit(); db.refresh(task)
    return {"id": task.id, "title": task.title, "is_done": task.is_done}

@app.patch("/changes/{change_id}/tasks/{task_id}")
def update_change_task(change_id: int, task_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change request not found")
    task = db.query(ChangeTask).filter(ChangeTask.id == task_id, ChangeTask.change_id == change_id).first()
    if not task: raise HTTPException(status_code=404, detail="Task not found")
    for k in ["title", "is_done", "assigned_to_id"]:
        if k in data: setattr(task, k, data[k])
    db.commit()
    return {"ok": True}

@app.delete("/changes/{change_id}/tasks/{task_id}")
def delete_change_task(change_id: int, task_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change:
        raise HTTPException(status_code=404, detail="Change request not found")
    task = db.query(ChangeTask).filter(ChangeTask.id == task_id, ChangeTask.change_id == change_id).first()
    if not task: raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task); db.commit()
    return {"ok": True}

# =============================================================================
# CHANGE COMMENTS
# =============================================================================

@app.get("/changes/{change_id}/comments")
def list_change_comments(change_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change: raise HTTPException(status_code=404, detail="Change not found")
    comments = db.query(ChangeComment).filter(ChangeComment.change_id == change_id).order_by(ChangeComment.created_at).all()
    return [{"id": c.id, "body": c.body, "is_internal": c.is_internal,
             "author_id": c.author_id,
             "author_name": c.author.full_name if c.author else "Unknown",
             "created_at": c.created_at} for c in comments]

@app.post("/changes/{change_id}/comments")
def add_change_comment(change_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    change = db.query(ChangeRequest).filter(ChangeRequest.id == change_id, ChangeRequest.tenant_id == current_user.tenant_id).first()
    if not change: raise HTTPException(status_code=404, detail="Change not found")
    comment = ChangeComment(change_id=change_id, author_id=current_user.id,
                            body=data.get("body", ""), is_internal=data.get("is_internal", False))
    db.add(comment); db.commit(); db.refresh(comment)
    return {"id": comment.id, "body": comment.body, "created_at": comment.created_at}

# =============================================================================
# CHANGE CALENDAR
# =============================================================================

# =============================================================================
# =============================================================================
# =============================================================================
# APPROVAL WORKFLOWS
# =============================================================================

@app.get("/approval-workflows/")
def list_workflows(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403)
    workflows = db.query(ApprovalWorkflow).filter(
        ApprovalWorkflow.tenant_id == current_user.tenant_id,
        ApprovalWorkflow.is_active == True
    ).all()
    result = []
    for w in workflows:
        steps = []
        for s in w.steps:
            approver = db.query(User).filter(User.id == s.approver_id).first() if s.approver_id else None
            steps.append({
                "id": s.id, "step_order": s.step_order, "name": s.name,
                "approver_id": s.approver_id,
                "approver_name": approver.full_name if approver else None,
                "approver_role": s.approver_role,
            })
        result.append({
            "id": w.id, "name": w.name, "category": w.category,
            "ticket_type": w.ticket_type, "steps": steps,
        })
    return result

@app.post("/approval-workflows/")
def create_workflow(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if tenant and not get_plan_limits(tenant.plan)["approval_workflows"]:
        raise HTTPException(status_code=403, detail="Approval workflows are available on the Pro plan and above. Please upgrade your plan.")
    workflow = ApprovalWorkflow(
        tenant_id=admin.tenant_id,
        name=data.get("name", ""),
        category=data.get("category") or None,
        ticket_type=data.get("ticket_type", "service_request"),
    )
    db.add(workflow)
    db.flush()
    for i, step in enumerate(data.get("steps", []), start=1):
        db.add(ApprovalStep(
            workflow_id=workflow.id,
            step_order=i,
            name=step.get("name", f"Step {i}"),
            approver_id=int(step["approver_id"]) if step.get("approver_id") else None,
            approver_role=step.get("approver_role") or None,
        ))
    log_system_event(db, admin, "workflow.created",
                     target_type="workflow", target_id=workflow.id, target_label=workflow.name)
    db.commit()
    return {"id": workflow.id, "name": workflow.name}

@app.put("/approval-workflows/{workflow_id}")
def update_workflow(workflow_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    workflow = db.query(ApprovalWorkflow).filter(
        ApprovalWorkflow.id == workflow_id,
        ApprovalWorkflow.tenant_id == admin.tenant_id
    ).first()
    if not workflow:
        raise HTTPException(status_code=404, detail="Workflow not found")

    workflow.name = data.get("name", workflow.name)
    workflow.category = data.get("category") or None
    workflow.ticket_type = data.get("ticket_type", workflow.ticket_type)

    # Replace all steps
    db.query(ApprovalStep).filter(ApprovalStep.workflow_id == workflow.id).delete()
    db.flush()
    for i, step in enumerate(data.get("steps", []), start=1):
        db.add(ApprovalStep(
            workflow_id=workflow.id,
            step_order=i,
            name=step.get("name", f"Step {i}"),
            approver_id=int(step["approver_id"]) if step.get("approver_id") else None,
            approver_role=step.get("approver_role") or None,
        ))
    log_system_event(db, admin, "workflow.updated",
                     target_type="workflow", target_id=workflow.id, target_label=workflow.name)
    db.commit()
    return {"id": workflow.id, "name": workflow.name}

@app.delete("/approval-workflows/{workflow_id}")
def delete_workflow(workflow_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    wf = db.query(ApprovalWorkflow).filter(
        ApprovalWorkflow.id == workflow_id,
        ApprovalWorkflow.tenant_id == admin.tenant_id
    ).first()
    if not wf:
        raise HTTPException(status_code=404)
    log_system_event(db, admin, "workflow.deleted",
                     target_type="workflow", target_id=wf.id, target_label=wf.name)
    wf.is_active = False
    db.commit()
    return {"ok": True}

@app.get("/tickets/{ticket_id}/approvals")
def get_ticket_approvals(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    approvals = db.query(TicketApproval).filter(
        TicketApproval.ticket_id == ticket_id
    ).order_by(TicketApproval.step_order).all()
    result = []
    for a in approvals:
        approver = db.query(User).filter(User.id == a.approver_id).first() if a.approver_id else None
        result.append({
            "id": a.id, "step_order": a.step_order, "step_name": a.step_name,
            "approver_id": a.approver_id,
            "approver_name": approver.full_name if approver else None,
            "approver_role": a.approver_role,
            "status": a.status, "comment": a.comment,
            "decided_at": a.decided_at,
        })
    return result

@app.post("/tickets/{ticket_id}/approvals/{approval_id}/decide")
def decide_approval(ticket_id: int, approval_id: int, data: dict,
                    db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Verify ticket belongs to current user's tenant before processing approval
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    approval = db.query(TicketApproval).filter(
        TicketApproval.id == approval_id,
        TicketApproval.ticket_id == ticket_id,
        TicketApproval.status == "pending"
    ).first()
    if not approval:
        raise HTTPException(status_code=404, detail="Approval step not found or already decided")

    # Check permission — must be the designated approver or have the right role
    can_approve = has_permission(current_user, Permission.EDIT_TICKETS)
    if approval.approver_id and approval.approver_id != current_user.id and not can_approve:
        raise HTTPException(status_code=403, detail="You are not the designated approver for this step")
    if approval.approver_role and (current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role)) != approval.approver_role and not can_approve:
        raise HTTPException(status_code=403, detail="You do not have the required role to approve this step")

    decision = data.get("decision")  # "approved" or "rejected"
    comment = data.get("comment", "")
    if decision not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Decision must be 'approved' or 'rejected'")

    approval.status = decision
    approval.comment = comment
    approval.decided_at = datetime.utcnow()

    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()

    log_ticket_event(db, ticket_id, ticket.tenant_id, current_user.id,
        action=decision,
        note=f"Step {approval.step_order} ({approval.step_name}): {decision}" + (f" — {comment}" if comment else ""))

    if decision == "rejected":
        ticket.status = "closed"
        # Mark remaining steps as skipped
        db.query(TicketApproval).filter(
            TicketApproval.ticket_id == ticket_id,
            TicketApproval.status.in_(["pending", "waiting"])
        ).update({"status": "skipped"})
        # Notify requester
        create_notification(db, ticket.requester_id, ticket.tenant_id,
            "approval_rejected",
            f"❌ Request rejected: {ticket.title}",
            f'Your request was rejected at step {approval.step_order} ({approval.step_name}).' + (f' Reason: {comment}' if comment else ''),
            f"/tickets/{ticket_id}")
    else:
        # Activate next step if exists
        next_approval = db.query(TicketApproval).filter(
            TicketApproval.ticket_id == ticket_id,
            TicketApproval.step_order == approval.step_order + 1
        ).first()

        if next_approval:
            next_approval.status = "pending"
            # Notify next approver
            if next_approval.approver_id:
                create_notification(db, next_approval.approver_id, ticket.tenant_id,
                    "approval_required",
                    f"✅ Approval required: {ticket.title}",
                    f'Step {next_approval.step_order}: {next_approval.step_name}',
                    f"/tickets/{ticket_id}")
            elif next_approval.approver_role:
                approvers = db.query(User).filter(
                    User.tenant_id == ticket.tenant_id,
                    User.role == next_approval.approver_role,
                    User.is_active == True
                ).all()
                for approver in approvers:
                    create_notification(db, approver.id, ticket.tenant_id,
                        "approval_required",
                        f"✅ Approval required: {ticket.title}",
                        f'Step {next_approval.step_order}: {next_approval.step_name}',
                        f"/tickets/{ticket_id}")
        else:
            # All steps approved — move ticket to open
            ticket.status = "open"
            create_notification(db, ticket.requester_id, ticket.tenant_id,
                "approval_approved",
                f"✅ Request approved: {ticket.title}",
                "All approval steps completed. Your request is now being processed.",
                f"/tickets/{ticket_id}")

    db.commit()
    return {"ok": True, "decision": decision}

# =============================================================================
# BULK TICKET OPERATIONS
# =============================================================================

@app.post("/tickets/bulk-update")
def bulk_update_tickets(
    payload: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Bulk update tickets. Payload:
      { "ticket_ids": [1,2,3], "action": "assign"|"status"|"priority",
        "value": "user_id"|"open"|"high" }
    Only agents and admins can bulk-update.
    """
    if not has_permission(current_user, Permission.EDIT_TICKETS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    ticket_ids = payload.get("ticket_ids", [])
    action = payload.get("action")
    value = payload.get("value")

    if not ticket_ids or not action or value is None:
        raise HTTPException(status_code=400, detail="ticket_ids, action and value are required")

    tickets = db.query(Ticket).filter(
        Ticket.id.in_(ticket_ids),
        Ticket.tenant_id == current_user.tenant_id
    ).all()

    if not tickets:
        raise HTTPException(status_code=404, detail="No tickets found")

    updated = 0
    for ticket in tickets:
        try:
            if action == "assign":
                new_assignee_id = int(value) if value else None
                old_name = db.query(User).filter(User.id == ticket.assigned_to_id).first()
                new_name = db.query(User).filter(User.id == new_assignee_id).first()
                ticket.assigned_to_id = new_assignee_id
                log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                    action="assigned", field="assigned_to",
                    old_value=old_name.full_name if old_name else "Unassigned",
                    new_value=new_name.full_name if new_name else "Unassigned")
                # Notify new assignee
                if new_assignee_id and new_assignee_id != current_user.id:
                    create_notification(db, new_assignee_id, ticket.tenant_id,
                        "ticket_assigned",
                        f"Ticket {ticket.id} assigned to you",
                        f'"{ticket.title}" has been assigned to you.',
                        f"/tickets/{ticket.id}")

            elif action == "status":
                old_status = (str(ticket.status) if hasattr(ticket.status, "value") else str(ticket.status))
                ticket.status = str(value).lower()
                log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                    action="status_changed", field="status",
                    old_value=old_status, new_value=value)

            elif action == "priority":
                ticket.priority = str(value).lower()
                log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                    action="status_changed", field="priority",
                    old_value=str(ticket.priority), new_value=value)

            elif action == "assign_group":
                new_group_id = int(value) if value else None
                ticket.group_id = new_group_id
                group = db.query(Group).filter(Group.id == new_group_id).first() if new_group_id else None
                log_ticket_event(db, ticket.id, ticket.tenant_id, current_user.id,
                    action="group_assigned", field="group_id",
                    new_value=group.name if group else "Unassigned")

            updated += 1
        except Exception:
            continue

    db.commit()
    return {"updated": updated, "total": len(ticket_ids)}

# NOTIFICATIONS
# =============================================================================

@app.get("/notifications/")
def list_notifications(
    unread_only: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.tenant_id == current_user.tenant_id
    )
    if unread_only:
        query = query.filter(Notification.is_read == False)
    total = query.count()
    unread_count = db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.is_read == False
    ).count()
    items = query.order_by(Notification.created_at.desc()).offset(skip).limit(limit).all()
    return {
        "items": [{"id": n.id, "type": n.type, "title": n.title, "body": n.body,
                   "link": n.link, "is_read": n.is_read, "created_at": n.created_at} for n in items],
        "total": total,
        "unread_count": unread_count
    }

@app.patch("/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, db: Session = Depends(get_db),
                            current_user: User = Depends(get_current_user)):
    notif = db.query(Notification).filter(
        Notification.id == notification_id,
        Notification.user_id == current_user.id
    ).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    notif.is_read = True
    db.commit()
    return {"ok": True}

@app.post("/notifications/read-all")
def mark_all_read(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.is_read == False
    ).update({"is_read": True})
    db.commit()
    return {"ok": True}

# =============================================================================
# BRANDING (ADMIN ONLY)
# =============================================================================

LOGO_DIR = "logos"
os.makedirs(LOGO_DIR, exist_ok=True)

@app.get("/ping")
def ping():
    """Ultra-lightweight keepalive for UptimeRobot — no DB, no auth, always 200."""
    return {"status": "ok"}

@app.get("/health")
def health_check(db: Session = Depends(get_db)):
    """Health check endpoint for Render.
    Verifies the API is running and the database is reachable.
    Render pings this every 30s — if it fails, Render auto-restarts the service.
    """
    try:
        db.execute(text("SELECT 1"))
        return {"status": "ok", "db": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database unavailable: {str(e)}")

@app.get("/branding/public")
def get_public_branding(db: Session = Depends(get_db)):
    """Public endpoint — returns DodoDesk platform branding for the login/signup page.
    Configurable via environment variables so you can change the platform name and
    colours without touching code. Does NOT leak any tenant's company name or logo."""
    from fastapi.responses import JSONResponse
    data = {
        "company_name":    os.getenv("PLATFORM_NAME", "DodoDesk"),
        "company_tagline": os.getenv("PLATFORM_TAGLINE", "IT Service Management"),
        "primary_color":   os.getenv("PLATFORM_PRIMARY_COLOR", "#1e1e2f"),
        "accent_color":    os.getenv("PLATFORM_ACCENT_COLOR", "#4f46e5"),
        "logo_url":        os.getenv("PLATFORM_LOGO_URL", None),
    }
    return JSONResponse(content=data, headers={"Cache-Control": "public, max-age=3600"})

@app.get("/admin/branding")
def get_branding(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return {
        "company_name": tenant.name,
        "company_tagline": tenant.company_tagline or "",
        "primary_color": tenant.primary_color or "#4f46e5",
        "accent_color": tenant.accent_color or "#818cf8",
        "logo_url": tenant.logo_url,
        "support_email": tenant.support_email or "",
        "plan": tenant.plan or "free",
        "plan_limits": get_plan_limits(tenant.plan),
    }

@app.put("/admin/branding")
def update_branding(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    if not get_plan_limits(tenant.plan)["branding"]:
        raise HTTPException(status_code=403, detail="Custom branding is available on the Pro plan and above. Please upgrade your plan.")
    if data.get("company_name"):
        tenant.name = data["company_name"]
    if "company_tagline" in data:
        tenant.company_tagline = data["company_tagline"]
    if data.get("primary_color"):
        tenant.primary_color = data["primary_color"]
    if data.get("accent_color"):
        tenant.accent_color = data["accent_color"]
    if "support_email" in data:
        tenant.support_email = data["support_email"]
    if data.get("logo_url") and "cloudinary.com" in str(data["logo_url"]):
        tenant.logo_url = data["logo_url"]
    elif "logo_url" in data and not data["logo_url"]:
        tenant.logo_url = None  # explicitly clearing the logo
    log_system_event(db, admin, "branding.updated",
                     target_type="tenant", target_id=tenant.id, target_label=tenant.name)
    db.commit()
    return {"ok": True}

@app.post("/admin/branding/logo")
async def upload_logo(file: UploadFile = File(...), db: Session = Depends(get_db),
                      admin: User = Depends(get_current_admin_user)):
    allowed = {"image/png", "image/jpeg", "image/jpg", "image/svg+xml", "image/webp"}
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, SVG and WebP images allowed")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo must be under 2 MB")

    if CLOUDINARY_CLOUD_NAME:
        ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "")
        _configure_cloudinary()
        # Logos must be public (type="upload") so they can be displayed in browser and emails
        import io as _io
        public_id = f"dodesk/tenants/{admin.tenant_id}/logos/logo{ext}"
        try:
            result = cloudinary.uploader.upload(
                _io.BytesIO(content),
                public_id=public_id,
                resource_type="image",
                type="upload",       # PUBLIC — not authenticated
                overwrite=True,
                invalidate=True,
            )
            logo_url = result.get("secure_url") or f"https://res.cloudinary.com/{cloud_name}/image/upload/{public_id}"
            print(f"✅ Logo uploaded (public): {logo_url}")
        except Exception as e:
            print(f"❌ Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail=f"Logo upload failed: {str(e)}")
    else:
        print(f"⚠️ CLOUDINARY_CLOUD_NAME not set — logo will be lost on next deploy")
        raise HTTPException(status_code=500, detail="Cloudinary not configured. Please set CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET in Render environment variables.")

    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    tenant.logo_url = logo_url
    db.commit()
    return {"logo_url": logo_url}

@app.get("/logos/{filename}")
def serve_logo(filename: str):
    path = os.path.join(LOGO_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Logo not found")
    return FileResponse(path)

# =============================================================================
# SLA CONFIGURATION (ADMIN ONLY)
# =============================================================================

@app.get("/admin/sla-config")
def get_sla_config(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    cfg = db.query(SLAConfig).filter(SLAConfig.tenant_id == admin.tenant_id).first()
    defaults = SLA_RULES
    return {
        "low_response":      cfg.low_response      if cfg else defaults["low"]["response"],
        "low_resolution":    cfg.low_resolution    if cfg else defaults["low"]["resolution"],
        "medium_response":   cfg.medium_response   if cfg else defaults["medium"]["response"],
        "medium_resolution": cfg.medium_resolution if cfg else defaults["medium"]["resolution"],
        "high_response":     cfg.high_response     if cfg else defaults["high"]["response"],
        "high_resolution":   cfg.high_resolution   if cfg else defaults["high"]["resolution"],
        "critical_response":     cfg.critical_response     if cfg else defaults["critical"]["response"],
        "critical_resolution":   cfg.critical_resolution   if cfg else defaults["critical"]["resolution"],
    }

@app.put("/admin/sla-config")
def update_sla_config(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if tenant and not get_plan_limits(tenant.plan)["sla"]:
        raise HTTPException(status_code=403, detail="SLA configuration is available on the Pro plan and above. Please upgrade your plan.")
    cfg = db.query(SLAConfig).filter(SLAConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = SLAConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.low_response      = int(data.get("low_response", 8))
    cfg.low_resolution    = int(data.get("low_resolution", 72))
    cfg.medium_response   = int(data.get("medium_response", 4))
    cfg.medium_resolution = int(data.get("medium_resolution", 48))
    cfg.high_response     = int(data.get("high_response", 2))
    cfg.high_resolution   = int(data.get("high_resolution", 24))
    cfg.critical_response     = int(data.get("critical_response", 1))
    cfg.critical_resolution   = int(data.get("critical_resolution", 8))
    log_system_event(db, admin, "sla_config.updated",
                     target_type="tenant", target_id=admin.tenant_id)
    db.commit()
    return {"ok": True}

# =============================================================================
# ESCALATION RULES (ADMIN ONLY)
# =============================================================================

@app.get("/admin/escalation-rules")
def list_escalation_rules(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rules = db.query(EscalationRule).filter(
        EscalationRule.tenant_id == admin.tenant_id,
        EscalationRule.is_active == True
    ).all()
    result = []
    for r in rules:
        agent = db.query(User).filter(User.id == r.escalate_to_id).first() if r.escalate_to_id else None
        result.append({
            "id": r.id, "name": r.name, "priority": r.priority,
            "idle_hours": r.idle_hours,
            "escalate_to_id": r.escalate_to_id,
            "escalate_to_name": agent.full_name if agent else None,
            "escalate_to_role": r.escalate_to_role,
            "created_at": r.created_at,
        })
    return result

@app.post("/admin/escalation-rules")
def create_escalation_rule(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = EscalationRule(
        tenant_id=admin.tenant_id,
        name=data.get("name", ""),
        priority=data.get("priority") or None,
        idle_hours=int(data.get("idle_hours", 24)),
        escalate_to_id=int(data["escalate_to_id"]) if data.get("escalate_to_id") else None,
        escalate_to_role=data.get("escalate_to_role") or None,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "name": rule.name}

@app.delete("/admin/escalation-rules/{rule_id}")
def delete_escalation_rule(rule_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = db.query(EscalationRule).filter(
        EscalationRule.id == rule_id,
        EscalationRule.tenant_id == admin.tenant_id
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    rule.is_active = False
    db.commit()
    return {"ok": True}

# =============================================================================
# BUSINESS HOURS CONFIGURATION (ADMIN ONLY)
# =============================================================================

@app.get("/admin/business-hours")
def get_business_hours(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    cfg = db.query(BusinessHoursConfig).filter(BusinessHoursConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        return {"enabled": False, "start_hour": 9, "end_hour": 17,
                "working_days": "0,1,2,3,4", "timezone": "UTC"}
    return {"enabled": cfg.enabled, "start_hour": cfg.start_hour,
            "end_hour": cfg.end_hour, "working_days": cfg.working_days,
            "timezone": cfg.timezone}

@app.put("/admin/business-hours")
def update_business_hours(data: dict, db: Session = Depends(get_db),
                          admin: User = Depends(get_current_admin_user)):
    cfg = db.query(BusinessHoursConfig).filter(BusinessHoursConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = BusinessHoursConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.enabled = data.get("enabled", False)
    cfg.start_hour = int(data.get("start_hour", 9))
    cfg.end_hour = int(data.get("end_hour", 17))
    cfg.working_days = data.get("working_days", "0,1,2,3,4")
    cfg.timezone = data.get("timezone", "UTC")
    db.commit()
    return {"ok": True}

# =============================================================================
# SECURITY CONFIGURATION (MFA + SSO) — ADMIN ONLY
# =============================================================================

@app.get("/admin/security-config")
def get_security_config(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return {
        "mfa_enabled":        bool(tenant.mfa_enabled),
        "mfa_required":       bool(tenant.mfa_required),
        "sso_enabled":        bool(tenant.sso_enabled),
        "sso_provider":       tenant.sso_provider or "saml",
        "sso_client_id":      tenant.sso_client_id or "",    # SAML Entity ID / Issuer
        "sso_client_secret":  "",                            # never return secrets
        "sso_sso_url":        getattr(tenant, "sso_sso_url", "") or "",  # IdP SSO URL
        "sso_domain":         tenant.sso_domain or "",       # allowed email domain
        "sso_tenant_id":      tenant.sso_tenant_id or "",    # Azure tenant ID (optional)
        "saml_cert":          getattr(tenant, "saml_cert", "") or "",    # IdP certificate
        "sp_metadata_url":    f"{API_URL}/auth/sso/metadata/{tenant.slug}",
        "sp_acs_url":         f"{API_URL}/auth/sso/callback/{tenant.slug}",
        "sp_entity_id":       f"{API_URL}/auth/sso/metadata/{tenant.slug}",
    }

@app.put("/admin/security-config")
def update_security_config(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    limits = get_plan_limits(tenant.plan)
    if data.get("mfa_enabled") and not limits["mfa"]:
        raise HTTPException(status_code=403, detail="Two-factor authentication is available on the Pro plan and above. Please upgrade your plan.")
    if data.get("sso_enabled") and not limits["sso"]:
        raise HTTPException(status_code=403, detail="Single sign-on (SSO) is available on the Pro plan and above. Please upgrade your plan.")

    tenant.mfa_enabled  = bool(data.get("mfa_enabled", False))
    tenant.mfa_required = bool(data.get("mfa_required", False)) if tenant.mfa_enabled else False
    tenant.sso_enabled  = bool(data.get("sso_enabled", False))
    tenant.sso_provider = data.get("sso_provider", "saml")
    tenant.sso_client_id = data.get("sso_client_id") or None       # SAML Entity ID
    if data.get("sso_client_secret"):
        tenant.sso_client_secret = data.get("sso_client_secret")
    tenant.sso_domain    = data.get("sso_domain") or None           # allowed email domain
    tenant.sso_tenant_id = data.get("sso_tenant_id") or None        # Azure tenant ID
    if hasattr(tenant, "sso_sso_url"):
        tenant.sso_sso_url = data.get("sso_sso_url") or None        # IdP SSO URL
    if hasattr(tenant, "saml_cert"):
        tenant.saml_cert   = data.get("saml_cert") or None          # IdP X.509 cert
    log_system_event(db, admin, "security_config.updated",
                     target_type="tenant", target_id=tenant.id, target_label=tenant.name,
                     new_value=f"mfa={tenant.mfa_enabled} mfa_required={tenant.mfa_required} sso={tenant.sso_enabled}")
    db.commit()
    return {"ok": True}

# =============================================================================
# IP WHITELIST — Enterprise plan only
# =============================================================================

@app.get("/admin/ip-whitelist")
def get_ip_whitelist(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("sso", tenant, "IP whitelisting is available on the Enterprise plan only.")
    raw = getattr(tenant, "ip_whitelist", None)
    try:
        cidrs = json.loads(raw) if raw else []
    except Exception:
        cidrs = []
    return {"cidrs": cidrs}

@app.put("/admin/ip-whitelist")
def update_ip_whitelist(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("sso", tenant, "IP whitelisting is available on the Enterprise plan only.")
    cidrs = data.get("cidrs", [])
    # Validate each CIDR
    import ipaddress as _ip
    valid = []
    errors = []
    for cidr in cidrs:
        cidr = cidr.strip()
        if not cidr:
            continue
        try:
            _ip.ip_network(cidr, strict=False)
            valid.append(cidr)
        except ValueError:
            errors.append(cidr)
    if errors:
        raise HTTPException(status_code=422, detail=f"Invalid CIDR(s): {', '.join(errors)}")
    tenant.ip_whitelist = json.dumps(valid) if valid else None
    db.commit()
    return {"ok": True, "cidrs": valid, "message": f"IP whitelist updated — {len(valid)} rule(s) active."}


# =============================================================================
# SCHEDULED REPORTS — Business plan and above
# =============================================================================

@app.get("/admin/scheduled-reports")
def get_scheduled_reports(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("custom_analytics", tenant, "Scheduled reports are available on the Business plan and above.")
    raw = getattr(tenant, "scheduled_reports", None)
    try:
        config = json.loads(raw) if raw else {"enabled": False, "frequency": "weekly", "day": "monday", "time": "08:00", "recipients": [], "include": ["summary", "sla", "agent_workload"]}
    except Exception:
        config = {"enabled": False, "frequency": "weekly", "day": "monday", "time": "08:00", "recipients": [], "include": ["summary"]}
    return config

@app.put("/admin/scheduled-reports")
def update_scheduled_reports(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("custom_analytics", tenant, "Scheduled reports are available on the Business plan and above.")
    config = {
        "enabled":    bool(data.get("enabled", False)),
        "frequency":  data.get("frequency", "weekly"),   # daily | weekly | monthly
        "day":        data.get("day", "monday"),          # day of week for weekly
        "time":       data.get("time", "08:00"),          # HH:MM UTC
        "recipients": data.get("recipients", []),         # list of email addresses
        "include":    data.get("include", ["summary"]),   # report sections to include
    }
    tenant.scheduled_reports = json.dumps(config)
    db.commit()
    return {"ok": True, "message": "Scheduled report settings saved.", **config}


def _send_scheduled_report(tenant_id: int):
    """Generate and email the scheduled report for a tenant. Called by APScheduler."""
    from sqlalchemy.orm import Session as _S
    db = next(get_db())
    try:
        tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        if not tenant:
            return
        raw = getattr(tenant, "scheduled_reports", None)
        if not raw:
            return
        config = json.loads(raw)
        if not config.get("enabled"):
            return
        recipients = config.get("recipients", [])
        if not recipients:
            return

        include = config.get("include", ["summary"])
        sections = []

        # Build report sections
        if "summary" in include:
            from sqlalchemy import text as _t
            row = db.execute(_t(
                "SELECT COUNT(*) FILTER (WHERE status NOT IN ('closed','resolved')) as open_count, "
                "COUNT(*) FILTER (WHERE status='resolved') as resolved_count, "
                "COUNT(*) FILTER (WHERE status='closed') as closed_count, "
                "COUNT(*) as total "
                "FROM tickets WHERE tenant_id=:tid AND created_at > NOW() - INTERVAL '7 days'"
            ), {"tid": tenant_id}).fetchone()
            if row:
                sections.append(
                    f"📊 Ticket Summary (last 7 days)\n"
                    f"  Open: {row[0]}  |  Resolved: {row[1]}  |  Closed: {row[2]}  |  Total: {row[3]}"
                )

        if "sla" in include:
            row2 = db.execute(_t(
                "SELECT COUNT(*) FILTER (WHERE sla_response_breached=true) as r_breach, "
                "COUNT(*) FILTER (WHERE sla_resolution_breached=true) as res_breach, "
                "COUNT(*) as total "
                "FROM tickets WHERE tenant_id=:tid AND created_at > NOW() - INTERVAL '7 days'"
            ), {"tid": tenant_id}).fetchone()
            if row2 and row2[2]:
                r_pct = round((1 - row2[0] / row2[2]) * 100, 1)
                sections.append(
                    f"\n⏱ SLA Performance (last 7 days)\n"
                    f"  Response SLA: {r_pct}%  |  Breaches: {row2[0]} response, {row2[1]} resolution"
                )

        if "agent_workload" in include:
            rows3 = db.execute(_t(
                "SELECT u.full_name, COUNT(*) as cnt "
                "FROM tickets t JOIN users u ON u.id=t.assigned_to_id "
                "WHERE t.tenant_id=:tid AND t.status NOT IN ('closed','resolved') "
                "GROUP BY u.full_name ORDER BY cnt DESC LIMIT 5"
            ), {"tid": tenant_id}).fetchall()
            if rows3:
                lines = "\n".join(f"  {r[0]}: {r[1]} open tickets" for r in rows3)
                sections.append(f"\n👥 Agent Workload (open tickets)\n{lines}")

        body = (
            f"Hi,\n\nHere is your {config.get('frequency', 'weekly')} DodoDesk report "
            f"for {tenant.name}.\n\n"
            + "\n".join(sections) +
            f"\n\nView full reports: {FRONTEND_URL}/reports\n\nDodoDesk"
        )

        notif_cfg = get_email_config(db, tenant_id)
        for recipient in recipients:
            try:
                send_email(
                    recipient,
                    f"📊 DodoDesk {config.get('frequency', 'Weekly').capitalize()} Report — {tenant.name}",
                    body,
                    db=None
                )
            except Exception as e:
                print(f"⚠️ Scheduled report email failed for {recipient}: {e}")

        print(f"✅ Scheduled report sent for tenant {tenant_id} to {len(recipients)} recipient(s)")
    except Exception as e:
        print(f"⚠️ Scheduled report error for tenant {tenant_id}: {e}")
    finally:
        db.close()


# =============================================================================
# EMAIL CONFIGURATION (ADMIN ONLY)
# =============================================================================

@app.get("/admin/email-config")
def get_email_config_endpoint(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Get email/webhook config for this tenant. Uses raw SQL to avoid ORM schema mismatch errors."""
    defaults = {
        "smtp_host": SMTP_HOST or "", "smtp_port": SMTP_PORT or 587,
        "smtp_user": SMTP_USER or "", "smtp_pass": "",
        "smtp_from": SMTP_FROM or "", "reply_to": "",
        "slack_webhook_url": SLACK_WEBHOOK_URL or "",
        "teams_webhook_url": TEAMS_WEBHOOK_URL or "",
        "email_signature": "", "email_footer": "",
    }
    try:
        from sqlalchemy import text as _t
        try:
            row = db.execute(_t(
                "SELECT smtp_host, smtp_port, smtp_user, smtp_from, reply_to, "
                "email_signature, email_footer, slack_webhook_url, teams_webhook_url "
                "FROM email_configs WHERE tenant_id = :tid LIMIT 1"
            ), {"tid": admin.tenant_id}).fetchone()
        except Exception:
            try:
                row = db.execute(_t(
                    "SELECT smtp_host, smtp_port, smtp_user, smtp_from, reply_to "
                    "FROM email_configs WHERE tenant_id = :tid LIMIT 1"
                ), {"tid": admin.tenant_id}).fetchone()
            except Exception:
                row = None
        if not row:
            return defaults
        result = dict(defaults)
        cols = row._fields if hasattr(row, '_fields') else row.keys()
        for col in cols:
            if col == "smtp_pass":
                continue
            val = row[col]
            if val is not None:
                result[col] = val
        result["smtp_pass"] = ""
        return result
    except Exception as e:
        print(f"⚠️ email-config error: {e}")
        return defaults  # always return defaults, never 500

@app.put("/admin/email-config")
def update_email_config(
    data: dict,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    cfg = db.query(EmailConfig).filter(EmailConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = EmailConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.smtp_host = data.get("smtp_host", "")
    cfg.smtp_port = int(data.get("smtp_port", 587))
    cfg.smtp_user = data.get("smtp_user", "")
    if data.get("smtp_pass"):
        cfg.smtp_pass = data.get("smtp_pass")
    cfg.smtp_from = data.get("smtp_from", "noreply@itsm.local")
    cfg.reply_to  = data.get("reply_to", "")
    # These columns may not exist yet in older DBs — safe assignment
    try: cfg.slack_webhook_url = data.get("slack_webhook_url", "")
    except Exception: pass
    try: cfg.teams_webhook_url = data.get("teams_webhook_url", "")
    except Exception: pass
    try: cfg.email_signature = data.get("email_signature", "")
    except Exception: pass
    try: cfg.email_footer = data.get("email_footer", "")
    except Exception: pass
    db.commit()
    return {"ok": True}

@app.post("/admin/email-config/test")
def test_email_config(
    data: dict,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """Send a test email using the provided config."""
    to = data.get("test_email") or admin.email
    cfg = {
        "smtp_host": data.get("smtp_host", ""),
        "smtp_port": int(data.get("smtp_port", 587)),
        "smtp_user": data.get("smtp_user", ""),
        "smtp_pass": data.get("smtp_pass", ""),
        "smtp_from": data.get("smtp_from", "noreply@itsm.local"),
    }
    try:
        send_email(to, "ITSM Test Email", "This is a test email from your ITSM portal.", cfg)
        return {"ok": True, "message": f"Test email sent to {to}"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/admin/email-config/test-slack")
def test_slack_webhook(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Send a test message to the configured Slack webhook."""
    cfg = get_email_config(db, admin.tenant_id)
    slack_url = cfg.get("slack_webhook_url", "")
    if not slack_url:
        raise HTTPException(status_code=400, detail="No Slack webhook URL configured. Add it in Webhooks settings first.")
    import httpx
    try:
        resp = httpx.post(slack_url, json={
            "text": f"✅ *DodoDesk test message* — Slack integration is working correctly for *{admin.full_name}*'s workspace."
        }, timeout=10.0)
        if resp.status_code == 200:
            return {"ok": True, "message": "Test message sent to Slack successfully."}
        raise HTTPException(status_code=400, detail=f"Slack returned {resp.status_code}: {resp.text}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not reach Slack: {str(e)}")

@app.post("/admin/email-config/test-teams")
def test_teams_webhook(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Send a test message to the configured Microsoft Teams webhook."""
    cfg = get_email_config(db, admin.tenant_id)
    teams_url = cfg.get("teams_webhook_url", "")
    if not teams_url:
        raise HTTPException(status_code=400, detail="No Teams webhook URL configured. Add it in Webhooks settings first.")
    import httpx
    try:
        resp = httpx.post(teams_url, json={
            "@type": "MessageCard",
            "@context": "http://schema.org/extensions",
            "summary": "DodoDesk Test",
            "themeColor": "059669",
            "title": "✅ DodoDesk — Test Message",
            "text": f"Teams integration is working correctly for **{admin.full_name}**'s workspace."
        }, timeout=10.0)
        if resp.status_code in (200, 202):
            return {"ok": True, "message": "Test message sent to Teams successfully."}
        raise HTTPException(status_code=400, detail=f"Teams returned {resp.status_code}: {resp.text}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not reach Teams: {str(e)}")

# USER MANAGEMENT (ADMIN ONLY, tenant‑scoped)
# =============================================================================

@app.get("/admin/users")
def admin_list_users(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=1000),
    search: str = Query("", alias="search"),
    role: str = Query("", alias="role"),
    tenant_id: int | None = Query(None, alias="tenant_id"),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    if str(admin.role) in ("super_admin", "platform_admin"):
        query = db.query(User)
    else:
        query = db.query(User).filter(User.tenant_id == admin.tenant_id)

    # Filter by tenant (super admin only)
    if tenant_id and str(admin.role) in ("super_admin", "platform_admin"):
        query = query.filter(User.tenant_id == tenant_id)

    # Filter by role
    if role:
        try:
            query = query.filter(User.role == UserRole(role))
        except ValueError:
            pass

    # Live search — ID, name, email, employee_id
    if search:
        from sqlalchemy import or_
        import re as _re
        s = f"%{search}%"
        id_match = _re.search(r'\d+', search)
        conditions = [
            User.full_name.ilike(s),
            User.email.ilike(s),
            User.employee_id.ilike(s),
            User.department.ilike(s),
            User.job_title.ilike(s),
        ]
        if id_match:
            conditions.append(User.id == int(id_match.group()))
        query = query.filter(or_(*conditions))

    total = query.count()
    users = query.order_by(User.tenant_id, User.id).offset(skip).limit(limit).all()
    # Pre-load tenants
    tenant_ids = {u.tenant_id for u in users}
    tenant_map = {t.id: t.name for t in db.query(Tenant).filter(Tenant.id.in_(tenant_ids)).all()}
    result = []
    for u in users:
        result.append({
            "id": u.id, "email": u.email, "full_name": u.full_name,
            "role": u.role, "is_active": u.is_active,
            "job_title": u.job_title, "department": u.department,
            "employee_id": getattr(u, 'employee_id', None),
            "tenant_id": u.tenant_id,
            "tenant_name": tenant_map.get(u.tenant_id, "—"),
            "created_at": u.created_at,
            "is_locked": bool(u.locked_until and u.locked_until > datetime.utcnow()),
            "status_changed_at": u.status_changed_at,
        })
    return {"items": result, "total": total, "skip": skip, "limit": limit}

@app.post("/admin/users", response_model=UserOut)
def admin_create_user(user_data: UserCreate, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    existing = db.query(User).filter(User.email == user_data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    validate_password_strength(user_data.password)
    # Allow super-admin to assign user to a different tenant
    target_tenant_id = admin.tenant_id
    if hasattr(user_data, 'tenant_id') and user_data.tenant_id:
        tenant = db.query(Tenant).filter(Tenant.id == user_data.tenant_id).first()
        if tenant:
            target_tenant_id = tenant.id
    check_user_limit(db, target_tenant_id, additional=1, role=user_data.role)
    new_user = User(
        tenant_id=target_tenant_id,
        email=user_data.email,
        hashed_password=get_password_hash(user_data.password),
        full_name=user_data.full_name,
        role=user_data.role,
        job_title=user_data.job_title,
        department=user_data.department,
        employee_id=getattr(user_data, 'employee_id', None),
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    log_system_event(db, admin, "user.created",
                     target_type="user", target_id=new_user.id,
                     target_label=new_user.email,
                     new_value=user_data.role if isinstance(user_data.role, str) else str(user_data.role))
    db.commit()
    return new_user

@app.post("/admin/users/invite")
def invite_user(invite: UserInvite, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Invite a new user to the admin's tenant by email — no password is set by the
    admin. The invitee receives an email with a link to set their own password,
    which also activates the account. This is the recommended way to add teammates,
    as opposed to /admin/users which requires the admin to choose a password directly."""
    email = invite.email.strip().lower()
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        if existing.is_active:
            raise HTTPException(status_code=400, detail="A user with this email already exists.")
        # Inactive account exists (e.g. a stale unverified signup or a previous
        # unactivated invite) — re-send the invite rather than blocking
        existing_user = existing
    else:
        existing_user = None

    check_user_limit(db, admin.tenant_id, additional=1, role=invite.role)

    if existing_user:
        existing_user.full_name = invite.full_name
        existing_user.role = invite.role
        existing_user.job_title = invite.job_title
        existing_user.department = invite.department
        existing_user.tenant_id = admin.tenant_id
        new_user = existing_user
    else:
        new_user = User(
            tenant_id=admin.tenant_id,
            email=email,
            hashed_password=get_password_hash(uuid.uuid4().hex),  # random unusable placeholder — invitee sets their own
            full_name=invite.full_name,
            role=invite.role,
            job_title=invite.job_title,
            department=invite.department,
            is_active=False,
            email_verified=False,
        )
        db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Generate a set-password token reusing the same mechanism as forgot-password,
    # tagged so reset_password() knows to activate the account on completion.
    from sqlalchemy import text as _text
    token = uuid.uuid4().hex
    invite_val = f"invite_{token}"
    expires_at = datetime.utcnow() + timedelta(days=7)  # invites get a longer window than a reset link
    with db.bind.connect() as conn:
        conn.execute(
            _text("UPDATE users SET password_reset_token = :tok, password_reset_expires_at = :exp WHERE id = :uid"),
            {"tok": invite_val, "uid": new_user.id, "exp": expires_at}
        )
        conn.commit()

    invite_url = f"{FRONTEND_URL}/reset-password?token={token}&invite=1"

    admin_tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    company_name = admin_tenant.name if admin_tenant else "DodoDesk"
    role_label = invite.role.value if hasattr(invite.role, "value") else str(invite.role)

    import threading
    _email, _name, _url, _company, _role = new_user.email, new_user.full_name, invite_url, company_name, role_label
    send_email_background(
        to=new_user.email,
        subject=f"You've been invited to {company_name} on DodoDesk",
        body=(
            f"Hi {new_user.full_name},\n\n"
            f"{admin.full_name} has invited you to join {company_name} on DodoDesk as a {role_label}.\n\n"
            f"Click the link below to set your password and activate your account:\n\n"
            f"{invite_url}\n\n"
            f"This invite link expires in 7 days."
        ),
        cta_url=invite_url,
        cta_label="Accept Invite & Set Password",
    )
    log_system_event(db, admin, "user.invited",
                     target_type="user", target_id=new_user.id,
                     target_label=new_user.email, new_value=role_label)
    db.commit()
    return {"ok": True, "message": f"Invite sent to {email}", "user_id": new_user.id}

@app.post("/admin/users/bulk-import")
async def bulk_import_users(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Bulk-create users from a CSV or XLSX file. Expected columns (header row required):
    full_name, email, role, job_title, department, password (optional), tenant (optional, super_admin only)
    If password is omitted, a random temporary password is generated.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    rows_data = []  # list of dicts, normalized lowercase keys

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel import is not available on this server (openpyxl not installed). Please use CSV instead, or contact support.")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(status_code=400, detail="Spreadsheet appears to be empty")
            headers = [str(h or "").strip().lower() for h in all_rows[0]]
            for r in all_rows[1:]:
                if all(c is None or str(c).strip() == "" for c in r):
                    continue  # skip fully empty rows
                row_dict = {headers[idx]: ("" if r[idx] is None else str(r[idx]).strip()) for idx in range(len(headers)) if idx < len(r)}
                rows_data.append(row_dict)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    else:
        try:
            text_content = content.decode("utf-8-sig")  # handle BOM from Excel CSV exports
        except UnicodeDecodeError:
            text_content = content.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text_content))
        if reader.fieldnames is None:
            raise HTTPException(status_code=400, detail="CSV file appears to be empty or invalid")
        reader.fieldnames = [(f or "").strip().lower() for f in reader.fieldnames]
        headers = reader.fieldnames
        for row in reader:
            rows_data.append({k: (v or "") for k, v in row.items()})

    required_cols = {"full_name", "email"}
    missing = required_cols - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"File is missing required column(s): {', '.join(missing)}")

    valid_roles = {r.value for r in UserRole}
    results = {"created": [], "skipped": [], "errors": []}

    for i, row in enumerate(rows_data, start=2):  # start=2 because row 1 is the header
        email = (row.get("email") or "").strip().lower()
        full_name = (row.get("full_name") or "").strip()

        if not email or not full_name:
            results["errors"].append({"row": i, "email": email, "reason": "Missing full_name or email"})
            continue

        if db.query(User).filter(User.email == email).first():
            results["skipped"].append({"row": i, "email": email, "reason": "Email already exists"})
            continue

        role_raw = (row.get("role") or "employee").strip().lower()
        if role_raw not in valid_roles or role_raw == "super_admin":
            role_raw = "employee"

        # Determine tenant
        target_tenant_id = admin.tenant_id
        tenant_identifier = (row.get("tenant") or row.get("tenant_slug") or "").strip()
        if str(admin.role) in ("super_admin", "platform_admin") and tenant_identifier:
            tenant = db.query(Tenant).filter(
                (sa_func.lower(Tenant.slug) == tenant_identifier.lower()) |
                (sa_func.lower(Tenant.name) == tenant_identifier.lower())
            ).first()
            if tenant:
                target_tenant_id = tenant.id
            else:
                results["errors"].append({"row": i, "email": email, "reason": f"Unknown tenant '{tenant_identifier}' (use exact company name or slug)"})
                continue

        password = (row.get("password") or "").strip()
        temp_password_generated = False
        if not password:
            password = _secrets.token_urlsafe(9)  # ~12 char random password
            temp_password_generated = True

        try:
            validate_password_strength(password)
        except HTTPException:
            password = _secrets.token_urlsafe(9)
            temp_password_generated = True

        try:
            check_user_limit(db, target_tenant_id, additional=1, role=role_raw)
        except HTTPException as e:
            results["errors"].append({"row": i, "email": email, "reason": e.detail})
            continue

        new_user = User(
            tenant_id=target_tenant_id,
            email=email,
            hashed_password=get_password_hash(password),
            full_name=full_name,
            role=UserRole(role_raw),
            job_title=(row.get("job_title") or "").strip() or None,
            department=(row.get("department") or "").strip() or None,
            employee_id=(row.get("employee_id") or "").strip() or None,
            is_active=True,
        )
        db.add(new_user)
        try:
            db.flush()
        except Exception as e:
            db.rollback()
            results["errors"].append({"row": i, "email": email, "reason": str(e)})
            continue

        results["created"].append({
            "row": i, "email": email, "full_name": full_name, "role": role_raw,
            "temp_password": password if temp_password_generated else None,
        })

    db.commit()
    return results

@app.delete("/admin/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Permanently delete a user. Super admin only.
    Handles all FK references using exact column names from the model definitions.
    """
    if str(admin.role) not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Only super admins can delete users.")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    email = user.email
    name  = user.full_name
    try:
        from sqlalchemy import text as _t
        u = user_id
        with db.bind.connect() as conn:
            # ── DELETE rows owned by this user (avoids NOT NULL violations) ──
            conn.execute(_t("DELETE FROM signup_verifications WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM group_members WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM notifications WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_watchers WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM time_entries WHERE agent_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM admin_tenant_access WHERE admin_user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM canned_responses WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM comments WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM change_comments WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_audit_logs WHERE actor_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_views WHERE created_by_id = :u"), {"u": u})
            # chat messages before chat sessions
            conn.execute(_t("DELETE FROM chat_messages WHERE session_id IN (SELECT id FROM chat_sessions WHERE user_id = :u)"), {"u": u})
            conn.execute(_t("DELETE FROM chat_sessions WHERE user_id = :u"), {"u": u})

            # ── NULLIFY shared entity references (keep historical records) ──
            conn.execute(_t("UPDATE tickets SET requester_id = NULL WHERE requester_id = :u"), {"u": u})
            conn.execute(_t("UPDATE tickets SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE kb_articles SET author_id = NULL WHERE author_id = :u"), {"u": u})
            conn.execute(_t("UPDATE kb_versions SET edited_by_id = NULL WHERE edited_by_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET requester_id = NULL WHERE requester_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET owner_id = NULL WHERE owner_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_tasks SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE ticket_tasks SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE assets SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE system_audit_logs SET actor_id = NULL WHERE actor_id = :u"), {"u": u})
            conn.execute(_t("UPDATE admin_tenant_access SET granted_by_id = NULL WHERE granted_by_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET from_user_id = NULL WHERE from_user_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET to_user_id = NULL WHERE to_user_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET changed_by_id = NULL WHERE changed_by_id = :u"), {"u": u})

            # ── DELETE the user row ───────────────────────────────────────
            conn.execute(_t("DELETE FROM users WHERE id = :u"), {"u": u})
            conn.commit()

        # Audit log after deletion (admin still exists)
        log_system_event(db, admin, "user.deleted",
                         target_type="user", target_id=user_id, target_label=email)
        db.commit()

        # Reduce Dodo Payments seat count if this was an agent/admin on a paid plan
        try:
            role_value = (user.role.value if hasattr(user.role, "value") else str(user.role)) if user.role else ""
            if role_value in ("agent", "admin", "super_admin"):
                tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first()
                if tenant and tenant.billing_status == "active":
                    new_count = db.query(User).filter(
                        User.tenant_id == tenant.id,
                        User.is_active == True,
                        User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
                    ).count()
                    _update_dodo_seat_count(tenant, max(1, new_count))
        except Exception as e:
            print(f"⚠️ Seat reduction after delete failed (user still deleted): {e}")

        return {"ok": True, "message": f"{name} ({email}) has been permanently deleted."}
    except Exception as e:
        print(f"❌ delete_user {user_id}: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not delete user: {str(e)}")

@app.post("/admin/users/{user_id}/unlock")
def unlock_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    query = db.query(User).filter(User.id == user_id)
    # Super admin can unlock users from any tenant; regular admins only their own
    if str(admin.role) not in ("super_admin", "platform_admin"):
        query = query.filter(User.tenant_id == admin.tenant_id)
    user = query.first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.locked_until = None
    user.failed_login_attempts = 0
    user.is_active = True   # also re-activate in case account was deactivated
    log_system_event(db, admin, "user.unlocked",
                     target_type="user", target_id=user.id, target_label=user.email)
    db.commit()
    return {"ok": True, "message": f"{user.full_name} has been unlocked."}

@app.get("/admin/users/{user_id}", response_model=UserOut)
def admin_get_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@app.patch("/admin/users/{user_id}", response_model=UserOut)
def admin_update_user(user_id: int, user_update: UserUpdate,
                      db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    query = db.query(User).filter(User.id == user_id)
    # Tenant admins may only manage users within their own tenant — super_admin can manage any tenant
    if str(admin.role) not in ("super_admin", "platform_admin"):
        query = query.filter(User.tenant_id == admin.tenant_id)
    user = query.first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_data = user_update.model_dump(exclude_unset=True)
    if "password" in update_data:
        validate_password_strength(update_data["password"])
        user.hashed_password = get_password_hash(update_data.pop("password"))
        log_system_event(db, admin, "user.password_reset",
                         target_type="user", target_id=user.id, target_label=user.email)
    if "is_active" in update_data and update_data["is_active"] != user.is_active:
        user.status_changed_at = datetime.utcnow()
        action = "user.activated" if update_data["is_active"] else "user.deactivated"
        log_system_event(db, admin, action,
                         target_type="user", target_id=user.id, target_label=user.email)
    if "role" in update_data and str(update_data["role"]) != str(user.role):
        # Only super_admin can grant or modify the super_admin role — prevents
        # a tenant admin from elevating themselves or another user beyond their tenant scope
        if str(update_data["role"]) in ("super_admin", "platform_admin") and str(admin.role) not in ("super_admin", "platform_admin"):
            raise HTTPException(status_code=403, detail="Only a super admin can grant super admin access")
        # Prevent an admin from demoting their own last admin account in a tenant —
        # avoids accidentally locking everyone out of tenant administration
        if user.id == admin.id and str(update_data["role"]) not in ("admin", "super_admin", "platform_admin"):
            other_admins = db.query(User).filter(
                User.tenant_id == admin.tenant_id, User.id != admin.id,
                User.role.in_(['admin', 'super_admin', 'platform_admin']), User.is_active == True
            ).count()
            if other_admins == 0:
                raise HTTPException(status_code=400, detail="You cannot remove your own admin access — you are the only admin on this account")
        log_system_event(db, admin, "user.role_changed",
                         target_type="user", target_id=user.id, target_label=user.email,
                         old_value=str(user.role), new_value=str(update_data["role"]))
    if "tenant_id" in update_data:
        if str(admin.role) not in ("super_admin", "platform_admin"):
            raise HTTPException(status_code=403, detail="Only a super admin can move a user between tenants")
        tenant = db.query(Tenant).filter(Tenant.id == update_data["tenant_id"]).first()
        if not tenant:
            raise HTTPException(status_code=400, detail="Invalid tenant")
    for key, value in update_data.items():
        setattr(user, key, value)
    db.commit()
    db.refresh(user)

    # Adjust Dodo seat count if activation status or role changed for agent/admin
    try:
        role_val = str((user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(user.role, 'value') else user.role)
        if role_val in ("agent", "admin", "super_admin"):
            tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first()
            if tenant and tenant.billing_status == "active":
                new_count = db.query(User).filter(
                    User.tenant_id == tenant.id,
                    User.is_active == True,
                    User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
                ).count()
                _update_dodo_seat_count(tenant, max(1, new_count))
    except Exception as e:
        print(f"⚠️ Seat adjustment after user update failed (change still saved): {e}")

    return user

# =============================================================================
# CANNED RESPONSES (permissions)
# =============================================================================

# =============================================================================
# AUTOMATION RULES
# =============================================================================

@app.get("/admin/automation-rules")
def list_automation_rules(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rules = db.query(AutomationRule).filter(AutomationRule.tenant_id == admin.tenant_id).order_by(AutomationRule.created_at.desc()).all()
    return [{"id": r.id, "name": r.name, "description": r.description, "is_active": r.is_active,
             "trigger": r.trigger, "conditions": json.loads(r.conditions) if r.conditions else [],
             "actions": json.loads(r.actions) if r.actions else [],
             "run_count": r.run_count or 0, "last_run_at": r.last_run_at, "created_at": r.created_at} for r in rules]

@app.post("/admin/automation-rules")
def create_automation_rule(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("workflow_automation", tenant, "Workflow Automation is available on the Growth plan and above. Please upgrade.")
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Rule name is required")
    trigger = data.get("trigger", "")
    if trigger not in ["on_create", "on_update", "on_status_change", "time_based"]:
        raise HTTPException(status_code=400, detail="Invalid trigger")
    actions = data.get("actions", [])
    if not actions:
        raise HTTPException(status_code=400, detail="At least one action is required")
    rule = AutomationRule(
        tenant_id=admin.tenant_id, name=name,
        description=data.get("description", ""),
        trigger=trigger, is_active=data.get("is_active", True),
        conditions=json.dumps(data.get("conditions", [])),
        actions=json.dumps(actions)
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "name": rule.name, "trigger": rule.trigger, "is_active": rule.is_active,
            "conditions": json.loads(rule.conditions) if rule.conditions else [],
            "actions": json.loads(rule.actions), "run_count": 0, "created_at": rule.created_at}

@app.patch("/admin/automation-rules/{rule_id}")
def update_automation_rule(rule_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = db.query(AutomationRule).filter(AutomationRule.id == rule_id, AutomationRule.tenant_id == admin.tenant_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    for field in ["name", "description", "trigger", "is_active"]:
        if field in data:
            setattr(rule, field, data[field])
    if "conditions" in data:
        rule.conditions = json.dumps(data["conditions"])
    if "actions" in data:
        rule.actions = json.dumps(data["actions"])
    db.commit()
    return {"id": rule.id, "name": rule.name, "is_active": rule.is_active, "trigger": rule.trigger}

@app.delete("/admin/automation-rules/{rule_id}")
def delete_automation_rule(rule_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = db.query(AutomationRule).filter(AutomationRule.id == rule_id, AutomationRule.tenant_id == admin.tenant_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    db.delete(rule)
    db.commit()
    return {"ok": True}

@app.post("/admin/automation-rules/{rule_id}/test")
def test_automation_rule(rule_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Test a rule against a specific ticket to see if it would fire."""
    rule = db.query(AutomationRule).filter(AutomationRule.id == rule_id, AutomationRule.tenant_id == admin.tenant_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    ticket_id = data.get("ticket_id")
    if not ticket_id:
        raise HTTPException(status_code=400, detail="ticket_id required")
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id, Ticket.tenant_id == admin.tenant_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    conditions = json.loads(rule.conditions) if rule.conditions else []
    results = []
    all_pass = True
    for c in conditions:
        passed = _evaluate_condition(ticket, c)
        results.append({"condition": c, "passed": passed})
        if not passed:
            all_pass = False
    return {"would_fire": all_pass, "condition_results": results,
            "actions": json.loads(rule.actions) if rule.actions else []}

# =============================================================================
# AGENT GROUPS
# =============================================================================

@app.get("/groups/")
def list_groups(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    groups = db.query(Group).filter(Group.tenant_id == current_user.tenant_id).all()
    result = []
    for g in groups:
        members = db.query(User).join(GroupMember, GroupMember.user_id == User.id)\
                    .filter(GroupMember.group_id == g.id).all()
        result.append({
            "id": g.id, "name": g.name, "description": g.description,
            "member_count": len(members),
            "members": [{"id": u.id, "full_name": u.full_name, "email": u.email} for u in members]
        })
    return result

@app.post("/groups/")
def create_group(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Group name is required")
    group = Group(tenant_id=admin.tenant_id, name=name, description=data.get("description", ""))
    db.add(group)
    db.commit()
    db.refresh(group)
    # Add initial members if provided
    for uid in data.get("member_ids", []):
        user = db.query(User).filter(User.id == uid, User.tenant_id == admin.tenant_id).first()
        if user:
            db.add(GroupMember(group_id=group.id, user_id=uid))
    db.commit()
    return {"id": group.id, "name": group.name, "description": group.description}

@app.patch("/groups/{group_id}")
def update_group(group_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    group = db.query(Group).filter(Group.id == group_id, Group.tenant_id == admin.tenant_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if "name" in data: group.name = data["name"]
    if "description" in data: group.description = data["description"]
    if "member_ids" in data:
        db.query(GroupMember).filter(GroupMember.group_id == group_id).delete()
        for uid in data["member_ids"]:
            user = db.query(User).filter(User.id == uid, User.tenant_id == admin.tenant_id).first()
            if user:
                db.add(GroupMember(group_id=group_id, user_id=uid))
    db.commit()
    db.refresh(group)
    return {"id": group.id, "name": group.name, "description": group.description}

@app.delete("/groups/{group_id}")
def delete_group(group_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    group = db.query(Group).filter(Group.id == group_id, Group.tenant_id == admin.tenant_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    # Unassign tickets from this group
    db.query(Ticket).filter(Ticket.group_id == group_id).update({"group_id": None})
    db.delete(group)
    db.commit()
    return {"ok": True}

# =============================================================================
# CUSTOM TICKET FIELDS
# =============================================================================

@app.get("/admin/custom-fields")
def list_custom_fields(applies_to: str | None = Query(None),
                       db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """List custom field definitions. Optionally filter by applies_to scope.
    e.g. ?applies_to=asset returns fields scoped to assets only."""
    query = db.query(CustomField).filter(CustomField.tenant_id == current_user.tenant_id)
    if applies_to:
        # Return fields that match explicitly OR fields that apply to 'all'
        # Exception: when filtering for ticket types (incident/service_request/change),
        # also include 'all' fields. For asset/kb_article, return only exact matches.
        if applies_to in ('asset', 'kb_article'):
            query = query.filter(CustomField.applies_to == applies_to)
        else:
            query = query.filter(
                (CustomField.applies_to == applies_to) | (CustomField.applies_to == 'all')
            )
    fields = query.order_by(CustomField.sort_order).all()
    return [{"id": f.id, "name": f.name, "field_key": f.field_key, "field_type": f.field_type,
             "options": json.loads(f.options) if f.options else [],
             "is_required": f.is_required, "applies_to": f.applies_to, "sort_order": f.sort_order} for f in fields]

@app.post("/admin/custom-fields")
def create_custom_field(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Field name is required")
    field_key = re.sub(r'[^a-z0-9_]', '_', name.lower().replace(' ', '_'))
    # ensure unique key per tenant
    existing = db.query(CustomField).filter(CustomField.tenant_id == admin.tenant_id, CustomField.field_key == field_key).first()
    if existing:
        field_key = f"{field_key}_{int(datetime.utcnow().timestamp())}"
    field = CustomField(
        tenant_id=admin.tenant_id, name=name, field_key=field_key,
        field_type=data.get("field_type", "text"),
        options=json.dumps(data.get("options", [])) if data.get("options") else None,
        is_required=data.get("is_required", False),
        applies_to=data.get("applies_to", "all"),
        sort_order=data.get("sort_order", 0)
    )
    db.add(field)
    db.commit()
    db.refresh(field)
    return {"id": field.id, "name": field.name, "field_key": field.field_key,
            "field_type": field.field_type, "options": json.loads(field.options) if field.options else [],
            "is_required": field.is_required, "applies_to": field.applies_to}

@app.put("/admin/custom-fields/{field_id}")
def update_custom_field(field_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    field = db.query(CustomField).filter(CustomField.id == field_id, CustomField.tenant_id == admin.tenant_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    for k in ["name", "field_type", "is_required", "applies_to", "sort_order"]:
        if k in data:
            setattr(field, k, data[k])
    if "options" in data:
        field.options = json.dumps(data["options"]) if data["options"] else None
    db.commit()
    return {"ok": True}

@app.delete("/admin/custom-fields/{field_id}")
def delete_custom_field(field_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    field = db.query(CustomField).filter(CustomField.id == field_id, CustomField.tenant_id == admin.tenant_id).first()
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    db.delete(field)
    db.commit()
    return {"ok": True}

# =============================================================================
# MACROS
# =============================================================================

@app.get("/macros/")
def list_macros(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(Macro).filter(Macro.tenant_id == current_user.tenant_id)
    query = query.filter((Macro.is_shared == True) | (Macro.created_by_id == current_user.id))
    macros = query.order_by(Macro.name).all()
    return [{"id": m.id, "name": m.name, "description": m.description,
             "actions": json.loads(m.actions) if m.actions else [],
             "is_shared": m.is_shared, "run_count": m.run_count or 0,
             "created_by": m.created_by.full_name if m.created_by else "Unknown"} for m in macros]

@app.post("/macros/")
def create_macro(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if not has_permission(current_user, Permission.MANAGE_SETTINGS):
        raise HTTPException(status_code=403, detail="Agents and admins only")
    macro = Macro(
        tenant_id=current_user.tenant_id, name=data.get("name", "New Macro"),
        description=data.get("description", ""),
        actions=json.dumps(data.get("actions", [])),
        is_shared=data.get("is_shared", True),
        created_by_id=current_user.id
    )
    db.add(macro)
    db.commit()
    db.refresh(macro)
    return {"id": macro.id, "name": macro.name}

@app.put("/macros/{macro_id}")
def update_macro(macro_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    macro = db.query(Macro).filter(Macro.id == macro_id, Macro.tenant_id == current_user.tenant_id).first()
    if not macro:
        raise HTTPException(status_code=404, detail="Macro not found")
    for k in ["name", "description", "is_shared"]:
        if k in data:
            setattr(macro, k, data[k])
    if "actions" in data:
        macro.actions = json.dumps(data["actions"])
    db.commit()
    return {"ok": True}

@app.delete("/macros/{macro_id}")
def delete_macro(macro_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    macro = db.query(Macro).filter(Macro.id == macro_id, Macro.tenant_id == current_user.tenant_id).first()
    if not macro:
        raise HTTPException(status_code=404, detail="Macro not found")
    db.delete(macro)
    db.commit()
    return {"ok": True}

@app.post("/macros/{macro_id}/apply/{ticket_id}")
def apply_macro(macro_id: int, ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Apply a macro to a ticket — executes all actions in sequence."""
    macro = db.query(Macro).filter(Macro.id == macro_id, Macro.tenant_id == current_user.tenant_id).first()
    if not macro:
        raise HTTPException(status_code=404, detail="Macro not found")
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    actions = json.loads(macro.actions) if macro.actions else []
    applied = []
    for action in actions:
        act_type = action.get("type")
        val = action.get("value")
        try:
            if act_type == "set_status" and val:
                ticket.status = str(val).lower()
                applied.append(f"Status → {val}")
            elif act_type == "set_priority" and val:
                ticket.priority = str(val).lower()
                applied.append(f"Priority → {val}")
            elif act_type == "assign_to" and val:
                agent = db.query(User).filter(User.id == int(val), User.tenant_id == current_user.tenant_id).first()
                if agent:
                    ticket.assigned_to_id = agent.id
                    applied.append(f"Assigned → {agent.full_name}")
            elif act_type == "add_tag" and val:
                tags = json.loads(ticket.tags) if ticket.tags else []
                if val not in tags:
                    tags.append(val)
                    ticket.tags = json.dumps(tags)
                applied.append(f"Tag → {val}")
            elif act_type == "add_comment" and val:
                db.add(Comment(ticket_id=ticket_id, author_id=current_user.id, body=val, is_internal=action.get("is_internal", False)))
                applied.append("Comment added")
            elif act_type == "set_category" and val:
                ticket.category = val
                applied.append(f"Category → {val}")
        except Exception:
            pass
    ticket.updated_at = datetime.utcnow()
    macro.run_count = (macro.run_count or 0) + 1
    db.commit()
    return {"ok": True, "applied": applied}

# =============================================================================
# SAVED TICKET VIEWS
# =============================================================================

@app.get("/ticket-views/")
def list_ticket_views(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    views = db.query(TicketView).filter(
        TicketView.tenant_id == current_user.tenant_id,
        (TicketView.is_shared == True) | (TicketView.created_by_id == current_user.id)
    ).order_by(TicketView.sort_order, TicketView.name).all()
    return [{"id": v.id, "name": v.name, "filters": json.loads(v.filters) if v.filters else {},
             "is_shared": v.is_shared, "is_mine": v.created_by_id == current_user.id,
             "created_by": v.created_by.full_name if v.created_by else ""} for v in views]

@app.post("/ticket-views/")
def create_ticket_view(data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    view = TicketView(
        tenant_id=current_user.tenant_id, created_by_id=current_user.id,
        name=data.get("name", "My View"),
        filters=json.dumps(data.get("filters", {})),
        is_shared=data.get("is_shared", False)
    )
    db.add(view)
    db.commit()
    db.refresh(view)
    return {"id": view.id, "name": view.name}

@app.put("/ticket-views/{view_id}")
def update_ticket_view(view_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    view = db.query(TicketView).filter(TicketView.id == view_id, TicketView.tenant_id == current_user.tenant_id,
                                       TicketView.created_by_id == current_user.id).first()
    if not view:
        raise HTTPException(status_code=404, detail="View not found or not yours")
    for k in ["name", "is_shared"]:
        if k in data:
            setattr(view, k, data[k])
    if "filters" in data:
        view.filters = json.dumps(data["filters"])
    db.commit()
    return {"ok": True}

@app.delete("/ticket-views/{view_id}")
def delete_ticket_view(view_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    view = db.query(TicketView).filter(TicketView.id == view_id, TicketView.tenant_id == current_user.tenant_id,
                                       TicketView.created_by_id == current_user.id).first()
    if not view:
        raise HTTPException(status_code=404, detail="View not found or not yours")
    db.delete(view)
    db.commit()
    return {"ok": True}

# =============================================================================
# TICKET TASKS
# =============================================================================

@app.get("/tickets/{ticket_id}/tasks")
def list_ticket_tasks(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    tasks = db.query(TicketTask).filter(TicketTask.ticket_id == ticket_id).order_by(TicketTask.created_at).all()
    return [{"id": t.id, "title": t.title, "is_done": t.is_done,
             "assigned_to_id": t.assigned_to_id,
             "assigned_to_name": t.assigned_to.full_name if t.assigned_to else None,
             "due_date": t.due_date, "created_at": t.created_at} for t in tasks]

@app.post("/tickets/{ticket_id}/tasks")
def create_ticket_task(ticket_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    task = TicketTask(
        ticket_id=ticket_id, title=data.get("title", "New Task"),
        assigned_to_id=data.get("assigned_to_id"),
        due_date=datetime.fromisoformat(data["due_date"]) if data.get("due_date") else None
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return {"id": task.id, "title": task.title, "is_done": task.is_done}

@app.patch("/tickets/{ticket_id}/tasks/{task_id}")
def update_ticket_task(ticket_id: int, task_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    task = db.query(TicketTask).filter(TicketTask.id == task_id, TicketTask.ticket_id == ticket_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    for k in ["title", "is_done", "assigned_to_id"]:
        if k in data:
            setattr(task, k, data[k])
    if "due_date" in data:
        task.due_date = datetime.fromisoformat(data["due_date"]) if data["due_date"] else None
    db.commit()
    return {"ok": True}

@app.delete("/tickets/{ticket_id}/tasks/{task_id}")
def delete_ticket_task(ticket_id: int, task_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    task = db.query(TicketTask).filter(TicketTask.id == task_id, TicketTask.ticket_id == ticket_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task)
    db.commit()
    return {"ok": True}

# =============================================================================
# TICKET TEMPLATES
# =============================================================================

@app.get("/ticket-templates/")
def list_ticket_templates(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    templates = db.query(TicketTemplate).filter(TicketTemplate.tenant_id == current_user.tenant_id).order_by(TicketTemplate.name).all()
    return [{"id": t.id, "name": t.name, "ticket_type": t.ticket_type, "title": t.title,
             "description": t.description, "category": t.category, "priority": t.priority,
             "tags": json.loads(t.tags) if t.tags else []} for t in templates]

@app.post("/ticket-templates/")
def create_ticket_template(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tmpl = TicketTemplate(
        tenant_id=admin.tenant_id, name=data.get("name", "New Template"),
        ticket_type=data.get("ticket_type", "incident"),
        title=data.get("title", ""), description=data.get("description", ""),
        category=data.get("category", ""), priority=data.get("priority", "medium"),
        tags=json.dumps(data.get("tags", []))
    )
    db.add(tmpl)
    db.commit()
    db.refresh(tmpl)
    return {"id": tmpl.id, "name": tmpl.name}

@app.put("/ticket-templates/{tmpl_id}")
def update_ticket_template(tmpl_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tmpl = db.query(TicketTemplate).filter(TicketTemplate.id == tmpl_id, TicketTemplate.tenant_id == admin.tenant_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    for k in ["name", "ticket_type", "title", "description", "category", "priority"]:
        if k in data:
            setattr(tmpl, k, data[k])
    if "tags" in data:
        tmpl.tags = json.dumps(data["tags"])
    db.commit()
    return {"ok": True}

@app.delete("/ticket-templates/{tmpl_id}")
def delete_ticket_template(tmpl_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tmpl = db.query(TicketTemplate).filter(TicketTemplate.id == tmpl_id, TicketTemplate.tenant_id == admin.tenant_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(tmpl)
    db.commit()
    return {"ok": True}

# =============================================================================
# PROBLEM MANAGEMENT
# =============================================================================

@app.get("/tickets/{ticket_id}/problem-links")
def get_problem_links(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ticket = _ticket_tenant_filter(db.query(Ticket), ticket_id, current_user).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    # This ticket as problem — show linked incidents
    as_problem = db.query(ProblemLink).filter(ProblemLink.problem_ticket_id == ticket_id).all()
    # This ticket as incident — show its problem
    as_incident = db.query(ProblemLink).filter(ProblemLink.incident_ticket_id == ticket_id).all()
    def fmt(t_id):
        t = db.query(Ticket).filter(Ticket.id == t_id).first()
        return {"id": t.id, "title": t.title, "status": str(t.status) if t and t.status else "", "ticket_type": str(t.ticket_type) if t else ""} if t else None
    return {
        "linked_incidents": [fmt(l.incident_ticket_id) for l in as_problem if fmt(l.incident_ticket_id)],
        "linked_problem": fmt(as_incident[0].problem_ticket_id) if as_incident else None
    }

@app.post("/tickets/{ticket_id}/problem-links")
def link_problem(ticket_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Link ticket_id (incident) to a problem ticket."""
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    plan_requires("problem_management", tenant, "Problem management is available on the Pro plan and above. Please upgrade.")
    problem_id = data.get("problem_ticket_id")
    if not problem_id:
        raise HTTPException(status_code=400, detail="problem_ticket_id required")
    # Verify both tickets belong to tenant
    for tid in [ticket_id, problem_id]:
        t = db.query(Ticket).filter(Ticket.id == tid, Ticket.tenant_id == current_user.tenant_id).first()
        if not t:
            raise HTTPException(status_code=404, detail=f"Ticket {tid} not found")
    existing = db.query(ProblemLink).filter(ProblemLink.incident_ticket_id == ticket_id).first()
    if existing:
        existing.problem_ticket_id = problem_id
    else:
        db.add(ProblemLink(problem_ticket_id=problem_id, incident_ticket_id=ticket_id))
    db.commit()
    return {"ok": True}

@app.delete("/tickets/{ticket_id}/problem-links")
def unlink_problem(ticket_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    link = db.query(ProblemLink).filter(ProblemLink.incident_ticket_id == ticket_id).first()
    if link:
        db.delete(link)
        db.commit()
    return {"ok": True}

# =============================================================================
# @MENTION NOTIFICATIONS — triggered from comment creation
# =============================================================================

def process_mentions(body: str, ticket_id: int, tenant_id: int, actor: User, db: Session):
    """Parse @Name mentions in comment body and create notifications."""
    mentions = re.findall(r'@([A-Za-z][A-Za-z0-9 ]{1,30}?)(?=\s|$|[,.])', body)
    for mention in mentions:
        mention = mention.strip()
        # Find user by first name or full name match
        users = db.query(User).filter(
            User.tenant_id == tenant_id,
            User.full_name.ilike(f"%{_sql_safe_search(mention)}%")
        ).all()
        for u in users:
            if u.id != actor.id:
                create_notification(
                    user_id=u.id, tenant_id=tenant_id,
                    type="mention",
                    title=f"You were mentioned in ticket #{ticket_id}",
                    body=f"{actor.full_name} mentioned you: {body[:100]}",
                    link=f"/tickets/{ticket_id}",
                    db=db
                )

# =============================================================================
# CANNED RESPONSES
# =============================================================================

def _cr_to_out(r, db):
    author = db.query(User).filter(User.id == r.author_id).first()
    return {
        "id": r.id, "title": r.title, "content": r.content,
        "category": r.category, "author_id": r.author_id,
        "author_name": author.full_name if author else "Unknown",
        "visibility": r.visibility or "all",
        "group_id": r.group_id,
        "use_count": r.use_count or 0,
        "sort_order": r.sort_order or 0,
        "created_at": r.created_at, "updated_at": r.updated_at,
    }

@app.get("/canned-responses/")
def list_canned_responses(
    category: str | None = Query(None),
    search: str | None = Query(None),
    skip: int = Query(0, ge=0), limit: int = Query(100, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(CannedResponse).filter(
        CannedResponse.tenant_id == current_user.tenant_id
    )
    # Visibility filter — personal only shows own, group shows group
    query = query.filter(
        (CannedResponse.visibility == "all") |
        ((CannedResponse.visibility == "personal") & (CannedResponse.author_id == current_user.id)) |
        (CannedResponse.visibility == "group")  # group filtering handled below
    )
    if category:
        query = query.filter(CannedResponse.category == category)
    if search:
        query = query.filter(
            CannedResponse.title.ilike(f"%{_sql_safe_search(search)}%") |
            CannedResponse.content.ilike(f"%{_sql_safe_search(search)}%")
        )
    total = query.count()
    responses = query.order_by(CannedResponse.sort_order, CannedResponse.title).offset(skip).limit(limit).all()
    return {"items": [_cr_to_out(r, db) for r in responses], "total": total}

@app.get("/canned-responses/categories")
def list_canned_categories(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Return all distinct categories (folders) used in canned responses."""
    rows = db.query(CannedResponse.category).filter(
        CannedResponse.tenant_id == current_user.tenant_id,
        CannedResponse.category != None,
        CannedResponse.category != ""
    ).distinct().all()
    return sorted([r[0] for r in rows if r[0]])

@app.post("/canned-responses/", response_model=CannedResponseOut)
def create_canned_response(
    response: CannedResponseCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not has_permission(current_user, Permission.MANAGE_CANNED):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_response = CannedResponse(
        tenant_id=current_user.tenant_id,
        title=response.title,
        content=response.content,
        category=response.category,
        author_id=current_user.id,
        visibility=getattr(response, "visibility", "all") or "all",
        group_id=getattr(response, "group_id", None),
    )
    db.add(db_response)
    db.commit()
    db.refresh(db_response)
    return _cr_to_out(db_response, db)

@app.put("/canned-responses/{response_id}", response_model=CannedResponseOut)
def update_canned_response(
    response_id: int,
    response_update: CannedResponseUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not has_permission(current_user, Permission.MANAGE_CANNED):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_response = db.query(CannedResponse).filter(
        CannedResponse.id == response_id,
        CannedResponse.tenant_id == current_user.tenant_id
    ).first()
    if not db_response:
        raise HTTPException(status_code=404, detail="Canned response not found")
    update_data = response_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_response, key, value)
    db.commit()
    db.refresh(db_response)
    return _cr_to_out(db_response, db)

@app.post("/canned-responses/{response_id}/use")
def record_canned_use(response_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Increment use_count when agent inserts a canned response."""
    r = db.query(CannedResponse).filter(
        CannedResponse.id == response_id,
        CannedResponse.tenant_id == current_user.tenant_id
    ).first()
    if r:
        r.use_count = (r.use_count or 0) + 1
        db.commit()
    return {"ok": True}

@app.delete("/canned-responses/{response_id}")
def delete_canned_response(
    response_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not has_permission(current_user, Permission.MANAGE_CANNED):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    db_response = db.query(CannedResponse).filter(
        CannedResponse.id == response_id,
        CannedResponse.tenant_id == current_user.tenant_id
    ).first()
    if not db_response:
        raise HTTPException(status_code=404, detail="Canned response not found")
    db.delete(db_response)
    db.commit()
    return {"detail": "Canned response deleted"}

# =============================================================================
# SETTINGS
# =============================================================================

@app.put("/users/me", response_model=UserOut)
def update_profile(
    update: UserProfileUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    update_data = update.model_dump(exclude_unset=True)

    # Email changes are handled separately via request-email-change flow
    # Never update email directly here — require confirmation to the new address
    if "email" in update_data:
        new_email = update_data.pop("email", "").strip().lower()
        if new_email and new_email != current_user.email:
            # Just inform the caller — they should use /users/me/request-email-change
            pass  # email field is ignored in direct profile save

    for field, value in update_data.items():
        setattr(current_user, field, value)
    db.commit()
    db.refresh(current_user)
    return {
        "id": current_user.id,
        "email": current_user.email,
        "pending_email": current_user.pending_email,
        "full_name": current_user.full_name,
        "role": (current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role)),
        "is_active": current_user.is_active,
        "language": current_user.language or "en",
        "theme": current_user.theme or "light",
        "profile_photo": current_user.profile_photo,
        "job_title": current_user.job_title,
        "department": current_user.department,
        "phone": current_user.phone,
        "timezone": current_user.timezone or "UTC",
        "availability": current_user.availability or "online",
        "notification_prefs": json.loads(current_user.notification_prefs) if current_user.notification_prefs else {},
        "created_at": current_user.created_at,
    }

@app.post("/users/me/request-email-change")
def request_email_change(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Step 1: user requests an email address change.
    Sends a confirmation link to the NEW email — existing email stays active until confirmed.
    """
    new_email = data.get("email", "").strip().lower()
    if not new_email:
        raise HTTPException(status_code=400, detail="Email address is required.")
    if new_email == current_user.email:
        raise HTTPException(status_code=400, detail="This is already your current email address.")
    # Check new email isn't taken by another account
    existing = db.query(User).filter(User.email == new_email, User.id != current_user.id).first()
    if existing:
        raise HTTPException(status_code=400, detail="That email address is already in use by another account.")

    # Generate a confirmation token
    token = uuid.uuid4().hex
    current_user.pending_email = new_email
    current_user.email_change_token = token
    current_user.email_change_expires_at = datetime.utcnow() + timedelta(hours=24)
    db.commit()

    confirm_url = f"{FRONTEND_URL}/confirm-email-change?token={token}"
    send_email_background(
        to=new_email,
        subject="Confirm your new email address for DodoDesk",
        body=(
            f"Hi {current_user.full_name},\n\n"
            f"You requested to change your DodoDesk login email to this address.\n\n"
            f"Click the link below to confirm. Your current email ({current_user.email}) "
            f"will remain active until you confirm.\n\n"
            f"{confirm_url}\n\n"
            f"This link expires in 24 hours. If you did not request this change, "
            f"you can safely ignore this email — your account is not affected."
        ),
        cta_url=confirm_url,
        cta_label="Confirm New Email Address",
    )
    return {"ok": True, "message": f"A confirmation link has been sent to {new_email}. "
                                    f"Your current email remains active until you confirm."}

@app.post("/users/me/cancel-email-change")
def cancel_email_change(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Cancel a pending email change request."""
    current_user.pending_email = None
    current_user.email_change_token = None
    current_user.email_change_expires_at = None
    db.commit()
    return {"ok": True, "message": "Email change cancelled."}

@app.get("/auth/confirm-email-change")
def confirm_email_change(token: str, db: Session = Depends(get_db)):
    """Step 2: user clicks the link in their new email inbox.
    Updates users.email to the new address and clears pending state.
    """
    user = db.query(User).filter(User.email_change_token == token).first()
    if not user:
        raise HTTPException(status_code=400, detail="Invalid or expired confirmation link.")
    if not user.email_change_expires_at or datetime.utcnow() > user.email_change_expires_at:
        user.pending_email = None
        user.email_change_token = None
        user.email_change_expires_at = None
        db.commit()
        raise HTTPException(status_code=400, detail="This confirmation link has expired. Please request a new email change from Settings.")

    new_email = user.pending_email
    # Double-check the new email isn't taken (race condition guard)
    taken = db.query(User).filter(User.email == new_email, User.id != user.id).first()
    if taken:
        user.pending_email = None
        user.email_change_token = None
        user.email_change_expires_at = None
        db.commit()
        raise HTTPException(status_code=400, detail="That email address has been taken by another account. Please choose a different email.")

    old_email = user.email
    user.email = new_email
    user.pending_email = None
    user.email_change_token = None
    user.email_change_expires_at = None
    db.commit()

    log_system_event(db, user, "user.email_changed",
                     target_type="user", target_id=user.id, target_label=new_email,
                     old_value=old_email, new_value=new_email)
    db.commit()

    # Redirect to login — their session token still has the old email so they must log in again
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=f"{FRONTEND_URL}/login?email_changed=1")

@app.patch("/users/me/availability")
def update_availability(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Update agent availability status — online | busy | away | offline."""
    status = data.get("availability", "online")
    if status not in ["online", "busy", "away", "offline"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    current_user.availability = status
    db.commit()
    return {"ok": True, "availability": status}

@app.get("/users/availability")
def list_team_availability(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Lightweight endpoint — returns availability status for all active agents/admins in the tenant.
    Used for the team availability panel and refreshed periodically."""
    if str(current_user.role) not in ["agent", "admin", "super_admin", "platform_admin"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    users = db.query(User).filter(
        User.tenant_id == current_user.tenant_id,
        User.is_active == True,
        User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
    ).all()
    order = {"online": 0, "busy": 1, "away": 2, "offline": 3}
    items = sorted(
        [{"id": u.id, "full_name": u.full_name, "profile_photo": u.profile_photo,
          "availability": u.availability or "online"} for u in users],
        key=lambda u: order.get(u["availability"], 4)
    )
    return items

@app.get("/users/me/notification-prefs")
def get_notification_prefs(current_user: User = Depends(get_current_user)):
    default_prefs = {
        "ticket_assigned": True,
        "ticket_commented": True,
        "ticket_status_changed": True,
        "ticket_sla_breach": True,
        "ticket_mentioned": True,
        "change_approved": True,
        "change_rejected": True,
        "email_ticket_assigned": True,
        "email_ticket_commented": True,
        "email_sla_breach": True,
    }
    if current_user.notification_prefs:
        try:
            stored = json.loads(current_user.notification_prefs)
            return {**default_prefs, **stored}
        except Exception:
            pass
    return default_prefs

@app.put("/users/me/notification-prefs")
def update_notification_prefs(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Re-query user within this db session to avoid detached instance error
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.notification_prefs = json.dumps(data)
    db.commit()
    return {"ok": True, "saved": True}


def _user_wants_notif(db, user_id: int, event_key: str) -> bool:
    """Check if user has enabled a notification event. Defaults to True if not set."""
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user or not user.notification_prefs:
            return True
        prefs = json.loads(user.notification_prefs)
        return prefs.get(event_key, True)
    except Exception:
        return True

@app.post("/admin/email-config/test")
def test_email_config(data: dict, admin: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    """Send a test email using the current SMTP configuration."""
    to_email = data.get("to_email", admin.email)
    cfg = get_email_config(db, admin.tenant_id)
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText("<p>This is a test email from DodoDesk. Your email configuration is working correctly.</p>", "html")
        msg["Subject"] = "DodoDesk — Test Email"
        msg["From"] = cfg.get("smtp_from") or cfg.get("smtp_user") or "noreply@dodoDesk.com"
        msg["To"] = to_email
        host = cfg.get("smtp_host", "")
        port = int(cfg.get("smtp_port", 587))
        user = cfg.get("smtp_user", "")
        password = cfg.get("smtp_pass", "")
        if not host:
            raise ValueError("SMTP host not configured")
        with smtplib.SMTP(host, port, timeout=10) as server:
            server.starttls()
            if user and password:
                server.login(user, password)
            server.send_message(msg)
        return {"ok": True, "message": f"Test email sent to {to_email}"}
    except Exception as e:
        return {"ok": False, "message": str(e)}

@app.get("/admin/integrations-status")
def get_integrations_status(admin: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    """Return status of all configured integrations for this tenant."""
    cfg = db.query(EmailConfig).filter(EmailConfig.tenant_id == admin.tenant_id).first()
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    return {
        "slack": {"configured": bool(cfg and cfg.slack_webhook_url), "url": cfg.slack_webhook_url if cfg else ""},
        "teams": {"configured": bool(cfg and cfg.teams_webhook_url), "url": cfg.teams_webhook_url if cfg else ""},
        "smtp": {"configured": bool(cfg and cfg.smtp_host), "host": cfg.smtp_host if cfg else ""},
        "sso": {"configured": bool(tenant and tenant.sso_enabled), "provider": tenant.sso_provider if tenant else ""},
    }

@app.put("/users/me/password")
def change_password(
    pwd: PasswordUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not verify_password(pwd.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    validate_password_strength(pwd.new_password)
    current_user.hashed_password = get_password_hash(pwd.new_password)
    db.commit()
    return {"detail": "Password updated successfully"}

# =============================================================================
# MFA (TOTP) — enrollment, verification, disable
# =============================================================================

@app.get("/users/me/mfa/status")
def mfa_status(current_user: User = Depends(get_current_user)):
    return {
        "mfa_enabled": bool(current_user.mfa_enabled),
        "backup_codes_remaining": len(json.loads(current_user.mfa_backup_codes or "[]")),
    }

@app.post("/users/me/mfa/setup")
def mfa_setup(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Step 1: generate a new secret and return QR provisioning URI + base64 QR image. Not yet enabled until confirmed."""
    if current_user.mfa_enabled:
        raise HTTPException(status_code=400, detail="MFA is already enabled. Disable it first to re-enroll.")
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if tenant and not get_plan_limits(tenant.plan)["mfa"]:
        raise HTTPException(status_code=403, detail="Two-factor authentication is available on the Pro plan and above. Please upgrade your plan.")
    secret = generate_totp_secret()
    current_user.mfa_secret = secret  # stored but mfa_enabled stays False until confirmed
    db.commit()
    uri = totp_provisioning_uri(secret, current_user.email, issuer="DodoDesk")
    # Generate QR code as base64 data URL — no external API calls needed
    qr_data_url = None
    try:
        import qrcode, base64, io
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,  # Medium — more reliable scanning
            box_size=10,
            border=4,
        )
        qr.add_data(uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        qr_data_url = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
        print(f"✅ MFA QR generated for {current_user.email}, URI length: {len(uri)}")
    except Exception as e:
        print(f"❌ QR generation failed: {type(e).__name__}: {e}")
        # qr_data_url stays None — frontend shows manual key entry fallback
    return {"secret": secret, "provisioning_uri": uri, "qr_data_url": qr_data_url}

@app.post("/users/me/mfa/confirm")
def mfa_confirm(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Step 2: user enters the 6-digit code from their app to confirm and enable MFA."""
    code = data.get("code", "")
    if not current_user.mfa_secret:
        raise HTTPException(status_code=400, detail="No MFA setup in progress. Call /mfa/setup first.")
    if not verify_totp(current_user.mfa_secret, code):
        raise HTTPException(status_code=400, detail="Invalid code. Please try again.")
    backup_codes = generate_backup_codes()
    current_user.mfa_enabled = True
    current_user.mfa_backup_codes = json.dumps(backup_codes)
    log_system_event(db, current_user, "user.mfa_enabled",
                     target_type="user", target_id=current_user.id, target_label=current_user.email)
    db.commit()
    return {"ok": True, "backup_codes": backup_codes}

@app.post("/users/me/mfa/disable")
def mfa_disable(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Disable MFA — requires current password for security."""
    password = data.get("password", "")
    if not verify_password(password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Incorrect password.")
    current_user.mfa_enabled = False
    current_user.mfa_secret = None
    current_user.mfa_backup_codes = None
    log_system_event(db, current_user, "user.mfa_disabled",
                     target_type="user", target_id=current_user.id, target_label=current_user.email)
    db.commit()
    return {"ok": True}

@app.post("/users/me/photo")
def upload_profile_photo(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {".png", ".jpg", ".jpeg"}:
        raise HTTPException(status_code=400, detail="Only PNG, JPG, or JPEG images are allowed")

    file_bytes = file.file.read()

    if CLOUDINARY_CLOUD_NAME:
        public_id = f"user_{current_user.id}_avatar{ext}"
        photo_url = upload_to_cloudinary(file_bytes, public_id,
            folder=_cloudinary_folder(current_user.tenant_id, "avatars"),
            filename=file.filename)
        current_user.profile_photo = photo_url
    else:
        # Fallback to local storage
        unique_name = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(AVATAR_DIR, unique_name)
        with open(file_path, "wb") as f:
            f.write(file_bytes)
        if current_user.profile_photo and not current_user.profile_photo.startswith('http'):
            old_path = os.path.join(AVATAR_DIR, current_user.profile_photo)
            if os.path.exists(old_path):
                os.remove(old_path)
        current_user.profile_photo = unique_name

    db.commit()
    return {"detail": "Photo updated"}

@app.get("/users/me/photo")
def get_profile_photo(current_user: User = Depends(get_current_user)):
    """Returns a 1-hour signed URL for the user's profile photo.
    The URL is time-limited — Cloudinary enforces expiry server-side.
    """
    if not current_user.profile_photo:
        raise HTTPException(status_code=404, detail="No photo")
    photo = current_user.profile_photo
    # If it's a Cloudinary public_id (stored after auth migration), sign it
    if photo and not photo.startswith("/") and not photo.startswith("http"):
        ext = os.path.splitext(photo)[1].lower()
        rtype = "image" if ext in {".png",".jpg",".jpeg",".gif",".webp",".svg"} else "raw"
        signed = get_signed_url(photo, resource_type=rtype)
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=signed)
    # Legacy local file
    if photo.startswith("http"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=photo)
    file_path = os.path.join(AVATAR_DIR, photo)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Photo not found")
    return FileResponse(file_path, media_type="image/jpeg")

@app.get("/users/{user_id}/photo")
def get_user_photo(user_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Returns a signed URL redirect for any user's profile photo within the same tenant."""
    user = db.query(User).filter(
        User.id == user_id,
        User.tenant_id == current_user.tenant_id
    ).first()
    if not user or not user.profile_photo:
        raise HTTPException(status_code=404, detail="No photo")
    photo = user.profile_photo
    if photo.startswith("http"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=photo)
    ext = os.path.splitext(photo)[1].lower()
    rtype = "image" if ext in {".png",".jpg",".jpeg",".gif",".webp",".svg"} else "raw"
    signed = get_signed_url(photo, resource_type=rtype)
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=signed)

@app.get("/users/me/photo-url")
def get_profile_photo_url(current_user: User = Depends(get_current_user)):
    """Returns a signed URL for the profile photo — for direct use in <img src>."""
    if not current_user.profile_photo:
        return {"url": None}
    photo = current_user.profile_photo
    if photo.startswith("http"):
        return {"url": photo}
    signed = get_signed_url(photo, resource_type="image")
    return {"url": signed, "expires_in": 3600}

# =============================================================================
# GDPR — Right to Erasure & Data Portability (Articles 17 & 20)
# =============================================================================

@app.post("/users/me/request-deletion")
def request_account_deletion(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """GDPR Art. 17 — Right to Erasure.

    IMPORTANT B2B GDPR DISTINCTION:
    - If the user is an admin/owner of their own tenant → DodoBay handles directly.
    - If the user is an employee/agent of a company tenant → the company (Controller)
      must handle the request. DodoBay notifies the tenant admin and forwards the request.
      DodoBay is a processor in this case and cannot act without Controller instruction.
    """
    reason = data.get("reason", "").strip()
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    tenant_name = tenant.name if tenant else f"Tenant {current_user.tenant_id}"

    # Find the tenant's admin/owner (the Data Controller)
    tenant_admin = db.query(User).filter(
        User.tenant_id == current_user.tenant_id,
        User.role == 'admin',
        User.is_active == True,
        User.id != current_user.id,
    ).first()

    is_own_account = current_user.role in ("admin", "super_admin", "platform_admin")

    if is_own_account:
        # User is the account owner — DodoBay handles directly
        send_email_background(
            to="privacy@dodobay.com",
            subject=f"[GDPR] Account deletion request — {current_user.email}",
            body=(
                f"Account owner has requested deletion.\n\n"
                f"User: {current_user.full_name} ({current_user.email})\n"
                f"Role: {(current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role))}\n"
                f"Tenant: {tenant_name} (ID: {current_user.tenant_id})\n"
                f"Reason: {reason or 'Not provided'}\n"
                f"Time: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n\n"
                f"Action required: Delete tenant data within 30 days per GDPR Art. 17."
            ),
        )
        user_message = (
            "Your deletion request has been received. As the account owner, "
            "we will delete your account and all associated data within 30 days. "
            "You will receive a confirmation email."
        )
    else:
        # User is an employee — notify their employer (the Controller)
        # DodoBay forwards the request but cannot act without Controller instruction
        if tenant_admin:
            send_email_background(
                to=tenant_admin.email,
                subject=f"[GDPR] Employee data deletion request — {current_user.full_name}",
                body=(
                    f"One of your team members has submitted a GDPR erasure request.\n\n"
                    f"Employee: {current_user.full_name} ({current_user.email})\n"
                    f"Reason: {reason or 'Not provided'}\n"
                    f"Time: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n\n"
                    f"As the data controller for your organisation's DodoDesk account, "
                    f"you are responsible for responding to this request within 30 days "
                    f"under GDPR Article 17.\n\n"
                    f"To delete this user's account: log in to DodoDesk → Users → "
                    f"find {current_user.full_name} → Delete.\n\n"
                    f"If you have questions, contact us at privacy@dodobay.com."
                ),
            )
        # Also notify DodoBay for our records
        send_email_background(
            to="privacy@dodobay.com",
            subject=f"[GDPR] Employee erasure request forwarded — {current_user.email}",
            body=(
                f"An employee erasure request has been forwarded to their employer.\n\n"
                f"Employee: {current_user.full_name} ({current_user.email})\n"
                f"Employer/Tenant: {tenant_name} (ID: {current_user.tenant_id})\n"
                f"Tenant admin notified: {tenant_admin.email if tenant_admin else 'None found'}\n"
                f"Reason: {reason or 'Not provided'}\n"
                f"Time: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n\n"
                f"Note: DodoBay is the processor in this case. The employer is the Controller "
                f"and must instruct DodoBay if deletion is to proceed."
            ),
        )
        user_message = (
            "Your request has been received and forwarded to your organisation's administrator. "
            "Under GDPR, your employer (as the data controller) is responsible for responding "
            "to your request within 30 days. If you do not receive a response, you may contact "
            "your national data protection authority."
        )

    # Send confirmation to user
    send_email_background(
        to=current_user.email,
        subject="Your data deletion request has been received — DodoDesk",
        body=(
            f"Hi {current_user.full_name},\n\n"
            f"We have received your request to delete your personal data from DodoDesk.\n\n"
            f"{user_message}\n\n"
            f"For further assistance, contact us at privacy@dodobay.com.\n\n"
            f"— The DodoDesk Team"
        ),
    )

    log_system_event(db, current_user, "user.deletion_requested",
                     target_type="user", target_id=current_user.id,
                     target_label=current_user.email, new_value=reason)
    db.commit()
    return {"ok": True, "message": user_message}

@app.get("/superadmin/users/{user_id}/files")
def get_user_files(user_id: int, db: Session = Depends(get_db),
                   admin: User = Depends(get_current_admin_user)):
    """List all files stored in Cloudinary for a specific user.
    Returns: their avatar, and all attachments on tickets they raised.
    Super admin only.
    """
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    files = []

    # 1. Profile photo
    if user.profile_photo:
        signed = get_signed_url(user.profile_photo, resource_type="image")
        files.append({
            "type": "Profile photo",
            "public_id": user.profile_photo,
            "signed_url": signed,
            "expires_in": "1 hour",
        })

    # 2. Ticket attachments — find tickets the user raised
    tickets = db.query(Ticket).filter(
        Ticket.tenant_id == user.tenant_id,
        Ticket.requester_id == user.id,
    ).all()

    ticket_ids = [t.id for t in tickets]
    if ticket_ids:
        attachments = []
        try:
            attachments = db.query(Attachment).filter(
                Attachment.ticket_id.in_(ticket_ids)
            ).all()
        except Exception:
            pass

        ticket_map = {t.id: t.title for t in tickets}
        for att in attachments:
            public_id = getattr(att, "url", "") or ""
            if public_id and not public_id.startswith("http"):
                ext = os.path.splitext(getattr(att, "filename", "") or "")[1].lower()
                rtype = "image" if ext in {".png",".jpg",".jpeg",".gif",".webp"} else "raw"
                signed = get_signed_url(public_id, resource_type=rtype)
            else:
                signed = public_id
            files.append({
                "type": "Ticket attachment",
                "ticket_id": att.ticket_id,
                "ticket_title": ticket_map.get(att.ticket_id, ""),
                "filename": getattr(att, "filename", ""),
                "public_id": public_id,
                "signed_url": signed,
                "expires_in": "1 hour",
                "size_kb": round(att.size / 1024, 1) if getattr(att, "size", None) else None,
            })

    return {
        "user_id": user.id,
        "user_name": user.full_name,
        "user_email": user.email,
        "tenant_id": user.tenant_id,
        "cloudinary_avatar_path": f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{user.tenant_id}/avatars/",
        "cloudinary_tickets_path": f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{user.tenant_id}/tickets/",
        "total_files": len(files),
        "files": files,
    }

@app.get("/users/me/export")
def export_my_data(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """GDPR Art. 20 — Right to Data Portability.
    Returns all personal data held about the current user in JSON format.
    """
    import io as _io
    # Gather all data about this user
    tickets_raised = db.query(Ticket).filter(
        Ticket.tenant_id == current_user.tenant_id,
        Ticket.requester_id == current_user.id
    ).all()
    comments = db.query(Comment).filter(Comment.author_id == current_user.id).all()
    audit = db.query(SystemAuditLog).filter(
        SystemAuditLog.actor_id == current_user.id
    ).limit(500).all()

    export = {
        "export_date": datetime.utcnow().isoformat(),
        "gdpr_article": "Article 20 — Right to Data Portability",
        "personal_data": {
            "id": current_user.id,
            "full_name": current_user.full_name,
            "email": current_user.email,
            "job_title": current_user.job_title,
            "department": current_user.department,
            "phone": current_user.phone,
            "country": getattr(current_user, "country", None),
            "language": current_user.language,
            "timezone": current_user.timezone,
            "role": (current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role)) if hasattr(current_user.role, "value") else current_user.role,
            "created_at": str(current_user.created_at),
            "email_verified": getattr(current_user, "email_verified", None),
            "mfa_enabled": getattr(current_user, "mfa_enabled", False),
        },
        "tickets_raised": [
            {"id": t.id, "title": t.title, "status": str(t.status) if hasattr(t.status, "value") else t.status,
             "priority": str(t.priority) if hasattr(t.priority, "value") else t.priority,
             "created_at": str(t.created_at)}
            for t in tickets_raised
        ],
        "comments": [
            {"id": c.id, "ticket_id": c.ticket_id, "body": c.body, "created_at": str(c.created_at)}
            for c in comments
        ],
        "audit_log": [
            {"action": a.action, "target_type": a.target_type, "created_at": str(a.created_at)}
            for a in audit
        ],
    }

    from fastapi.responses import JSONResponse
    return JSONResponse(
        content=export,
        headers={"Content-Disposition": f'attachment; filename="dododesk_my_data_{current_user.id}.json"'}
    )



@app.get("/admin/branding/logo-url")
def get_logo_signed_url(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Returns a 1-hour signed URL for the tenant's logo."""
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if not tenant or not tenant.logo_url:
        return {"url": None}
    logo = tenant.logo_url
    if logo.startswith("http"):
        return {"url": logo}   # legacy public URL
    signed = get_signed_url(logo, resource_type="image")
    return {"url": signed, "expires_in": 3600}

# =============================================================================
# BILLING ENDPOINTS
# =============================================================================

@app.get("/billing/config")
def billing_config(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Returns billing configuration and current plan/trial status for this tenant."""
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    limits = get_plan_limits(tenant.plan if tenant else "free")
    trial  = get_trial_status(tenant) if tenant else {"on_trial": False, "trial_days_remaining": None, "trial_expired": False}
    staff_count = db.query(User).filter(
        User.tenant_id == admin.tenant_id,
        User.role.in_(['admin', 'agent', 'super_admin']),
        User.is_active == True,
    ).count()
    max_users = limits.get("max_agents")
    return {
        "plan": tenant.plan if tenant else "free",
        "plan_label": limits.get("label", "Free"),
        "billing_status": getattr(tenant, "billing_status", None) if tenant else None,
        "plan_renews_at": str(tenant.plan_renews_at)[:10] if tenant and getattr(tenant, "plan_renews_at", None) else None,
        "plan_limits": limits,
        "staff_count": staff_count,
        "max_users": max_users,
        "seats_over_limit": max(staff_count - max_users, 0) if max_users is not None else 0,
        **trial,
    }

@app.post("/billing/checkout")
def billing_create_checkout(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Create a Dodo Payments hosted checkout session using the official Python SDK."""
    try:
        plan     = data.get("plan", "essentials")
        interval = data.get("interval", "month")
        print(f"📦 Checkout: plan={plan} interval={interval} admin={admin.email} tenant={admin.tenant_id}")

        if not DODO_API_KEY:
            raise HTTPException(status_code=500, detail="DODO_PAYMENTS_API_KEY is not configured on Render. Please add it.")

        tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant not found")

        plan_products = DODO_PRODUCTS.get(plan)
        if not plan_products:
            raise HTTPException(status_code=400, detail=f"Unknown plan: {plan}. Valid plans: {list(DODO_PRODUCTS.keys())}")

        product_id = plan_products.get(interval)
        if not product_id:
            raise HTTPException(status_code=400, detail=f"No product configured for {plan}/{interval}")

        # Count current agents to set initial seat quantity
        current_agents = db.query(User).filter(
            User.tenant_id == tenant.id,
            User.is_active == True,
            User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
        ).count()
        initial_seats = max(1, current_agents)

        # Get addon ID for per-seat billing
        addon_id = DODO_ADDONS.get(plan, {}).get(interval)

        print(f"📦 Checkout: product={product_id} addon={addon_id} seats={initial_seats} "
              f"tenant={tenant.id} plan={plan}/{interval} environment={DODO_ENVIRONMENT}")
        print(f"📊 Seat breakdown: {initial_seats} active agent(s)/admin(s) in tenant {tenant.id}")

        # Use the official Dodo Payments Python SDK
        from dodopayments import DodoPayments
        client = DodoPayments(
            bearer_token=DODO_API_KEY,
            environment=DODO_ENVIRONMENT,
        )

        # Per-seat billing: quantity on base product = number of agents
        # No addon needed at checkout — addons are used for mid-cycle seat changes
        product_cart_item = {"product_id": product_id, "quantity": initial_seats}

        session = client.checkout_sessions.create(
            product_cart=[product_cart_item],
            customer={"email": admin.email, "name": admin.full_name},
            return_url=f"{FRONTEND_URL}/settings?billing=success&plan={plan}",
            metadata={"tenant_id": str(tenant.id), "plan": plan, "interval": interval},
        )

        checkout_url = getattr(session, "checkout_url", None) or getattr(session, "url", None)
        print(f"✅ Dodo checkout session created: {checkout_url}")

        if not checkout_url:
            raise HTTPException(status_code=502, detail=f"Dodo Payments did not return a checkout URL. Session: {session}")

        return {"checkout_url": checkout_url}

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Checkout failed: {type(e).__name__}: {str(e)}")

@app.post("/billing/portal")
def billing_customer_portal(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Return the Dodo Payments customer portal URL for this tenant."""
    business_id = os.getenv("DODO_BUSINESS_ID", "")
    if not business_id:
        raise HTTPException(
            status_code=500,
            detail="DODO_BUSINESS_ID is not configured on Render. Please add it."
        )
    # Correct URL format per Dodo docs:
    # Test: https://test.customer.dodopayments.com/login/{business_id}
    # Live: https://customer.dodopayments.com/login/{business_id}
    if DODO_ENVIRONMENT == "test_mode":
        portal_url = f"https://test.customer.dodopayments.com/login/{business_id}"
    else:
        portal_url = f"https://customer.dodopayments.com/login/{business_id}"

    print(f"✅ Portal URL: {portal_url}")
    return {"url": portal_url}

@app.post("/billing/webhook")
async def billing_webhook(request: Request, db: Session = Depends(get_db)):
    """Receives subscription lifecycle events from Dodo Payments."""
    raw_body  = await request.body()
    signature = request.headers.get("webhook-signature", "")
    timestamp = request.headers.get("webhook-timestamp", "")

    print(f"📦 Dodo webhook received: sig={'yes' if signature else 'no'} ts={timestamp}")
    print(f"📦 Raw body preview: {raw_body.decode()[:300]}")

    # Verify signature only if secret is configured
    if DODO_WEBHOOK_SECRET and signature:
        import hmac as _hmac, base64 as _b64, hashlib as _hs
        signed_payload = f"{timestamp}.{raw_body.decode()}"
        expected = _b64.b64encode(
            _hmac.new(DODO_WEBHOOK_SECRET.encode(), signed_payload.encode(), _hs.sha256).digest()
        ).decode()
        provided = signature.split(",")[1] if "," in signature else signature
        if not _hmac.compare_digest(expected, provided):
            print(f"❌ Webhook signature mismatch. Expected: {expected[:20]}... Got: {provided[:20]}...")
            raise HTTPException(status_code=401, detail="Invalid webhook signature")

    try:
        event = json.loads(raw_body.decode())
    except Exception as e:
        print(f"❌ Webhook JSON parse error: {e}")
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    event_type = event.get("type", "")
    data       = event.get("data", {})
    print(f"📦 Dodo event type: '{event_type}'")
    print(f"📦 Dodo event data keys: {list(data.keys()) if isinstance(data, dict) else type(data)}")

    def upgrade_tenant(tenant_id_str, subscription_id, customer_id, status, plan, next_billing=None):
        """Helper to find tenant and update plan."""
        tenant = None
        if tenant_id_str:
            try:
                tenant = db.query(Tenant).filter(Tenant.id == int(tenant_id_str)).first()
            except (ValueError, TypeError):
                pass
        if not tenant and customer_id:
            tenant = db.query(Tenant).filter(Tenant.dodo_customer_id == customer_id).first()
        if not tenant:
            print(f"⚠️ Webhook: no tenant found for tenant_id={tenant_id_str} customer_id={customer_id}")
            return
        old_plan = tenant.plan
        if customer_id:
            tenant.dodo_customer_id = customer_id
        if subscription_id:
            tenant.dodo_subscription_id = subscription_id
        tenant.billing_status = status
        if status in ("active", "trialing", "succeeded"):
            valid = ("essentials", "business", "pro", "enterprise")
            tenant.plan = plan if plan in valid else "essentials"
        elif status in ("cancelled", "failed", "on_hold", "past_due"):
            tenant.plan = "free"
        if next_billing:
            try:
                tenant.plan_renews_at = datetime.fromisoformat(next_billing.replace("Z", "+00:00"))
            except Exception:
                pass
        db.commit()
        print(f"✅ Tenant {tenant.id} ({tenant.name}): plan {old_plan} → {tenant.plan}, status={status}")

    # Handle subscription events
    if event_type in ("subscription.active", "subscription.activated",
                      "subscription.renewed", "subscription.updated",
                      "subscription.created"):
        subscription_id = data.get("subscription_id") or data.get("id")
        customer        = data.get("customer") or {}
        customer_id     = customer.get("customer_id") or data.get("customer_id")
        status          = data.get("status", "active")
        metadata        = data.get("metadata") or {}
        tenant_id_str   = metadata.get("tenant_id")
        plan            = metadata.get("plan", "essentials")
        next_billing    = data.get("next_billing_date") or data.get("current_period_end")
        upgrade_tenant(tenant_id_str, subscription_id, customer_id, status, plan, next_billing)

    # Handle payment.succeeded (fires when checkout completes for subscriptions too)
    elif event_type == "payment.succeeded":
        customer        = data.get("customer") or {}
        customer_id     = customer.get("customer_id") or data.get("customer_id")
        metadata        = data.get("metadata") or {}
        tenant_id_str   = metadata.get("tenant_id")
        plan            = metadata.get("plan", "essentials")
        subscription_id = data.get("subscription_id") or data.get("payment_id")
        upgrade_tenant(tenant_id_str, subscription_id, customer_id, "active", plan)

    elif event_type in ("subscription.cancelled", "subscription.on_hold"):
        subscription_id = data.get("subscription_id") or data.get("id")
        tenant = db.query(Tenant).filter(Tenant.dodo_subscription_id == subscription_id).first()
        if tenant:
            tenant.billing_status = "cancelled"
            tenant.plan = "free"
            db.commit()
            print(f"✅ Tenant {tenant.id} downgraded to free: {event_type}")

    elif event_type == "subscription.plan_changed":
        # Fires after a seat count change — update local record
        subscription_id = data.get("subscription_id") or data.get("id")
        quantity        = data.get("quantity", 1)
        tenant = db.query(Tenant).filter(Tenant.dodo_subscription_id == subscription_id).first()
        if tenant:
            print(f"✅ Seat count confirmed by Dodo: tenant {tenant.id} → {quantity} seats")
            db.commit()

    elif event_type == "payment.failed":
        # Seat update payment failed — log and potentially notify admin
        subscription_id = data.get("subscription_id") or data.get("payment", {}).get("subscription_id")
        customer        = data.get("customer") or {}
        customer_id     = customer.get("customer_id") or data.get("customer_id")
        tenant = None
        if subscription_id:
            tenant = db.query(Tenant).filter(Tenant.dodo_subscription_id == subscription_id).first()
        if not tenant and customer_id:
            tenant = db.query(Tenant).filter(Tenant.dodo_customer_id == customer_id).first()
        if tenant:
            print(f"⚠️ Payment failed for tenant {tenant.id} ({tenant.name}) — subscription may go on hold")
            # Notify tenant admin by email
            try:
                admin = db.query(User).filter(
                    User.tenant_id == tenant.id,
                    User.role.in_(['admin', 'super_admin', 'platform_admin']),
                    User.is_active == True
                ).first()
                if admin:
                    send_email(
                        admin.email,
                        "⚠️ DodoDesk — Payment failed",
                        f"Hi {admin.full_name},\n\n"
                        f"A payment for your DodoDesk subscription failed. "
                        f"Please update your payment method to avoid service interruption.\n\n"
                        f"Update payment: https://customer.dodopayments.com/login/{os.getenv('DODO_BUSINESS_ID', '')}\n\n"
                        f"Thank you.",
                        db=db
                    )
            except Exception as e:
                print(f"⚠️ Failed to send payment failure email: {e}")

    else:
        print(f"📦 Unhandled Dodo event type: {event_type} — ignoring")

    return {"ok": True}

# =============================================================================
# SERVICE CATALOG ENDPOINTS
# =============================================================================

@app.get("/catalog/")
def list_catalog_items(
    search: str | None = Query(None),
    category: str | None = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.tenant_id == current_user.tenant_id,
        ServiceCatalogItem.is_active == True
    )
    # Visibility filter
    if str(current_user.role) == "employee":
        query = query.filter(ServiceCatalogItem.visibility.in_(["all", "employees_only"]))
    elif str(current_user.role) in ["agent", "admin", "super_admin", "platform_admin"]:
        query = query.filter(ServiceCatalogItem.visibility.in_(["all", "agents_only"]))
    if search:
        query = query.filter(
            ServiceCatalogItem.name.ilike(f"%{_sql_safe_search(search)}%") |
            ServiceCatalogItem.description.ilike(f"%{_sql_safe_search(search)}%")
        )
    if category:
        query = query.filter(ServiceCatalogItem.category == category)
    items = query.order_by(ServiceCatalogItem.sort_order, ServiceCatalogItem.name).all()
    return [_catalog_to_out(i) for i in items]

def _catalog_to_out(item):
    return {
        "id": item.id, "tenant_id": item.tenant_id, "name": item.name, "description": item.description,
        "category": item.category, "estimated_cost": item.estimated_cost,
        "delivery_time_days": item.delivery_time_days, "approval_required": item.approval_required,
        "ticket_title": item.ticket_title or item.name,
        "ticket_description": item.ticket_description or item.description or "",
        "ticket_type": item.ticket_type or "service_request",
        "priority": item.priority or "medium",
        "is_onboarding": item.is_onboarding or False,
        "onboarding_tasks": json.loads(item.onboarding_tasks) if item.onboarding_tasks else [],
        "is_active": item.is_active, "is_featured": item.is_featured or False,
        "sort_order": item.sort_order or 0,
        "icon": item.icon or "📦",
        "request_form_fields": json.loads(item.request_form_fields) if item.request_form_fields else [],
        "visibility": item.visibility or "all",
        "sla_hours": item.sla_hours,
        "request_count": item.request_count or 0,
        "fulfillment_checklist": json.loads(item.fulfillment_checklist) if item.fulfillment_checklist else [],
        "approval_workflow_id": item.approval_workflow_id,
        "created_at": item.created_at,
    }

@app.get("/catalog/{item_id}")
def get_catalog_item(item_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.id == item_id,
        ServiceCatalogItem.tenant_id == current_user.tenant_id,
        ServiceCatalogItem.is_active == True
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Catalog item not found")
    return _catalog_to_out(item)

@app.post("/catalog/")
def create_catalog_item(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_CATALOG):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    plan_requires("service_catalog", tenant, "Service Catalog is available on the Starter plan and above. Please upgrade.")
    if not data.get("category", "").strip():
        raise HTTPException(status_code=422, detail="Category is required")
    db_item = ServiceCatalogItem(
        tenant_id=current_user.tenant_id,
        name=data.get("name", ""),
        description=data.get("description", ""),
        category=data.get("category", ""),
        estimated_cost=data.get("estimated_cost"),
        delivery_time_days=data.get("delivery_time_days"),
        approval_required=data.get("approval_required", True),
        ticket_title=data.get("ticket_title", ""),
        ticket_description=data.get("ticket_description", ""),
        ticket_type=data.get("ticket_type", "service_request"),
        priority=data.get("priority", "medium"),
        is_onboarding=data.get("is_onboarding", False),
        onboarding_tasks=json.dumps(data.get("onboarding_tasks", [])),
        is_featured=data.get("is_featured", False),
        sort_order=data.get("sort_order", 0),
        icon=data.get("icon", "📦"),
        request_form_fields=json.dumps(data.get("request_form_fields", [])) if data.get("request_form_fields") else None,
        visibility=data.get("visibility", "all"),
        sla_hours=data.get("sla_hours"),
        fulfillment_checklist=json.dumps(data.get("fulfillment_checklist", [])) if data.get("fulfillment_checklist") else None,
        approval_workflow_id=data.get("approval_workflow_id"),
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return _catalog_to_out(db_item)

@app.put("/catalog/{item_id}")
def update_catalog_item(item_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_CATALOG):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    item = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.id == item_id,
        ServiceCatalogItem.tenant_id == current_user.tenant_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Catalog item not found")
    if "category" in data and not (data.get("category") or "").strip():
        raise HTTPException(status_code=422, detail="Category is required")
    for field in ["name","description","category","estimated_cost","delivery_time_days",
                  "approval_required","ticket_title","ticket_description","ticket_type",
                  "priority","is_onboarding","is_featured","sort_order","icon","visibility",
                  "sla_hours","approval_workflow_id"]:
        if field in data:
            setattr(item, field, data[field])
    for json_field in ["onboarding_tasks","request_form_fields","fulfillment_checklist"]:
        if json_field in data:
            setattr(item, json_field, json.dumps(data[json_field]) if data[json_field] else None)
    db.commit()
    return _catalog_to_out(item)

@app.post("/catalog/{item_id}/onboard")
def trigger_onboarding(item_id: int, data: dict,
                       current_user: User = Depends(get_current_user),
                       db: Session = Depends(get_db)):
    """
    Trigger onboarding for a new joiner.
    data: { employee_name, employee_email, start_date, manager_name, department }
    Creates one ticket per onboarding task, all linked by a shared reference.
    """
    item = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.id == item_id,
        ServiceCatalogItem.tenant_id == current_user.tenant_id,
        ServiceCatalogItem.is_active == True,
        ServiceCatalogItem.is_onboarding == True,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Onboarding catalog item not found")

    tasks = json.loads(item.onboarding_tasks) if item.onboarding_tasks else []
    if not tasks:
        raise HTTPException(status_code=400, detail="No onboarding tasks defined")

    employee_name = data.get("employee_name", "New Employee")
    employee_email = data.get("employee_email", "")
    start_date = data.get("start_date", "")
    manager_name = data.get("manager_name", "")
    department = data.get("department", "")

    now = datetime.utcnow()
    created_tickets = []

    for task in tasks:
        # Find assignee
        assignee_id = None
        if task.get("assign_to_id"):
            assignee_id = int(task["assign_to_id"])
        elif task.get("assign_to_role"):
            assignee = db.query(User).filter(
                User.tenant_id == current_user.tenant_id,
                User.role == task["assign_to_role"],
                User.is_active == True
            ).first()
            if assignee:
                assignee_id = assignee.id

        resp, reso = compute_sla_deadlines(
            task.get("priority", "medium"), now, db, current_user.tenant_id)

        description = task.get("description", "")
        # Substitute placeholders
        for placeholder, value in [
            ("{employee_name}", employee_name),
            ("{employee_email}", employee_email),
            ("{start_date}", start_date),
            ("{manager_name}", manager_name),
            ("{department}", department),
        ]:
            description = description.replace(placeholder, value)

        title = task.get("title", "Onboarding task").replace("{employee_name}", employee_name)

        ticket = Ticket(
            tenant_id=current_user.tenant_id,
            ticket_type="service_request",
            title=title,
            description=description,
            category=task.get("category", "Onboarding"),
            priority=str(task.get("priority", "medium")).lower(),
            requester_id=current_user.id,
            assigned_to_id=assignee_id,
            status="open",
            sla_response_deadline=resp,
            sla_resolution_deadline=reso,
            created_at=now,
        )
        db.add(ticket)
        db.flush()

        log_ticket_event(db, ticket.id, current_user.tenant_id, current_user.id,
            action="created",
            note=f'Onboarding ticket for {employee_name}: {title}')

        if assignee_id:
            create_notification(db, assignee_id, current_user.tenant_id,
                "ticket_assigned",
                f"🎉 Onboarding task: {title}",
                f"New joiner: {employee_name} · Start date: {start_date}",
                f"/tickets/{ticket.id}")

        created_tickets.append({"id": ticket.id, "title": title})

    db.commit()
    return {"created": len(created_tickets), "tickets": created_tickets}

@app.delete("/catalog/{item_id}")
def delete_catalog_item(item_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_CATALOG):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    item = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.id == item_id,
        ServiceCatalogItem.tenant_id == current_user.tenant_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Catalog item not found")
    item.is_active = False
    db.commit()
    return {"ok": True}

@app.get("/catalog/categories")
def get_catalog_categories(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Get all distinct categories used in the catalog."""
    items = db.query(ServiceCatalogItem.category).filter(
        ServiceCatalogItem.tenant_id == current_user.tenant_id,
        ServiceCatalogItem.is_active == True,
        ServiceCatalogItem.category != None,
        ServiceCatalogItem.category != ""
    ).distinct().all()
    return sorted([i[0] for i in items if i[0]])

@app.patch("/catalog/{item_id}/sort")
def update_catalog_sort(item_id: int, data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Update sort order of a catalog item."""
    if not has_permission(current_user, Permission.MANAGE_CATALOG):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    item = db.query(ServiceCatalogItem).filter(
        ServiceCatalogItem.id == item_id, ServiceCatalogItem.tenant_id == current_user.tenant_id
    ).first()
    if item:
        item.sort_order = data.get("sort_order", 0)
        db.commit()
    return {"ok": True}

# =============================================================================
# CUSTOM ROLES (ADMIN)
# =============================================================================

@app.get("/admin/roles", response_model=list[CustomRoleOut])
def list_custom_roles(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_USERS):
        raise HTTPException(status_code=403)
    roles = db.query(CustomRole).filter(CustomRole.tenant_id == current_user.tenant_id).all()
    return roles

@app.post("/admin/roles", response_model=CustomRoleOut)
def create_custom_role(
    role: CustomRoleCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not has_permission(current_user, Permission.MANAGE_USERS):
        raise HTTPException(status_code=403)
    db_role = CustomRole(
        tenant_id=current_user.tenant_id,
        name=role.name,
        permissions=json.dumps([p.value for p in role.permissions])
    )
    db.add(db_role)
    db.commit()
    db.refresh(db_role)
    return db_role

# =============================================================================
# BILLING (Paddle)
# =============================================================================



@app.post("/superadmin/tenants/{tenant_id}/logo")
async def upload_tenant_logo(tenant_id: int, file: UploadFile = File(...),
                              db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    if str(admin.role) not in ("super_admin", "platform_admin") and admin.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only update your own tenant's logo")
    allowed = {"image/png", "image/jpeg", "image/jpg", "image/svg+xml", "image/webp"}
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, SVG and WebP images allowed")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo must be under 2 MB")
    if CLOUDINARY_CLOUD_NAME:
        _configure_cloudinary()
        import io as _io2
        ext2 = os.path.splitext(file.filename)[1].lower() or ".jpg"
        public_id2 = f"dodesk/tenants/{tenant_id}/logos/logo{ext2}"
        try:
            result2 = cloudinary.uploader.upload(
                _io2.BytesIO(content),
                public_id=public_id2,
                resource_type="image",
                type="upload",  # PUBLIC
                overwrite=True,
                invalidate=True,
            )
            logo_url = result2.get("secure_url") or f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/image/upload/{public_id2}"
            print(f"✅ Tenant logo uploaded (public): {logo_url}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Logo upload failed: {str(e)}")
    else:
        ext = file.filename.rsplit(".", 1)[-1].lower()
        filename = f"tenant_{tenant_id}_logo.{ext}"
        path = os.path.join(LOGO_DIR, filename)
        with open(path, "wb") as f:
            f.write(content)
        logo_url = f"/logos/{filename}"
    tenant.logo_url = logo_url
    db.commit()
    return {"logo_url": logo_url}

@app.patch("/superadmin/tenants/{tenant_id}")
def update_tenant(tenant_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    if str(admin.role) in ("super_admin", "platform_admin"):
        allowed_fields = ["name", "support_email", "company_tagline", "primary_color", "accent_color", "is_active", "plan"]
        if "plan" in data and data["plan"] not in PLAN_LIMITS:
            raise HTTPException(status_code=400, detail=f"Invalid plan. Must be one of: {', '.join(PLAN_LIMITS.keys())}")
    elif admin.tenant_id == tenant_id:
        allowed_fields = ["support_email", "company_tagline", "primary_color", "accent_color"]
    else:
        raise HTTPException(status_code=403, detail="You can only update your own tenant")

    for field in allowed_fields:
        if field in data:
            old_val = getattr(tenant, field, None)
            new_val = data[field]
            if str(old_val) != str(new_val):
                log_system_event(db, admin, f"tenant.{field}.changed",
                                 target_type="tenant", target_id=tenant.id,
                                 target_label=tenant.name,
                                 old_value=str(old_val), new_value=str(new_val))
            setattr(tenant, field, new_val)

    # If super_admin manually sets a plan, mark billing as active (clears trial state)
    if str(admin.role) in ("super_admin", "platform_admin") and "plan" in data:
        new_plan = data["plan"]
        if new_plan == "free":
            tenant.billing_status = "cancelled"
        else:
            tenant.billing_status = "active"

    db.commit()
    return {"ok": True, "plan": tenant.plan, "billing_status": tenant.billing_status}

@app.get("/admin/tenant", response_model=TenantOut)
def get_tenant(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not has_permission(current_user, Permission.MANAGE_TENANT):
        raise HTTPException(status_code=403)
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    return tenant

# =============================================================================
# PLATFORM ADMIN — MSP CLIENT TENANT ASSIGNMENT
# Only platform_admin can assign/remove client tenants to/from MSP super_admins
# =============================================================================

@app.get("/platform/msp/{super_admin_id}/clients")
def list_msp_clients(super_admin_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """List all client tenants assigned to a specific MSP super_admin. Platform admin only."""
    role = admin.role.value if hasattr(admin.role, 'value') else str(admin.role)
    if role != 'platform_admin':
        raise HTTPException(status_code=403, detail="Only platform admin can manage MSP client assignments")
    msp_user = db.query(User).filter(User.id == super_admin_id).first()
    if not msp_user:
        raise HTTPException(status_code=404, detail="MSP admin not found")
    granted = db.query(AdminTenantAccess).filter(AdminTenantAccess.admin_user_id == super_admin_id).all()
    result = []
    for g in granted:
        t = db.query(Tenant).filter(Tenant.id == g.tenant_id).first()
        if t:
            result.append({
                "id": t.id, "name": t.name, "slug": t.slug,
                "plan": t.plan, "is_active": t.is_active,
                "granted_at": str(g.granted_at)[:10]
            })
    return {"msp_user": {"id": msp_user.id, "name": msp_user.full_name, "email": msp_user.email}, "clients": result}

@app.post("/platform/msp/{super_admin_id}/clients")
def assign_client_to_msp(super_admin_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Assign a client tenant to an MSP super_admin. Platform admin only."""
    role = admin.role.value if hasattr(admin.role, 'value') else str(admin.role)
    if role != 'platform_admin':
        raise HTTPException(status_code=403, detail="Only platform admin can assign client tenants to MSPs")
    msp_user = db.query(User).filter(User.id == super_admin_id).first()
    if not msp_user:
        raise HTTPException(status_code=404, detail="MSP admin not found")
    msp_role = msp_user.role.value if hasattr(msp_user.role, 'value') else str(msp_user.role)
    if msp_role != 'super_admin':
        raise HTTPException(status_code=400, detail="Target user must be a super_admin (MSP admin)")
    tenant_id = data.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=422, detail="tenant_id is required")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id, Tenant.is_active == True).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    existing = db.query(AdminTenantAccess).filter(
        AdminTenantAccess.admin_user_id == super_admin_id,
        AdminTenantAccess.tenant_id == tenant_id
    ).first()
    if existing:
        return {"ok": True, "message": f"{tenant.name} is already assigned to this MSP"}
    db.add(AdminTenantAccess(admin_user_id=super_admin_id, tenant_id=tenant_id, granted_by_id=admin.id))
    log_system_event(db, admin, "platform.msp_client_assigned",
                     target_type="tenant", target_id=tenant_id,
                     target_label=f"{tenant.name} → {msp_user.full_name}")
    db.commit()
    return {"ok": True, "message": f"{tenant.name} assigned to {msp_user.full_name}"}

@app.delete("/platform/msp/{super_admin_id}/clients/{tenant_id}")
def remove_client_from_msp(super_admin_id: int, tenant_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Remove a client tenant from an MSP super_admin's scope. Platform admin only."""
    role = admin.role.value if hasattr(admin.role, 'value') else str(admin.role)
    if role != 'platform_admin':
        raise HTTPException(status_code=403, detail="Only platform admin can remove client tenant assignments")
    access = db.query(AdminTenantAccess).filter(
        AdminTenantAccess.admin_user_id == super_admin_id,
        AdminTenantAccess.tenant_id == tenant_id
    ).first()
    if not access:
        raise HTTPException(status_code=404, detail="Assignment not found")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    msp_user = db.query(User).filter(User.id == super_admin_id).first()
    db.delete(access)
    log_system_event(db, admin, "platform.msp_client_removed",
                     target_type="tenant", target_id=tenant_id,
                     target_label=f"{tenant.name if tenant else tenant_id} → {msp_user.full_name if msp_user else super_admin_id}")
    db.commit()
    return {"ok": True, "message": "Client tenant removed from MSP scope"}


@app.get("/superadmin/tenants")
def list_all_tenants(
    search: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """List tenants visible to this admin.
    - platform_admin: ALL tenants in the system
    - super_admin (MSP): own tenant + explicitly granted client tenants only
    - regular admin: own tenant only
    """
    role = admin.role.value if hasattr(admin.role, 'value') else str(admin.role)

    if role == 'platform_admin':
        # DodoDesk owner — sees everything
        query = db.query(Tenant)
        if search:
            query = query.filter(Tenant.name.ilike(f"%{_sql_safe_search(search)}%"))
        tenants = query.order_by(Tenant.created_at.desc()).all()

    else:
        # super_admin (MSP) or regular admin — own tenant + explicitly granted only
        own_ids = {admin.tenant_id}
        granted = db.query(AdminTenantAccess).filter(
            AdminTenantAccess.admin_user_id == admin.id
        ).all()
        for g in granted:
            own_ids.add(g.tenant_id)
        query = db.query(Tenant).filter(Tenant.id.in_(own_ids))
        if search:
            query = query.filter(Tenant.name.ilike(f"%{_sql_safe_search(search)}%"))
        tenants = query.order_by(Tenant.name).all()

    def tenant_row(t):
        is_own = (t.id == admin.tenant_id)
        is_granted = not is_own
        return {
            "id": t.id,
            "name": t.name,
            "slug": t.slug,
            "plan": t.plan or "free",
            "is_active": t.is_active,
            "primary_color": t.primary_color or "#4f46e5",
            "accent_color": t.accent_color or "#818cf8",
            "logo_url": t.logo_url,
            "company_tagline": t.company_tagline,
            "support_email": t.support_email,
            "billing_status": getattr(t, "billing_status", None),
            "plan_renews_at": str(t.plan_renews_at)[:10] if getattr(t, "plan_renews_at", None) else None,
            "created_at": str(t.created_at)[:10] if t.created_at else None,
            "is_own": is_own,
            "is_granted": is_granted,
            "user_count": db.query(User).filter(
                User.tenant_id == t.id, User.is_active == True
            ).count(),
        }

    return [tenant_row(t) for t in tenants]

@app.get("/superadmin/tenants/{tenant_id}")
def get_tenant_by_id(tenant_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Get a single tenant. Super admin can fetch any; regular admin only their own."""
    query = db.query(Tenant).filter(Tenant.id == tenant_id)
    if str(admin.role) not in ("super_admin", "platform_admin"):
        query = query.filter(Tenant.id == admin.tenant_id)
    tenant = query.first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return tenant

@app.post("/superadmin/tenants")
def create_tenant_superadmin(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Create a new tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Tenant name is required")
    slug = data.get("slug") or name.lower().replace(" ", "-")
    existing = db.query(Tenant).filter(Tenant.slug == slug).first()
    if existing:
        slug = f"{slug}-{db.query(Tenant).count()}"
    tenant = Tenant(
        name=name, slug=slug,
        plan=data.get("plan", "free"),
        is_active=True,
        primary_color=data.get("primary_color", "#4f46e5"),
    )
    db.add(tenant)
    db.commit()
    db.refresh(tenant)
    return {"id": tenant.id, "name": tenant.name, "slug": tenant.slug, "plan": tenant.plan}

@app.delete("/superadmin/tenants/{tenant_id}")
def delete_tenant(tenant_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Permanently delete a tenant and ALL its data. Super admin only.
    Cleans up all related tables before deleting the tenant row.
    """
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    if tenant_id == admin.tenant_id:
        raise HTTPException(status_code=400, detail="You cannot delete your own tenant.")

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    tenant_name = tenant.name
    tenant_slug = tenant.slug

    # Log BEFORE deletion (audit logs for target tenant will be wiped)
    # Log to super admin's own tenant so the record persists
    try:
        log_system_event(db, admin, "tenant.deleted",
                         target_type="tenant", target_id=tenant_id,
                         target_label=f"{tenant_name} ({tenant_slug})",
                         old_value=f"plan={tenant.plan}, users=deleted")
        db.commit()
    except Exception as e:
        print(f"⚠️ Pre-deletion audit log failed: {e}")

    try:
        from sqlalchemy import text as _t
        with db.bind.connect() as conn:
            tid = tenant_id
            # ── Delete in FK dependency order ─────────────────────────────
            # Ticket children
            conn.execute(_t("DELETE FROM ticket_watchers WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_audit_logs WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM time_entries WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM comments WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM attachments WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_tasks WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            # ticket_views has tenant_id directly
            conn.execute(_t("DELETE FROM ticket_views WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM tickets WHERE tenant_id = :t"), {"t": tid})
            # Change request children (use change_id not change_request_id)
            conn.execute(_t("DELETE FROM change_comments WHERE change_id IN (SELECT id FROM change_requests WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM change_tasks WHERE change_id IN (SELECT id FROM change_requests WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM change_requests WHERE tenant_id = :t"), {"t": tid})
            # Asset children
            conn.execute(_t("DELETE FROM asset_history WHERE asset_id IN (SELECT id FROM assets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM assets WHERE tenant_id = :t"), {"t": tid})
            # KB children
            conn.execute(_t("DELETE FROM kb_versions WHERE article_id IN (SELECT id FROM kb_articles WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM kb_articles WHERE tenant_id = :t"), {"t": tid})
            # Service catalog (depends on approval_workflows)
            conn.execute(_t("UPDATE service_catalog_items SET approval_workflow_id = NULL WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM service_catalog_items WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM approval_workflows WHERE tenant_id = :t"), {"t": tid})
            # Tenant-level config
            conn.execute(_t("DELETE FROM automation_rules WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM escalation_rules WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM sla_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM business_hours_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM email_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM canned_responses WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM custom_fields WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_templates WHERE tenant_id = :t"), {"t": tid})
            # Approval workflow children before the workflow itself
            conn.execute(_t("DELETE FROM ticket_approvals WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM approval_steps WHERE workflow_id IN (SELECT id FROM approval_workflows WHERE tenant_id = :t)"), {"t": tid})
            # Groups — table is 'groups' not 'agent_groups'
            conn.execute(_t("DELETE FROM group_members WHERE group_id IN (SELECT id FROM groups WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM groups WHERE tenant_id = :t"), {"t": tid})
            # User-level cleanup
            conn.execute(_t("DELETE FROM notifications WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM admin_tenant_access WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM system_audit_logs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM chat_messages WHERE session_id IN (SELECT id FROM chat_sessions WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM chat_sessions WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM time_entries WHERE agent_id IN (SELECT id FROM users WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM signup_verifications WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM users WHERE tenant_id = :t"), {"t": tid})
            # Finally the tenant itself
            conn.execute(_t("DELETE FROM tenants WHERE id = :t"), {"t": tid})
            conn.commit()

        print(f"✅ Tenant deleted: {tenant_name} (id={tenant_id})")
        return {"ok": True, "message": f"Tenant \"{tenant_name}\" and all its data have been permanently deleted."}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not delete tenant: {type(e).__name__}: {str(e)[:300]}")


# =============================================================================

@app.get("/superadmin/admin-access")
def list_admin_access(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """List all admin-to-tenant access grants. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    records = db.query(AdminTenantAccess).all()
    return [{
        "id": r.id,
        "admin_user_id": r.admin_user_id,
        "admin_name": r.admin_user.full_name if r.admin_user else "",
        "admin_email": r.admin_user.email if r.admin_user else "",
        "tenant_id": r.tenant_id,
        "tenant_name": r.tenant.name if r.tenant else "",
        "granted_at": r.granted_at,
    } for r in records]

@app.post("/superadmin/admin-access")
def grant_admin_access(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Grant an admin access to an additional tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    admin_user_id = data.get("admin_user_id")
    tenant_id = data.get("tenant_id")
    if not admin_user_id or not tenant_id:
        raise HTTPException(status_code=400, detail="admin_user_id and tenant_id are required")
    # Verify target user is an admin
    target = db.query(User).filter(User.id == admin_user_id).first()
    if not target or str(target.role) not in ["admin"]:
        raise HTTPException(status_code=400, detail="Target user must be an Admin role")
    # Check tenant exists
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    # Check not already granted
    existing = db.query(AdminTenantAccess).filter(
        AdminTenantAccess.admin_user_id == admin_user_id,
        AdminTenantAccess.tenant_id == tenant_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Access already granted")
    access = AdminTenantAccess(admin_user_id=admin_user_id, tenant_id=tenant_id, granted_by_id=admin.id)
    db.add(access)
    db.commit()
    return {"ok": True, "admin_user_id": admin_user_id, "tenant_id": tenant_id}

@app.delete("/superadmin/admin-access/{access_id}")
def revoke_admin_access(access_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Revoke an admin's access to a tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    record = db.query(AdminTenantAccess).filter(AdminTenantAccess.id == access_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Access record not found")
    db.delete(record)
    db.commit()
    return {"ok": True}

@app.get("/superadmin/tenants/{tenant_id}/export")
def export_tenant_data(tenant_id: int, db: Session = Depends(get_db),
                       admin: User = Depends(get_current_admin_user)):
    """Export all data for a tenant as a multi-sheet Excel file."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        HEADER_FILL  = PatternFill("solid", fgColor="4F46E5")
        HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

        def make_sheet(title, headers, rows):
            ws = wb.create_sheet(title=title[:31])
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
            for row in rows:
                ws.append([str(v) if v is not None else "" for v in row])
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
            ws.freeze_panes = "A2"
            return ws

        # Sheet 1: Tenant Info
        make_sheet("Tenant Info", ["Field", "Value"], [
            ["ID", tenant.id], ["Name", tenant.name], ["Slug", tenant.slug],
            ["Plan", tenant.plan], ["Active", tenant.is_active],
            ["Primary Color", tenant.primary_color],
            ["Support Email", getattr(tenant, "support_email", "") or ""],
            ["Billing Status", getattr(tenant, "billing_status", "") or ""],
            ["Created At", str(tenant.created_at)[:19] if tenant.created_at else ""],
        ])

        # Sheet 2: Users
        users = db.query(User).filter(User.tenant_id == tenant_id).all()
        make_sheet("Users",
            ["ID", "Full Name", "Email", "Role", "Job Title", "Department", "Active", "MFA", "Created At"],
            [(u.id, u.full_name, u.email, str(u.role) if u.role else "",
              getattr(u, "job_title", "") or "", getattr(u, "department", "") or "",
              u.is_active, getattr(u, "mfa_enabled", False),
              str(u.created_at)[:19] if u.created_at else "") for u in users]
        )

        # Sheet 3: Tickets
        tickets = db.query(Ticket).filter(Ticket.tenant_id == tenant_id).all()
        user_map = {u.id: u.full_name for u in users}
        make_sheet("Tickets",
            ["ID", "Ref", "Type", "Title", "Status", "Priority", "Category", "Requester", "Assigned To", "Created At"],
            [(t.id,
              f"{'INC' if t.ticket_type and 'incident' in str(t.ticket_type).lower() else 'REQ'}-{t.id:04d}",
              str(str(t.ticket_type)) if t.ticket_type else "",
              t.title, str(str(t.status)) if t.status else "",
              str(str(t.priority)) if t.priority else "",
              t.category or "",
              user_map.get(t.requester_id, str(t.requester_id) if t.requester_id else ""),
              user_map.get(t.assigned_to_id, "") if t.assigned_to_id else "Unassigned",
              str(t.created_at)[:19] if t.created_at else "") for t in tickets]
        )

        # Sheet 4: Assets
        assets = db.query(Asset).filter(Asset.tenant_id == tenant_id).all()
        make_sheet("Assets",
            ["ID", "Name", "Type", "Serial Number", "Status", "Assigned To"],
            [(a.id, a.name, str(getattr(a, "asset_type", {}).value if hasattr(getattr(a, "asset_type", None), "value") else ""),
              getattr(a, "serial_number", "") or "",
              str(getattr(a, "status", {}).value if hasattr(getattr(a, "status", None), "value") else ""),
              user_map.get(getattr(a, "assigned_to_id", None), "")) for a in assets]
        )

        # Sheet 5: Knowledge Base
        articles = db.query(KBArticle).filter(KBArticle.tenant_id == tenant_id).all()
        make_sheet("Knowledge Base",
            ["ID", "Title", "Category", "Author", "Created At"],
            [(a.id, a.title, getattr(a, "category", "") or "",
              user_map.get(getattr(a, "author_id", None), ""),
              str(a.created_at)[:19] if getattr(a, "created_at", None) else "") for a in articles]
        )

        # Sheet 6: Audit Log
        logs = db.query(SystemAuditLog).filter(SystemAuditLog.tenant_id == tenant_id)                  .order_by(SystemAuditLog.id.desc()).limit(5000).all()
        make_sheet("Audit Log",
            ["ID", "Timestamp", "Action", "Actor", "Target Type"],
            [(l.id,
              str(l.created_at)[:19] if getattr(l, "created_at", None) else "",
              getattr(l, "action", "") or "",
              user_map.get(getattr(l, "actor_id", None), ""),
              getattr(l, "target_type", "") or "") for l in logs]
        )

        # Sheet 7: Attachments
        ticket_ids = [t.id for t in tickets]
        attachments = []
        if ticket_ids:
            try:
                attachments = db.query(Attachment).filter(
                    Attachment.ticket_id.in_(ticket_ids)
                ).all()
            except Exception:
                pass
        ticket_ref_map = {t.id: f"INC-{t.id:04d}" for t in tickets}
        make_sheet("Attachments",
            ["ID", "Ticket Ref", "Filename", "Size (KB)", "Uploaded At"],
            [(a.id,
              ticket_ref_map.get(a.ticket_id, str(a.ticket_id)),
              getattr(a, "filename", "") or "",
              round(a.size / 1024, 1) if getattr(a, "size", None) else 0,
              str(a.uploaded_at)[:19] if getattr(a, "uploaded_at", None) else "") for a in attachments]
        )

        # Sheet 8: Cloudinary Files
        cloudinary_files = []
        try:
            _configure_cloudinary()
            import cloudinary.api as _capi
            folder_prefix = f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{tenant_id}"
            for resource_type in ["image", "raw"]:
                try:
                    result = _capi.resources(
                        type="authenticated", prefix=folder_prefix,
                        max_results=500, resource_type=resource_type,
                    )
                    cloudinary_files += [
                        (r["public_id"], resource_type, round(r.get("bytes", 0)/1024, 1), r.get("created_at", ""))
                        for r in result.get("resources", [])
                    ]
                except Exception:
                    pass
        except Exception as e:
            print(f"⚠️ Cloudinary listing: {e}")

        make_sheet("Cloudinary Files",
            ["Path", "Type", "Size (KB)", "Uploaded At"],
            cloudinary_files if cloudinary_files else [["No files found", "", "", ""]]
        )

        # Stream Excel
        output = _io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"dodesk_export_{tenant.slug}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"✅ Export: {tenant.name} ({len(tickets)} tickets, {len(users)} users, {len(attachments)} attachments)")
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {type(e).__name__}: {str(e)[:300]}")


# =============================================================================
# CSAT ENDPOINTS
# =============================================================================

@app.get("/csat/{token}")
def get_csat_survey(token: str, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.csat_token == token).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Survey not found")
    return {"id": ticket.id, "title": ticket.title, "rating": ticket.csat_rating, "comment": ticket.csat_comment}

@app.post("/csat/{token}")
def submit_csat_survey(token: str, data: CSATSubmit, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.csat_token == token).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Survey not found")
    if ticket.csat_rating is not None:
        raise HTTPException(status_code=400, detail="Survey already submitted")
    if data.rating < 1 or data.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    ticket.csat_rating = data.rating
    ticket.csat_comment = data.comment
    db.commit()
    return {"detail": "Thank you for your feedback"}

@app.get("/reports/csat")
def csat_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    results = db.query(Ticket.csat_rating, sa_func.count(Ticket.id)).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    ).group_by(Ticket.csat_rating).all()
    distribution = {str(k): v for k, v in results}
    avg = db.query(sa_func.avg(Ticket.csat_rating)).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    ).scalar()
    count = sum(distribution.values())
    positive = sum(v for k, v in distribution.items() if int(float(k)) >= 4)
    negative = sum(v for k, v in distribution.items() if int(float(k)) <= 2)
    satisfaction_rate = round(positive / count * 100, 1) if count > 0 else 0
    return {
        "average": round(avg, 2) if avg else None,
        "avg_rating": round(avg, 2) if avg else None,
        "count": count,
        "total_responses": count,
        "distribution": distribution,
        "satisfaction_rate": satisfaction_rate,
        "negative_count": negative,
        "positive_count": positive,
    }

# =============================================================================
# AI CHATBOT — Enterprise plan only (DodoBot)
# =============================================================================

def _check_enterprise(current_user: User, db: Session):
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if not tenant or tenant.plan != "enterprise":
        raise HTTPException(
            status_code=403,
            detail="The AI assistant is available on the Enterprise plan. Contact us to upgrade."
        )

def _build_system_prompt(current_user: User, tenant: Tenant) -> str:
    return f"""You are DodoBot, an AI IT support assistant for {tenant.name} powered by DodoDesk.

You help employees and IT staff with:
- Raising and tracking support tickets
- Searching the knowledge base for solutions
- Looking up asset information
- Answering IT policy and procedure questions

Current user: {current_user.full_name} (role: {(current_(user.role.value if hasattr(user.role, "value") else str(user.role)) if hasattr(current_user.role, "value") else str(current_user.role))})
Company: {tenant.name}

Guidelines:
- Be concise, friendly and professional
- When user asks to "see", "track", "show", or "list" their tickets — use list_my_tickets, not search_tickets
- Use search_tickets only when the user provides a specific keyword to search for
- Always confirm ticket details before creating one
- Cite KB article titles when referencing knowledge base content
- Never fabricate ticket IDs or asset data — use tools only
- Format ticket IDs as INC-XXXX or REQ-XXXX
- If you cannot help, suggest the user raise a ticket
"""

CHAT_TOOLS = [
    {
        "name": "list_my_tickets",
        "description": "List the current user's tickets, optionally filtered by status. Use when the user asks to see, track, or check their tickets.",
        "input_schema": {
            "type": "object",
            "properties": {
                "status": {"type": "string", "enum": ["open", "in_progress", "resolved", "closed", "all"]},
                "limit":  {"type": "integer", "description": "Max tickets to return (default 10)"}
            }
        }
    },
    {
        "name": "search_tickets",
        "description": "Search the user's tickets by keyword. Returns up to 5 matching tickets.",
        "input_schema": {
            "type": "object",
            "properties": {"query": {"type": "string"}},
            "required": ["query"]
        }
    },
    {
        "name": "get_ticket",
        "description": "Get full details of a specific ticket by its numeric ID.",
        "input_schema": {
            "type": "object",
            "properties": {"ticket_id": {"type": "integer"}},
            "required": ["ticket_id"]
        }
    },
    {
        "name": "create_ticket",
        "description": "Create a new support ticket on behalf of the user.",
        "input_schema": {
            "type": "object",
            "properties": {
                "title":       {"type": "string"},
                "description": {"type": "string"},
                "priority":    {"type": "string", "enum": ["low", "medium", "high", "critical"]},
                "ticket_type": {"type": "string", "enum": ["incident", "service_request"]},
                "category":    {"type": "string"}
            },
            "required": ["title", "description"]
        }
    },
    {
        "name": "update_ticket",
        "description": "Update a ticket's status, priority, or add a comment. Use when the user asks to close, resolve, reopen, or update a ticket.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ticket_id": {"type": "integer"},
                "status":    {"type": "string", "enum": ["open", "in_progress", "resolved", "closed"], "description": "New status (optional)"},
                "priority":  {"type": "string", "enum": ["low", "medium", "high", "critical"], "description": "New priority (optional)"},
                "comment":   {"type": "string", "description": "Comment to add to the ticket (optional)"}
            },
            "required": ["ticket_id"]
        }
    },
    {
        "name": "search_kb",
        "description": "Search the knowledge base for articles matching a query. Always search KB before suggesting the user raise a ticket.",
        "input_schema": {
            "type": "object",
            "properties": {"query": {"type": "string"}},
            "required": ["query"]
        }
    },
    {
        "name": "list_kb_articles",
        "description": "List published KB articles, optionally filtered by category.",
        "input_schema": {
            "type": "object",
            "properties": {
                "category": {"type": "string", "description": "Filter by category (optional)"},
                "limit":    {"type": "integer", "description": "Max articles to return (default 8)"}
            }
        }
    },
    {
        "name": "get_asset",
        "description": "Look up details of an IT asset by its numeric ID.",
        "input_schema": {
            "type": "object",
            "properties": {"asset_id": {"type": "integer"}},
            "required": ["asset_id"]
        }
    },
    {
        "name": "list_my_assets",
        "description": "List assets assigned to the current user.",
        "input_schema": {
            "type": "object",
            "properties": {"limit": {"type": "integer", "description": "Max assets to return (default 10)"}},
        }
    },
    {
        "name": "check_sla",
        "description": "Check SLA status for the current user's open tickets — which are overdue, near breach, or on track.",
        "input_schema": {"type": "object", "properties": {}}
    },
]

def _execute_tool(tool_name: str, tool_input: dict, current_user: User, db: Session) -> str:
    import json as _json

    def _ticket_prefix(t):
        return "INC" if t.ticket_type and "incident" in str(t.ticket_type) else "REQ"

    if tool_name == "list_my_tickets":
        status_filter = tool_input.get("status", "all")
        limit = min(int(tool_input.get("limit", 10)), 20)
        query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)
        if str(current_user.role) == "employee":
            query = query.filter(Ticket.requester_id == current_user.id)
        if status_filter and status_filter != "all":
            try: query = query.filter(Ticket.status == str(status_filter).lower())
            except ValueError: pass
        tickets = query.order_by(Ticket.created_at.desc()).limit(limit).all()
        if not tickets:
            return f"No tickets found{' with status ' + status_filter if status_filter != 'all' else ''}."
        return "\n".join(f"{_ticket_prefix(t)}-{t.id:04d}: {t.title} [{str(t.status)}] [{str(t.priority)}]" for t in tickets)

    elif tool_name == "search_tickets":
        q = f"%{tool_input.get('query', '')}%"
        tickets = db.query(Ticket).filter(
            Ticket.tenant_id == current_user.tenant_id,
            (Ticket.title.ilike(q)) | (Ticket.description.ilike(q))
        ).order_by(Ticket.created_at.desc()).limit(5).all()
        if not tickets:
            return f"No tickets found matching '{tool_input.get('query')}'."
        return "\n".join(f"{_ticket_prefix(t)}-{t.id:04d}: {t.title} [{str(t.status)}] [{str(t.priority)}]" for t in tickets)

    elif tool_name == "get_ticket":
        tid = tool_input.get("ticket_id")
        t = db.query(Ticket).filter(Ticket.id == tid, Ticket.tenant_id == current_user.tenant_id).first()
        if not t: return f"Ticket #{tid} not found."
        assignee = db.query(User).filter(User.id == t.assigned_to_id).first() if t.assigned_to_id else None
        sla_info = ""
        if t.sla_resolution_deadline:
            diff = (t.sla_resolution_deadline - datetime.utcnow()).total_seconds()
            if diff < 0: sla_info = f"\nSLA: ⚠️ OVERDUE by {abs(int(diff//3600))}h"
            elif diff < 3600: sla_info = f"\nSLA: ⏰ {int(diff//60)}m remaining"
            else: sla_info = f"\nSLA: ✅ {int(diff//3600)}h remaining"
        return (f"Ticket {_ticket_prefix(t)}-{t.id:04d}\n"
                f"Title: {t.title}\nStatus: {str(t.status)}\nPriority: {str(t.priority)}\n"
                f"Category: {t.category or 'Uncategorised'}\n"
                f"Assigned to: {assignee.full_name if assignee else 'Unassigned'}{sla_info}\n"
                f"Description: {t.description[:300]}")

    elif tool_name == "create_ticket":
        new_t = Ticket(
            tenant_id=current_user.tenant_id, requester_id=current_user.id,
            title=tool_input.get("title", ""), description=tool_input.get("description", ""),
            priority=str(tool_input.get("priority", "medium")).lower(),
            ticket_type=str(tool_input.get("ticket_type", "service_request")).lower(),
            category=tool_input.get("category", "Other"), status="open",
        )
        db.add(new_t); db.commit(); db.refresh(new_t)
        prefix = "INC" if new_t.ticket_type == "incident" else "REQ"
        return f"✅ Ticket created: {prefix}-{new_t.id:04d} — \"{new_t.title}\"\nYou can track it on your dashboard."

    elif tool_name == "update_ticket":
        tid = tool_input.get("ticket_id")
        t = db.query(Ticket).filter(Ticket.id == tid, Ticket.tenant_id == current_user.tenant_id).first()
        if not t: return f"Ticket #{tid} not found."
        changes = []
        if "status" in tool_input and tool_input["status"]:
            try:
                t.status = str(tool_input["status"]).lower()
                changes.append(f"status → {tool_input['status']}")
            except ValueError: pass
        if "priority" in tool_input and tool_input["priority"]:
            try:
                t.priority = str(tool_input["priority"]).lower()
                changes.append(f"priority → {tool_input['priority']}")
            except ValueError: pass
        if "comment" in tool_input and tool_input["comment"]:
            comment = Comment(ticket_id=t.id, author_id=current_user.id,
                              body=tool_input["comment"], is_internal=False)
            db.add(comment)
            changes.append("comment added")
        db.commit()
        if not changes: return f"No changes made to ticket #{tid}."
        return f"✅ Ticket {_ticket_prefix(t)}-{t.id:04d} updated: {', '.join(changes)}"

    elif tool_name == "search_kb":
        q = f"%{tool_input.get('query', '')}%"
        articles = db.query(KBArticle).filter(
            KBArticle.tenant_id == current_user.tenant_id,
            (KBArticle.title.ilike(q)) | (KBArticle.content.ilike(q))
        ).limit(4).all()
        if not articles: return f"No knowledge base articles found for '{tool_input.get('query')}'."
        return "\n\n".join(f"**{a.title}**: {(a.content or '')[:250]}..." for a in articles)

    elif tool_name == "list_kb_articles":
        limit = min(int(tool_input.get("limit", 8)), 20)
        query = db.query(KBArticle).filter(
            KBArticle.tenant_id == current_user.tenant_id,
            KBArticle.status == "published"
        )
        if tool_input.get("category"):
            query = query.filter(KBArticle.category.ilike(f"%{_sql_safe_search(tool_input['category'])}%"))
        articles = query.order_by(KBArticle.view_count.desc()).limit(limit).all()
        if not articles: return "No published knowledge base articles found."
        return "\n".join(f"• {a.title} [{a.category or 'General'}]" for a in articles)

    elif tool_name == "get_asset":
        aid = tool_input.get("asset_id")
        a = db.query(Asset).filter(Asset.id == aid, Asset.tenant_id == current_user.tenant_id).first()
        if not a: return f"Asset #{aid} not found."
        expiry = f"\nExpiry: {a.expiry_date}" if a.expiry_date else ""
        warranty = f"\nWarranty: {a.warranty_expiry}" if getattr(a, 'warranty_expiry', None) else ""
        return (f"Asset: {a.name}\nType: {str(a.type)}\nStatus: {str(a.status)}\n"
                f"Serial: {a.serial_number or 'N/A'}\nAssigned to: {a.assigned_to_id or 'Unassigned'}"
                f"{expiry}{warranty}")

    elif tool_name == "list_my_assets":
        limit = min(int(tool_input.get("limit", 10)), 20)
        assets = db.query(Asset).filter(
            Asset.tenant_id == current_user.tenant_id,
            Asset.assigned_to_id == current_user.id
        ).limit(limit).all()
        if not assets: return "No assets are assigned to you."
        return "\n".join(f"• #{a.id} {a.name} [{str(a.type)}] — {str(a.status)}" for a in assets)

    elif tool_name == "check_sla":
        now = datetime.utcnow()
        open_statuses = ["open", "in_progress"]
        tickets = db.query(Ticket).filter(
            Ticket.tenant_id == current_user.tenant_id,
            Ticket.status.in_(open_statuses),
            Ticket.sla_resolution_deadline.isnot(None)
        )
        if str(current_user.role) == "employee":
            tickets = tickets.filter(Ticket.requester_id == current_user.id)
        tickets = tickets.all()
        if not tickets: return "No open tickets with SLA deadlines found."
        overdue = [t for t in tickets if t.sla_resolution_deadline < now]
        warning = [t for t in tickets if t.sla_resolution_deadline >= now and (t.sla_resolution_deadline - now).total_seconds() < 3600*2]
        ok      = [t for t in tickets if t not in overdue and t not in warning]
        lines = []
        if overdue: lines.append(f"⚠️ OVERDUE ({len(overdue)}):\n" + "\n".join(f"  {_ticket_prefix(t)}-{t.id:04d}: {t.title}" for t in overdue[:5]))
        if warning: lines.append(f"⏰ Breaching soon ({len(warning)}):\n" + "\n".join(f"  {_ticket_prefix(t)}-{t.id:04d}: {t.title}" for t in warning[:5]))
        if ok:      lines.append(f"✅ On track: {len(ok)} ticket(s)")
        return "\n\n".join(lines)

    return f"Unknown tool: {tool_name}"


def _run_agentic_loop(messages: list, system: str, db: Session, current_user: User):
    """Run the Claude agentic loop. Returns (final_reply, tool_summary)."""
    import urllib.request as _urllib, urllib.error as _urllib_error, json as _json
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="AI chatbot is not configured. Please add ANTHROPIC_API_KEY on Render.")

    if not messages:
        raise HTTPException(status_code=400, detail="No messages to send.")

    loop_messages = list(messages)
    tool_summary = []

    for _ in range(5):  # max 5 tool-call iterations
        payload = _json.dumps({
            "model": ANTHROPIC_MODEL,
            "max_tokens": 1024,
            "system": system,
            "messages": loop_messages,
            "tools": CHAT_TOOLS,
        }).encode()
        req = _urllib.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            method="POST"
        )
        try:
            with _urllib.urlopen(req) as resp:
                response = _json.loads(resp.read().decode())
        except _urllib_error.HTTPError as e:
            error_body = e.read().decode() if e.fp else str(e)
            raise HTTPException(status_code=502, detail=f"Anthropic API error {e.code}: {error_body}")

        stop_reason = response.get("stop_reason")
        content_blocks = response.get("content", [])

        if stop_reason == "tool_use":
            tool_blocks = [b for b in content_blocks if b.get("type") == "tool_use"]
            tool_results = []
            for tb in tool_blocks:
                result = _execute_tool(tb["name"], tb["input"], current_user, db)
                tool_summary.append(tb["name"])
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tb["id"],
                    "content": result
                })
            loop_messages.append({"role": "assistant", "content": content_blocks})
            loop_messages.append({"role": "user", "content": tool_results})
        else:
            text_parts = [b["text"] for b in content_blocks if b.get("type") == "text" and b.get("text")]
            return "\n".join(text_parts).strip(), tool_summary

    return "I was unable to complete that request. Please try again.", tool_summary


def _get_or_create_session(session_id, current_user: User, first_message: str, db: Session):
    if session_id:
        session = db.query(ChatSession).filter(
            ChatSession.id == session_id,
            ChatSession.user_id == current_user.id,
            ChatSession.tenant_id == current_user.tenant_id
        ).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        return session, False
    title = first_message[:60] + ("..." if len(first_message) > 60 else "")
    session = ChatSession(tenant_id=current_user.tenant_id, user_id=current_user.id, title=title)
    db.add(session)
    db.flush()
    return session, True


def _build_anthropic_history(session_id: int, db: Session) -> list:
    history = db.query(ChatMessage).filter(
        ChatMessage.session_id == session_id
    ).order_by(ChatMessage.created_at.asc()).all()
    return [{"role": m.role, "content": m.content} for m in history if m.role in ("user", "assistant")]


# ── Session management endpoints ─────────────────────────────────────────

@app.get("/api/chat/sessions")
def list_chat_sessions(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_enterprise(current_user, db)
    sessions = db.query(ChatSession).filter(
        ChatSession.tenant_id == current_user.tenant_id,
        ChatSession.user_id == current_user.id
    ).order_by(ChatSession.updated_at.desc()).limit(20).all()
    return [{"id": s.id, "title": s.title, "created_at": s.created_at, "updated_at": s.updated_at}
            for s in sessions]

@app.get("/api/chat/sessions/{session_id}")
def get_chat_session(session_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_enterprise(current_user, db)
    session = db.query(ChatSession).filter(
        ChatSession.id == session_id,
        ChatSession.user_id == current_user.id,
        ChatSession.tenant_id == current_user.tenant_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    messages = db.query(ChatMessage).filter(
        ChatMessage.session_id == session_id
    ).order_by(ChatMessage.created_at.asc()).all()
    return {
        "id": session.id, "title": session.title,
        "messages": [{"id": m.id, "role": m.role, "content": m.content, "created_at": m.created_at}
                     for m in messages]
    }

@app.delete("/api/chat/sessions/{session_id}")
def delete_chat_session(session_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _check_enterprise(current_user, db)
    session = db.query(ChatSession).filter(
        ChatSession.id == session_id,
        ChatSession.user_id == current_user.id,
        ChatSession.tenant_id == current_user.tenant_id
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    db.delete(session)
    db.commit()
    return {"ok": True}


# ── Non-streaming chat endpoint ───────────────────────────────────────────

@app.post("/api/chat")
def chat(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Non-streaming chat. Body: {message, session_id?, attachment?}
    attachment: {name, media_type, data} where data is base64-encoded
    """
    import json as _json, base64 as _b64
    _check_enterprise(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    user_message = (data.get("message") or "").strip()
    attachment   = data.get("attachment")  # {name, media_type, data (base64)}

    if not user_message and not attachment:
        raise HTTPException(status_code=400, detail="Message or attachment required.")

    # Build display message for saving (text only)
    display_message = user_message or f"[Attached file: {attachment.get('name', 'file')}]"

    session, is_new = _get_or_create_session(data.get("session_id"), current_user, display_message, db)
    existing_history = _build_anthropic_history(session.id, db)

    db.add(ChatMessage(session_id=session.id, role="user", content=display_message))
    db.flush()

    # Build Anthropic user message content — text + optional file
    user_content = []
    if attachment:
        media_type = attachment.get("media_type", "image/jpeg")
        file_data  = attachment.get("data", "")
        file_name  = attachment.get("name", "file")
        if media_type == "application/pdf":
            user_content.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": file_data},
                "title": file_name,
            })
        elif media_type.startswith("image/"):
            user_content.append({
                "type": "image",
                "source": {"type": "base64", "media_type": media_type, "data": file_data},
            })
        else:
            # For Word/other docs — tell Claude what it is
            user_content.append({
                "type": "text",
                "text": f"[The user has attached a file: {file_name} ({media_type}). Unfortunately this file type cannot be read directly — please let the user know.]"
            })

    if user_message:
        user_content.append({"type": "text", "text": user_message})

    if not user_content:
        user_content = [{"type": "text", "text": display_message}]

    history = existing_history + [{"role": "user", "content": user_content}]
    system  = _build_system_prompt(current_user, tenant)
    reply, tool_summary = _run_agentic_loop(history, system, db, current_user)

    db.add(ChatMessage(
        session_id=session.id, role="assistant", content=reply,
        tool_calls=_json.dumps(tool_summary) if tool_summary else None
    ))
    session.updated_at = datetime.utcnow()
    db.commit()

    return {"reply": reply, "session_id": session.id, "session_title": session.title, "tools_used": tool_summary}


# ── SSE Streaming chat endpoint ───────────────────────────────────────────

@app.post("/api/chat/stream")
def chat_stream(data: dict, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    SSE streaming chat.
    Body: {message, session_id?}
    Yields SSE events:
      data: {"type":"delta","text":"..."}
      data: {"type":"tool","name":"..."}
      data: {"type":"done","session_id":N,"session_title":"...","tools_used":[...]}
      data: {"type":"error","message":"..."}
    """
    import json as _json, urllib.request as _urllib

    _check_enterprise(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    user_message = (data.get("message") or "").strip()
    attachment   = data.get("attachment")
    if not user_message and not attachment:
        raise HTTPException(status_code=400, detail="Message or attachment required.")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="AI chatbot is not configured.")

    display_message = user_message or f"[Attached file: {attachment.get('name', 'file')}]"

    session, _ = _get_or_create_session(data.get("session_id"), current_user, display_message, db)
    existing_history = _build_anthropic_history(session.id, db)

    db.add(ChatMessage(session_id=session.id, role="user", content=display_message))
    db.flush()
    db.commit()

    session_id    = session.id
    session_title = session.title
    system = _build_system_prompt(current_user, tenant)

    # Build user content with optional attachment
    user_content = []
    if attachment:
        media_type = attachment.get("media_type", "image/jpeg")
        file_data  = attachment.get("data", "")
        file_name  = attachment.get("name", "file")
        if media_type == "application/pdf":
            user_content.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": file_data}, "title": file_name})
        elif media_type.startswith("image/"):
            user_content.append({"type": "image", "source": {"type": "base64", "media_type": media_type, "data": file_data}})
        else:
            user_content.append({"type": "text", "text": f"[User attached: {file_name} ({media_type}) — this file type cannot be read directly]"})
    if user_message:
        user_content.append({"type": "text", "text": user_message})
    if not user_content:
        user_content = [{"type": "text", "text": display_message}]

    initial_messages = existing_history + [{"role": "user", "content": user_content}]

    def event_stream():
        import json as _j, urllib.request as _ur
        tool_summary = []
        full_reply   = []
        loop_messages = list(initial_messages)

        for iteration in range(5):
            payload = _j.dumps({
                "model": ANTHROPIC_MODEL,
                "max_tokens": 1024,
                "system": system,
                "messages": loop_messages,
                "tools": CHAT_TOOLS,
                "stream": True,
            }).encode()

            req = _ur.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                method="POST"
            )

            # Accumulate full streamed response
            current_text   = []
            current_tools  = []
            stop_reason    = None
            response_id    = None
            response_content_for_loop = []

            try:
                with _ur.urlopen(req) as resp:
                    for raw_line in resp:
                        line = raw_line.decode("utf-8").strip()
                        if not line or not line.startswith("data:"):
                            continue
                        event_data = line[5:].strip()
                        if event_data == "[DONE]":
                            break
                        try:
                            event = _j.loads(event_data)
                        except Exception:
                            continue

                        etype = event.get("type")

                        if etype == "message_start":
                            response_id = event.get("message", {}).get("id")

                        elif etype == "content_block_start":
                            block = event.get("content_block", {})
                            if block.get("type") == "tool_use":
                                current_tools.append({
                                    "id": block.get("id"),
                                    "name": block.get("name"),
                                    "input_str": ""
                                })
                                # Notify frontend a tool is being called
                                yield f"data: {_j.dumps({'type': 'tool', 'name': block.get('name')})}\n\n"

                        elif etype == "content_block_delta":
                            delta = event.get("delta", {})
                            dtype = delta.get("type")
                            if dtype == "text_delta":
                                chunk = delta.get("text", "")
                                if chunk:
                                    current_text.append(chunk)
                                    full_reply.append(chunk)
                                    # Stream text token to frontend
                                    yield f"data: {_j.dumps({'type': 'delta', 'text': chunk})}\n\n"
                            elif dtype == "input_json_delta":
                                if current_tools:
                                    current_tools[-1]["input_str"] += delta.get("partial_json", "")

                        elif etype == "message_delta":
                            stop_reason = event.get("delta", {}).get("stop_reason")

            except Exception as e:
                import urllib.error as _ue
                if isinstance(e, _ue.HTTPError):
                    body = e.read().decode() if e.fp else str(e)
                    yield f"data: {_j.dumps({'type': 'error', 'message': f'Anthropic API error {e.code}: {body}'})}\n\n"
                else:
                    yield f"data: {_j.dumps({'type': 'error', 'message': str(e)})}\n\n"
                return

            # Build content blocks for loop continuation
            if current_text:
                response_content_for_loop.append({"type": "text", "text": "".join(current_text)})
            for t in current_tools:
                try:
                    parsed_input = _j.loads(t["input_str"]) if t["input_str"] else {}
                except Exception:
                    parsed_input = {}
                response_content_for_loop.append({
                    "type": "tool_use",
                    "id": t["id"],
                    "name": t["name"],
                    "input": parsed_input
                })

            if stop_reason == "tool_use" and current_tools:
                # Execute tools and continue loop
                tool_results = []
                for t in current_tools:
                    try:
                        parsed_input = _j.loads(t["input_str"]) if t["input_str"] else {}
                    except Exception:
                        parsed_input = {}
                    result = _execute_tool(t["name"], parsed_input, current_user, db)
                    tool_summary.append(t["name"])
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": t["id"],
                        "content": result
                    })
                loop_messages.append({"role": "assistant", "content": response_content_for_loop})
                loop_messages.append({"role": "user", "content": tool_results})
            else:
                # Done — save final reply and close
                break

        # Persist assistant message
        final_text = "".join(full_reply).strip() or "I was unable to complete that request."
        with next(get_db()) as save_db:
            save_db.add(ChatMessage(
                session_id=session_id, role="assistant", content=final_text,
                tool_calls=_j.dumps(tool_summary) if tool_summary else None
            ))
            s = save_db.query(ChatSession).filter(ChatSession.id == session_id).first()
            if s:
                s.updated_at = datetime.utcnow()
            save_db.commit()

        yield f"data: {_j.dumps({'type': 'done', 'session_id': session_id, 'session_title': session_title, 'tools_used': tool_summary})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",       # disables Nginx buffering on Render
            "Connection": "keep-alive",
        }
    )
