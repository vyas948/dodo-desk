"""
Admin — branding, SLA config, security, email, superadmin router.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, BackgroundTasks, Query, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, text
from datetime import datetime, timedelta, date
from typing import Optional
import json, uuid, os, re, io, csv, threading

# Import everything from shared utilities (avoids circular import with main.py)
from shared import (
    get_db, SessionLocal,
    get_current_user, get_current_admin_user, has_permission, Permission,
    oauth2_scheme, OAuth2PasswordRequestForm,
    User, Tenant, Ticket, Comment, Notification, KBArticle, Asset,
    ChangeRequest, SystemAuditLog, TicketAuditLog, Macro, AutomationRule,
    EscalationRule, Group, CustomField, TicketTemplate, ApprovalWorkflow,
    ServiceCatalogItem, CannedResponse, TicketView, TicketTask, Attachment,
    TicketCreate, TicketUpdate, TicketOut, CommentCreate, CommentOut,
    KBArticleCreate, KBArticleUpdate, KBArticleOut,
    AssetCreate, AssetUpdate, AssetOut, LinkAssetRequest,
    UserOut, UserCreate, UserInvite, UserUpdate, UserProfileUpdate, PasswordUpdate,
    CannedResponseCreate, CannedResponseUpdate, CannedResponseOut,
    ChangeCreate, ChangeUpdate, ChangeOut,
    ServiceCatalogItemCreate, ServiceCatalogItemOut,
    BulkTicketAction, InboundEmail, CSATSubmit, CSATStats, AttachmentOut,
    SignupRequest, TenantOut, CustomRoleCreate, CustomRoleOut,
    get_email_config, send_email, send_email_background, send_notification,
    create_notification, build_html_email, translate_email,
    log_ticket_event, _ticket_tenant_filter, compute_sla_deadlines,
    pause_sla, resume_sla, get_sla_rules, compute_sla_status,
    run_automation_rules, trigger_approval_workflow,
    get_plan_limits, get_user_language, check_tenant_limit, check_user_limit,
    plan_requires, get_trial_status, validate_password_strength,
    get_password_hash, verify_password, create_access_token,
    create_access_token_with_expiry, decode_access_token,
    check_ip_rate_limit, log_system_event,
    generate_totp_secret, totp_provisioning_uri, verify_totp, generate_backup_codes,
    upload_to_cloudinary, get_signed_url,
    get_business_hours_config, add_business_hours,
    _sql_safe_search, _cr_to_out, _safe_json,
    _asset_to_out, _ticket_to_out, _catalog_to_out, _change_to_out,
    _notify_watchers, _round_robin_assign, _user_wants_notif,
    _get_oauth_redirect_uri, _build_anthropic_history, _build_system_prompt,
    _check_enterprise, _execute_tool, _get_or_create_session,
    _run_agentic_loop, _send_scheduled_report,
    limiter, Limiter, get_remote_address,
    API_URL, FRONTEND_URL, SECRET_KEY, ALGORITHM,
    ACCESS_TOKEN_EXPIRE_MINUTES, MAX_FAILED_ATTEMPTS,
    PLAN_LIMITS, SLA_RULES,
)

router = APIRouter()


@router.get("/admin/audit-log")
def get_system_audit_log(
    search: str | None = Query(None),
    action: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """System-wide audit log for the tenant. Shows all admin actions."""
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("audit_log", tenant, "Full audit log is available on the Business plan and above. Please upgrade.")
    query = db.query(SystemAuditLog).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    )
    if action:
        query = query.filter(SystemAuditLog.action.ilike(f"{_sql_safe_search(action)}%"))
    if search:
        s = f"%{_sql_safe_search(search)}%"
        query = query.filter(
            SystemAuditLog.action.ilike(s) |
            SystemAuditLog.target_label.ilike(s) |
            SystemAuditLog.actor_email.ilike(s)
        )
    if start_date:
        try:
            query = query.filter(SystemAuditLog.created_at >= datetime.fromisoformat(start_date))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(SystemAuditLog.created_at <= datetime.fromisoformat(end_date))
        except Exception:
            pass

    total = query.count()
    logs  = query.order_by(SystemAuditLog.created_at.desc()).offset(skip).limit(limit).all()

    # Build by_category counts for the frontend
    all_logs = db.query(SystemAuditLog.action).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    ).all()
    by_category: dict = {}
    for (act,) in all_logs:
        if act:
            cat = act.split('.')[0]
            by_category[cat] = by_category.get(cat, 0) + 1

    # Resolve actor names
    actor_ids = {l.actor_id for l in logs if l.actor_id}
    actor_map  = {}
    if actor_ids:
        actors = db.query(User).filter(User.id.in_(actor_ids)).all()
        actor_map = {u.id: u.full_name for u in actors}

    items = [
        {
            "id": l.id,
            "action": l.action or "",
            "target_type": getattr(l, "target_type", "") or "",
            "target_id": getattr(l, "target_id", None),
            "target_label": getattr(l, "target_label", "") or "",
            "old_value": getattr(l, "old_value", "") or "",
            "new_value": getattr(l, "new_value", "") or "",
            "actor_id": l.actor_id,
            "actor_name": actor_map.get(l.actor_id, ""),
            "actor_email": getattr(l, "actor_email", "") or "",
            "created_at": str(l.created_at)[:19] if l.created_at else "",
        }
        for l in logs
    ]
    return {"items": items, "total": total, "by_category": by_category}


@router.get("/admin/audit-log/export/csv")
def export_audit_log_csv(
    search: str | None = Query(None),
    action: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Export the full audit log as CSV."""
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("audit_log", tenant, "Audit log export is available on the Business plan and above.")
    query = db.query(SystemAuditLog).filter(
        SystemAuditLog.tenant_id == admin.tenant_id
    )
    if action:
        query = query.filter(SystemAuditLog.action.ilike(f"{_sql_safe_search(action)}%"))
    if search:
        s = f"%{_sql_safe_search(search)}%"
        query = query.filter(
            SystemAuditLog.action.ilike(s) |
            SystemAuditLog.actor_email.ilike(s)
        )
    logs = query.order_by(SystemAuditLog.created_at.desc()).limit(10000).all()

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "Action", "Target Type", "Target", "Actor Email", "Old Value", "New Value"])
    for l in logs:
        writer.writerow([
            str(l.created_at)[:19] if l.created_at else "",
            l.action or "",
            getattr(l, "target_type", "") or "",
            getattr(l, "target_label", "") or "",
            getattr(l, "actor_email", "") or "",
            getattr(l, "old_value", "") or "",
            getattr(l, "new_value", "") or "",
        ])
    output.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=\"audit_log.csv\""}
    )



@router.get("/branding/public")
def get_public_branding(db: Session = Depends(get_db)):
    """Public endpoint — returns DodoDesk platform branding for the login/signup page.
    Configurable via environment variables so you can change the platform name and
    colours without touching code. Does NOT leak any tenant's company name or logo."""
    from fastapi.responses import JSONResponse
    data = {
        "company_name":    os.getenv("PLATFORM_NAME", "DodoDesk"),
        "company_tagline": os.getenv("PLATFORM_TAGLINE", "IT Service Management"),
        "primary_color":   os.getenv("PLATFORM_PRIMARY_COLOR", "#1e1e2f"),
        "accent_color":    os.getenv("PLATFORM_ACCENT_COLOR", "#4f46e5"),
        "logo_url":        os.getenv("PLATFORM_LOGO_URL", None),
    }
    return JSONResponse(content=data, headers={"Cache-Control": "public, max-age=3600"})


@router.get("/admin/branding")
def get_branding(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return {
        "company_name": tenant.name,
        "company_tagline": tenant.company_tagline or "",
        "primary_color": tenant.primary_color or "#4f46e5",
        "accent_color": tenant.accent_color or "#818cf8",
        "logo_url": tenant.logo_url,
        "support_email": tenant.support_email or "",
        "plan": tenant.plan or "free",
        "plan_limits": get_plan_limits(tenant.plan),
    }


@router.put("/admin/branding")
def update_branding(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    if not get_plan_limits(tenant.plan)["branding"]:
        raise HTTPException(status_code=403, detail="Custom branding is available on the Pro plan and above. Please upgrade your plan.")
    if data.get("company_name"):
        tenant.name = data["company_name"]
    if "company_tagline" in data:
        tenant.company_tagline = data["company_tagline"]
    if data.get("primary_color"):
        tenant.primary_color = data["primary_color"]
    if data.get("accent_color"):
        tenant.accent_color = data["accent_color"]
    if "support_email" in data:
        tenant.support_email = data["support_email"]
    if data.get("logo_url") and "cloudinary.com" in str(data["logo_url"]):
        tenant.logo_url = data["logo_url"]
    elif "logo_url" in data and not data["logo_url"]:
        tenant.logo_url = None  # explicitly clearing the logo
    log_system_event(db, admin, "branding.updated",
                     target_type="tenant", target_id=tenant.id, target_label=tenant.name)
    db.commit()
    return {"ok": True}


@router.post("/admin/branding/logo")
async def upload_logo(file: UploadFile = File(...), db: Session = Depends(get_db),
                      admin: User = Depends(get_current_admin_user)):
    allowed = {"image/png", "image/jpeg", "image/jpg", "image/svg+xml", "image/webp"}
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, SVG and WebP images allowed")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo must be under 2 MB")

    if CLOUDINARY_CLOUD_NAME:
        ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "")
        _configure_cloudinary()
        # Logos must be public (type="upload") so they can be displayed in browser and emails
        import io as _io
        public_id = f"dodesk/tenants/{admin.tenant_id}/logos/logo{ext}"
        try:
            result = cloudinary.uploader.upload(
                _io.BytesIO(content),
                public_id=public_id,
                resource_type="image",
                type="upload",       # PUBLIC — not authenticated
                overwrite=True,
                invalidate=True,
            )
            logo_url = result.get("secure_url") or f"https://res.cloudinary.com/{cloud_name}/image/upload/{public_id}"
            print(f"✅ Logo uploaded (public): {logo_url}")
        except Exception as e:
            print(f"❌ Logo upload failed: {e}")
            raise HTTPException(status_code=500, detail=f"Logo upload failed: {str(e)}")
    else:
        print(f"⚠️ CLOUDINARY_CLOUD_NAME not set — logo will be lost on next deploy")
        raise HTTPException(status_code=500, detail="Cloudinary not configured. Please set CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET in Render environment variables.")

    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    tenant.logo_url = logo_url
    db.commit()
    return {"logo_url": logo_url}


@router.get("/logos/{filename}")
def serve_logo(filename: str):
    path = os.path.join(LOGO_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Logo not found")
    return FileResponse(path)



@router.get("/admin/sla-config")
def get_sla_config(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    cfg = db.query(SLAConfig).filter(SLAConfig.tenant_id == admin.tenant_id).first()
    defaults = SLA_RULES
    return {
        "low_response":      cfg.low_response      if cfg else defaults["low"]["response"],
        "low_resolution":    cfg.low_resolution    if cfg else defaults["low"]["resolution"],
        "medium_response":   cfg.medium_response   if cfg else defaults["medium"]["response"],
        "medium_resolution": cfg.medium_resolution if cfg else defaults["medium"]["resolution"],
        "high_response":     cfg.high_response     if cfg else defaults["high"]["response"],
        "high_resolution":   cfg.high_resolution   if cfg else defaults["high"]["resolution"],
        "critical_response":     cfg.critical_response     if cfg else defaults["critical"]["response"],
        "critical_resolution":   cfg.critical_resolution   if cfg else defaults["critical"]["resolution"],
    }


@router.put("/admin/sla-config")
def update_sla_config(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if tenant and not get_plan_limits(tenant.plan)["sla"]:
        raise HTTPException(status_code=403, detail="SLA configuration is available on the Pro plan and above. Please upgrade your plan.")
    cfg = db.query(SLAConfig).filter(SLAConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = SLAConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.low_response      = int(data.get("low_response", 8))
    cfg.low_resolution    = int(data.get("low_resolution", 72))
    cfg.medium_response   = int(data.get("medium_response", 4))
    cfg.medium_resolution = int(data.get("medium_resolution", 48))
    cfg.high_response     = int(data.get("high_response", 2))
    cfg.high_resolution   = int(data.get("high_resolution", 24))
    cfg.critical_response     = int(data.get("critical_response", 1))
    cfg.critical_resolution   = int(data.get("critical_resolution", 8))
    log_system_event(db, admin, "sla_config.updated",
                     target_type="tenant", target_id=admin.tenant_id)
    db.commit()
    return {"ok": True}



@router.get("/admin/escalation-rules")
def list_escalation_rules(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rules = db.query(EscalationRule).filter(
        EscalationRule.tenant_id == admin.tenant_id,
        EscalationRule.is_active == True
    ).all()
    result = []
    for r in rules:
        agent = db.query(User).filter(User.id == r.escalate_to_id).first() if r.escalate_to_id else None
        result.append({
            "id": r.id, "name": r.name, "priority": r.priority,
            "idle_hours": r.idle_hours,
            "escalate_to_id": r.escalate_to_id,
            "escalate_to_name": agent.full_name if agent else None,
            "escalate_to_role": r.escalate_to_role,
            "created_at": r.created_at,
        })
    return result


@router.post("/admin/escalation-rules")
def create_escalation_rule(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = EscalationRule(
        tenant_id=admin.tenant_id,
        name=data.get("name", ""),
        priority=data.get("priority") or None,
        idle_hours=int(data.get("idle_hours", 24)),
        escalate_to_id=int(data["escalate_to_id"]) if data.get("escalate_to_id") else None,
        escalate_to_role=data.get("escalate_to_role") or None,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return {"id": rule.id, "name": rule.name}


@router.delete("/admin/escalation-rules/{rule_id}")
def delete_escalation_rule(rule_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    rule = db.query(EscalationRule).filter(
        EscalationRule.id == rule_id,
        EscalationRule.tenant_id == admin.tenant_id
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    rule.is_active = False
    db.commit()
    return {"ok": True}



@router.get("/admin/business-hours")
def get_business_hours(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    cfg = db.query(BusinessHoursConfig).filter(BusinessHoursConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        return {"enabled": False, "start_hour": 9, "end_hour": 17,
                "working_days": "0,1,2,3,4", "timezone": "UTC"}
    return {"enabled": cfg.enabled, "start_hour": cfg.start_hour,
            "end_hour": cfg.end_hour, "working_days": cfg.working_days,
            "timezone": cfg.timezone}


@router.put("/admin/business-hours")
def update_business_hours(data: dict, db: Session = Depends(get_db),
                          admin: User = Depends(get_current_admin_user)):
    cfg = db.query(BusinessHoursConfig).filter(BusinessHoursConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = BusinessHoursConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.enabled = data.get("enabled", False)
    cfg.start_hour = int(data.get("start_hour", 9))
    cfg.end_hour = int(data.get("end_hour", 17))
    cfg.working_days = data.get("working_days", "0,1,2,3,4")
    cfg.timezone = data.get("timezone", "UTC")
    db.commit()
    return {"ok": True}



@router.get("/admin/security-config")
def get_security_config(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return {
        "mfa_enabled":        bool(tenant.mfa_enabled),
        "mfa_required":       bool(tenant.mfa_required),
        "sso_enabled":        bool(tenant.sso_enabled),
        "sso_provider":       tenant.sso_provider or "saml",
        "sso_client_id":      tenant.sso_client_id or "",    # SAML Entity ID / Issuer
        "sso_client_secret":  "",                            # never return secrets
        "sso_sso_url":        getattr(tenant, "sso_sso_url", "") or "",  # IdP SSO URL
        "sso_domain":         tenant.sso_domain or "",       # allowed email domain
        "sso_tenant_id":      tenant.sso_tenant_id or "",    # Azure tenant ID (optional)
        "saml_cert":          getattr(tenant, "saml_cert", "") or "",    # IdP certificate
        "sp_metadata_url":    f"{API_URL}/auth/sso/metadata/{tenant.slug}",
        "sp_acs_url":         f"{API_URL}/auth/sso/callback/{tenant.slug}",
        "sp_entity_id":       f"{API_URL}/auth/sso/metadata/{tenant.slug}",
        "session_timeout_minutes": getattr(tenant, "session_timeout_minutes", 60) or 60,
        "max_login_attempts": getattr(tenant, "max_login_attempts", 0) or 0,
    }


@router.put("/admin/security-config")
def update_security_config(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    limits = get_plan_limits(tenant.plan)
    if data.get("mfa_enabled") and not limits["mfa"]:
        raise HTTPException(status_code=403, detail="Two-factor authentication is available on the Pro plan and above. Please upgrade your plan.")
    if data.get("sso_enabled") and not limits["sso"]:
        raise HTTPException(status_code=403, detail="Single sign-on (SSO) is available on the Pro plan and above. Please upgrade your plan.")

    # Session & login policy
    if data.get("session_timeout_minutes") is not None:
        timeout_val = int(data.get("session_timeout_minutes") or 0)
        tenant.session_timeout_minutes = timeout_val if timeout_val > 0 else 60
    if data.get("max_login_attempts") is not None:
        tenant.max_login_attempts = max(0, int(data.get("max_login_attempts") or 0))
    tenant.mfa_enabled  = bool(data.get("mfa_enabled", False))
    tenant.mfa_required = bool(data.get("mfa_required", False)) if tenant.mfa_enabled else False
    tenant.sso_enabled  = bool(data.get("sso_enabled", False))
    tenant.sso_provider = data.get("sso_provider", "saml")
    tenant.sso_client_id = data.get("sso_client_id") or None       # SAML Entity ID
    if data.get("sso_client_secret"):
        tenant.sso_client_secret = data.get("sso_client_secret")
    tenant.sso_domain    = data.get("sso_domain").split("@")[-1].lower().strip() if data.get("sso_domain") else None           # allowed email domain
    tenant.sso_tenant_id = data.get("sso_tenant_id") or None        # Azure tenant ID
    if hasattr(tenant, "sso_sso_url"):
        tenant.sso_sso_url = data.get("sso_sso_url") or None        # IdP SSO URL
    if hasattr(tenant, "saml_cert"):
        tenant.saml_cert   = data.get("saml_cert") or None          # IdP X.509 cert
    log_system_event(db, admin, "security_config.updated",
                     target_type="tenant", target_id=tenant.id, target_label=tenant.name,
                     new_value=f"mfa={tenant.mfa_enabled} mfa_required={tenant.mfa_required} sso={tenant.sso_enabled}")
    db.commit()
    return {"ok": True}




@router.get("/admin/ip-whitelist")
def get_ip_whitelist(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("sso", tenant, "IP whitelisting is available on the Enterprise plan only.")
    raw = getattr(tenant, "ip_whitelist", None)
    try:
        cidrs = json.loads(raw) if raw else []
    except Exception:
        cidrs = []
    return {"cidrs": cidrs}


@router.put("/admin/ip-whitelist")
def update_ip_whitelist(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("sso", tenant, "IP whitelisting is available on the Enterprise plan only.")
    cidrs = data.get("cidrs", [])
    # Validate each CIDR
    import ipaddress as _ip
    valid = []
    errors = []
    for cidr in cidrs:
        cidr = cidr.strip()
        if not cidr:
            continue
        try:
            _ip.ip_network(cidr, strict=False)
            valid.append(cidr)
        except ValueError:
            errors.append(cidr)
    if errors:
        raise HTTPException(status_code=422, detail=f"Invalid CIDR(s): {', '.join(errors)}")
    tenant.ip_whitelist = json.dumps(valid) if valid else None
    db.commit()
    return {"ok": True, "cidrs": valid, "message": f"IP whitelist updated — {len(valid)} rule(s) active."}




@router.get("/admin/scheduled-reports")
def get_scheduled_reports(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("custom_analytics", tenant, "Scheduled reports are available on the Business plan and above.")
    raw = getattr(tenant, "scheduled_reports", None)
    try:
        config = json.loads(raw) if raw else {"enabled": False, "frequency": "weekly", "day": "monday", "time": "08:00", "recipients": [], "include": ["summary", "sla", "agent_workload"]}
    except Exception:
        config = {"enabled": False, "frequency": "weekly", "day": "monday", "time": "08:00", "recipients": [], "include": ["summary"]}
    return config


@router.put("/admin/scheduled-reports")
def update_scheduled_reports(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    plan_requires("custom_analytics", tenant, "Scheduled reports are available on the Business plan and above.")
    config = {
        "enabled":    bool(data.get("enabled", False)),
        "frequency":  data.get("frequency", "weekly"),   # daily | weekly | monthly
        "day":        data.get("day", "monday"),          # day of week for weekly
        "time":       data.get("time", "08:00"),          # HH:MM UTC
        "recipients": data.get("recipients", []),         # list of email addresses
        "include":    data.get("include", ["summary"]),   # report sections to include
    }
    tenant.scheduled_reports = json.dumps(config)
    db.commit()
    return {"ok": True, "message": "Scheduled report settings saved.", **config}


def _send_scheduled_report(tenant_id: int):
    """Generate and email the scheduled report for a tenant. Called by APScheduler."""
    from sqlalchemy.orm import Session as _S
    db = next(get_db())
    try:
        tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        if not tenant:
            return
        raw = getattr(tenant, "scheduled_reports", None)
        if not raw:
            return
        config = json.loads(raw)
        if not config.get("enabled"):
            return
        recipients = config.get("recipients", [])
        if not recipients:
            return

        include = config.get("include", ["summary"])
        sections = []

        # Build report sections
        if "summary" in include:
            from sqlalchemy import text as _t
            row = db.execute(_t(
                "SELECT COUNT(*) FILTER (WHERE status NOT IN ('closed','resolved')) as open_count, "
                "COUNT(*) FILTER (WHERE status='resolved') as resolved_count, "
                "COUNT(*) FILTER (WHERE status='closed') as closed_count, "
                "COUNT(*) as total "
                "FROM tickets WHERE tenant_id=:tid AND created_at > NOW() - INTERVAL '7 days'"
            ), {"tid": tenant_id}).fetchone()
            if row:
                sections.append(
                    f"📊 Ticket Summary (last 7 days)\n"
                    f"  Open: {row[0]}  |  Resolved: {row[1]}  |  Closed: {row[2]}  |  Total: {row[3]}"
                )

        if "sla" in include:
            row2 = db.execute(_t(
                "SELECT COUNT(*) FILTER (WHERE sla_response_breached=true) as r_breach, "
                "COUNT(*) FILTER (WHERE sla_resolution_breached=true) as res_breach, "
                "COUNT(*) as total "
                "FROM tickets WHERE tenant_id=:tid AND created_at > NOW() - INTERVAL '7 days'"
            ), {"tid": tenant_id}).fetchone()
            if row2 and row2[2]:
                r_pct = round((1 - row2[0] / row2[2]) * 100, 1)
                sections.append(
                    f"\n⏱ SLA Performance (last 7 days)\n"
                    f"  Response SLA: {r_pct}%  |  Breaches: {row2[0]} response, {row2[1]} resolution"
                )

        if "agent_workload" in include:
            rows3 = db.execute(_t(
                "SELECT u.full_name, COUNT(*) as cnt "
                "FROM tickets t JOIN users u ON u.id=t.assigned_to_id "
                "WHERE t.tenant_id=:tid AND t.status NOT IN ('closed','resolved') "
                "GROUP BY u.full_name ORDER BY cnt DESC LIMIT 5"
            ), {"tid": tenant_id}).fetchall()
            if rows3:
                lines = "\n".join(f"  {r[0]}: {r[1]} open tickets" for r in rows3)
                sections.append(f"\n👥 Agent Workload (open tickets)\n{lines}")

        body = (
            f"Hi,\n\nHere is your {config.get('frequency', 'weekly')} DodoDesk report "
            f"for {tenant.name}.\n\n"
            + "\n".join(sections) +
            f"\n\nView full reports: {FRONTEND_URL}/reports\n\nDodoDesk"
        )

        notif_cfg = get_email_config(db, tenant_id)
        for recipient in recipients:
            try:
                send_email(
                    recipient,
                    f"📊 DodoDesk {config.get('frequency', 'Weekly').capitalize()} Report — {tenant.name}",
                    body,
                    db=None
                )
            except Exception as e:
                print(f"⚠️ Scheduled report email failed for {recipient}: {e}")

        print(f"✅ Scheduled report sent for tenant {tenant_id} to {len(recipients)} recipient(s)")
    except Exception as e:
        print(f"⚠️ Scheduled report error for tenant {tenant_id}: {e}")
    finally:
        db.close()




@router.get("/admin/email-config")
def get_email_config_endpoint(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Get email/webhook config for this tenant. Uses raw SQL to avoid ORM schema mismatch errors."""
    defaults = {
        "smtp_host": SMTP_HOST or "", "smtp_port": SMTP_PORT or 587,
        "smtp_user": SMTP_USER or "", "smtp_pass": "",
        "smtp_from": SMTP_FROM or "", "reply_to": "",
        "slack_webhook_url": SLACK_WEBHOOK_URL or "",
        "teams_webhook_url": TEAMS_WEBHOOK_URL or "",
        "email_signature": "", "email_footer": "",
    }
    try:
        from sqlalchemy import text as _t
        try:
            row = db.execute(_t(
                "SELECT smtp_host, smtp_port, smtp_user, smtp_from, reply_to, "
                "email_signature, email_footer, slack_webhook_url, teams_webhook_url "
                "FROM email_configs WHERE tenant_id = :tid LIMIT 1"
            ), {"tid": admin.tenant_id}).fetchone()
        except Exception:
            try:
                row = db.execute(_t(
                    "SELECT smtp_host, smtp_port, smtp_user, smtp_from, reply_to "
                    "FROM email_configs WHERE tenant_id = :tid LIMIT 1"
                ), {"tid": admin.tenant_id}).fetchone()
            except Exception:
                row = None
        if not row:
            return defaults
        result = dict(defaults)
        cols = row._fields if hasattr(row, '_fields') else row.keys()
        for col in cols:
            if col == "smtp_pass":
                continue
            val = row[col]
            if val is not None:
                result[col] = val
        result["smtp_pass"] = ""
        return result
    except Exception as e:
        print(f"⚠️ email-config error: {e}")
        return defaults  # always return defaults, never 500


@router.put("/admin/email-config")
def update_email_config(
    data: dict,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    cfg = db.query(EmailConfig).filter(EmailConfig.tenant_id == admin.tenant_id).first()
    if not cfg:
        cfg = EmailConfig(tenant_id=admin.tenant_id)
        db.add(cfg)
    cfg.smtp_host = data.get("smtp_host", "")
    cfg.smtp_port = int(data.get("smtp_port", 587))
    cfg.smtp_user = data.get("smtp_user", "")
    if data.get("smtp_pass"):
        cfg.smtp_pass = data.get("smtp_pass")
    cfg.smtp_from = data.get("smtp_from", "noreply@itsm.local")
    cfg.reply_to  = data.get("reply_to", "")
    # These columns may not exist yet in older DBs — safe assignment
    try: cfg.slack_webhook_url = data.get("slack_webhook_url", "")
    except Exception: pass
    try: cfg.teams_webhook_url = data.get("teams_webhook_url", "")
    except Exception: pass
    try: cfg.email_signature = data.get("email_signature", "")
    except Exception: pass
    try: cfg.email_footer = data.get("email_footer", "")
    except Exception: pass
    db.commit()
    return {"ok": True}


@router.post("/admin/email-config/test")
def test_email_config(
    data: dict,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """Send a test email using the provided config."""
    to = data.get("test_email") or admin.email
    cfg = {
        "smtp_host": data.get("smtp_host", ""),
        "smtp_port": int(data.get("smtp_port", 587)),
        "smtp_user": data.get("smtp_user", ""),
        "smtp_pass": data.get("smtp_pass", ""),
        "smtp_from": data.get("smtp_from", "noreply@itsm.local"),
    }
    try:
        send_email(to, "ITSM Test Email", "This is a test email from your ITSM portal.", cfg)
        return {"ok": True, "message": f"Test email sent to {to}"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Email test failed: {str(e)}")


@router.post("/admin/email-config/test-slack")
def test_slack_webhook(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Send a test message to the configured Slack webhook."""
    cfg = get_email_config(db, admin.tenant_id)
    slack_url = cfg.get("slack_webhook_url", "")
    if not slack_url:
        raise HTTPException(status_code=400, detail="No Slack webhook URL configured. Add it in Webhooks settings first.")
    import httpx
    try:
        resp = httpx.post(slack_url, json={
            "text": f"✅ *DodoDesk test message* — Slack integration is working correctly for *{admin.full_name}*'s workspace."
        }, timeout=10.0)
        if resp.status_code == 200:
            return {"ok": True, "message": "Test message sent to Slack successfully."}
        raise HTTPException(status_code=400, detail=f"Slack returned {resp.status_code}: {resp.text}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not reach Slack: {str(e)}")


@router.post("/admin/email-config/test-teams")
def test_teams_webhook(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Send a test message to the configured Microsoft Teams webhook."""
    cfg = get_email_config(db, admin.tenant_id)
    teams_url = cfg.get("teams_webhook_url", "")
    if not teams_url:
        raise HTTPException(status_code=400, detail="No Teams webhook URL configured. Add it in Webhooks settings first.")
    import httpx
    try:
        resp = httpx.post(teams_url, json={
            "@type": "MessageCard",
            "@context": "http://schema.org/extensions",
            "summary": "DodoDesk Test",
            "themeColor": "059669",
            "title": "✅ DodoDesk — Test Message",
            "text": f"Teams integration is working correctly for **{admin.full_name}**'s workspace."
        }, timeout=10.0)
        if resp.status_code in (200, 202):
            return {"ok": True, "message": "Test message sent to Teams successfully."}
        raise HTTPException(status_code=400, detail=f"Teams returned {resp.status_code}: {resp.text}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not reach Teams: {str(e)}")

# USER MANAGEMENT (ADMIN ONLY, tenant‑scoped)


@router.get("/admin/users")
def admin_list_users(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=1000),
    search: str = Query("", alias="search"),
    role: str = Query("", alias="role"),
    tenant_id: int | None = Query(None, alias="tenant_id"),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    if str(admin.role) in ("super_admin", "platform_admin"):
        query = db.query(User)
    else:
        query = db.query(User).filter(User.tenant_id == admin.tenant_id)

    # Filter by tenant (super admin only)
    if tenant_id and str(admin.role) in ("super_admin", "platform_admin"):
        query = query.filter(User.tenant_id == tenant_id)

    # Filter by role
    if role:
        try:
            query = query.filter(User.role == UserRole(role))
        except ValueError:
            pass

    # Live search — ID, name, email, employee_id
    if search:
        from sqlalchemy import or_
        import re as _re
        s = f"%{search}%"
        id_match = _re.search(r'\d+', search)
        conditions = [
            User.full_name.ilike(s),
            User.email.ilike(s),
            User.employee_id.ilike(s),
            User.department.ilike(s),
            User.job_title.ilike(s),
        ]
        if id_match:
            conditions.append(User.id == int(id_match.group()))
        query = query.filter(or_(*conditions))

    total = query.count()
    users = query.order_by(User.tenant_id, User.id).offset(skip).limit(limit).all()
    # Pre-load tenants
    tenant_ids = {u.tenant_id for u in users}
    tenant_map = {t.id: t.name for t in db.query(Tenant).filter(Tenant.id.in_(tenant_ids)).all()}
    result = []
    for u in users:
        result.append({
            "id": u.id, "email": u.email, "full_name": u.full_name,
            "role": u.role, "is_active": u.is_active,
            "job_title": u.job_title, "department": u.department,
            "employee_id": getattr(u, 'employee_id', None),
            "tenant_id": u.tenant_id,
            "tenant_name": tenant_map.get(u.tenant_id, "—"),
            "created_at": u.created_at,
            "is_locked": bool(u.locked_until and u.locked_until > datetime.utcnow()),
            "status_changed_at": u.status_changed_at,
        })
    return {"items": result, "total": total, "skip": skip, "limit": limit}


@router.post("/admin/users/invite")
def invite_user(invite: UserInvite, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Invite a new user to the admin's tenant by email — no password is set by the
    admin. The invitee receives an email with a link to set their own password,
    which also activates the account. This is the recommended way to add teammates,
    as opposed to /admin/users which requires the admin to choose a password directly."""
    email = invite.email.strip().lower()
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        if existing.is_active:
            raise HTTPException(status_code=400, detail="A user with this email already exists.")
        # Inactive account exists (e.g. a stale unverified signup or a previous
        # unactivated invite) — re-send the invite rather than blocking
        existing_user = existing
    else:
        existing_user = None

    check_user_limit(db, admin.tenant_id, additional=1, role=invite.role)

    if existing_user:
        existing_user.full_name = invite.full_name
        existing_user.role = invite.role
        existing_user.job_title = invite.job_title
        existing_user.department = invite.department
        existing_user.tenant_id = admin.tenant_id
        new_user = existing_user
    else:
        new_user = User(
            tenant_id=admin.tenant_id,
            email=email,
            hashed_password=get_password_hash(uuid.uuid4().hex),  # random unusable placeholder — invitee sets their own
            full_name=invite.full_name,
            role=invite.role,
            job_title=invite.job_title,
            department=invite.department,
            is_active=False,
            email_verified=False,
        )
        db.add(new_user)
    db.commit()
    db.refresh(new_user)

    # Generate a set-password token reusing the same mechanism as forgot-password,
    # tagged so reset_password() knows to activate the account on completion.
    from sqlalchemy import text as _text
    token = uuid.uuid4().hex
    invite_val = f"invite_{token}"
    expires_at = datetime.utcnow() + timedelta(days=7)  # invites get a longer window than a reset link
    with db.bind.connect() as conn:
        conn.execute(
            _text("UPDATE users SET password_reset_token = :tok, password_reset_expires_at = :exp WHERE id = :uid"),
            {"tok": invite_val, "uid": new_user.id, "exp": expires_at}
        )
        conn.commit()

    invite_url = f"{FRONTEND_URL}/reset-password?token={token}&invite=1"

    admin_tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    company_name = admin_tenant.name if admin_tenant else "DodoDesk"
    role_label = invite.role.value if hasattr(invite.role, "value") else str(invite.role)

    import threading
    _email, _name, _url, _company, _role = new_user.email, new_user.full_name, invite_url, company_name, role_label
    send_email_background(
        to=new_user.email,
        subject=f"You've been invited to {company_name} on DodoDesk",
        body=(
            f"Hi {new_user.full_name},\n\n"
            f"{admin.full_name} has invited you to join {company_name} on DodoDesk as a {role_label}.\n\n"
            f"Click the link below to set your password and activate your account:\n\n"
            f"{invite_url}\n\n"
            f"This invite link expires in 7 days."
        ),
        cta_url=invite_url,
        cta_label="Accept Invite & Set Password",
    )
    log_system_event(db, admin, "user.invited",
                     target_type="user", target_id=new_user.id,
                     target_label=new_user.email, new_value=role_label)
    db.commit()
    return {"ok": True, "message": f"Invite sent to {email}", "user_id": new_user.id}


@router.post("/admin/users/bulk-import")
async def bulk_import_users(file: UploadFile = File(...), db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Bulk-create users from a CSV or XLSX file. Expected columns (header row required):
    full_name, email, role, job_title, department, password (optional), tenant (optional, super_admin only)
    If password is omitted, a random temporary password is generated.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    rows_data = []  # list of dicts, normalized lowercase keys

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel import is not available on this server (openpyxl not installed). Please use CSV instead, or contact support.")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(status_code=400, detail="Spreadsheet appears to be empty")
            headers = [str(h or "").strip().lower() for h in all_rows[0]]
            for r in all_rows[1:]:
                if all(c is None or str(c).strip() == "" for c in r):
                    continue  # skip fully empty rows
                row_dict = {headers[idx]: ("" if r[idx] is None else str(r[idx]).strip()) for idx in range(len(headers)) if idx < len(r)}
                rows_data.append(row_dict)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")
    else:
        try:
            text_content = content.decode("utf-8-sig")  # handle BOM from Excel CSV exports
        except UnicodeDecodeError:
            text_content = content.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text_content))
        if reader.fieldnames is None:
            raise HTTPException(status_code=400, detail="CSV file appears to be empty or invalid")
        reader.fieldnames = [(f or "").strip().lower() for f in reader.fieldnames]
        headers = reader.fieldnames
        for row in reader:
            rows_data.append({k: (v or "") for k, v in row.items()})

    required_cols = {"full_name", "email"}
    missing = required_cols - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"File is missing required column(s): {', '.join(missing)}")

    valid_roles = {r.value for r in UserRole}
    results = {"created": [], "skipped": [], "errors": []}

    for i, row in enumerate(rows_data, start=2):  # start=2 because row 1 is the header
        email = (row.get("email") or "").strip().lower()
        full_name = (row.get("full_name") or "").strip()

        if not email or not full_name:
            results["errors"].append({"row": i, "email": email, "reason": "Missing full_name or email"})
            continue

        if db.query(User).filter(User.email == email).first():
            results["skipped"].append({"row": i, "email": email, "reason": "Email already exists"})
            continue

        role_raw = (row.get("role") or "employee").strip().lower()
        if role_raw not in valid_roles or role_raw == "super_admin":
            role_raw = "employee"

        # Determine tenant
        target_tenant_id = admin.tenant_id
        tenant_identifier = (row.get("tenant") or row.get("tenant_slug") or "").strip()
        if str(admin.role) in ("super_admin", "platform_admin") and tenant_identifier:
            tenant = db.query(Tenant).filter(
                (sa_func.lower(Tenant.slug) == tenant_identifier.lower()) |
                (sa_func.lower(Tenant.name) == tenant_identifier.lower())
            ).first()
            if tenant:
                target_tenant_id = tenant.id
            else:
                results["errors"].append({"row": i, "email": email, "reason": f"Unknown tenant '{tenant_identifier}' (use exact company name or slug)"})
                continue

        password = (row.get("password") or "").strip()
        temp_password_generated = False
        if not password:
            password = _secrets.token_urlsafe(9)  # ~12 char random password
            temp_password_generated = True

        try:
            validate_password_strength(password)
        except HTTPException:
            password = _secrets.token_urlsafe(9)
            temp_password_generated = True

        try:
            check_user_limit(db, target_tenant_id, additional=1, role=role_raw)
        except HTTPException as e:
            results["errors"].append({"row": i, "email": email, "reason": e.detail})
            continue

        new_user = User(
            tenant_id=target_tenant_id,
            email=email,
            hashed_password=get_password_hash(password),
            full_name=full_name,
            role=UserRole(role_raw),
            job_title=(row.get("job_title") or "").strip() or None,
            department=(row.get("department") or "").strip() or None,
            employee_id=(row.get("employee_id") or "").strip() or None,
            is_active=True,
        )
        db.add(new_user)
        try:
            db.flush()
        except Exception as e:
            db.rollback()
            results["errors"].append({"row": i, "email": email, "reason": str(e)})
            continue

        results["created"].append({
            "row": i, "email": email, "full_name": full_name, "role": role_raw,
            "temp_password": password if temp_password_generated else None,
        })

    db.commit()
    return results


@router.delete("/admin/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Permanently delete a user. Super admin only.
    Handles all FK references using exact column names from the model definitions.
    """
    if str(admin.role) not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Only super admins can delete users.")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    email = user.email
    name  = user.full_name
    try:
        from sqlalchemy import text as _t
        u = user_id
        with db.bind.connect() as conn:
            # ── DELETE rows owned by this user (avoids NOT NULL violations) ──
            conn.execute(_t("DELETE FROM signup_verifications WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM group_members WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM notifications WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_watchers WHERE user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM time_entries WHERE agent_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM admin_tenant_access WHERE admin_user_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM canned_responses WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM comments WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM change_comments WHERE author_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_audit_logs WHERE actor_id = :u"), {"u": u})
            conn.execute(_t("DELETE FROM ticket_views WHERE created_by_id = :u"), {"u": u})
            # chat messages before chat sessions
            conn.execute(_t("DELETE FROM chat_messages WHERE session_id IN (SELECT id FROM chat_sessions WHERE user_id = :u)"), {"u": u})
            conn.execute(_t("DELETE FROM chat_sessions WHERE user_id = :u"), {"u": u})

            # ── NULLIFY shared entity references (keep historical records) ──
            conn.execute(_t("UPDATE tickets SET requester_id = NULL WHERE requester_id = :u"), {"u": u})
            conn.execute(_t("UPDATE tickets SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE kb_articles SET author_id = NULL WHERE author_id = :u"), {"u": u})
            conn.execute(_t("UPDATE kb_versions SET edited_by_id = NULL WHERE edited_by_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET requester_id = NULL WHERE requester_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET owner_id = NULL WHERE owner_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_requests SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE change_tasks SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE ticket_tasks SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE assets SET assigned_to_id = NULL WHERE assigned_to_id = :u"), {"u": u})
            conn.execute(_t("UPDATE system_audit_logs SET actor_id = NULL WHERE actor_id = :u"), {"u": u})
            conn.execute(_t("UPDATE admin_tenant_access SET granted_by_id = NULL WHERE granted_by_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET from_user_id = NULL WHERE from_user_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET to_user_id = NULL WHERE to_user_id = :u"), {"u": u})
            conn.execute(_t("UPDATE asset_history SET changed_by_id = NULL WHERE changed_by_id = :u"), {"u": u})

            # ── DELETE the user row ───────────────────────────────────────
            conn.execute(_t("DELETE FROM users WHERE id = :u"), {"u": u})
            conn.commit()

        # Audit log after deletion (admin still exists)
        log_system_event(db, admin, "user.deleted",
                         target_type="user", target_id=user_id, target_label=email)
        db.commit()

        # Reduce Dodo Payments seat count if this was an agent/admin on a paid plan
        try:
            role_value = (user.role.value if hasattr(user.role, "value") else str(user.role)) if user.role else ""
            if role_value in ("agent", "admin", "super_admin"):
                tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first()
                if tenant and tenant.billing_status == "active":
                    new_count = db.query(User).filter(
                        User.tenant_id == tenant.id,
                        User.is_active == True,
                        User.role.in_(['agent', 'admin', 'super_admin', 'platform_admin'])
                    ).count()
                    _update_dodo_seat_count(tenant, max(1, new_count))
        except Exception as e:
            print(f"⚠️ Seat reduction after delete failed (user still deleted): {e}")

        return {"ok": True, "message": f"{name} ({email}) has been permanently deleted."}
    except Exception as e:
        print(f"❌ delete_user {user_id}: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not delete user: {str(e)}")


@router.post("/admin/users/{user_id}/unlock")
def unlock_user(user_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    query = db.query(User).filter(User.id == user_id)
    # Super admin can unlock users from any tenant; regular admins only their own
    if str(admin.role) not in ("super_admin", "platform_admin"):
        query = query.filter(User.tenant_id == admin.tenant_id)
    user = query.first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.locked_until = None
    user.failed_login_attempts = 0
    user.is_active = True   # also re-activate in case account was deactivated
    log_system_event(db, admin, "user.unlocked",
                     target_type="user", target_id=user.id, target_label=user.email)
    db.commit()
    return {"ok": True, "message": f"{user.full_name} has been unlocked."}


@router.get("/admin/integrations-status")
def get_integrations_status(admin: User = Depends(get_current_admin_user), db: Session = Depends(get_db)):
    """Return status of all configured integrations for this tenant."""
    cfg = db.query(EmailConfig).filter(EmailConfig.tenant_id == admin.tenant_id).first()
    tenant = db.query(Tenant).filter(Tenant.id == admin.tenant_id).first()
    return {
        "slack": {"configured": bool(cfg and cfg.slack_webhook_url), "url": cfg.slack_webhook_url if cfg else ""},
        "teams": {"configured": bool(cfg and cfg.teams_webhook_url), "url": cfg.teams_webhook_url if cfg else ""},
        "smtp": {"configured": bool(cfg and cfg.smtp_host), "host": cfg.smtp_host if cfg else ""},
        "sso": {"configured": bool(tenant and tenant.sso_enabled), "provider": tenant.sso_provider if tenant else ""},
    }


@router.get("/superadmin/users/{user_id}/files")
def get_user_files(user_id: int, db: Session = Depends(get_db),
                   admin: User = Depends(get_current_admin_user)):
    """List all files stored in Cloudinary for a specific user.
    Returns: their avatar, and all attachments on tickets they raised.
    Super admin only.
    """
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    files = []

    # 1. Profile photo
    if user.profile_photo:
        signed = get_signed_url(user.profile_photo, resource_type="image")
        files.append({
            "type": "Profile photo",
            "public_id": user.profile_photo,
            "signed_url": signed,
            "expires_in": "1 hour",
        })

    # 2. Ticket attachments — find tickets the user raised
    tickets = db.query(Ticket).filter(
        Ticket.tenant_id == user.tenant_id,
        Ticket.requester_id == user.id,
    ).all()

    ticket_ids = [t.id for t in tickets]
    if ticket_ids:
        attachments = []
        try:
            attachments = db.query(Attachment).filter(
                Attachment.ticket_id.in_(ticket_ids)
            ).all()
        except Exception:
            pass

        ticket_map = {t.id: t.title for t in tickets}
        for att in attachments:
            public_id = getattr(att, "url", "") or ""
            if public_id and not public_id.startswith("http"):
                ext = os.path.splitext(getattr(att, "filename", "") or "")[1].lower()
                rtype = "image" if ext in {".png",".jpg",".jpeg",".gif",".webp"} else "raw"
                signed = get_signed_url(public_id, resource_type=rtype)
            else:
                signed = public_id
            files.append({
                "type": "Ticket attachment",
                "ticket_id": att.ticket_id,
                "ticket_title": ticket_map.get(att.ticket_id, ""),
                "filename": getattr(att, "filename", ""),
                "public_id": public_id,
                "signed_url": signed,
                "expires_in": "1 hour",
                "size_kb": round(att.size / 1024, 1) if getattr(att, "size", None) else None,
            })

    return {
        "user_id": user.id,
        "user_name": user.full_name,
        "user_email": user.email,
        "tenant_id": user.tenant_id,
        "cloudinary_avatar_path": f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{user.tenant_id}/avatars/",
        "cloudinary_tickets_path": f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{user.tenant_id}/tickets/",
        "total_files": len(files),
        "files": files,
    }


@router.get("/admin/branding/logo-url")
def get_logo_signed_url(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Returns a 1-hour signed URL for the tenant's logo."""
    tenant = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
    if not tenant or not tenant.logo_url:
        return {"url": None}
    logo = tenant.logo_url
    if logo.startswith("http"):
        return {"url": logo}   # legacy public URL
    signed = get_signed_url(logo, resource_type="image")
    return {"url": signed, "expires_in": 3600}



@router.post("/superadmin/tenants/{tenant_id}/logo")
async def upload_tenant_logo(tenant_id: int, file: UploadFile = File(...),
                              db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    if str(admin.role) not in ("super_admin", "platform_admin") and admin.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="You can only update your own tenant's logo")
    allowed = {"image/png", "image/jpeg", "image/jpg", "image/svg+xml", "image/webp"}
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, SVG and WebP images allowed")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Logo must be under 2 MB")
    if CLOUDINARY_CLOUD_NAME:
        _configure_cloudinary()
        import io as _io2
        ext2 = os.path.splitext(file.filename)[1].lower() or ".jpg"
        public_id2 = f"dodesk/tenants/{tenant_id}/logos/logo{ext2}"
        try:
            result2 = cloudinary.uploader.upload(
                _io2.BytesIO(content),
                public_id=public_id2,
                resource_type="image",
                type="upload",  # PUBLIC
                overwrite=True,
                invalidate=True,
            )
            logo_url = result2.get("secure_url") or f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/image/upload/{public_id2}"
            print(f"✅ Tenant logo uploaded (public): {logo_url}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Logo upload failed: {str(e)}")
    else:
        ext = file.filename.rsplit(".", 1)[-1].lower()
        filename = f"tenant_{tenant_id}_logo.{ext}"
        path = os.path.join(LOGO_DIR, filename)
        with open(path, "wb") as f:
            f.write(content)
        logo_url = f"/logos/{filename}"
    tenant.logo_url = logo_url
    db.commit()
    return {"logo_url": logo_url}


@router.patch("/superadmin/tenants/{tenant_id}")
def update_tenant(tenant_id: int, data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    if str(admin.role) in ("super_admin", "platform_admin"):
        allowed_fields = ["name", "support_email", "company_tagline", "primary_color", "accent_color", "is_active", "plan"]
        if "plan" in data and data["plan"] not in PLAN_LIMITS:
            raise HTTPException(status_code=400, detail=f"Invalid plan. Must be one of: {', '.join(PLAN_LIMITS.keys())}")
    elif admin.tenant_id == tenant_id:
        allowed_fields = ["support_email", "company_tagline", "primary_color", "accent_color"]
    else:
        raise HTTPException(status_code=403, detail="You can only update your own tenant")

    for field in allowed_fields:
        if field in data:
            old_val = getattr(tenant, field, None)
            new_val = data[field]
            if str(old_val) != str(new_val):
                log_system_event(db, admin, f"tenant.{field}.changed",
                                 target_type="tenant", target_id=tenant.id,
                                 target_label=tenant.name,
                                 old_value=str(old_val), new_value=str(new_val))
            setattr(tenant, field, new_val)

    # If super_admin manually sets a plan, mark billing as active (clears trial state)
    if str(admin.role) in ("super_admin", "platform_admin") and "plan" in data:
        new_plan = data["plan"]
        if new_plan == "free":
            tenant.billing_status = "cancelled"
        else:
            tenant.billing_status = "active"

    db.commit()
    return {"ok": True, "plan": tenant.plan, "billing_status": tenant.billing_status}


@router.get("/superadmin/tenants")
def list_all_tenants(
    search: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user)
):
    """List tenants visible to this admin.
    - platform_admin: ALL tenants in the system
    - super_admin (MSP): own tenant + explicitly granted client tenants only
    - regular admin: own tenant only
    """
    role = admin.role.value if hasattr(admin.role, 'value') else str(admin.role)

    if role == 'platform_admin':
        # DodoDesk owner — sees everything
        query = db.query(Tenant)
        if search:
            query = query.filter(Tenant.name.ilike(f"%{_sql_safe_search(search)}%"))
        tenants = query.order_by(Tenant.created_at.desc()).all()

    else:
        # super_admin (MSP) or regular admin — own tenant + explicitly granted only
        own_ids = {admin.tenant_id}
        granted = db.query(AdminTenantAccess).filter(
            AdminTenantAccess.admin_user_id == admin.id
        ).all()
        for g in granted:
            own_ids.add(g.tenant_id)
        query = db.query(Tenant).filter(Tenant.id.in_(own_ids))
        if search:
            query = query.filter(Tenant.name.ilike(f"%{_sql_safe_search(search)}%"))
        tenants = query.order_by(Tenant.name).all()

    def tenant_row(t):
        is_own = (t.id == admin.tenant_id)
        is_granted = not is_own
        return {
            "id": t.id,
            "name": t.name,
            "slug": t.slug,
            "plan": t.plan or "free",
            "is_active": t.is_active,
            "primary_color": t.primary_color or "#4f46e5",
            "accent_color": t.accent_color or "#818cf8",
            "logo_url": t.logo_url,
            "company_tagline": t.company_tagline,
            "support_email": t.support_email,
            "billing_status": getattr(t, "billing_status", None),
            "plan_renews_at": str(t.plan_renews_at)[:10] if getattr(t, "plan_renews_at", None) else None,
            "created_at": str(t.created_at)[:10] if t.created_at else None,
            "is_own": is_own,
            "is_granted": is_granted,
            "user_count": db.query(User).filter(
                User.tenant_id == t.id, User.is_active == True
            ).count(),
        }

    return [tenant_row(t) for t in tenants]


@router.get("/superadmin/tenants/{tenant_id}")
def get_tenant_by_id(tenant_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Get a single tenant. Super admin can fetch any; regular admin only their own."""
    query = db.query(Tenant).filter(Tenant.id == tenant_id)
    if str(admin.role) not in ("super_admin", "platform_admin"):
        query = query.filter(Tenant.id == admin.tenant_id)
    tenant = query.first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return tenant


@router.post("/superadmin/tenants")
def create_tenant_superadmin(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Create a new tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Tenant name is required")
    slug = data.get("slug") or name.lower().replace(" ", "-")
    existing = db.query(Tenant).filter(Tenant.slug == slug).first()
    if existing:
        slug = f"{slug}-{db.query(Tenant).count()}"
    tenant = Tenant(
        name=name, slug=slug,
        plan=data.get("plan", "free"),
        is_active=True,
        primary_color=data.get("primary_color", "#4f46e5"),
    )
    db.add(tenant)
    db.commit()
    db.refresh(tenant)
    return {"id": tenant.id, "name": tenant.name, "slug": tenant.slug, "plan": tenant.plan}


@router.delete("/superadmin/tenants/{tenant_id}")
def delete_tenant(tenant_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Permanently delete a tenant and ALL its data. Super admin only.
    Cleans up all related tables before deleting the tenant row.
    """
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    if tenant_id == admin.tenant_id:
        raise HTTPException(status_code=400, detail="You cannot delete your own tenant.")

    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    tenant_name = tenant.name
    tenant_slug = tenant.slug

    # Log BEFORE deletion (audit logs for target tenant will be wiped)
    # Log to super admin's own tenant so the record persists
    try:
        log_system_event(db, admin, "tenant.deleted",
                         target_type="tenant", target_id=tenant_id,
                         target_label=f"{tenant_name} ({tenant_slug})",
                         old_value=f"plan={tenant.plan}, users=deleted")
        db.commit()
    except Exception as e:
        print(f"⚠️ Pre-deletion audit log failed: {e}")

    try:
        from sqlalchemy import text as _t
        with db.bind.connect() as conn:
            tid = tenant_id
            # ── Delete in FK dependency order ─────────────────────────────
            # Ticket children
            conn.execute(_t("DELETE FROM ticket_watchers WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_audit_logs WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM time_entries WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM comments WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM attachments WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_tasks WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            # ticket_views has tenant_id directly
            conn.execute(_t("DELETE FROM ticket_views WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM tickets WHERE tenant_id = :t"), {"t": tid})
            # Change request children (use change_id not change_request_id)
            conn.execute(_t("DELETE FROM change_comments WHERE change_id IN (SELECT id FROM change_requests WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM change_tasks WHERE change_id IN (SELECT id FROM change_requests WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM change_requests WHERE tenant_id = :t"), {"t": tid})
            # Asset children
            conn.execute(_t("DELETE FROM asset_history WHERE asset_id IN (SELECT id FROM assets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM assets WHERE tenant_id = :t"), {"t": tid})
            # KB children
            conn.execute(_t("DELETE FROM kb_versions WHERE article_id IN (SELECT id FROM kb_articles WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM kb_articles WHERE tenant_id = :t"), {"t": tid})
            # Service catalog (depends on approval_workflows)
            conn.execute(_t("UPDATE service_catalog_items SET approval_workflow_id = NULL WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM service_catalog_items WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM approval_workflows WHERE tenant_id = :t"), {"t": tid})
            # Tenant-level config
            conn.execute(_t("DELETE FROM automation_rules WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM escalation_rules WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM sla_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM business_hours_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM email_configs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM canned_responses WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM custom_fields WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM ticket_templates WHERE tenant_id = :t"), {"t": tid})
            # Approval workflow children before the workflow itself
            conn.execute(_t("DELETE FROM ticket_approvals WHERE ticket_id IN (SELECT id FROM tickets WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM approval_steps WHERE workflow_id IN (SELECT id FROM approval_workflows WHERE tenant_id = :t)"), {"t": tid})
            # Groups — table is 'groups' not 'agent_groups'
            conn.execute(_t("DELETE FROM group_members WHERE group_id IN (SELECT id FROM groups WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM groups WHERE tenant_id = :t"), {"t": tid})
            # User-level cleanup
            conn.execute(_t("DELETE FROM notifications WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM admin_tenant_access WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM system_audit_logs WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM chat_messages WHERE session_id IN (SELECT id FROM chat_sessions WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM chat_sessions WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM time_entries WHERE agent_id IN (SELECT id FROM users WHERE tenant_id = :t)"), {"t": tid})
            conn.execute(_t("DELETE FROM signup_verifications WHERE tenant_id = :t"), {"t": tid})
            conn.execute(_t("DELETE FROM users WHERE tenant_id = :t"), {"t": tid})
            # Finally the tenant itself
            conn.execute(_t("DELETE FROM tenants WHERE id = :t"), {"t": tid})
            conn.commit()

        print(f"✅ Tenant deleted: {tenant_name} (id={tenant_id})")
        return {"ok": True, "message": f"Tenant \"{tenant_name}\" and all its data have been permanently deleted."}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not delete tenant: {type(e).__name__}: {str(e)[:300]}")




@router.get("/superadmin/admin-access")
def list_admin_access(db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """List all admin-to-tenant access grants. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    records = db.query(AdminTenantAccess).all()
    return [{
        "id": r.id,
        "admin_user_id": r.admin_user_id,
        "admin_name": r.admin_user.full_name if r.admin_user else "",
        "admin_email": r.admin_user.email if r.admin_user else "",
        "tenant_id": r.tenant_id,
        "tenant_name": r.tenant.name if r.tenant else "",
        "granted_at": r.granted_at,
    } for r in records]


@router.post("/superadmin/admin-access")
def grant_admin_access(data: dict, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Grant an admin access to an additional tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    admin_user_id = data.get("admin_user_id")
    tenant_id = data.get("tenant_id")
    if not admin_user_id or not tenant_id:
        raise HTTPException(status_code=400, detail="admin_user_id and tenant_id are required")
    # Verify target user is an admin
    target = db.query(User).filter(User.id == admin_user_id).first()
    if not target or str(target.role) not in ["admin"]:
        raise HTTPException(status_code=400, detail="Target user must be an Admin role")
    # Check tenant exists
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    # Check not already granted
    existing = db.query(AdminTenantAccess).filter(
        AdminTenantAccess.admin_user_id == admin_user_id,
        AdminTenantAccess.tenant_id == tenant_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Access already granted")
    access = AdminTenantAccess(admin_user_id=admin_user_id, tenant_id=tenant_id, granted_by_id=admin.id)
    db.add(access)
    db.commit()
    return {"ok": True, "admin_user_id": admin_user_id, "tenant_id": tenant_id}


@router.delete("/superadmin/admin-access/{access_id}")
def revoke_admin_access(access_id: int, db: Session = Depends(get_db), admin: User = Depends(get_current_admin_user)):
    """Revoke an admin's access to a tenant. Super admin only."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    record = db.query(AdminTenantAccess).filter(AdminTenantAccess.id == access_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Access record not found")
    db.delete(record)
    db.commit()
    return {"ok": True}


@router.get("/superadmin/tenants/{tenant_id}/export")
def export_tenant_data(tenant_id: int, db: Session = Depends(get_db),
                       admin: User = Depends(get_current_admin_user)):
    """Export all data for a tenant as a multi-sheet Excel file."""
    role = admin.role.value if hasattr(admin.role, "value") else str(admin.role)
    if role not in ("super_admin", "platform_admin"):
        raise HTTPException(status_code=403, detail="Super admin only")
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        HEADER_FILL  = PatternFill("solid", fgColor="4F46E5")
        HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

        def make_sheet(title, headers, rows):
            ws = wb.create_sheet(title=title[:31])
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
            for row in rows:
                ws.append([str(v) if v is not None else "" for v in row])
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
            ws.freeze_panes = "A2"
            return ws

        # Sheet 1: Tenant Info
        make_sheet("Tenant Info", ["Field", "Value"], [
            ["ID", tenant.id], ["Name", tenant.name], ["Slug", tenant.slug],
            ["Plan", tenant.plan], ["Active", tenant.is_active],
            ["Primary Color", tenant.primary_color],
            ["Support Email", getattr(tenant, "support_email", "") or ""],
            ["Billing Status", getattr(tenant, "billing_status", "") or ""],
            ["Created At", str(tenant.created_at)[:19] if tenant.created_at else ""],
        ])

        # Sheet 2: Users
        users = db.query(User).filter(User.tenant_id == tenant_id).all()
        make_sheet("Users",
            ["ID", "Full Name", "Email", "Role", "Job Title", "Department", "Active", "MFA", "Created At"],
            [(u.id, u.full_name, u.email, str(u.role) if u.role else "",
              getattr(u, "job_title", "") or "", getattr(u, "department", "") or "",
              u.is_active, getattr(u, "mfa_enabled", False),
              str(u.created_at)[:19] if u.created_at else "") for u in users]
        )

        # Sheet 3: Tickets
        tickets = db.query(Ticket).filter(Ticket.tenant_id == tenant_id).all()
        user_map = {u.id: u.full_name for u in users}
        make_sheet("Tickets",
            ["ID", "Ref", "Type", "Title", "Status", "Priority", "Category", "Requester", "Assigned To", "Created At"],
            [(t.id,
              f"{'INC' if t.ticket_type and 'incident' in str(t.ticket_type).lower() else 'REQ'}-{t.id:04d}",
              str(str(t.ticket_type)) if t.ticket_type else "",
              t.title, str(str(t.status)) if t.status else "",
              str(str(t.priority)) if t.priority else "",
              t.category or "",
              user_map.get(t.requester_id, str(t.requester_id) if t.requester_id else ""),
              user_map.get(t.assigned_to_id, "") if t.assigned_to_id else "Unassigned",
              str(t.created_at)[:19] if t.created_at else "") for t in tickets]
        )

        # Sheet 4: Assets
        assets = db.query(Asset).filter(Asset.tenant_id == tenant_id).all()
        make_sheet("Assets",
            ["ID", "Name", "Type", "Serial Number", "Status", "Assigned To"],
            [(a.id, a.name, str(getattr(a, "asset_type", {}).value if hasattr(getattr(a, "asset_type", None), "value") else ""),
              getattr(a, "serial_number", "") or "",
              str(getattr(a, "status", {}).value if hasattr(getattr(a, "status", None), "value") else ""),
              user_map.get(getattr(a, "assigned_to_id", None), "")) for a in assets]
        )

        # Sheet 5: Knowledge Base
        articles = db.query(KBArticle).filter(KBArticle.tenant_id == tenant_id).all()
        make_sheet("Knowledge Base",
            ["ID", "Title", "Category", "Author", "Created At"],
            [(a.id, a.title, getattr(a, "category", "") or "",
              user_map.get(getattr(a, "author_id", None), ""),
              str(a.created_at)[:19] if getattr(a, "created_at", None) else "") for a in articles]
        )

        # Sheet 6: Audit Log
        logs = db.query(SystemAuditLog).filter(SystemAuditLog.tenant_id == tenant_id)                  .order_by(SystemAuditLog.id.desc()).limit(5000).all()
        make_sheet("Audit Log",
            ["ID", "Timestamp", "Action", "Actor", "Target Type"],
            [(l.id,
              str(l.created_at)[:19] if getattr(l, "created_at", None) else "",
              getattr(l, "action", "") or "",
              user_map.get(getattr(l, "actor_id", None), ""),
              getattr(l, "target_type", "") or "") for l in logs]
        )

        # Sheet 7: Attachments
        ticket_ids = [t.id for t in tickets]
        attachments = []
        if ticket_ids:
            try:
                attachments = db.query(Attachment).filter(
                    Attachment.ticket_id.in_(ticket_ids)
                ).all()
            except Exception:
                pass
        ticket_ref_map = {t.id: f"INC-{t.id:04d}" for t in tickets}
        make_sheet("Attachments",
            ["ID", "Ticket Ref", "Filename", "Size (KB)", "Uploaded At"],
            [(a.id,
              ticket_ref_map.get(a.ticket_id, str(a.ticket_id)),
              getattr(a, "filename", "") or "",
              round(a.size / 1024, 1) if getattr(a, "size", None) else 0,
              str(a.uploaded_at)[:19] if getattr(a, "uploaded_at", None) else "") for a in attachments]
        )

        # Sheet 8: Cloudinary Files
        cloudinary_files = []
        try:
            _configure_cloudinary()
            import cloudinary.api as _capi
            folder_prefix = f"{CLOUDINARY_PRODUCT_PREFIX}/tenants/{tenant_id}"
            for resource_type in ["image", "raw"]:
                try:
                    result = _capi.resources(
                        type="authenticated", prefix=folder_prefix,
                        max_results=500, resource_type=resource_type,
                    )
                    cloudinary_files += [
                        (r["public_id"], resource_type, round(r.get("bytes", 0)/1024, 1), r.get("created_at", ""))
                        for r in result.get("resources", [])
                    ]
                except Exception:
                    pass
        except Exception as e:
            print(f"⚠️ Cloudinary listing: {e}")

        make_sheet("Cloudinary Files",
            ["Path", "Type", "Size (KB)", "Uploaded At"],
            cloudinary_files if cloudinary_files else [["No files found", "", "", ""]]
        )

        # Stream Excel
        output = _io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"dodesk_export_{tenant.slug}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"✅ Export: {tenant.name} ({len(tickets)} tickets, {len(users)} users, {len(attachments)} attachments)")
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {type(e).__name__}: {str(e)[:300]}")


