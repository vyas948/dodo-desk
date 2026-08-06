"""
Reports — SLA, agent workload, CSAT, ticket analytics router.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, BackgroundTasks, Query, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, text
from datetime import datetime, timedelta
from typing import Optional
import json, uuid, os, re, io, csv, threading

# Import everything from shared utilities (avoids circular import with main.py)
from shared import (
    # Database
    get_db, SessionLocal,
    # Auth dependencies
    get_current_user, get_current_admin_user, has_permission, Permission,
    oauth2_scheme, OAuth2PasswordRequestForm,
    # Models - SQLAlchemy
    User, Tenant, Ticket, Comment, Notification, KBArticle, Asset,
    ChangeRequest, SystemAuditLog, TicketAuditLog, Macro, AutomationRule,
    EscalationRule, Group, CustomField, TicketTemplate, ApprovalWorkflow,
    ServiceCatalogItem, CannedResponse, TicketView, TicketTask, Attachment,
    # Pydantic models
    TicketCreate, TicketUpdate, TicketOut, CommentCreate, CommentOut,
    KBArticleCreate, KBArticleUpdate, KBArticleOut,
    AssetCreate, AssetUpdate, AssetOut, LinkAssetRequest,
    UserOut, UserCreate, UserInvite, UserUpdate, UserProfileUpdate, PasswordUpdate,
    CannedResponseCreate, CannedResponseUpdate, CannedResponseOut,
    ChangeCreate, ChangeUpdate, ChangeOut,
    ServiceCatalogItemCreate, ServiceCatalogItemOut,
    BulkTicketAction, InboundEmail, CSATSubmit, CSATStats, AttachmentOut,
    SignupRequest, TenantOut, CustomRoleCreate, CustomRoleOut,
    # Email & notifications
    get_email_config, send_email, send_email_background, send_notification,
    create_notification, build_html_email, translate_email,
    # Ticket helpers
    log_ticket_event, _ticket_tenant_filter, compute_sla_deadlines,
    pause_sla, resume_sla, get_sla_rules, compute_sla_status,
    run_automation_rules, trigger_approval_workflow,
    # User helpers
    get_plan_limits, get_user_language, check_tenant_limit, check_user_limit,
    plan_requires, get_trial_status, validate_password_strength,
    # Auth helpers
    get_password_hash, verify_password, create_access_token,
    create_access_token_with_expiry, decode_access_token,
    check_ip_rate_limit, log_system_event,
    # MFA/TOTP
    generate_totp_secret, totp_provisioning_uri, verify_totp,
    generate_backup_codes,
    # File/media
    upload_to_cloudinary, get_signed_url,
    # Business hours & SLA
    get_business_hours_config, add_business_hours,
    # Audit
    _sql_safe_search, _cr_to_out, _safe_json,
    _asset_to_out, _ticket_to_out, _catalog_to_out, _change_to_out,
    _notify_watchers, _round_robin_assign, _user_wants_notif,
    # Rate limiting
    limiter, Limiter, get_remote_address,
    # Constants
    API_URL, FRONTEND_URL, SECRET_KEY, ALGORITHM,
    ACCESS_TOKEN_EXPIRE_MINUTES, MAX_FAILED_ATTEMPTS,
    PLAN_LIMITS, SLA_RULES,
)

router = APIRouter()


@router.get("/reports/my-clients")
def get_my_clients(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns the list of client tenants accessible to the current super_admin/platform_admin.
    Used by the Reports page to show a client selector for MSPs."""
    role = str(current_user.role)
    
    # platform_admin sees all tenants
    if role == 'platform_admin':
        tenants = db.query(Tenant).filter(Tenant.is_active == True).all()
        return [{"id": t.id, "name": t.name, "plan": t.plan} for t in tenants]
    
    # super_admin sees own tenant + assigned client tenants
    if role == 'super_admin':
        # Own tenant first
        own = db.query(Tenant).filter(Tenant.id == current_user.tenant_id).first()
        result = [{"id": own.id, "name": own.name + " (own)", "plan": own.plan}] if own else []
        
        # Assigned client tenants
        granted = db.query(AdminTenantAccess).filter(
            AdminTenantAccess.admin_user_id == current_user.id
        ).all()
        for g in granted:
            t = db.query(Tenant).filter(Tenant.id == g.tenant_id, Tenant.is_active == True).first()
            if t and t.id != current_user.tenant_id:
                result.append({"id": t.id, "name": t.name, "plan": t.plan})
        return result
    
    # Other roles — no client selector
    return []



@router.get("/reports/summary")
def report_summary(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    client_tenant_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    try:
        from sqlalchemy import text as _t
        # MSP: allow viewing client tenant reports
        if client_tenant_id and str(current_user.role) in ("super_admin", "platform_admin"):
            tid = client_tenant_id
        else:
            tid = current_user.tenant_id

        # Use raw SQL to avoid ORM enum deserialisation issues
        def count_q(where_extra=""):
            sql = f"SELECT COUNT(*) FROM tickets WHERE tenant_id=:tid {where_extra}"
            return db.execute(_t(sql), {"tid": tid}).scalar() or 0

        total        = count_q()
        open_count   = count_q("AND status IN ('open','in_progress','pending_approval')")
        overdue      = count_q(f"AND sla_resolution_deadline < NOW() AND status IN ('open','in_progress')")
        today_start  = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        resolved_today = db.execute(_t(
            "SELECT COUNT(*) FROM tickets WHERE tenant_id=:tid AND status='resolved' AND updated_at>=:ts"
        ), {"tid": tid, "ts": today_start}).scalar() or 0

        # Avg resolution hours
        try:
            rows = db.execute(_t(
                "SELECT created_at, updated_at FROM tickets "
                "WHERE tenant_id=:tid AND status='resolved' AND created_at IS NOT NULL AND updated_at IS NOT NULL"
            ), {"tid": tid}).fetchall()
            avg_resolution_hours = round(
                sum((r[1]-r[0]).total_seconds()/3600 for r in rows if r[1] and r[0]) / len(rows), 1
            ) if rows else 0
        except Exception:
            avg_resolution_hours = 0

        # Avg first response hours
        try:
            rows2 = db.execute(_t(
                "SELECT created_at, first_response_at FROM tickets "
                "WHERE tenant_id=:tid AND first_response_at IS NOT NULL AND created_at IS NOT NULL"
            ), {"tid": tid}).fetchall()
            avg_first_response_hours = round(
                sum((r[1]-r[0]).total_seconds()/3600 for r in rows2 if r[1] and r[0]) / len(rows2), 1
            ) if rows2 else 0
        except Exception:
            avg_first_response_hours = 0

        # Open changes — all non-completed statuses
        try:
            open_changes = db.execute(_t(
                "SELECT COUNT(*) FROM change_requests WHERE tenant_id=:tid "
                "AND status NOT IN ('completed','cancelled','failed')"
            ), {"tid": tid}).scalar() or 0
        except Exception:
            open_changes = 0

        return {
            "total": total, "open": open_count, "overdue": overdue,
            "resolved_today": resolved_today,
            "avg_resolution_hours": avg_resolution_hours,
            "avg_first_response_hours": avg_first_response_hours,
            "open_changes": open_changes,
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Summary error: {str(e)[:200]}")


@router.get("/reports/sla-compliance")
def sla_compliance(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    base_query = db.query(Ticket).filter(Ticket.tenant_id == _eff_tid)
    base_query = apply_filters(base_query, ticket_type, start_date, end_date)
    resolved_total = base_query.filter(Ticket.status == 'resolved').count()
    if resolved_total == 0:
        return {"compliance_percent": 100.0, "total_resolved": 0}
    on_time = base_query.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at <= Ticket.sla_resolution_deadline
    ).count()
    compliance = round((on_time / resolved_total) * 100, 1)
    return {"compliance_percent": compliance, "total_resolved": resolved_total, "on_time": on_time}


@router.get("/reports/tickets-by-priority")
def tickets_by_priority(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    query = db.query(Ticket.priority, sa_func.count(Ticket.id)).filter(Ticket.tenant_id == _eff_tid)
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(Ticket.priority).all()
    return [{"priority": (p.value if hasattr(p, "value") else str(p)) if p else "unknown", "count": c} for p, c in results]


@router.get("/reports/tickets-by-status")
def tickets_by_status(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    query = db.query(Ticket.status, sa_func.count(Ticket.id)).filter(Ticket.tenant_id == _eff_tid)
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(Ticket.status).all()
    return [{"status": (s.value if hasattr(s, "value") else str(s)) if s else "unknown", "count": c} for s, c in results]


@router.get("/reports/tickets-created-daily")
def tickets_created_daily(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    if start_date and end_date:
        start = start_date
        end = end_date
    else:
        today = datetime.utcnow().date()
        start = today - timedelta(days=6)
        end = today
    days = []
    current = start
    while current <= end:
        day_start = datetime(current.year, current.month, current.day)
        day_end = day_start + timedelta(days=1)
        query = db.query(Ticket).filter(
            Ticket.tenant_id == _eff_tid,
            Ticket.created_at >= day_start,
            Ticket.created_at < day_end
        )
        query = apply_filters(query, ticket_type, None, None)
        count = query.count()
        days.append({"date": current.isoformat(), "count": count})
        current += timedelta(days=1)
    return days


@router.get("/reports/my-stats")
def my_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Personal stats for the current agent: assigned, due today, overdue, resolved this week."""
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end   = today_start + timedelta(days=1)
    week_start  = today_start - timedelta(days=today_start.weekday())
    now         = datetime.utcnow()
    base = db.query(Ticket).filter(
        Ticket.tenant_id == current_user.tenant_id,
        Ticket.assigned_to_id == current_user.id
    )
    assigned_open = base.filter(Ticket.status.in_(['open','in_progress','pending_approval'])).count()
    due_today = base.filter(
        Ticket.status.in_(['open','in_progress']),
        (
            (Ticket.due_date >= today_start) & (Ticket.due_date < today_end)
        ) | (
            (Ticket.sla_resolution_deadline >= today_start) & (Ticket.sla_resolution_deadline < today_end)
        )
    ).count()
    overdue_mine = base.filter(
        Ticket.sla_resolution_deadline < now,
        Ticket.status.in_(['open','in_progress'])
    ).count()
    resolved_week = base.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at >= week_start
    ).count()
    # Avg resolution time this week
    resolved_tix = base.filter(
        Ticket.status == 'resolved',
        Ticket.updated_at >= week_start,
        Ticket.updated_at.isnot(None)
    ).with_entities(Ticket.created_at, Ticket.updated_at).all()
    avg_res = 0
    if resolved_tix:
        avg_res = round(sum((t.updated_at - t.created_at).total_seconds() / 3600 for t in resolved_tix if t.updated_at and t.created_at) / len(resolved_tix), 1)
    return {
        "assigned_open": assigned_open,
        "due_today": due_today,
        "overdue_mine": overdue_mine,
        "resolved_week": resolved_week,
        "avg_resolution_hours": avg_res,
    }


@router.get("/reports/agent-workload")
def agent_workload(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    agents = db.query(User).filter(User.tenant_id == _eff_tid, User.role == 'agent').all()
    if not agents:
        return []
    agent_ids = [a.id for a in agents]

    # Single query for all assigned/resolved counts, grouped by agent — replaces N×2 per-agent count() calls
    base_q = db.query(Ticket).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.assigned_to_id.in_(agent_ids)
    )
    base_q = apply_filters(base_q, ticket_type, start_date, end_date)
    tickets = base_q.with_entities(Ticket.id, Ticket.assigned_to_id, Ticket.status).all()

    assigned_counts = {}
    resolved_counts = {}
    ticket_ids_by_agent = {}
    for tid, agent_id, status in tickets:
        assigned_counts[agent_id] = assigned_counts.get(agent_id, 0) + 1
        ticket_ids_by_agent.setdefault(agent_id, []).append(tid)
        if status == "resolved":
            resolved_counts[agent_id] = resolved_counts.get(agent_id, 0) + 1

    # Single query for all time entries across all agents — replaces N separate queries
    all_ticket_ids = [tid for ids in ticket_ids_by_agent.values() for tid in ids]
    minutes_by_agent = {}
    if all_ticket_ids:
        time_rows = db.query(TimeEntry.agent_id, TimeEntry.minutes).filter(
            TimeEntry.agent_id.in_(agent_ids),
            TimeEntry.ticket_id.in_(all_ticket_ids)
        ).all()
        for agent_id, minutes in time_rows:
            minutes_by_agent[agent_id] = minutes_by_agent.get(agent_id, 0) + minutes

    result = []
    for agent in agents:
        result.append({
            "agent_name": agent.full_name,
            "assigned": assigned_counts.get(agent.id, 0),
            "resolved": resolved_counts.get(agent.id, 0),
            "total_hours": round(minutes_by_agent.get(agent.id, 0) / 60, 1),
        })
    return result


@router.get("/reports/changes-summary")
def changes_summary(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    client_tenant_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Summary stats for change requests."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    q = db.query(ChangeRequest).filter(ChangeRequest.tenant_id == _eff_tid)
    if start_date:
        q = q.filter(ChangeRequest.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        q = q.filter(ChangeRequest.created_at <= datetime.combine(end_date, datetime.max.time()))
    total = q.count()
    # By status
    by_status = {}
    for row in q.with_entities(ChangeRequest.status, sa_func.count()).group_by(ChangeRequest.status).all():
        by_status[str(row[0].value if hasattr(row[0],'value') else row[0]) if row[0] else 'unknown'] = row[1]
    # By risk
    by_risk = {}
    for row in q.with_entities(ChangeRequest.risk_level, sa_func.count()).group_by(ChangeRequest.risk_level).all():
        by_risk[str(row[0].value if hasattr(row[0],'value') else row[0]) if row[0] else 'unknown'] = row[1]
    # Daily trend (last 30 days)
    from sqlalchemy import cast, Date as SADate
    daily = []
    try:
        rows = (q.with_entities(cast(ChangeRequest.created_at, SADate).label('day'), sa_func.count())
                  .group_by(cast(ChangeRequest.created_at, SADate))
                  .order_by(cast(ChangeRequest.created_at, SADate)).all())
        daily = [{"date": str(r[0]), "count": r[1]} for r in rows]
    except Exception:
        pass
    # Open count
    open_statuses = ["pending_approval", "approved"]
    open_count = q.filter(ChangeRequest.status.in_(open_statuses)).count()
    implemented = by_status.get('implemented', 0)
    rejected    = by_status.get('rejected', 0)
    return {
        "total": total,
        "open": open_count,
        "implemented": implemented,
        "rejected": rejected,
        "by_status": by_status,
        "by_risk": by_risk,
        "daily": daily,
    }



@router.get("/reports/export/csv")
def export_csv(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Type", "Title", "Category", "Priority", "Status", "Requester", "Assigned To", "Created", "SLA Status", "Attachments"])

    if ticket_type == "change":
        # Export change requests
        query = db.query(ChangeRequest).filter(ChangeRequest.tenant_id == current_user.tenant_id)
        if start_date:
            query = query.filter(ChangeRequest.created_at >= datetime(start_date.year, start_date.month, start_date.day))
        if end_date:
            end_dt = datetime(end_date.year, end_date.month, end_date.day) + timedelta(days=1)
            query = query.filter(ChangeRequest.created_at < end_dt)
        changes = query.order_by(ChangeRequest.id).all()
        req_ids = {c.requester_id for c in changes if c.requester_id}
        req_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(req_ids)).all()} if req_ids else {}
        for c in changes:
            try:
                writer.writerow([
                    f"CHG-{c.id:04d}", "change_request", c.title or "",
                    getattr(c, 'category', '') or "",
                    str(c.risk_level) if c.risk_level else "",
                    str(c.status) if c.status else "",
                    req_map.get(c.requester_id, ""),
                    "", c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "", ""
                ])
            except Exception:
                continue
    else:
        # Export tickets (incidents and/or service requests)
        query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)
        query = apply_filters(query, ticket_type, start_date, end_date)
        tickets = query.order_by(Ticket.id).all()
        # Pre-load users to avoid lazy loading issues
        user_ids = set()
        for t in tickets:
            if t.requester_id: user_ids.add(t.requester_id)
            if t.assigned_to_id: user_ids.add(t.assigned_to_id)
        user_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(user_ids)).all()} if user_ids else {}

        for t in tickets:
            if t.ticket_type == "incident":
                ticket_ref = f"INC-{t.id:04d}"
            elif t.ticket_type == "service_request":
                ticket_ref = f"REQ-{t.id:04d}"
            else:
                ticket_ref = f"CHG-{t.id:04d}"
            try:
                att_list = db.query(Attachment).filter(Attachment.ticket_id == t.id).all()
                att_urls = " | ".join([a.url or a.file_path or "" for a in att_list if (a.url or a.file_path)])
                writer.writerow([
                    ticket_ref,
                    str(t.ticket_type) if t.ticket_type else "",
                    t.title or "",
                    t.category or "",
                    str(t.priority) if t.priority else "",
                    str(t.status) if t.status else "",
                    user_map.get(t.requester_id, ""),
                    user_map.get(t.assigned_to_id, "Unassigned"),
                    t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                    compute_sla_status(t),
                    att_urls,
                ])
            except Exception:
                continue

    csv_content = output.getvalue()
    output.close()
    from fastapi.responses import Response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=dodesk_export.csv"}
    )


@router.get("/reports/tickets-by-category")
def tickets_by_category(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Category breakdown with volume, open count, and avg resolution time —
    used to identify which categories need the most operational focus."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id

    base = db.query(Ticket).filter(Ticket.tenant_id == _eff_tid)
    base = apply_filters(base, ticket_type, start_date, end_date)
    tickets = base.all()

    by_cat = {}
    for t in tickets:
        cat = t.category or "Uncategorised"
        if cat not in by_cat:
            by_cat[cat] = {"category": cat, "count": 0, "open": 0, "overdue": 0, "res_hours": [], "critical": 0}
        entry = by_cat[cat]
        entry["count"] += 1
        if t.status in ("open", "in_progress", "pending_approval"):
            entry["open"] += 1
            if t.sla_resolution_deadline and t.sla_resolution_deadline < datetime.utcnow():
                entry["overdue"] += 1
        if t.priority == "critical":
            entry["critical"] += 1
        if t.status == "resolved" and t.updated_at and t.created_at:
            entry["res_hours"].append((t.updated_at - t.created_at).total_seconds() / 3600)

    results = []
    for cat, e in by_cat.items():
        avg_res = round(sum(e["res_hours"]) / len(e["res_hours"]), 1) if e["res_hours"] else None
        # Focus score: weighted combination of volume, overdue count, and critical tickets
        # Higher score = needs more attention
        focus_score = e["count"] + (e["overdue"] * 3) + (e["critical"] * 2)
        results.append({
            "category": cat,
            "count": e["count"],
            "open": e["open"],
            "overdue": e["overdue"],
            "critical": e["critical"],
            "avg_resolution_hours": avg_res,
            "focus_score": focus_score,
        })

    results.sort(key=lambda r: r["focus_score"], reverse=True)
    return results


@router.get("/reports/resolution-time-trend")
def resolution_time_trend(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Average resolution time per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.updated_at, SADate).label("day"),
        sa_func.avg(
            sa_func.extract("epoch", Ticket.updated_at - Ticket.created_at) / 3600
        ).label("avg_hours")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.status == 'resolved',
        Ticket.updated_at.isnot(None)
    )
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(cast(Ticket.updated_at, SADate)).order_by(cast(Ticket.updated_at, SADate)).all()
    return [{"date": str(r.day), "avg_hours": round(float(r.avg_hours or 0), 1)} for r in results]


@router.get("/reports/first-response-trend")
def first_response_trend(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Average first response time per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.created_at, SADate).label("day"),
        sa_func.avg(
            sa_func.extract("epoch", Ticket.first_response_at - Ticket.created_at) / 3600
        ).label("avg_hours")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.first_response_at.isnot(None)
    )
    query = apply_filters(query, ticket_type, start_date, end_date)
    results = query.group_by(cast(Ticket.created_at, SADate)).order_by(cast(Ticket.created_at, SADate)).all()
    return [{"date": str(r.day), "avg_hours": round(float(r.avg_hours or 0), 1)} for r in results]


@router.get("/reports/tickets-aging")
def tickets_aging(
    ticket_type: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Open tickets bucketed by age: <1d, 1-3d, 3-7d, 7-30d, >30d."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    now = datetime.utcnow()
    open_statuses = ["open", "in_progress", "pending_approval"]
    query = db.query(Ticket).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.status.in_(open_statuses)
    )
    if ticket_type:
        try: query = query.filter(Ticket.ticket_type == ticket_type.lower())
        except ValueError: pass
    tickets = query.all()
    buckets = {"<1 day": 0, "1-3 days": 0, "3-7 days": 0, "7-30 days": 0, ">30 days": 0}
    for t in tickets:
        if not t.created_at: continue
        age = (now - t.created_at).days
        if age < 1:   buckets["<1 day"] += 1
        elif age < 3:  buckets["1-3 days"] += 1
        elif age < 7:  buckets["3-7 days"] += 1
        elif age < 30: buckets["7-30 days"] += 1
        else:          buckets[">30 days"] += 1
    return [{"bucket": k, "count": v} for k, v in buckets.items()]


@router.get("/reports/csat-trend")
def csat_trend(
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """CSAT average score per day over the selected period."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    from sqlalchemy import cast, Date as SADate
    query = db.query(
        cast(Ticket.updated_at, SADate).label("day"),
        sa_func.avg(Ticket.csat_rating).label("avg_rating"),
        sa_func.count(Ticket.id).label("count")
    ).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    )
    if start_date:
        query = query.filter(Ticket.updated_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(Ticket.updated_at <= datetime.combine(end_date, datetime.max.time()))
    results = query.group_by(cast(Ticket.updated_at, SADate)).order_by(cast(Ticket.updated_at, SADate)).all()
    return [{"date": str(r.day), "avg_rating": round(float(r.avg_rating or 0), 2), "count": r.count} for r in results]


@router.get("/reports/kb-analytics")
def kb_analytics(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """KB article analytics — views, feedback, by category."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    articles = db.query(KBArticle).filter(
        KBArticle.tenant_id == _eff_tid,
        KBArticle.status == "published"
    ).all()
    total_views = sum(a.view_count or 0 for a in articles)
    total_helpful = sum(a.helpful_count or 0 for a in articles)
    total_not_helpful = sum(a.not_helpful_count or 0 for a in articles)
    by_category = {}
    for a in articles:
        cat = a.category or "General"
        if cat not in by_category:
            by_category[cat] = {"articles": 0, "views": 0}
        by_category[cat]["articles"] += 1
        by_category[cat]["views"] += a.view_count or 0
    most_viewed = sorted(articles, key=lambda a: a.view_count or 0, reverse=True)[:10]
    return {
        "total_articles": len(articles),
        "total_views": total_views,
        "total_helpful": total_helpful,
        "total_not_helpful": total_not_helpful,
        "satisfaction_rate": round(total_helpful / max(total_helpful + total_not_helpful, 1) * 100, 1),
        "by_category": [{"category": k, **v} for k, v in by_category.items()],
        "most_viewed": [{"id": a.id, "title": a.title, "category": a.category, "views": a.view_count or 0, "helpful": a.helpful_count or 0} for a in most_viewed],
    }


@router.get("/reports/asset-summary")
def asset_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    """Asset report — by type, status, expiry alerts."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    assets = db.query(Asset).filter(Asset.tenant_id == _eff_tid).all()
    today = date.today()
    by_type = {}
    by_status = {}
    for a in assets:
        t = str(a.type) if a.type else "other"
        s = str(a.status) if a.status else "unknown"
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
    expiring_30 = [a for a in assets if a.expiry_date and 0 <= (a.expiry_date - today).days <= 30]
    return {
        "total": len(assets),
        "by_type": [{"type": k, "count": v} for k, v in by_type.items()],
        "by_status": [{"status": k, "count": v} for k, v in by_status.items()],
        "expiring_30_days": len(expiring_30),
        "expiring_soon": [{"id": a.id, "name": a.name, "expiry_date": str(a.expiry_date)} for a in expiring_30[:10]],
        "total_cost": round(sum(a.purchase_cost or 0 for a in assets), 2),
    }


@router.get("/reports/export/excel")
def export_excel(
    ticket_type: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export tickets as Excel file."""
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["ID", "Type", "Title", "Category", "Priority", "Status", "Requester", "Assigned To", "Created", "SLA Deadline", "Resolution Time (hrs)", "Attachments"]
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    query = db.query(Ticket).filter(Ticket.tenant_id == current_user.tenant_id)
    query = apply_filters(query, ticket_type, start_date, end_date)
    tickets = query.order_by(Ticket.id).all()
    req_ids = {t.requester_id for t in tickets if t.requester_id}
    asgn_ids = {t.assigned_to_id for t in tickets if t.assigned_to_id}
    all_ids = req_ids | asgn_ids
    user_map = {u.id: u.full_name for u in db.query(User).filter(User.id.in_(all_ids)).all()} if all_ids else {}
    # Load attachments for all tickets
    ticket_ids = [t.id for t in tickets]
    attachments_map = {}
    if ticket_ids:
        all_attachments = db.query(Attachment).filter(Attachment.ticket_id.in_(ticket_ids)).all()
        for att in all_attachments:
            url = att.url or att.file_path or ""
            if url:
                attachments_map.setdefault(att.ticket_id, []).append(url)
    for row, t in enumerate(tickets, 2):
        prefix = {"incident": "INC", "service_request": "REQ", "change": "CHG"}.get(str(t.ticket_type) if t.ticket_type else "incident", "INC")
        res_hours = ""
        if t.status == "resolved" and t.updated_at and t.created_at:
            res_hours = round((t.updated_at - t.created_at).total_seconds() / 3600, 1)
        att_urls = " | ".join(attachments_map.get(t.id, []))
        ws.append([
            f"{prefix}{t.id:06d}",
            str(t.ticket_type) if t.ticket_type else "",
            t.title or "",
            t.category or "",
            str(t.priority) if t.priority else "",
            str(t.status) if t.status else "",
            user_map.get(t.requester_id, ""),
            user_map.get(t.assigned_to_id, ""),
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
            str(t.sla_resolution_deadline.date()) if t.sla_resolution_deadline else "",
            res_hours,
            att_urls,
        ])
    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dodesk_tickets.xlsx"}
    )



@router.get("/reports/csat")
def csat_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    client_tenant_id: int | None = Query(None)
):
    if not has_permission(current_user, Permission.VIEW_REPORTS):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    _eff_tid = client_tenant_id if (client_tenant_id and str(current_user.role) in ("super_admin","platform_admin")) else current_user.tenant_id
    results = db.query(Ticket.csat_rating, sa_func.count(Ticket.id)).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    ).group_by(Ticket.csat_rating).all()
    distribution = {str(k): v for k, v in results}
    avg = db.query(sa_func.avg(Ticket.csat_rating)).filter(
        Ticket.tenant_id == _eff_tid,
        Ticket.csat_rating.isnot(None)
    ).scalar()
    count = sum(distribution.values())
    positive = sum(v for k, v in distribution.items() if int(float(k)) >= 4)
    negative = sum(v for k, v in distribution.items() if int(float(k)) <= 2)
    satisfaction_rate = round(positive / count * 100, 1) if count > 0 else 0
    return {
        "average": round(avg, 2) if avg else None,
        "avg_rating": round(avg, 2) if avg else None,
        "count": count,
        "total_responses": count,
        "distribution": distribution,
        "satisfaction_rate": satisfaction_rate,
        "negative_count": negative,
        "positive_count": positive,
    }

